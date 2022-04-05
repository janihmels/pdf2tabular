from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, colors, NamedStyle
from datetime import date
from pandas_utils.pandas_cursor import pandas_cursor


def summary(database, df, filename):
    mycursor = pandas_cursor(df=df)

#Build workbook
    wb = Workbook()
    ws1 = wb.active
    ws1.title = '1.1. Summary By Year'
    ws1_1 = wb.create_sheet(title = '1.2. Summary By Half')
    ws1_2 = wb.create_sheet(title = '1.3. Summary By Quarter')
    ws2 = wb.create_sheet(title = '2.1. Top Earners By Year')
    ws2_1 = wb.create_sheet(title = '2.2. Top Earners By Half')
    ws2_2 = wb.create_sheet(title = '2.3. Top Earners By Quarter')
    ws3 = wb.create_sheet(title = '3.1. By Type By Year')
    ws3_1 = wb.create_sheet(title='3.2. By Type By Half')
    ws3_2 = wb.create_sheet(title='3.3. By Type By Quarter')
    ws3_3 = wb.create_sheet(title='3.4. Top 80% Earners By Type')
    ws4 = wb.create_sheet(title = '4.1. By Source By Year')
    ws4_1 = wb.create_sheet(title='4.2. By Source By Half')
    ws4_2 = wb.create_sheet(title='4.3. By Source By Quarter')
    ws4_3 = wb.create_sheet(title='4.4. Top 80% Earners By Source')

    #Current year
    todays_date = date.today()
    current_year = todays_date.year

    # Find most recent year in data
    find_recent_year = '''SELECT Year_Statement_9LC FROM Master GROUP BY Year_Statement_9LC'''
    mycursor.execute(find_recent_year)
    recent_years = [i[0] for i in mycursor.fetchall()]
    recent_year = recent_years[-1]
    print(recent_year)
    print(current_year)
    if recent_year == current_year:
        find_cut_off_year = current_year - 1
    else:
        find_cut_off_year = int(recent_year) - 1

    #Find cut off
    find_period = 'SELECT DISTINCT Statement_Period_Half_9LC FROM Master WHERE Year_Statement_9LC <= "{}" ORDER BY Statement_Period_Half_9LC'.format(
        find_cut_off_year)
    mycursor.execute(find_period)
    statement_period_half_blank = [i[0] for i in mycursor.fetchall()]
    if len(statement_period_half_list) > 10:
        statement_period_half = statement_period_half_list[-10:]
        cut_off = statement_period_half_list[-10]
    else:
        statement_period_half = statement_period_half_list
        cut_off = statement_period_half_list[0]
    year_list = []
    for period in statement_period_half:
        year_list.append(period[0:4])
    year_list=list(dict.fromkeys(year_list))

#List of statement quarter periods
    find_quarter_period = '''SELECT Year_Statement_9LC, Quarter_Statement_9LC FROM Master 
                             WHERE Quarter_Statement_9LC <> "" AND Statement_Period_Half_9LC >= "{}" GROUP BY Year_Statement_9LC, Quarter_Statement_9LC'''.format(cut_off)
    mycursor.execute(find_quarter_period)
    complete_list = [i for i in mycursor.fetchall()]
    statement_period_quarter_list = [i[0] + ' ' + i[1] for i in complete_list]
    year_statement_list = [i[0] for i in complete_list]
    quarter_statement_list = [i[1] for i in complete_list]
    print(statement_period_quarter_list)
    if len(statement_period_quarter_list) > 24:
        statement_period_quarter = statement_period_quarter_list[-24:]
        quarter_cut_off = statement_period_quarter_list[-24]
    else:
        statement_period_quarter = statement_period_quarter_list
        quarter_cut_off = statement_period_quarter_list[0]
    find_year_statement_list = '''SELECT Year_Statement_9LC FROM Master WHERE Statement_Period_Half_9LC >= "{}" 
                                  GROUP BY Year_Statement_9LC'''.format(cut_off)
    mycursor.execute(find_year_statement_list)
    #year_statement_list = [i[0] for i in mycursor.fetchall()]
    find_quarter_statement_list = '''SELECT Quarter_Statement_9LC FROM Master WHERE Quarter_Statement_9LC <> ""
                                     GROUP BY Quarter_Statement_9LC'''
    mycursor.execute(find_quarter_statement_list)
    #quarter_statement_list = [i[0] for i in mycursor.fetchall()]
    print(year_statement_list)
    print(quarter_statement_list)

#Find H1/H2 for last year
    penultimate_half_period = statement_period_half[-2]
    last_half_period = statement_period_half[-1]
    if penultimate_half_period[0:4] == last_half_period[0:4]:
        final_half = '(H2)'
    else:
        final_half = '(H1)'

#Find total royalty amount in period
    find_total = '''SELECT Normalized_Income_Type_9LC, sum( CASE WHEN Statement_Period_Half_9LC >= '{}'
                    THEN Adjusted_Royalty_SB ELSE NULL END) AS `Total`
                    FROM Master GROUP BY Normalized_Income_Type_9LC'''.format(cut_off)
    mycursor.execute(find_total)
    total_list = [i[1] for i in mycursor.fetchall()]
    royalty_total = 0
    for a in total_list:
        royalty_total += a
    eighty_percent = royalty_total*0.8

#Find list of songs
    find_songs = '''SELECT Song_Name_9LC,
                    sum( CASE WHEN Statement_Period_Half_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `Total`
                    FROM Master WHERE Song_Name_9LC <> "Pool Revenue"
                    GROUP BY Song_Name_9LC ORDER BY `TOTAL` DESC'''.format(cut_off)
    mycursor.execute(find_songs)
    song_list = [i[0] for i in mycursor.fetchall()]

#Find list of PROs - CHANGE TO WHERE Is_PRO_SB = 1 LATER
    find_PRO = '''SELECT Third_Party_9LC FROM Master WHERE Third_Party_9LC IN ('ASCAP', 'BMI', 'PRS', 'MCPS', 
                         'SoundExchange', 'SOCAN', 'APRA', 'AMCOS', 'SESAC', 'AMRA', 'GEMA', 'SUISA', 'ZAiKS') GROUP BY Third_Party_9LC'''
    mycursor.execute(find_PRO)
    PRO_list = [i[0] for i in mycursor.fetchall()]
    PRO_string = ', '.join('"{}"'.format(str(x)) for x in PRO_list)
    print(PRO_list)

#Find list of publishers - CHANGE TO WHERE Is_PRO_SB = 0 LATER
    find_publishers = '''SELECT Third_Party_9LC FROM Master WHERE Third_Party_9LC NOT IN ('ASCAP', 'BMI', 'PRS', 'MCPS', 
                         'SoundExchange', 'SOCAN', 'APRA', 'AMCOS', 'SESAC', 'AMRA', 'GEMA', 'SUISA', 'ZAiKS') GROUP BY Third_Party_9LC'''
    mycursor.execute(find_publishers)
    publisher_list = [i[0] for i in mycursor.fetchall()]
    publisher_string = ', '.join('"{}"'.format(str(x)) for x in publisher_list)

#Find third party list
    third_party_list = publisher_list + PRO_list
    print(third_party_list)

#Find where quarterly data exists
    find_quarterly_data_PRO = '''SELECT DISTINCT Third_Party_9LC FROM Master 
                                     WHERE Quarter_Statement_9LC <> "" 
                                     AND Third_Party_9LC IN ({})
                                     GROUP BY Third_Party_9LC'''.format(PRO_string)
    mycursor.execute(find_quarterly_data_PRO)
    quarterly_PRO_list = [i[0] for i in mycursor.fetchall()]
    if len(quarterly_PRO_list) != 0:
        quarterly_PRO_string = ', '.join('"{}"'.format(str(x)) for x in quarterly_PRO_list)
    else:
        quarterly_PRO_string = "''"
    print(quarterly_PRO_list)


    find_quarterly_data_publisher = '''SELECT DISTINCT Third_Party_9LC FROM Master
                                           WHERE Quarter_Statement_9LC <> ""
                                           AND Third_Party_9LC IN ({})
                                           GROUP BY Third_Party_9LC'''.format(publisher_string)
    mycursor.execute(find_quarterly_data_publisher)
    quarterly_publisher_list = [i[0] for i in mycursor.fetchall()]
    if len(quarterly_publisher_list) != 0:
        quarterly_publisher_string = ', '.join('"{}"'.format(str(x)) for x in quarterly_publisher_list)
    else:
        quarterly_publisher_string = "''"
    print(quarterly_publisher_list)
    print(quarterly_publisher_string)
    if quarterly_publisher_string == '' or quarterly_PRO_string == '':
        quarterly_third_party = quarterly_publisher_string + quarterly_PRO_string
    else:
        quarterly_third_party = quarterly_publisher_string + ',' + quarterly_PRO_string
    print(quarterly_third_party)

#Find where quarterly data does not exist
    not_quarterly_PRO_list = [x for x in PRO_list if x not in quarterly_PRO_list]
    print(not_quarterly_PRO_list)

    not_quarterly_publisher_list = [x for x in publisher_list if x not in quarterly_publisher_list]
    print(not_quarterly_publisher_list)

    not_quarterly_third_party = not_quarterly_publisher_list + not_quarterly_PRO_list
    not_quarterly_third_party_string = ','.join('"{}"'.format(str(x)) for x in not_quarterly_third_party)
    print(not_quarterly_third_party_string)

    #PRO_overlap
    if len(quarterly_publisher_list) == len(publisher_list):
        print('yes')
    if len(quarterly_publisher_list) != len(publisher_list):
        print('no')



#Column letters
    column_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S',
                      'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ',
                      'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ']

#Create styles
    #Main numbers
    number_style = NamedStyle(name="number_style")
    number_style.alignment = Alignment(horizontal="right", vertical="center")
    number_style.font = Font(name="Calibri", size="11")
    number_style.number_format = '#,##0.00'

    #Title cells
    title_style = NamedStyle(name="title_style")
    title_style.alignment = Alignment(horizontal="center", vertical="center")
    title_style.font = Font(name="Calibri", size="11", bold=True)
    title_style.fill = PatternFill("solid", fgColor="A6ACAF")

    #Header cells
    header_style = NamedStyle(name="header_style")
    header_style.alignment = Alignment(horizontal="center", vertical="center")
    header_style.font = Font(name="Calibri", size="11", bold=True)
    header_style.fill = PatternFill("solid", fgColor="C9CDCF")

    #Sub header cells
    sub_header_style = NamedStyle(name="sub_header_style")
    sub_header_style.alignment = Alignment(horizontal="left", vertical="center")
    sub_header_style.font = Font(name="Calibri", size="11", bold=True, italic=True)

    #Name cells
    name_style = NamedStyle(name="name_style")
    name_style.alignment = Alignment(horizontal="left", vertical="center")
    name_style.font = Font(name="Calibri", size="11")

    #Total cells
    total_style = NamedStyle(name="total_style")
    total_style.alignment = Alignment(horizontal="right", vertical="center")
    total_style.font = Font(name="Calibri", size="11", bold=True)
    total_style.number_format = '#,##0.00'

    #Lined total row
    thin = Side(border_style="thin", color="000000")
    lined_total_style = NamedStyle(name="lined_total_style")
    lined_total_style.alignment = Alignment(horizontal="right", vertical="center")
    lined_total_style.font = Font(name="Calibri", size="11", bold=True)
    lined_total_style.number_format = '#,##0.00'
    lined_total_style.border = Border(top=thin, left=None, right=None, bottom=thin)

    #Total label style
    total_label_style = NamedStyle(name="total_label_style")
    total_label_style.alignment = Alignment(horizontal="center", vertical="center")
    total_label_style.font = Font(name="Calibri", size="11", bold=True)
    total_label_style.fill = PatternFill("solid", fgColor="C9CDCF")
    total_label_style.border = Border(top=thin, left=None, right=None, bottom=thin)

    #Publisher/PRO label style
    thick = Side(border_style="medium", color="000000")
    publisher_label_style = NamedStyle(name="publisher_label_style")
    publisher_label_style.alignment = Alignment(horizontal="center", vertical="center")
    publisher_label_style.font = Font(name="Calibri", size="11", bold=True, italic=True)
    publisher_label_style.border = Border(bottom=thick)

    #Not available style
    not_available_style = NamedStyle(name="not_available_style")
    not_available_style.alignment = Alignment(horizontal="center", vertical="center")
    not_available_style.font = Font(name="Calibri", size="11")

    #Add styles to workbook
    wb.add_named_style(number_style)
    wb.add_named_style(header_style)
    wb.add_named_style(title_style)
    wb.add_named_style(name_style)
    wb.add_named_style(total_style)
    wb.add_named_style(sub_header_style)
    wb.add_named_style(lined_total_style)
    wb.add_named_style(total_label_style)
    wb.add_named_style(publisher_label_style)
    wb.add_named_style(not_available_style)

#1. Executive summary by year
    #Build publisher table
    select_table_1_1_1 = '''SELECT Third_Party_9LC,'''
    select_table_1_1_2 = ""
    for j in year_list:
        select_table_1_1_2 += ('''sum( CASE WHEN Statement_Period_Half_9LC LIKE "{} H%" 
                                  THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(j,j))
    select_table_1_1_3 = '''sum( CASE WHEN Statement_Period_Half_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END) 
                            AS `Total` FROM Master WHERE Third_Party_9LC NOT IN ('ASCAP', 'BMI', 'PRS', 'MCPS', 
                            'SoundExchange', 'SOCAN', 'APRA', 'AMCOS', 'SESAC', 'AMRA')
                            GROUP BY Third_Party_9LC ORDER BY `Total` DESC'''.format(cut_off)
    select_table_1_1 = select_table_1_1_1 + select_table_1_1_2 + select_table_1_1_3
    mycursor.execute(select_table_1_1)
    table_1_1 = mycursor.fetchall()

    #Build PRO table
    select_table_1_2_1 = '''SELECT Third_Party_9LC,'''
    select_table_1_2_2 = ""
    for k in year_list:
        select_table_1_2_2 += ('''sum( CASE WHEN Statement_Period_Half_9LC LIKE "{} H%"
                                  THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(k,k))
    select_table_1_2_3 = '''sum( CASE WHEN Statement_Period_Half_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END)
                            AS `Total` FROM Master WHERE Third_Party_9LC IN ('ASCAP', 'BMI', 'PRS', 'MCPS', 
                            'SoundExchange', 'SOCAN', 'APRA', 'AMCOS', 'SESAC', 'AMRA')
                            GROUP BY Third_Party_9LC ORDER BY `Total` DESC'''.format(cut_off)
    select_table_1_2 = select_table_1_2_1 + select_table_1_2_2 + select_table_1_2_3
    mycursor.execute(select_table_1_2)
    table_1_2 = mycursor.fetchall()

    #Publisher and PRO labels
    ws1.merge_cells('A4:{}4'.format(column_letters[len(year_list) + 1]))
    ws1.cell(row=4, column=1).value = 'Publishers'
    for w in range(1, len(year_list) + 3):
        ws1.cell(row=4, column=w).style = 'publisher_label_style'

    ws1.merge_cells('A{}:{}{}'.format(len(table_1_1)+7, column_letters[len(year_list) + 1],
                                      len(table_1_1)+7))
    ws1.cell(row=len(table_1_1)+7, column=1).value = 'PROs'
    for x in range(1, len(year_list) + 3):
        ws1.cell(row=len(table_1_1)+7, column=x).style = 'publisher_label_style'

    #Column name list
    column_names_1 = ['Third Party']
    for d in year_list:
        column_names_1.append(d)
    column_names_1.append('Total')

    #Add title row to sheet
    ws1.merge_cells('B1:{}1'.format(column_letters[len(year_list) + 1]))
    ws1.cell(row=1, column=1).value = '{}'.format(database[:-25])
    ws1.cell(row=1, column=1).style = 'title_style'
    ws1.cell(row=1, column=2).value = 'Summary'
    ws1.cell(row=1, column=2).style = 'title_style'

    #Add column names to sheet
    column_name_row = ws1[2]
    for cell, e in zip(column_name_row, range(len(column_names_1))):
        cell.value = column_names_1[e]
        cell.style = 'header_style'

    #Specify H1/H2 for most recent year
    ws1.cell(row=3, column=len(year_list) + 1).value = final_half
    for col in range(1, len(column_names_1) + 1):
        ws1.cell(row=3, column=col).style = 'header_style'

    #Add publisher table to sheet
    publisher_rows = ws1[5:len(table_1_1)+5]
    #ws1.cell(row=4, column=1).value = 'Publishers'
    #ws1.cell(row=4, column=1).style = 'sub_header_style'
    for row, f in zip(publisher_rows, table_1_1):
        for cell, g in zip(row, range(len(f))):
            cell.value = f[g]
            cell.style = 'number_style'
    for row in range(5, len(table_1_1) + 5):
        ws1.cell(row=row, column=1).style = 'name_style'
        ws1.cell(row=row, column=len(year_list)+2).style = 'total_style'

    #Add total row
    for n, o in zip(range(2, len(year_list)+3), column_letters[1:]):
        ws1.cell(row=len(table_1_1)+5, column=n).value = "=SUM({}{}:{}{})".format(o,5,o,len(table_1_1)+4)
        ws1.cell(row=len(table_1_1)+5, column=n).style = 'lined_total_style'
    ws1.cell(row=len(table_1_1)+5, column=1).value = 'Subtotal'
    ws1.cell(row=len(table_1_1)+5, column=1).style = 'total_label_style'

    #Add PRO table to sheet
    PRO_rows = ws1[len(table_1_1)+8:len(table_1_1)+len(table_1_2)+8]
    #ws1.cell(row=len(table_1_1)+6, column=1).value = 'PROs'
    #ws1.cell(row=len(table_1_1)+6, column=1).style = 'sub_header_style'
    for row, p in zip(PRO_rows, table_1_2):
        for cell, q in zip(row, range(len(p))):
            cell.value = p[q]
            cell.style = 'number_style'
    for row in range(len(table_1_1)+8, len(table_1_1)+len(table_1_2)+8):
        ws1.cell(row=row, column=1).style = 'name_style'
        ws1.cell(row=row, column=len(year_list)+2).style = 'total_style'

    #Add total row
    for r, s in zip(range(2, len(year_list)+3), column_letters[1:]):
        ws1.cell(row=len(table_1_1)+len(table_1_2)+8, column=r).value = "=SUM({}{}:{}{})".format(s,len(table_1_1)+8,s,
                                                                                    len(table_1_1)+len(table_1_2)+7)
        ws1.cell(row=len(table_1_1) + len(table_1_2) + 8, column=r).style = 'lined_total_style'
    ws1.cell(row=len(table_1_1)+len(table_1_2)+8, column=1).value = 'Subtotal'
    ws1.cell(row=len(table_1_1)+len(table_1_2)+8, column=1).style = 'total_label_style'

    #Add grand total row
    for t, u in zip(range(2, len(year_list)+3), column_letters[1:]):
        ws1.cell(row=len(table_1_1)+len(table_1_2)+10, column=t).value = "={}{}+{}{}".format(u,len(table_1_1)+5,u,
                                                                                       len(table_1_1)+len(table_1_2)+8)
        ws1.cell(row=len(table_1_1)+len(table_1_2)+10, column=t).style = 'lined_total_style'
    ws1.cell(row=len(table_1_1)+len(table_1_2)+10, column=1).value = 'Total'
    ws1.cell(row=len(table_1_1)+len(table_1_2)+10, column=1).style = 'total_label_style'

#1.1. Executive summary by half
    select_table_1_1_1_1 = '''SELECT Third_Party_9LC,'''
    select_table_1_1_1_2 = ""
    for j in statement_period_half:
        select_table_1_1_1_2 += ('''sum( CASE WHEN Statement_Period_Half_9LC = "{}" 
                                      THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(j, j))
    select_table_1_1_1_3 = '''sum( CASE WHEN Statement_Period_Half_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END) 
                                AS `Total` FROM Master WHERE Third_Party_9LC NOT IN ('ASCAP', 'BMI', 'PRS', 'MCPS', 
                                'SoundExchange', 'SOCAN', 'APRA', 'AMCOS', 'SESAC', 'AMRA')
                                GROUP BY Third_Party_9LC ORDER BY `Total` DESC'''.format(cut_off)
    select_table_1_1_1 = select_table_1_1_1_1 + select_table_1_1_1_2 + select_table_1_1_1_3
    mycursor.execute(select_table_1_1_1)
    table_1_1_1 = mycursor.fetchall()

    #Build PRO table
    select_table_1_1_2_1 = '''SELECT Third_Party_9LC,'''
    select_table_1_1_2_2 = ""
    for k in statement_period_half:
        select_table_1_1_2_2 += ('''sum( CASE WHEN Statement_Period_Half_9LC = "{}"
                                      THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(k, k))
    select_table_1_1_2_3 = '''sum( CASE WHEN Statement_Period_Half_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END)
                                AS `Total` FROM Master WHERE Third_Party_9LC IN ('ASCAP', 'BMI', 'PRS', 'MCPS', 
                                'SoundExchange', 'SOCAN', 'APRA', 'AMCOS', 'SESAC', 'AMRA')
                                GROUP BY Third_Party_9LC ORDER BY `Total` DESC'''.format(cut_off)
    select_table_1_1_2 = select_table_1_1_2_1 + select_table_1_1_2_2 + select_table_1_1_2_3
    mycursor.execute(select_table_1_1_2)
    table_1_1_2 = mycursor.fetchall()


    #Publisher and PRO labels
    ws1_1.merge_cells('A4:{}4'.format(column_letters[len(statement_period_half) + 1]))
    ws1_1.cell(row=4, column=1).value = 'Publishers'
    for w in range(1, len(statement_period_half) + 3):
        ws1_1.cell(row=4, column=w).style = 'publisher_label_style'

    ws1_1.merge_cells('A{}:{}{}'.format(len(table_1_1) + 7, column_letters[len(statement_period_half) + 1],
                                      len(table_1_1) + 7))
    ws1_1.cell(row=len(table_1_1) + 7, column=1).value = 'PROs'
    for x in range(1, len(statement_period_half) + 3):
        ws1_1.cell(row=len(table_1_1) + 7, column=x).style = 'publisher_label_style'

    #Column name list
    column_names_1_1 = ['Third Party']
    for d in statement_period_half:
        column_names_1_1.append(d)
    column_names_1_1.append('Total')

    #Add title row to sheet
    ws1_1.merge_cells('B1:{}1'.format(column_letters[len(statement_period_half) + 1]))
    ws1_1.cell(row=1, column=1).value = '{}'.format(database[:-25])
    ws1_1.cell(row=1, column=1).style = 'title_style'
    ws1_1.cell(row=1, column=2).value = 'Summary'
    ws1_1.cell(row=1, column=2).style = 'title_style'

    #Add column names to sheet
    column_name_row_1 = ws1_1[2]
    for cell, e in zip(column_name_row_1, range(len(column_names_1_1))):
        cell.value = column_names_1_1[e]
        cell.style = 'header_style'

    #Add publisher table to sheet
    publisher_rows = ws1_1[5:len(table_1_1_1) + 5]
    # ws1.cell(row=4, column=1).value = 'Publishers'
    # ws1.cell(row=4, column=1).style = 'sub_header_style'
    for row, f in zip(publisher_rows, table_1_1_1):
        for cell, g in zip(row, range(len(f))):
            cell.value = f[g]
            cell.style = 'number_style'
    for row in range(5, len(table_1_1_1) + 5):
        ws1_1.cell(row=row, column=1).style = 'name_style'
        ws1_1.cell(row=row, column=len(statement_period_half) + 2).style = 'total_style'

    #Add total row
    for n, o in zip(range(2, len(statement_period_half) + 3), column_letters[1:]):
        ws1_1.cell(row=len(table_1_1_1) + 5, column=n).value = "=SUM({}{}:{}{})".format(o, 5, o, len(table_1_1_1) + 4)
        ws1_1.cell(row=len(table_1_1_1) + 5, column=n).style = 'lined_total_style'
    ws1_1.cell(row=len(table_1_1_1) + 5, column=1).value = 'Subtotal'
    ws1_1.cell(row=len(table_1_1_1) + 5, column=1).style = 'total_label_style'

    #Add PRO table to sheet
    PRO_rows = ws1_1[len(table_1_1_1) + 8:len(table_1_1_1) + len(table_1_1_2) + 8]
    # ws1.cell(row=len(table_1_1)+6, column=1).value = 'PROs'
    # ws1.cell(row=len(table_1_1)+6, column=1).style = 'sub_header_style'
    for row, p in zip(PRO_rows, table_1_1_2):
        for cell, q in zip(row, range(len(p))):
            cell.value = p[q]
            cell.style = 'number_style'
    for row in range(len(table_1_1_1) + 8, len(table_1_1_1) + len(table_1_1_2) + 8):
        ws1_1.cell(row=row, column=1).style = 'name_style'
        ws1_1.cell(row=row, column=len(statement_period_half) + 2).style = 'total_style'

    #Add total row
    for r, s in zip(range(2, len(statement_period_half) + 3), column_letters[1:]):
        ws1_1.cell(row=len(table_1_1_1) + len(table_1_1_2) + 8, column=r).value = "=SUM({}{}:{}{})".format(s,
                                                                                            len(table_1_1_1) + 8, s,
                                                                                len(table_1_1_1) + len(table_1_1_2) + 7)
        ws1_1.cell(row=len(table_1_1_1) + len(table_1_1_2) + 8, column=r).style = 'lined_total_style'
    ws1_1.cell(row=len(table_1_1_1) + len(table_1_1_2) + 8, column=1).value = 'Subtotal'
    ws1_1.cell(row=len(table_1_1_1) + len(table_1_1_2) + 8, column=1).style = 'total_label_style'

    #Add grand total row
    for t, u in zip(range(2, len(statement_period_half) + 3), column_letters[1:]):
        ws1_1.cell(row=len(table_1_1_1) + len(table_1_1_2) + 10, column=t).value = "={}{}+{}{}".format(u,
                                                                                            len(table_1_1_1) + 5, u,
                                                                                len(table_1_1_1) + len(table_1_1_2) + 8)
        ws1_1.cell(row=len(table_1_1_1) + len(table_1_1_2) + 10, column=t).style = 'lined_total_style'
    ws1_1.cell(row=len(table_1_1_1) + len(table_1_1_2) + 10, column=1).value = 'Total'
    ws1_1.cell(row=len(table_1_1_1) + len(table_1_1_2) + 10, column=1).style = 'total_label_style'

    #Third row
    for col in range(1, len(column_names_1_1) + 1):
        ws1_1.cell(row=3, column=col).style = 'header_style'

#1.2.Executive Summary By Quarter
    #Build publisher table
    select_table_1_2_1_1 = '''SELECT Third_Party_9LC,'''
    select_table_1_2_1_2 = ""
    for i,j in zip(year_statement_list, quarter_statement_list):
        #for j in quarter_statement_list:
        select_table_1_2_1_2 += ('''sum( CASE WHEN Year_Statement_9LC = "{}" AND Quarter_Statement_9LC = "{}"
                                            THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(i, j, i, j))
    select_table_1_2_1_3 = '''sum( CASE WHEN Statement_Period_Half_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END) 
                                    AS `Total` FROM Master WHERE Third_Party_9LC IN ({})
                                    GROUP BY Third_Party_9LC ORDER BY `Total` DESC'''.format(cut_off, quarterly_publisher_string)
    select_table_1_2_1 = select_table_1_2_1_1 + select_table_1_2_1_2 + select_table_1_2_1_3
    mycursor.execute(select_table_1_2_1)
    table_1_2_1 = mycursor.fetchall()


    #Build PRO table
    select_table_1_2_2_1 = '''SELECT Third_Party_9LC,'''
    select_table_1_2_2_2 = ""
    for i,k in zip(year_statement_list, quarter_statement_list):
        select_table_1_2_2_2 += ('''sum( CASE WHEN Year_Statement_9LC = "{}" AND Quarter_Statement_9LC = "{}"
                                         THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(i, k, i, k))
    select_table_1_2_2_3 = '''sum( CASE WHEN Statement_Period_Half_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END)
                                    AS `Total` FROM Master WHERE Third_Party_9LC IN ({})
                                    GROUP BY Third_Party_9LC ORDER BY `Total` DESC'''.format(cut_off, quarterly_PRO_string)
    select_table_1_2_2 = select_table_1_2_2_1 + select_table_1_2_2_2 + select_table_1_2_2_3
    mycursor.execute(select_table_1_2_2)
    table_1_2_2 = mycursor.fetchall()


    #Publisher and PRO labels
    ws1_2.merge_cells('A4:{}4'.format(column_letters[len(statement_period_quarter) + 1]))
    ws1_2.cell(row=4, column=1).value = 'Publishers'
    for w in range(1, len(statement_period_quarter) + 3):
        ws1_2.cell(row=4, column=w).style = 'publisher_label_style'

    ws1_2.merge_cells('A{}:{}{}'.format(len(table_1_2_1) + 7, column_letters[len(statement_period_quarter) + 1],
                                        len(table_1_2_1) + 7))
    ws1_2.cell(row=len(table_1_2_1) + 7, column=1).value = 'PROs'
    for x in range(1, len(statement_period_quarter) + 3):
        ws1_2.cell(row=len(table_1_2_1) + 7, column=x).style = 'publisher_label_style'

    #Column name list
    column_names_1_2 = ['Third Party']
    for d in statement_period_quarter:
        column_names_1_2.append(d)
    column_names_1_2.append('Total')

    #Add title row to sheet
    ws1_2.merge_cells('B1:{}1'.format(column_letters[len(statement_period_quarter) + 1]))
    ws1_2.cell(row=1, column=1).value = '{}'.format(database[:-25])
    ws1_2.cell(row=1, column=1).style = 'title_style'
    ws1_2.cell(row=1, column=2).value = 'Summary'
    ws1_2.cell(row=1, column=2).style = 'title_style'

    #Add column names to sheet
    column_name_row_1_2 = ws1_2[2]
    for cell, e in zip(column_name_row_1_2, range(len(column_names_1_2))):
        cell.value = column_names_1_2[e]
        cell.style = 'header_style'

    #Add publisher table to sheet
    first_sub_total_row_1_2 = 0
    if len(quarterly_publisher_list) == len(publisher_list):
        publisher_rows = ws1_2[5:len(table_1_2_1) + 5]
        #ws1.cell(row=4, column=1).value = 'Publishers'
        #ws1.cell(row=4, column=1).style = 'sub_header_style'
        for row, f in zip(publisher_rows, table_1_2_1):
            for cell, g in zip(row, range(len(f))):
                cell.value = f[g]
                cell.style = 'number_style'
        for row in range(5, len(table_1_2_1) + 5):
            ws1_2.cell(row=row, column=1).style = 'name_style'
            ws1_2.cell(row=row, column=len(statement_period_quarter) + 2).style = 'total_style'

    #Add total row
        for n, o in zip(range(2, len(statement_period_quarter) + 3), column_letters[1:]):
            if len(table_1_2_1) > 0:
                ws1_2.cell(row=len(table_1_2_1) + 5, column=n).value = "=SUM({}{}:{}{})".format(o, 5,
                                                                                                o, len(table_1_2_1) + 4)
                ws1_2.cell(row=len(table_1_2_1) + 5, column=n).style = 'lined_total_style'
            else:
                ws1_2.cell(row=len(table_1_2_1) + 5, column=n).value = '0.00'
                ws1_2.cell(row=len(table_1_2_1) + 5, column=n).style = 'lined_total_style'
        ws1_2.cell(row=len(table_1_2_1) + 5, column=1).value = 'Subtotal'
        ws1_2.cell(row=len(table_1_2_1) + 5, column=1).style = 'total_label_style'
        first_sub_total_row_1_2 = (len(table_1_2_1) + 5)

    else:
        ws1_2.merge_cells('A5:{}5'.format(column_letters[len(statement_period_quarter)+1]))
        if len(not_quarterly_publisher_list) == 1:
            ws1_2.cell(row=5, column=1).value = 'Quarterly data is not available for {}'.format(not_quarterly_publisher_list[0])
        if len(not_quarterly_publisher_list) == 2:
            ws1_2.cell(row=5, column=1).value  = 'Quarterly data is not available for {} and {}'.format(not_quarterly_publisher_list[0], not_quarterly_publisher_list[1])
        if len(not_quarterly_publisher_list) > 2:
            not_available_statement = 'Quarterly data is not available for ' + '{}, '*(len(not_quarterly_publisher_list)-1) + 'and {}'
            ws1_2.cell(row=5, column=1).value = not_available_statement.format(*not_quarterly_publisher_list)
        ws1_2.cell(row=5, column=1).style = 'not_available_style'
        if len(quarterly_publisher_list) > 0:
            publisher_rows = ws1_2[6:len(table_1_2_1)+6]
            for row, f in zip(publisher_rows, table_1_2_1):
                for cell, g in zip(row, range(len(f))):
                    cell.value = f[g]
                    cell.style = 'number_style'
            for row in range(6, len(table_1_2_1) + 6):
                ws1_2.cell(row=row, column=1).style = 'name_style'
                ws1_2.cell(row=row, column=len(statement_period_quarter) + 2).style = 'total_style'
            for n, o in zip(range(2, len(statement_period_quarter) + 3), column_letters[1:]):
                ws1_2.cell(row=len(table_1_2_1) + 6, column=n).value = "=SUM({}{}:{}{})".format(o, 6, o,
                                                                                                len(table_1_2_1) + 5)
                ws1_2.cell(row=len(table_1_2_1) + 6, column=n).style = 'lined_total_style'
            ws1_2.cell(row=len(table_1_2_1) + 6, column=1).value = 'Subtotal'
            ws1_2.cell(row=len(table_1_2_1) + 6, column=1).style = 'total_label_style'
            first_sub_total_row_1_2 = (len(table_1_2_1) + 6)



    #Add PRO table to sheet
    second_sub_total_row_1_2 = 0
    if len(quarterly_PRO_list) == len(PRO_list):
        PRO_rows = ws1_2[len(table_1_2_1) + 8:len(table_1_2_1) + len(table_1_2_2) + 8]
        # ws1.cell(row=len(table_1_1)+6, column=1).value = 'PROs'
        # ws1.cell(row=len(table_1_1)+6, column=1).style = 'sub_header_style'
        for row, p in zip(PRO_rows, table_1_2_2):
            for cell, q in zip(row, range(len(p))):
                cell.value = p[q]
                cell.style = 'number_style'
        for row in range(len(table_1_2_1) + 8, len(table_1_2_1) + len(table_1_2_2) + 8):
            ws1_2.cell(row=row, column=1).style = 'name_style'
            ws1_2.cell(row=row, column=len(statement_period_quarter) + 2).style = 'total_style'

        #Add total row
        for r, s in zip(range(2, len(statement_period_quarter) + 3), column_letters[1:]):
            if len(table_1_2_2) > 0:
                ws1_2.cell(row=len(table_1_2_1) + len(table_1_2_2) + 8, column=r).value = "=SUM({}{}:{}{})".format(s,
                                                                                                    len(table_1_2_1) + 8,
                                                                                                               s,
                                                                                    len(table_1_2_1) + len(table_1_2_2) + 7)
                ws1_2.cell(row=len(table_1_2_1) + len(table_1_2_2) + 8, column=r).style = 'lined_total_style'
            else:
                ws1_2.cell(row=len(table_1_2_1) + len(table_1_2_2) + 8, column=r).value = '0.00'
                ws1_2.cell(row=len(table_1_2_1) + len(table_1_2_2) + 8, column=r).style = 'lined_total_style'
        ws1_2.cell(row=len(table_1_2_1) + len(table_1_2_2) + 8, column=1).value = 'Subtotal'
        ws1_2.cell(row=len(table_1_2_1) + len(table_1_2_2) + 8, column=1).style = 'total_label_style'
        second_sub_total_row_1_2 = len(table_1_2_1) + len(table_1_2_2) + 8

    else:
        ws1_2.merge_cells('A{}:{}{}'.format(len(table_1_2_1) + 8,column_letters[len(statement_period_quarter)+1],len(table_1_2_1) + 8))
        if len(not_quarterly_PRO_list) == 1:
            ws1_2.cell(row=len(table_1_2_1)+ 8, column=1).value = 'Quarterly data is not available for {}'.format(not_quarterly_PRO_list[0])
        if len(not_quarterly_PRO_list) == 2:
            ws1_2.cell(row=len(table_1_2_1) + 8, column=1).value  = 'Quarterly data is not available for {} and {}'.format(not_quarterly_PRO_list[0], not_quarterly_PRO_list[1])
        if len(not_quarterly_PRO_list) > 2:
            not_available_statement = 'Quarterly data is not available for' + '{}, '*(len(not_quarterly_PRO_list)-1)
            + 'and {}'
            ws1_2.cell(row=len(table_1_2_1) + 8, column=1).value = not_available_statement.format(not_quarterly_PRO_list)
        ws1_2.cell(row=len(table_1_2_1) + 8, column=1).style = 'not_available_style'
        if len(quarterly_PRO_list) > 0:
            PRO_rows = ws1_2[len(table_1_2_1) + 9:len(table_1_2_1)+ len(table_1_2_2)+9]
            for row, f in zip(PRO_rows, table_1_2_2):
                for cell, g in zip(row, range(len(f))):
                    cell.value = f[g]
                    cell.style = 'number_style'
            for row in range(len(table_1_2_1) + 9, len(table_1_2_1) + len(table_1_2_2) + 9):
                ws1_2.cell(row=row, column=1).style = 'name_style'
                ws1_2.cell(row=row, column=len(statement_period_quarter) + 2).style = 'total_style'
            for n, o in zip(range(2, len(statement_period_quarter) + 3), column_letters[1:]):
                ws1_2.cell(row=len(table_1_2_1) + len(table_1_2_2) + 9, column=n).value = "=SUM({}{}:{}{})".format(o, len(table_1_2_1) + 9, o,
                                                                                                len(table_1_2_1)+ len(table_1_2_2)+8)
                ws1_2.cell(row=len(table_1_2_1) + len(table_1_2_2) + 9, column=n).style = 'lined_total_style'
            ws1_2.cell(row=len(table_1_2_1)+ len(table_1_2_2)+9, column=1).value = 'Subtotal'
            ws1_2.cell(row=len(table_1_2_1)+ len(table_1_2_2)+9, column=1).style = 'total_label_style'
            second_sub_total_row_1_2 = len(table_1_2_1) + len(table_1_2_2) + 9



    #Add grand total row
    for t, u in zip(range(2, len(statement_period_quarter) + 3), column_letters[1:]):
        if first_sub_total_row_1_2 != 0 and second_sub_total_row_1_2 != 0:
            ws1_2.cell(row=second_sub_total_row_1_2 + 2, column=t).value = "={}{}+{}{}".format(u, first_sub_total_row_1_2,
                                                                                                 u,second_sub_total_row_1_2)
        if first_sub_total_row_1_2 == 0 and second_sub_total_row_1_2 != 0:
            ws1_2.cell(row=second_sub_total_row_1_2 + 2, column=t).value = "={}{}".format(u, second_sub_total_row_1_2)
        if first_sub_total_row_1_2 != 0 and second_sub_total_row_1_2 == 0:
            ws1_2.cell(row=second_sub_total_row_1_2 + 2, column=t).value = "={}{}".format(u, first_sub_total_row_1_2)
        ws1_2.cell(row=second_sub_total_row_1_2 + 2, column=t).style = 'lined_total_style'
    ws1_2.cell(row=second_sub_total_row_1_2 + 2, column=1).value = 'Total'
    ws1_2.cell(row=second_sub_total_row_1_2 + 2, column=1).style = 'total_label_style'

    #Third row
    for col in range(1, len(column_names_1_2) + 1):
        ws1_2.cell(row=3, column=col).style = 'header_style'


#2. Top earners by year

    #Build main pivot table (song x rev x half)
    select_table_2_1_1 = '''SELECT Song_Name_9LC,'''
    select_table_2_1_2 = ""
    for b in year_list:
        select_table_2_1_2 += (''' sum( CASE WHEN Statement_Period_Half_9LC LIKE "{} H%" 
                                 AND Song_Name_9LC <> "Pool Revenue" 
                                 THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(b,b))
    select_table_2_1_3 = '''sum( CASE WHEN Song_Name_9LC <> "Pool Revenue" AND Statement_Period_Half_9LC >= "{}"
                          THEN Adjusted_Royalty_SB ELSE "" END) AS `Total`
                          FROM Master WHERE Song_Name_9LC <> "Pool Revenue" AND Statement_Period_Half_9LC <> ""
                          GROUP BY Song_Name_9LC ORDER BY `Total` DESC'''.format(cut_off)
    select_table_2_1 = select_table_2_1_1 + select_table_2_1_2 + select_table_2_1_3
    mycursor.execute(select_table_2_1)
    table_2_1 = mycursor.fetchall()

    #Find eighty percent cut off
    cumulative_total = 0
    song_cut_off_no = 0
    for i in table_2_1:
        if cumulative_total <= eighty_percent:
            cumulative_total += i[len(year_list)+1]
            song_cut_off = i[0]
            song_cut_off_no += 1

    #Pool revenue line
    pool_rev_2_1_1 = '''SELECT Song_Name_9LC,'''
    pool_rev_2_1_2 = ""
    for c in year_list:
        pool_rev_2_1_2 += ('''sum( CASE WHEN Statement_Period_Half_9LC LIKE "{} H%" AND Song_Name_9LC = "Pool Revenue"
                            THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(c,c))
    pool_rev_2_1_3 = '''sum( CASE WHEN Song_Name_9LC = "Pool Revenue" AND Statement_Period_Half_9LC >= "{}"
                           THEN Adjusted_Royalty_SB ELSE "" END) AS `Total` 
                           FROM Master WHERE Song_Name_9LC = "Pool Revenue" AND Statement_Period_Half_9LC <> ""
                           GROUP BY Song_Name_9LC ORDER BY `Total` DESC'''.format(cut_off)
    pool_rev_2_1 = pool_rev_2_1_1 + pool_rev_2_1_2 + pool_rev_2_1_3
    mycursor.execute(pool_rev_2_1)
    pool_revenue_2_1 = mycursor.fetchall()

    #Column name list
    column_names_2 = ['Song Title']
    for d in year_list:
        column_names_2.append(d)
    column_names_2.append('Total')

    #Add title row to worksheet
    ws2.merge_cells('B1:{}1'.format(column_letters[len(year_list)+1]))
    ws2.cell(row=1, column=1).value = '{}'.format(database[:-25])
    ws2.cell(row=1, column=1).style = 'title_style'
    ws2.cell(row=1, column=2).value = 'Top Earners'
    ws2.cell(row=1, column=2).style = 'title_style'

    #Add column names to worksheet
    column_name_row_2 = ws2[2]
    for cell, e in zip(column_name_row_2, range(len(column_names_2))):
        cell.value = column_names_2[e]
        cell.style = 'header_style'

    #Specify H1/H2 for most recent year
    ws2.cell(row=3, column=len(year_list)+1).value = final_half
    for col in range(1,len(column_names_2)+1):
        ws2.cell(row=3, column=col).style = 'header_style'

    #Add top 80% songs
    top_80_songs_rows = ws2[4:3+song_cut_off_no]
    top_80_songs = table_2_1[0:song_cut_off_no]
    for row, f in zip(top_80_songs_rows,top_80_songs):
        for cell, g in zip(row, range(len(f))):
            cell.value = f[g]
            cell.style = 'number_style'
    for row in range(4,song_cut_off_no+4):
        ws2.cell(row=row, column=1).style = 'name_style'

    #Add top 80% song total row
    for h, i in zip(range(2, len(column_names_2)+1), column_letters[1:]):
        ws2.cell(row=song_cut_off_no+4, column=h).value = "=SUM({}{}:{}{})".format(i, 4, i, song_cut_off_no+3)
        ws2.cell(row=song_cut_off_no+4, column=h).style = 'lined_total_style'
    ws2.cell(row=song_cut_off_no+4, column=1).value = 'Subtotal (Top 80%)'
    ws2.cell(row=song_cut_off_no+4, column=1).style = 'total_label_style'

    #Add pool revenue line
    pool_rev_row_2 = ws2[song_cut_off_no+5]
    pool_rev_element_2 = pool_revenue_2_1[0]
    for cell, j in zip(pool_rev_row_2, range(len(pool_rev_element_2))):
        cell.value = pool_rev_element_2[j]
        cell.style = 'number_style'
    ws2.cell(row=song_cut_off_no+5, column=1).style = 'name_style'

    #Add other songs row
    other_totals_2 = ['Other Songs']
    other_songs_2 = table_2_1[song_cut_off_no:]
    other_songs_row_2 = ws2[song_cut_off_no+6]
    year_total_2 = 0
    for k in range(1,len(column_names_2)):
        for l in other_songs_2:
            year_total_2 += l[k]
        other_totals_2.append(year_total_2)
        year_total_2 = 0
    for cell, m in zip(other_songs_row_2, range(len(other_totals_2))):
        cell.value = other_totals_2[m]
        cell.style = 'number_style'
    ws2.cell(row=song_cut_off_no+6, column=1).style = 'name_style'

    #Add total row
    for n, o in zip(range(2, len(column_names_2)+1), column_letters[1:]):
        ws2.cell(row=song_cut_off_no+7, column=n).value = "=SUM({}{}:{}{})".format(o,song_cut_off_no+4,
                                                                                   o,song_cut_off_no+6)
        ws2.cell(row=song_cut_off_no+7, column=n).style = 'lined_total_style'
    ws2.cell(row=song_cut_off_no+7, column=1).value = 'Total'
    ws2.cell(row=song_cut_off_no + 7, column=1).style = 'total_label_style'

    #Format total column
    for p in range(4, song_cut_off_no+4):
        ws2.cell(row=p, column=len(year_list)+2).style = 'total_style'
    ws2.cell(row=song_cut_off_no+5, column=len(year_list) + 2).style = 'total_style'
    ws2.cell(row=song_cut_off_no+6, column=len(year_list) + 2).style = 'total_style'


#2.1. Top earners by half
    #Build main pivot table (song x rev x half)
    select_table_2_2_1 = '''SELECT Song_Name_9LC,'''
    select_table_2_2_2 = ""
    for b in statement_period_half:
        select_table_2_2_2 += (''' sum( CASE WHEN Statement_Period_Half_9LC = "{}"
                                    AND Song_Name_9LC <> "Pool Revenue" 
                                    THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(b, b))
    select_table_2_2_3 = '''sum( CASE WHEN Song_Name_9LC <> "Pool Revenue" AND Statement_Period_Half_9LC >= "{}"
                             THEN Adjusted_Royalty_SB ELSE "" END) AS `Total`
                             FROM Master WHERE Song_Name_9LC <> "Pool Revenue" AND Statement_Period_Half_9LC <> ""
                             GROUP BY Song_Name_9LC ORDER BY `Total` DESC'''.format(cut_off)
    select_table_2_2 = select_table_2_2_1 + select_table_2_2_2 + select_table_2_2_3
    mycursor.execute(select_table_2_2)
    table_2_2 = mycursor.fetchall()

    #Pool revenue line
    pool_rev_2_2_1 = '''SELECT Song_Name_9LC,'''
    pool_rev_2_2_2 = ""
    for c in statement_period_half:
        pool_rev_2_2_2 += ('''sum( CASE WHEN Statement_Period_Half_9LC = "{}" AND Song_Name_9LC = "Pool Revenue"
                               THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(c, c))
    pool_rev_2_2_3 = '''sum( CASE WHEN Song_Name_9LC = "Pool Revenue" AND Statement_Period_Half_9LC >= "{}"
                              THEN Adjusted_Royalty_SB ELSE "" END) AS `Total` 
                              FROM Master WHERE Song_Name_9LC = "Pool Revenue" AND Statement_Period_Half_9LC <> ""
                              GROUP BY Song_Name_9LC ORDER BY `Total` DESC'''.format(cut_off)
    pool_rev_2_2 = pool_rev_2_2_1 + pool_rev_2_2_2 + pool_rev_2_2_3
    mycursor.execute(pool_rev_2_2)
    pool_revenue_2_2 = mycursor.fetchall()

    #Column name list
    column_names_2_1 = ['Song Title']
    for d in statement_period_half:
        column_names_2_1.append(d)
    column_names_2_1.append('Total')

    #Add title row to worksheet
    ws2_1.merge_cells('B1:{}1'.format(column_letters[len(statement_period_half) + 1]))
    ws2_1.cell(row=1, column=1).value = '{}'.format(database[:-25])
    ws2_1.cell(row=1, column=1).style = 'title_style'
    ws2_1.cell(row=1, column=2).value = 'Top Earners'
    ws2_1.cell(row=1, column=2).style = 'title_style'

    #Add column names to worksheet
    column_name_row_2_1 = ws2_1[2]
    for cell, e in zip(column_name_row_2_1, range(len(column_names_2_1))):
        cell.value = column_names_2_1[e]
        cell.style = 'header_style'

    #Add top 80% songs
    top_80_songs_rows_2_1 = ws2_1[4:3 + song_cut_off_no]
    top_80_songs_2_1 = table_2_2[0:song_cut_off_no]
    for row, f in zip(top_80_songs_rows_2_1, top_80_songs_2_1):
        for cell, g in zip(row, range(len(f))):
            cell.value = f[g]
            cell.style = 'number_style'
    for row in range(4, song_cut_off_no + 4):
        ws2_1.cell(row=row, column=1).style = 'name_style'

    #Add top 80% song total row
    for h, i in zip(range(2, len(column_names_2_1) + 1), column_letters[1:]):
        ws2_1.cell(row=song_cut_off_no + 4, column=h).value = "=SUM({}{}:{}{})".format(i, 4, i, song_cut_off_no + 3)
        ws2_1.cell(row=song_cut_off_no + 4, column=h).style = 'lined_total_style'
    ws2_1.cell(row=song_cut_off_no + 4, column=1).value = 'Subtotal (Top 80%)'
    ws2_1.cell(row=song_cut_off_no + 4, column=1).style = 'total_label_style'

    #Add pool revenue line
    pool_rev_row_2_1 = ws2_1[song_cut_off_no + 5]
    pool_rev_element_2_1 = pool_revenue_2_2[0]
    for cell, j in zip(pool_rev_row_2_1, range(len(pool_rev_element_2_1))):
        cell.value = pool_rev_element_2_1[j]
        cell.style = 'number_style'
    ws2_1.cell(row=song_cut_off_no + 5, column=1).style = 'name_style'

    #Add other songs row
    other_totals_2_1 = ['Other Songs']
    other_songs_2_1 = table_2_2[song_cut_off_no:]
    other_songs_row_2_1 = ws2_1[song_cut_off_no + 6]
    half_total_2 = 0
    for k in range(1, len(column_names_2_1)):
        for l in other_songs_2_1:
            half_total_2 += l[k]
        other_totals_2_1.append(half_total_2)
        half_total_2 = 0
    for cell, m in zip(other_songs_row_2_1, range(len(other_totals_2_1))):
        cell.value = other_totals_2_1[m]
        cell.style = 'number_style'
    ws2_1.cell(row=song_cut_off_no + 6, column=1).style = 'name_style'

    #Add total row
    for n, o in zip(range(2, len(column_names_2_1) + 1), column_letters[1:]):
        ws2_1.cell(row=song_cut_off_no + 7, column=n).value = "=SUM({}{}:{}{})".format(o, song_cut_off_no + 4,
                                                                                     o, song_cut_off_no + 6)
        ws2_1.cell(row=song_cut_off_no + 7, column=n).style = 'lined_total_style'
    ws2_1.cell(row=song_cut_off_no + 7, column=1).value = 'Total'
    ws2_1.cell(row=song_cut_off_no + 7, column=1).style = 'total_label_style'

    #Format total column
    for p in range(4, song_cut_off_no + 4):
        ws2_1.cell(row=p, column=len(statement_period_half) + 2).style = 'total_style'
    ws2_1.cell(row=song_cut_off_no + 5, column=len(statement_period_half) + 2).style = 'total_style'
    ws2_1.cell(row=song_cut_off_no + 6, column=len(statement_period_half) + 2).style = 'total_style'

    #Third row
    for col in range(1, len(column_names_2_1) + 1):
        ws2_1.cell(row=3, column=col).style = 'header_style'

#2.2.Top earners by quarter
    #Build main pivot table
    select_table_2_3_1 = '''SELECT Song_Name_9LC,'''
    select_table_2_3_2 = ""
    for b, c in zip(year_statement_list, quarter_statement_list):
        select_table_2_3_2 += (''' sum( CASE WHEN Year_Statement_9LC = "{}" AND Quarter_Statement_9LC = "{}"
                                        AND Song_Name_9LC <> "Pool Revenue" 
                                        THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(b,c,b,c))
    select_table_2_3_3 = '''sum( CASE WHEN Song_Name_9LC <> "Pool Revenue" AND Statement_Period_Half_9LC >= "{}"
                                 THEN Adjusted_Royalty_SB ELSE "" END) AS `Total`
                                 FROM Master WHERE Song_Name_9LC <> "Pool Revenue" AND Statement_Period_Half_9LC <> ""
                                 AND Third_Party_9LC IN ({})
                                 GROUP BY Song_Name_9LC ORDER BY `Total` DESC'''.format(cut_off, quarterly_third_party)
    select_table_2_3 = select_table_2_3_1 + select_table_2_3_2 + select_table_2_3_3
    mycursor.execute(select_table_2_3)
    table_2_3 = mycursor.fetchall()

    #Pool revenue line
    pool_rev_2_3_1 = '''SELECT Song_Name_9LC,'''
    pool_rev_2_3_2 = ""
    for b, c in zip(year_statement_list, quarter_statement_list):
        pool_rev_2_3_2 += ('''sum( CASE WHEN Year_Statement_9LC = "{}" AND Quarter_Statement_9LC = "{}"
                                   AND Song_Name_9LC = "Pool Revenue"
                                   THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(b,c,b,c))
    pool_rev_2_3_3 = '''sum( CASE WHEN Song_Name_9LC = "Pool Revenue" AND Statement_Period_Half_9LC >= "{}"
                                  THEN Adjusted_Royalty_SB ELSE "" END) AS `Total` 
                                  FROM Master WHERE Song_Name_9LC = "Pool Revenue" AND Statement_Period_Half_9LC <> ""
                                  AND Third_Party_9LC IN ({})
                                  GROUP BY Song_Name_9LC ORDER BY `Total` DESC'''.format(cut_off, quarterly_third_party)
    pool_rev_2_3 = pool_rev_2_3_1 + pool_rev_2_3_2 + pool_rev_2_3_3
    mycursor.execute(pool_rev_2_3)
    pool_revenue_2_3 = mycursor.fetchall()

    #Column name list
    column_names_2_2 = ['Song Title']
    for d in statement_period_quarter:
        column_names_2_2.append(d)
    column_names_2_2.append('Total')

    #Add title row to worksheet
    ws2_2.merge_cells('B1:{}1'.format(column_letters[len(statement_period_quarter) + 1]))
    ws2_2.cell(row=1, column=1).value = '{}'.format(database[:-25])
    ws2_2.cell(row=1, column=1).style = 'title_style'
    ws2_2.cell(row=1, column=2).value = 'Top Earners'
    ws2_2.cell(row=1, column=2).style = 'title_style'

    #Add column names to worksheet
    column_name_row_2_2 = ws2_2[2]
    for cell, e in zip(column_name_row_2_2, range(len(column_names_2_2))):
        cell.value = column_names_2_2[e]
        cell.style = 'header_style'

    #Add top 80% songs
    if len(quarterly_PRO_list) + len(quarterly_publisher_list) == len(publisher_list) + len(PRO_list):
        top_80_songs_rows_2_2 = ws2_2[4:3 + song_cut_off_no]
        top_80_songs_2_2 = table_2_3[0:song_cut_off_no]
        for row, f in zip(top_80_songs_rows_2_2, top_80_songs_2_2):
            for cell, g in zip(row, range(len(f))):
                cell.value = f[g]
                cell.style = 'number_style'
        for row in range(4, song_cut_off_no + 4):
            ws2_2.cell(row=row, column=1).style = 'name_style'

    #Add top 80% song total row
        for h, i in zip(range(2, len(column_names_2_2) + 1), column_letters[1:]):
            if len(top_80_songs_2_2) > 0:
                ws2_2.cell(row=song_cut_off_no + 4, column=h).value = "=SUM({}{}:{}{})".format(i, 4, i, song_cut_off_no + 3)
                ws2_2.cell(row=song_cut_off_no + 4, column=h).style = 'lined_total_style'
            else:
                ws2_2.cell(row=song_cut_off_no + 4, column=h).value = '0.00'
                ws2_2.cell(row=song_cut_off_no + 4, column=h).style = 'lined_total_style'
        ws2_2.cell(row=song_cut_off_no + 4, column=1).value = 'Subtotal (Top 80%)'
        ws2_2.cell(row=song_cut_off_no + 4, column=1).style = 'total_label_style'

    #Add pool revenue line
        pool_rev_row_2_2 = ws2_2[song_cut_off_no + 5]
        pool_rev_element_2_2 = pool_revenue_2_3[0]
        for cell, j in zip(pool_rev_row_2_2, range(len(pool_rev_element_2_2))):
            cell.value = pool_rev_element_2_2[j]
            cell.style = 'number_style'
        ws2_2.cell(row=song_cut_off_no + 5, column=1).style = 'name_style'

    #Add other songs row
        other_totals_2_2 = ['Other Songs']
        other_songs_2_2 = table_2_3[song_cut_off_no:]
        other_songs_row_2_2 = ws2_2[song_cut_off_no + 6]
        quarter_total_2 = 0
        for k in range(1, len(column_names_2_2)):
            for l in other_songs_2_2:
                quarter_total_2 += l[k]
            other_totals_2_2.append(quarter_total_2)
            quarter_total_2 = 0
        for cell, m in zip(other_songs_row_2_2, range(len(other_totals_2_2))):
            cell.value = other_totals_2_2[m]
            cell.style = 'number_style'
        ws2_2.cell(row=song_cut_off_no + 6, column=1).style = 'name_style'

    #Add total row
        for n, o in zip(range(2, len(column_names_2_2) + 1), column_letters[1:]):
            ws2_2.cell(row=song_cut_off_no + 7, column=n).value = "=SUM({}{}:{}{})".format(o, song_cut_off_no + 4,
                                                                                           o, song_cut_off_no + 6)
            ws2_2.cell(row=song_cut_off_no + 7, column=n).style = 'lined_total_style'
        ws2_2.cell(row=song_cut_off_no + 7, column=1).value = 'Total'
        ws2_2.cell(row=song_cut_off_no + 7, column=1).style = 'total_label_style'

    else:
        ws2_2.merge_cells('A4:{}4'.format(column_letters[len(statement_period_quarter)+1]))
        if len(not_quarterly_third_party) == 1:
            ws2_2.cell(row=4, column=1).value = 'Quarterly data is not available for {} - data for top earners by quarter is incomplete'.format(not_quarterly_third_party[0])
        if len(not_quarterly_third_party) == 2:
            ws2_2.cell(row=4, column=1).value = 'Quarterly data is not available for {} and {} - data for top earners by quarter is incomplete'.format(
                not_quarterly_third_party[0], not_quarterly_third_party[1])
        if len(not_quarterly_third_party) > 2:
            not_available_statement = 'Quarterly data is not available for ' + '{}, '*(len(not_quarterly_third_party)-1) + 'and {} - data for top earners by quarter is incomplete'
            ws2_2.cell(row=4, column=1).value = not_available_statement.format(*not_quarterly_third_party)
        ws2_2.cell(row=4, column=1).style = 'not_available_style'
        if len(quarterly_third_party) > 0:
            top_80_songs_rows_2_2 = ws2_2[5:4 + song_cut_off_no]
            top_80_songs_2_2 = table_2_3[0:song_cut_off_no]
            for row, f in zip(top_80_songs_rows_2_2, top_80_songs_2_2):
                for cell, g in zip(row, range(len(f))):
                    cell.value = f[g]
                    cell.style = 'number_style'
            for row in range(5, song_cut_off_no + 4):
                ws2_2.cell(row=row, column=1).style = 'name_style'

        #Add top 80% song total row
            for h, i in zip(range(2, len(column_names_2_2) + 1), column_letters[1:]):
                    ws2_2.cell(row=song_cut_off_no + 5, column=h).value = "=SUM({}{}:{}{})".format(i, 5, i,
                                                                                                   song_cut_off_no + 4)
                    ws2_2.cell(row=song_cut_off_no + 5, column=h).style = 'lined_total_style'

            ws2_2.cell(row=song_cut_off_no + 5, column=1).value = 'Subtotal (Top 80%)'
            ws2_2.cell(row=song_cut_off_no + 5, column=1).style = 'total_label_style'

        #Add pool revenue line
            pool_rev_row_2_2 = ws2_2[song_cut_off_no + 6]
            pool_rev_element_2_2 = pool_revenue_2_3[0]
            for cell, j in zip(pool_rev_row_2_2, range(len(pool_rev_element_2_2))):
                cell.value = pool_rev_element_2_2[j]
                cell.style = 'number_style'
            ws2_2.cell(row=song_cut_off_no + 6, column=1).style = 'name_style'

        #Add other songs row
            other_totals_2_2 = ['Other Songs']
            other_songs_2_2 = table_2_3[song_cut_off_no:]
            other_songs_row_2_2 = ws2_2[song_cut_off_no + 7]
            quarter_total_2 = 0
            for k in range(1, len(column_names_2_2)):
                for l in other_songs_2_2:
                    quarter_total_2 += l[k]
                other_totals_2_2.append(quarter_total_2)
                quarter_total_2 = 0
            for cell, m in zip(other_songs_row_2_2, range(len(other_totals_2_2))):
                cell.value = other_totals_2_2[m]
                cell.style = 'number_style'
            ws2_2.cell(row=song_cut_off_no + 7, column=1).style = 'name_style'

        #Add total row
            for n, o in zip(range(2, len(column_names_2_2) + 1), column_letters[1:]):
                ws2_2.cell(row=song_cut_off_no + 8, column=n).value = "=SUM({}{}:{}{})".format(o, song_cut_off_no + 5,
                                                                                               o, song_cut_off_no + 7)
                ws2_2.cell(row=song_cut_off_no + 8, column=n).style = 'lined_total_style'
            ws2_2.cell(row=song_cut_off_no + 8, column=1).value = 'Total'
            ws2_2.cell(row=song_cut_off_no + 8, column=1).style = 'total_label_style'

        #Format total column
            for p in range(5, song_cut_off_no + 5):
                ws2_2.cell(row=p, column=len(statement_period_quarter) + 2).style = 'total_style'
            ws2_2.cell(row=song_cut_off_no + 6, column=len(statement_period_quarter) + 2).style = 'total_style'
            ws2_2.cell(row=song_cut_off_no + 7, column=len(statement_period_quarter) + 2).style = 'total_style'

    #Third row
    for col in range(1, len(column_names_2_2) + 1):
        ws2_2.cell(row=3, column=col).style = 'header_style'

#3. By Type By Year
    #Build publisher tables
    publisher_total_row_no_3 = 3
    sub_total_row_list_3 = []
    ws3.insert_cols(1, len(year_list)+2)
    for a in publisher_list:
        select_table_3_1_1_1 = '''SELECT Normalized_Income_Type_9LC,'''
        select_table_3_1_1_2 = ""
        for b in year_list:
            select_table_3_1_1_2 += ('''sum( CASE WHEN Statement_Period_Half_9LC LIKE "{} H%"
                                   THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(b,b))
        select_table_3_1_1_3 = '''sum( CASE WHEN Statement_Period_Half_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END)
                              AS `TOTAL`
                              FROM Master WHERE Third_Party_9LC = "{}" AND Statement_Period_Half_9LC <> ""
                              GROUP BY Third_Party_9LC, Normalized_Income_Type_9LC 
                              ORDER BY `Total` DESC'''.format(cut_off, a)
        select_table_3_1_1 = select_table_3_1_1_1 + select_table_3_1_1_2 + select_table_3_1_1_3
        mycursor.execute(select_table_3_1_1)
        table_3_1_1 = mycursor.fetchall()

        publisher_total_row_no_3 += len(table_3_1_1)+3
        publisher_space_row_no_3 = publisher_total_row_no_3 +1
        publisher_header_row_no_3 = publisher_total_row_no_3 - len(table_3_1_1)
        publisher_table_rows_3 = ws3[publisher_header_row_no_3:publisher_space_row_no_3]
        for row, c in zip(publisher_table_rows_3, table_3_1_1):
            for cell, d in zip(row, range(len(c))):
                cell.value =c[d]
                cell.style = 'number_style'
        ws3.cell(row=publisher_header_row_no_3-1, column=1).value = '{}'.format(a)
        ws3.cell(row=publisher_header_row_no_3-1, column=1).style = 'sub_header_style'
        for d in range(2, len(year_list)+3):
            ws3.cell(row=publisher_total_row_no_3, column=d).value = '=SUM({}{}:{}{})'.format(column_letters[d-1],
                                                                                            publisher_header_row_no_3,
                                                                                            column_letters[d-1],
                                                                                            publisher_total_row_no_3-1)
            ws3.cell(row=publisher_total_row_no_3, column=d).style = 'lined_total_style'
        for e in range(publisher_header_row_no_3, publisher_total_row_no_3):
            ws3.cell(row=e, column=1).style = 'name_style'
            ws3.cell(row=e, column=len(year_list)+2).style = 'total_style'
        ws3.cell(row=publisher_total_row_no_3, column=1).value = 'Subtotal'
        ws3.cell(row=publisher_total_row_no_3, column=1).style = 'total_label_style'
        sub_total_row_list_3.append(publisher_total_row_no_3)

    #Build PRO tables
    PRO_total_row_no_3 = publisher_total_row_no_3 + 1
    for a in PRO_list:
        select_table_3_1_2_1 = '''SELECT Normalized_Income_Type_9LC,'''
        select_table_3_1_2_2 = ""
        for b in year_list:
            select_table_3_1_2_2 += ('''sum( CASE WHEN Statement_Period_Half_9LC LIKE "{} H%"
                                      THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(b,b))
        select_table_3_1_2_3 = '''sum( CASE WHEN Statement_Period_Half_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END)
                                AS `TOTAL`
                                FROM Master WHERE Third_Party_9LC = "{}" AND Statement_Period_Half_9LC <> ""
                                GROUP BY Third_Party_9LC, Normalized_Income_Type_9LC 
                                ORDER BY `Total` DESC'''.format(cut_off, a)
        select_table_3_1_2 = select_table_3_1_2_1 + select_table_3_1_2_2 + select_table_3_1_2_3
        mycursor.execute(select_table_3_1_2)
        table_3_1_2 = mycursor.fetchall()

        PRO_total_row_no_3 += len(table_3_1_2) + 3
        PRO_space_row_no_3 = PRO_total_row_no_3 + 1
        PRO_header_row_no_3 = PRO_total_row_no_3 - len(table_3_1_2)
        PRO_table_rows_3 = ws3[PRO_header_row_no_3:PRO_space_row_no_3]
        for row, c in zip(PRO_table_rows_3, table_3_1_2):
            for cell, d in zip(row, range(len(c))):
                cell.value = c[d]
                cell.style = 'number_style'
        ws3.cell(row=PRO_header_row_no_3 - 1, column=1).value = '{}'.format(a)
        ws3.cell(row=PRO_header_row_no_3 - 1, column=1).style = 'sub_header_style'
        for d in range(2, len(year_list) + 3):
            ws3.cell(row=PRO_total_row_no_3, column=d).value = '=SUM({}{}:{}{})'.format(column_letters[d - 1],
                                                                                            PRO_header_row_no_3,
                                                                                            column_letters[d - 1],
                                                                                            PRO_total_row_no_3 - 1)
            ws3.cell(row=PRO_total_row_no_3, column=d).style = 'lined_total_style'
        for e in range(PRO_header_row_no_3, PRO_total_row_no_3):
            ws3.cell(row=e, column=1).style = 'name_style'
            ws3.cell(row=e, column=len(year_list) + 2).style = 'total_style'
        ws3.cell(row=PRO_total_row_no_3, column=1).value = 'Subtotal'
        ws3.cell(row=PRO_total_row_no_3, column=1).style = 'total_label_style'
        sub_total_row_list_3.append(PRO_total_row_no_3)

    #Publisher and PRO labels
    ws3.merge_cells('A4:{}4'.format(column_letters[len(year_list)+1]))
    ws3.cell(row=4, column=1).value = 'Publishers'
    for w in range(1, len(year_list)+3):
        ws3.cell(row=4, column=w).style = 'publisher_label_style'

    ws3.merge_cells('A{}:{}{}'.format(publisher_total_row_no_3+2,column_letters[len(year_list)+1],
                                      publisher_total_row_no_3+2))
    ws3.cell(row=publisher_total_row_no_3+2, column=1).value = 'PROs'
    for x in range(1, len(year_list)+3):
        ws3.cell(row=publisher_total_row_no_3+2, column=x).style = 'publisher_label_style'

    #Column name list
    column_names_3 = ['Income Type']
    for c in year_list:
        column_names_3.append(c)
    column_names_3.append('Total')

    #Add title row to worksheet
    ws3.merge_cells('B1:{}1'.format(column_letters[len(year_list)+1]))
    ws3.cell(row=1, column=1).value = '{}'.format(database[:-25])
    ws3.cell(row=1, column=1).style = 'title_style'
    ws3.cell(row=1, column=2).value = 'By Income Type'
    ws3.cell(row=1, column=2).style = 'title_style'

    #Add column names to worksheet
    column_name_row_3 = ws3[2]
    for cell, e in zip(column_name_row_3, range(len(column_names_3))):
        cell.value = column_names_3[e]
        cell.style = 'header_style'

    #Specify H1/H2 for most recent year
    ws3.cell(row=3, column=len(year_list) + 1).value = final_half
    for col in range(1, len(column_names_3) + 1):
        ws3.cell(row=3, column=col).style = 'header_style'

    #Add grand total row
    grand_total_formula_initial_3 = '={}{}'
    grand_total_formula_3 = grand_total_formula_initial_3 + '+{}{}'*(len(sub_total_row_list_3)-1)
    for col, f in zip(range(2, len(year_list)+3), column_letters[1:len(year_list)+2]):
        format_list_3 = []
        for g in sub_total_row_list_3:
            format_list_3.append(f)
            format_list_3.append(g)
        ws3.cell(row=PRO_total_row_no_3+2, column=col).value = grand_total_formula_3.format(*format_list_3)
        ws3.cell(row=PRO_total_row_no_3+2, column=col).style = 'lined_total_style'
    ws3.cell(row=PRO_total_row_no_3+2, column=1).value = 'Grand Total'
    ws3.cell(row=PRO_total_row_no_3+2, column=1).style = 'total_label_style'

    #Build percentage income table
    select_table_3_1_3_1 = '''SELECT Normalized_Income_Type_9LC,'''
    select_table_3_1_3_2 = ""
    for j in year_list:
        select_table_3_1_3_2 += '''sum( CASE WHEN Statement_Period_Half_9LC LIKE "{} H%"
                                 THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(j,j)
    select_table_3_1_3_3 = '''sum( CASE WHEN Statement_Period_Half_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END)
                            AS `Total`
                            FROM Master WHERE Statement_Period_Half_9LC <> "" AND Normalized_Income_Type_9LC <> ""
                            GROUP BY Normalized_Income_Type_9LC ORDER BY `Total` DESC'''.format(cut_off)
    select_table_3_1_3 = select_table_3_1_3_1 + select_table_3_1_3_2 + select_table_3_1_3_3
    mycursor.execute(select_table_3_1_3)
    table_3_1_3 = mycursor.fetchall()
    total_per_year_list_3 = []
    for y in range(1,len(year_list)+1):
        year_total_3 = 0
        for z in table_3_1_3:
            year_total_3 += z[y]
        total_per_year_list_3.append(year_total_3)

    income_percentage_table_1 = []
    for x in table_3_1_3:
        income_percentage_element_1 = [x[0]]
        for v,w in zip(range(1,len(year_list)+1),range(len(total_per_year_list_3))):
            income_percentage_element_1.append(x[v]/total_per_year_list_3[w])
        income_percentage_table_1.append(income_percentage_element_1)

    for c, d in zip(range(4,len(income_percentage_table_1)+4), income_percentage_table_1):
        for e, f in zip((range(len(year_list)+4, 2*len(year_list)+5)), range(len(total_per_year_list_3)+1)):
            ws3.cell(row=c, column=e).value = d[f]
            ws3.cell(row=c, column=e).style = 'Percent'

    #Add title row to income percentage table
    left_merge_percentage_income_1 = column_letters[len(year_list)+4]
    right_merge_percentage_income_1 = column_letters[2*len(year_list)+3]
    ws3.merge_cells('{}1:{}1'.format(left_merge_percentage_income_1, right_merge_percentage_income_1))
    ws3.cell(row=1, column=len(year_list)+4).value = '{}'.format(database[:-25])
    ws3.cell(row=1, column=len(year_list)+4).style = 'title_style'
    ws3.cell(row=1, column=len(year_list)+5).value = 'By Income Type (Percentage)'
    ws3.cell(row=1, column=len(year_list)+5).style = 'title_style'

    #Add column names to income percentage table
    top_left_percentage_income_1 =column_letters[len(year_list)+3]
    for col, e in zip(range(len(year_list)+4,2*len(year_list)+5),range(len(column_names_3)+1)):
        ws3.cell(row=2, column=col).value = column_names_3[e]
        ws3.cell(row=2, column=col).style = 'header_style'

    #Specify H1/H2 for most recent year
    ws3.cell(row=3, column=2*len(year_list)+4).value = final_half
    for col in range(len(year_list)+4,2*len(year_list)+5):
        ws3.cell(row=3, column=col).style = 'header_style'

#3.1.By Type By Half
    #Build publisher tables
    publisher_total_row_no_3_1 = 3
    sub_total_row_list_3_1 = []
    ws3_1.insert_cols(1, len(statement_period_half) + 2)
    for a in publisher_list:
        select_table_3_2_1_1 = '''SELECT Normalized_Income_Type_9LC,'''
        select_table_3_2_1_2 = ""
        for b in statement_period_half:
            select_table_3_2_1_2 += ('''sum( CASE WHEN Statement_Period_Half_9LC = "{}"
                                       THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(b, b))
        select_table_3_2_1_3 = '''sum( CASE WHEN Statement_Period_Half_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END)
                                  AS `TOTAL`
                                  FROM Master WHERE Third_Party_9LC = "{}" AND Statement_Period_Half_9LC <> ""
                                  GROUP BY Third_Party_9LC, Normalized_Income_Type_9LC 
                                  ORDER BY `Total` DESC'''.format(cut_off, a)
        select_table_3_2_1 = select_table_3_2_1_1 + select_table_3_2_1_2 + select_table_3_2_1_3
        mycursor.execute(select_table_3_2_1)
        table_3_2_1 = mycursor.fetchall()

        publisher_total_row_no_3_1 += len(table_3_2_1) + 3
        publisher_space_row_no_3_1 = publisher_total_row_no_3_1 + 1
        publisher_header_row_no_3_1 = publisher_total_row_no_3_1 - len(table_3_2_1)
        publisher_table_rows_3_1 = ws3_1[publisher_header_row_no_3_1:publisher_space_row_no_3_1]
        for row, c in zip(publisher_table_rows_3_1, table_3_2_1):
            for cell, d in zip(row, range(len(c))):
                cell.value = c[d]
                cell.style = 'number_style'
        ws3_1.cell(row=publisher_header_row_no_3_1 - 1, column=1).value = '{}'.format(a)
        ws3_1.cell(row=publisher_header_row_no_3_1 - 1, column=1).style = 'sub_header_style'
        for d in range(2, len(statement_period_half) + 3):
            ws3_1.cell(row=publisher_total_row_no_3_1, column=d).value = '=SUM({}{}:{}{})'.format(column_letters[d-1],
                                                                                            publisher_header_row_no_3_1,
                                                                                            column_letters[d - 1],
                                                                                        publisher_total_row_no_3_1 - 1)
            ws3_1.cell(row=publisher_total_row_no_3_1, column=d).style = 'lined_total_style'
        for e in range(publisher_header_row_no_3_1, publisher_total_row_no_3_1):
            ws3_1.cell(row=e, column=1).style = 'name_style'
            ws3_1.cell(row=e, column=len(statement_period_half) + 2).style = 'total_style'
        ws3_1.cell(row=publisher_total_row_no_3_1, column=1).value = 'Subtotal'
        ws3_1.cell(row=publisher_total_row_no_3_1, column=1).style = 'total_label_style'
        sub_total_row_list_3_1.append(publisher_total_row_no_3_1)

    #Build PRO tables
    PRO_total_row_no_3_1 = publisher_total_row_no_3_1 + 1
    for a in PRO_list:
        select_table_3_2_2_1 = '''SELECT Normalized_Income_Type_9LC,'''
        select_table_3_2_2_2 = ""
        for b in statement_period_half:
            select_table_3_2_2_2 += ('''sum( CASE WHEN Statement_Period_Half_9LC = "{}"
                                          THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(b, b))
        select_table_3_2_2_3 = '''sum( CASE WHEN Statement_Period_Half_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END)
                                    AS `TOTAL`
                                    FROM Master WHERE Third_Party_9LC = "{}" AND Statement_Period_Half_9LC <> ""
                                    GROUP BY Third_Party_9LC, Normalized_Income_Type_9LC 
                                    ORDER BY `Total` DESC'''.format(cut_off, a)
        select_table_3_2_2 = select_table_3_2_2_1 + select_table_3_2_2_2 + select_table_3_2_2_3
        mycursor.execute(select_table_3_2_2)
        table_3_2_2 = mycursor.fetchall()

        PRO_total_row_no_3_1 += len(table_3_2_2) + 3
        PRO_space_row_no_3_1 = PRO_total_row_no_3_1 + 1
        PRO_header_row_no_3_1 = PRO_total_row_no_3_1 - len(table_3_2_2)
        PRO_table_rows_3_1 = ws3_1[PRO_header_row_no_3_1:PRO_space_row_no_3_1]
        for row, c in zip(PRO_table_rows_3_1, table_3_2_2):
            for cell, d in zip(row, range(len(c))):
                cell.value = c[d]
                cell.style = 'number_style'
        ws3_1.cell(row=PRO_header_row_no_3_1 - 1, column=1).value = '{}'.format(a)
        ws3_1.cell(row=PRO_header_row_no_3_1 - 1, column=1).style = 'sub_header_style'
        for d in range(2, len(statement_period_half) + 3):
            ws3_1.cell(row=PRO_total_row_no_3_1, column=d).value = '=SUM({}{}:{}{})'.format(column_letters[d - 1],
                                                                                            PRO_header_row_no_3_1,
                                                                                            column_letters[d - 1],
                                                                                            PRO_total_row_no_3_1 - 1)
            ws3_1.cell(row=PRO_total_row_no_3_1, column=d).style = 'lined_total_style'
        for e in range(PRO_header_row_no_3_1, PRO_total_row_no_3_1):
            ws3_1.cell(row=e, column=1).style = 'name_style'
            ws3_1.cell(row=e, column=len(statement_period_half) + 2).style = 'total_style'
        ws3_1.cell(row=PRO_total_row_no_3_1, column=1).value = 'Subtotal'
        ws3_1.cell(row=PRO_total_row_no_3_1, column=1).style = 'total_label_style'
        sub_total_row_list_3_1.append(PRO_total_row_no_3_1)

    #Publisher and PRO labels
    ws3_1.merge_cells('A4:{}4'.format(column_letters[len(statement_period_half) + 1]))
    ws3_1.cell(row=4, column=1).value = 'Publishers'
    for w in range(1, len(statement_period_half) + 3):
        ws3_1.cell(row=4, column=w).style = 'publisher_label_style'

    ws3_1.merge_cells('A{}:{}{}'.format(publisher_total_row_no_3_1 + 2, column_letters[len(statement_period_half) + 1],
                                          publisher_total_row_no_3_1 + 2))
    ws3_1.cell(row=publisher_total_row_no_3_1 + 2, column=1).value = 'PROs'
    for x in range(1, len(statement_period_half) + 3):
        ws3_1.cell(row=publisher_total_row_no_3_1 + 2, column=x).style = 'publisher_label_style'

    #Column name list
    column_names_3_1 = ['Income Type']
    for c in statement_period_half:
        column_names_3_1.append(c)
    column_names_3_1.append('Total')

    #Add title row to worksheet
    ws3_1.merge_cells('B1:{}1'.format(column_letters[len(statement_period_half) + 1]))
    ws3_1.cell(row=1, column=1).value = '{}'.format(database[:-25])
    ws3_1.cell(row=1, column=1).style = 'title_style'
    ws3_1.cell(row=1, column=2).value = 'By Income Type'
    ws3_1.cell(row=1, column=2).style = 'title_style'

    #Add column names to worksheet
    column_name_row_3_1 = ws3_1[2]
    for cell, e in zip(column_name_row_3_1, range(len(column_names_3_1))):
        cell.value = column_names_3_1[e]
        cell.style = 'header_style'

    #Third row
    for col in range(1, len(column_names_3_1) + 1):
        ws3_1.cell(row=3, column=col).style = 'header_style'

    #Add grand total row
    grand_total_formula_initial_3_1 = '={}{}'
    grand_total_formula_3_1 = grand_total_formula_initial_3_1 + '+{}{}' * (len(sub_total_row_list_3_1) - 1)
    for col, f in zip(range(2, len(statement_period_half) + 3), column_letters[1:len(statement_period_half) + 2]):
        format_list_3_1 = []
        for g in sub_total_row_list_3_1:
            format_list_3_1.append(f)
            format_list_3_1.append(g)
        ws3_1.cell(row=PRO_total_row_no_3_1 + 2, column=col).value = grand_total_formula_3_1.format(*format_list_3_1)
        ws3_1.cell(row=PRO_total_row_no_3_1 + 2, column=col).style = 'lined_total_style'
    ws3_1.cell(row=PRO_total_row_no_3_1 + 2, column=1).value = 'Grand Total'
    ws3_1.cell(row=PRO_total_row_no_3_1 + 2, column=1).style = 'total_label_style'

    #Build percentage income table
    select_table_3_2_3_1 = '''SELECT Normalized_Income_Type_9LC,'''
    select_table_3_2_3_2 = ""
    for j in statement_period_half:
        select_table_3_2_3_2 += '''sum( CASE WHEN Statement_Period_Half_9LC = "{}"
                                     THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(j, j)
    select_table_3_2_3_3 = '''sum( CASE WHEN Statement_Period_Half_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END)
                                AS `Total`
                                FROM Master WHERE Statement_Period_Half_9LC <> "" AND Normalized_Income_Type_9LC <> ""
                                GROUP BY Normalized_Income_Type_9LC ORDER BY `Total` DESC'''.format(cut_off)
    select_table_3_2_3 = select_table_3_2_3_1 + select_table_3_2_3_2 + select_table_3_2_3_3
    mycursor.execute(select_table_3_2_3)
    table_3_2_3 = mycursor.fetchall()
    total_per_year_list_3_1 = []
    for y in range(1, len(statement_period_half) + 1):
        year_total_3_1 = 0
        for z in table_3_2_3:
            year_total_3_1 += z[y]
        total_per_year_list_3_1.append(year_total_3_1)

    income_percentage_table_2 = []
    for x in table_3_2_3:
        income_percentage_element_2 = [x[0]]
        for v, w in zip(range(1, len(statement_period_half) + 1), range(len(total_per_year_list_3_1))):
            income_percentage_element_2.append(x[v] / total_per_year_list_3_1[w])
        income_percentage_table_2.append(income_percentage_element_2)

    for c, d in zip(range(4, len(income_percentage_table_2) + 4), income_percentage_table_2):
        for e, f in zip((range(len(statement_period_half) + 4, 2 * len(statement_period_half) + 5)),
                        range(len(total_per_year_list_3_1) + 1)):
            ws3_1.cell(row=c, column=e).value = d[f]
            ws3_1.cell(row=c, column=e).style = 'Percent'

    #Add title row to income percentage table
    left_merge_percentage_income_2 = column_letters[len(statement_period_half) + 4]
    right_merge_percentage_income_2 = column_letters[2 * len(statement_period_half) + 3]
    ws3_1.merge_cells('{}1:{}1'.format(left_merge_percentage_income_2, right_merge_percentage_income_2))
    ws3_1.cell(row=1, column=len(statement_period_half) + 4).value = '{}'.format(database[:-25])
    ws3_1.cell(row=1, column=len(statement_period_half) + 4).style = 'title_style'
    ws3_1.cell(row=1, column=len(statement_period_half) + 5).value = 'By Income Type (Percentage)'
    ws3_1.cell(row=1, column=len(statement_period_half) + 5).style = 'title_style'

    #Add column names to income percentage table
    top_left_percentage_income_2 = column_letters[len(statement_period_half) + 3]
    for col, e in zip(range(len(statement_period_half) + 4, 2 * len(statement_period_half) + 5), range(len(column_names_3_1) + 1)):
        ws3_1.cell(row=2, column=col).value = column_names_3_1[e]
        ws3_1.cell(row=2, column=col).style = 'header_style'

    #Third row
    for col in range(len(statement_period_half) + 4, 2 * len(statement_period_half) + 5):
        ws3_1.cell(row=3, column=col).style = 'header_style'

#3.2. By Type By Quarter
    #Build publisher tables
    ws3_2.insert_cols(1, len(statement_period_quarter) + 2)
    if len(quarterly_publisher_list) == len(publisher_list):
        publisher_total_row_no_3_2 = 3
    else:
        publisher_total_row_no_3_2 = 4
        ws3_2.merge_cells('A5:{}5'.format(column_letters[len(statement_period_quarter)+1]))
        if len(not_quarterly_publisher_list) == 1:
            ws3_2.cell(row=5, column=1).value = 'Quarterly data not available for {}'.format(not_quarterly_publisher_list[0])
        if len(not_quarterly_publisher_list) == 2:
            ws3_2.cell(row=5, column=1).value = 'Quarterly data not available for {} and {}'.format(not_quarterly_publisher_list[0], not_quarterly_publisher_list[1])
        if len(not_quarterly_publisher_list) > 2:
            not_available_statement = 'Quarterly data not available for ' + '{}, '*(len(not_quarterly_publisher_list)-1) + 'and {}'
            ws3_2.cell(row=5, column=1).value = not_available_statement.format(*not_quarterly_publisher_list)
        ws3_2.cell(row=5, column=1).style = 'not_available_style'
    sub_total_row_list_3_2 = []
    for a in quarterly_publisher_list:
        select_table_3_3_1_1 = '''SELECT Normalized_Income_Type_9LC,'''
        select_table_3_3_1_2 = ""
        for b, c in zip(year_statement_list, quarter_statement_list):
            select_table_3_3_1_2 += ('''sum( CASE WHEN Year_Statement_9LC = "{}" AND Quarter_Statement_9LC = "{}"
                                        THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(b,c,b,c))
        select_table_3_3_1_3 = '''sum( CASE WHEN Statement_Period_Half_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END)
                                      AS `TOTAL`
                                      FROM Master WHERE Third_Party_9LC = "{}" AND Statement_Period_Half_9LC <> ""
                                      GROUP BY Third_Party_9LC, Normalized_Income_Type_9LC 
                                      ORDER BY `Total` DESC'''.format(cut_off, a)
        select_table_3_3_1 = select_table_3_3_1_1 + select_table_3_3_1_2 + select_table_3_3_1_3
        mycursor.execute(select_table_3_3_1)
        table_3_3_1 = mycursor.fetchall()
        print(table_3_3_1)


        publisher_total_row_no_3_2 += len(table_3_3_1) + 3
        publisher_space_row_no_3_2 = publisher_total_row_no_3_2 + 1
        publisher_header_row_no_3_2 = publisher_total_row_no_3_2 - len(table_3_3_1)
        publisher_table_rows_3_2 = ws3_2[publisher_header_row_no_3_2:publisher_space_row_no_3_2]
        for row, c in zip(publisher_table_rows_3_2, table_3_3_1):
            for cell, d in zip(row, range(len(c))):
                cell.value = c[d]
                cell.style = 'number_style'
        ws3_2.cell(row=publisher_header_row_no_3_2 - 1, column=1).value = '{}'.format(a)
        ws3_2.cell(row=publisher_header_row_no_3_2 - 1, column=1).style = 'sub_header_style'
        for d in range(2, len(statement_period_quarter) + 3):
            ws3_2.cell(row=publisher_total_row_no_3_2, column=d).value = '=SUM({}{}:{}{})'.format(
                        column_letters[d - 1],
                        publisher_header_row_no_3_2,
                        column_letters[d - 1],
                        publisher_total_row_no_3_2 - 1)
            ws3_2.cell(row=publisher_total_row_no_3_2, column=d).style = 'lined_total_style'
        for e in range(publisher_header_row_no_3_2, publisher_total_row_no_3_2):
            ws3_2.cell(row=e, column=1).style = 'name_style'
            ws3_2.cell(row=e, column=len(statement_period_quarter) + 2).style = 'total_style'
        ws3_2.cell(row=publisher_total_row_no_3_2, column=1).value = 'Subtotal'
        ws3_2.cell(row=publisher_total_row_no_3_2, column=1).style = 'total_label_style'
        sub_total_row_list_3_2.append(publisher_total_row_no_3_2)





    #Build PRO tables
    if len(quarterly_PRO_list) == len(PRO_list):
        PRO_total_row_no_3_2 = publisher_total_row_no_3_2 + 1
    else:
        PRO_total_row_no_3_2 = publisher_total_row_no_3_2 + 2
        ws3_2.merge_cells('A{}:{}{}'.format(PRO_total_row_no_3_2 + 1, column_letters[len(statement_period_quarter)+1],PRO_total_row_no_3_2 + 1))
        if len(not_quarterly_PRO_list) == 1:
            ws3_2.cell(row=PRO_total_row_no_3_2 + 1, column=1).value = 'Quarterly data not available for {}'.format(not_quarterly_PRO_list[0])
        if len(not_quarterly_PRO_list) == 2:
            ws3_2.cell(row=PRO_total_row_no_3_2 + 1, column=1).value = 'Quarterly data not available for {} and {}'.format(not_quarterly_PRO_list[0], not_quarterly_PRO_list[1])
        if len(not_quarterly_PRO_list) > 2:
            not_available_statement = 'Quarterly data not available for ' + '{}, '*(len(not_quarterly_PRO_list)-1) + 'and {}'
            ws3_2.cell(row=PRO_total_row_no_3_2 + 1, column=1).value = not_available_statement.format(*not_quarterly_PRO_list)
        ws3_2.cell(row=PRO_total_row_no_3_2 + 1, column=1).style = 'not_available_style'

    for a in quarterly_PRO_list:
        select_table_3_3_2_1 = '''SELECT Normalized_Income_Type_9LC,'''
        select_table_3_3_2_2 = ""
        for b, c in zip(year_statement_list, quarter_statement_list):
            select_table_3_3_2_2 += ('''sum( CASE WHEN Year_Statement_9LC = "{}" AND Quarter_Statement_9LC = "{}"
                                              THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(b,c,b,c))
        select_table_3_3_2_3 = '''sum( CASE WHEN Statement_Period_Half_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END)
                                        AS `TOTAL`
                                        FROM Master WHERE Third_Party_9LC = "{}" AND Statement_Period_Half_9LC <> ""
                                        GROUP BY Third_Party_9LC, Normalized_Income_Type_9LC 
                                        ORDER BY `Total` DESC'''.format(cut_off, a)
        select_table_3_3_2 = select_table_3_3_2_1 + select_table_3_3_2_2 + select_table_3_3_2_3
        mycursor.execute(select_table_3_3_2)
        table_3_3_2 = mycursor.fetchall()
        print(table_3_3_2)


        PRO_total_row_no_3_2 += len(table_3_3_2) + 3
        PRO_space_row_no_3_2 = PRO_total_row_no_3_2 + 1
        PRO_header_row_no_3_2 = PRO_total_row_no_3_2 - len(table_3_3_2)
        PRO_table_rows_3_2 = ws3_2[PRO_header_row_no_3_2:PRO_space_row_no_3_2]
        for row, c in zip(PRO_table_rows_3_2, table_3_3_2):
            for cell, d in zip(row, range(len(c))):
                cell.value = c[d]
                cell.style = 'number_style'
        ws3_2.cell(row=PRO_header_row_no_3_2 - 1, column=1).value = '{}'.format(a)
        ws3_2.cell(row=PRO_header_row_no_3_2 - 1, column=1).style = 'sub_header_style'
        for d in range(2, len(statement_period_quarter) + 3):
            ws3_2.cell(row=PRO_total_row_no_3_2, column=d).value = '=SUM({}{}:{}{})'.format(column_letters[d - 1],
                                                                                                PRO_header_row_no_3_2,
                                                                                                column_letters[d - 1],
                                                                                                PRO_total_row_no_3_2 - 1)
            ws3_2.cell(row=PRO_total_row_no_3_2, column=d).style = 'lined_total_style'
        for e in range(PRO_header_row_no_3_2, PRO_total_row_no_3_2):
            ws3_2.cell(row=e, column=1).style = 'name_style'
            ws3_2.cell(row=e, column=len(statement_period_quarter) + 2).style = 'total_style'
        ws3_2.cell(row=PRO_total_row_no_3_2, column=1).value = 'Subtotal'
        ws3_2.cell(row=PRO_total_row_no_3_2, column=1).style = 'total_label_style'
        sub_total_row_list_3_2.append(PRO_total_row_no_3_2)

    #Publisher and PRO labels
    ws3_2.merge_cells('A4:{}4'.format(column_letters[len(statement_period_quarter) + 1]))
    ws3_2.cell(row=4, column=1).value = 'Publishers'
    for w in range(1, len(statement_period_quarter) + 3):
        ws3_2.cell(row=4, column=w).style = 'publisher_label_style'

    ws3_2.merge_cells('A{}:{}{}'.format(publisher_total_row_no_3_2 + 2, column_letters[len(statement_period_quarter) + 1],
                              publisher_total_row_no_3_2 + 2))
    ws3_2.cell(row=publisher_total_row_no_3_2 + 2, column=1).value = 'PROs'
    for x in range(1, len(statement_period_quarter) + 3):
        ws3_2.cell(row=publisher_total_row_no_3_2 + 2, column=x).style = 'publisher_label_style'

    #Column name list
    column_names_3_2 = ['Income Type']
    for c in statement_period_quarter:
        column_names_3_2.append(c)
    column_names_3_2.append('Total')

    #Add title row to worksheet
    ws3_2.merge_cells('B1:{}1'.format(column_letters[len(statement_period_quarter) + 1]))
    ws3_2.cell(row=1, column=1).value = '{}'.format(database[:-25])
    ws3_2.cell(row=1, column=1).style = 'title_style'
    ws3_2.cell(row=1, column=2).value = 'By Income Type'
    ws3_2.cell(row=1, column=2).style = 'title_style'

    #Add column names to worksheet
    column_name_row_3_2 = ws3_2[2]
    for cell, e in zip(column_name_row_3_2, range(len(column_names_3_2))):
        cell.value = column_names_3_2[e]
        cell.style = 'header_style'

    #Third row
    for col in range(1, len(column_names_3_2) + 1):
        ws3_2.cell(row=3, column=col).style = 'header_style'

    #Add grand total row
    grand_total_formula_initial_3_2 = '={}{}'
    grand_total_formula_3_2 = grand_total_formula_initial_3_2 + '+{}{}' * (len(sub_total_row_list_3_2) - 1)
    for col, f in zip(range(2, len(statement_period_quarter) + 3), column_letters[1:len(statement_period_quarter) + 2]):
        format_list_3_2 = []
        for g in sub_total_row_list_3_2:
            format_list_3_2.append(f)
            format_list_3_2.append(g)
        ws3_2.cell(row=PRO_total_row_no_3_2 + 2, column=col).value = grand_total_formula_3_2.format(*format_list_3_2)
        ws3_2.cell(row=PRO_total_row_no_3_2 + 2, column=col).style = 'lined_total_style'
    ws3_2.cell(row=PRO_total_row_no_3_2 + 2, column=1).value = 'Grand Total'
    ws3_2.cell(row=PRO_total_row_no_3_2 + 2, column=1).style = 'total_label_style'

    #Build percentage income table
    select_table_3_3_3_1 = '''SELECT Normalized_Income_Type_9LC,'''
    select_table_3_3_3_2 = ""
    for i,j in zip(year_statement_list, quarter_statement_list):
        select_table_3_3_3_2 += '''sum( CASE WHEN Year_Statement_9LC = "{}" AND Quarter_Statement_9LC = "{}"
                                         THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(i,j,i,j)
    select_table_3_3_3_3 = '''sum( CASE WHEN Statement_Period_Half_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END)
                                    AS `Total`
                                    FROM Master WHERE Statement_Period_Half_9LC <> "" 
                                    AND Normalized_Income_Type_9LC <> ""
                                    GROUP BY Normalized_Income_Type_9LC ORDER BY `Total` DESC'''.format(cut_off)
    select_table_3_3_3 = select_table_3_3_3_1 + select_table_3_3_3_2 + select_table_3_3_3_3
    mycursor.execute(select_table_3_3_3)
    table_3_3_3 = mycursor.fetchall()
    total_per_year_list_3_2 = []
    for y in range(1, len(statement_period_quarter) + 1):
        year_total_3_2 = 0
        for z in table_3_3_3:
            year_total_3_2 += z[y]
        total_per_year_list_3_2.append(year_total_3_2)

    income_percentage_table_3 = []
    for x in table_3_3_3:
        income_percentage_element_3 = [x[0]]
        for v, w in zip(range(1, len(statement_period_quarter) + 1), range(len(total_per_year_list_3_2))):
            income_percentage_element_3.append(x[v] / total_per_year_list_3_2[w])
        income_percentage_table_3.append(income_percentage_element_3)

    for c, d in zip(range(4, len(income_percentage_table_3) + 4), income_percentage_table_3):
        for e, f in zip((range(len(statement_period_quarter) + 4, 2 * len(statement_period_quarter) + 5)),
                            range(len(total_per_year_list_3_2) + 1)):
            ws3_2.cell(row=c, column=e).value = d[f]
            ws3_2.cell(row=c, column=e).style = 'Percent'

    #Add title row to income percentage table
    left_merge_percentage_income_3 = column_letters[len(statement_period_quarter) + 4]
    right_merge_percentage_income_3 = column_letters[2 * len(statement_period_quarter) + 3]
    ws3_2.merge_cells('{}1:{}1'.format(left_merge_percentage_income_3, right_merge_percentage_income_3))
    ws3_2.cell(row=1, column=len(statement_period_quarter) + 4).value = '{}'.format(database[:-25])
    ws3_2.cell(row=1, column=len(statement_period_quarter) + 4).style = 'title_style'
    ws3_2.cell(row=1, column=len(statement_period_quarter) + 5).value = 'By Income Type (Percentage)'
    ws3_2.cell(row=1, column=len(statement_period_quarter) + 5).style = 'title_style'

    #Add column names to income percentage table
    top_left_percentage_income_3 = column_letters[len(statement_period_quarter) + 3]
    for col, e in zip(range(len(statement_period_quarter) + 4, 2 * len(statement_period_quarter) + 5),
                        range(len(column_names_3_2) + 1)):
        ws3_2.cell(row=2, column=col).value = column_names_3_2[e]
        ws3_2.cell(row=2, column=col).style = 'header_style'

    #Third row
    for col in range(len(statement_period_quarter) + 4, 2 * len(statement_period_quarter) + 5):
        ws3_2.cell(row=3, column=col).style = 'header_style'


#3.3. By Type - Top Earner Breakdown
    song_total_row_no_3_3 = 2
    sub_total_row_list_3_3 = []

    #Top eighty song list
    top_80_song_list = song_list[:song_cut_off_no]

    if len(third_party_list) == len(quarterly_third_party):
        ws3_3.insert_cols(1, len(statement_period_quarter) + 3)
        for a in top_80_song_list:
            select_table_3_4_1_1 = '''SELECT Normalized_Income_Type_9LC,'''
            select_table_3_4_1_2 = ""
            for b, c in zip(year_statement_list, quarter_statement_list):
                select_table_3_4_1_2 += ('''sum( CASE WHEN Year_Statement_9LC = "{}" AND Quarter_Statement_9LC = "{}"
                                                        THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(b,c,b,c))
            select_table_3_4_1_3 = '''sum( CASE WHEN Statement_Period_Half_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END)
                                                      AS `TOTAL`
                                                      FROM Master WHERE Song_Name_9LC = "{}" AND Statement_Period_Half_9LC <> ""
                                                      GROUP BY Song_Name_9LC, Normalized_Income_Type_9LC 
                                                      ORDER BY `Total` DESC'''.format(cut_off, a)
            select_table_3_4_1 = select_table_3_4_1_1 + select_table_3_4_1_2 + select_table_3_4_1_3
            mycursor.execute(select_table_3_4_1)
            table_3_4_1 = mycursor.fetchall()

            song_total_row_no_3_3 += len(table_3_4_1) + 3
            song_space_row_no_3_3 = song_total_row_no_3_3 + 1
            song_header_row_no_3_3 = song_total_row_no_3_3 - len(table_3_4_1)
            song_table_rows_3_3 = ws3_3[song_header_row_no_3_3:song_space_row_no_3_3]
            for row, c in zip(song_table_rows_3_3, table_3_4_1):
                for cell, d in zip(row, range(len(c))):
                    cell.value = c[d]
                    cell.style = 'number_style'
            for d in range(2, len(statement_period_quarter) + 3):
                ws3_3.cell(row=song_total_row_no_3_3, column=d).value = '=SUM({}{}:{}{})'.format(column_letters[d - 1],
                                                                                                 song_header_row_no_3_3,
                                                                                                 column_letters[d - 1],
                                                                                                 song_total_row_no_3_3 - 1)
                ws3_3.cell(row=song_total_row_no_3_3, column=d).style = 'lined_total_style'
            for e in range(song_header_row_no_3_3, song_total_row_no_3_3):
                ws3_3.cell(row=e, column=1).style = 'name_style'
                ws3_3.cell(row=e, column=len(statement_period_quarter) + 2).style = 'total_style'
            ws3_3.cell(row=song_header_row_no_3_3 - 1, column=1).value = "{}".format(a)
            ws3_3.cell(row=song_header_row_no_3_3 - 1, column=1).style = 'sub_header_style'
            ws3_3.cell(row=song_total_row_no_3_3, column=1).value = 'Subtotal'
            ws3_3.cell(row=song_total_row_no_3_3, column=1).style = 'total_label_style'
            sub_total_row_list_3_3.append(song_total_row_no_3_3)

        # Column name list
        column_names_3_3 = ['Income Type']
        for c in statement_period_quarter:
            column_names_3_3.append(c)
        column_names_3_3.append('Total')

        #Add title row to worksheet
        ws3_3.merge_cells('B1:{}1'.format(column_letters[len(statement_period_quarter) + 1]))
        ws3_3.cell(row=1, column=1).value = '{}'.format(database[:-25])
        ws3_3.cell(row=1, column=1).style = 'title_style'
        ws3_3.cell(row=1, column=2).value = 'By Income Type'
        ws3_3.cell(row=1, column=2).style = 'title_style'

        #Add column names to worksheet
        column_name_row_3_3 = ws3_3[2]
        for cell, e in zip(column_name_row_3_3, range(len(column_names_3_3))):
            cell.value = column_names_3_3[e]
            cell.style = 'header_style'

            #Third row
            for col in range(1, len(column_names_3_3) + 1):
                ws3_3.cell(row=3, column=col).style = 'header_style'

        #Add grand total row
        grand_total_formula_initial_3_3 = '={}{}'
        grand_total_formula_3_3 = grand_total_formula_initial_3_3 + '+{}{}' * (len(sub_total_row_list_3_3) - 1)
        for col, f in zip(range(2, len(statement_period_quarter) + 4), column_letters[1:len(statement_period_quarter) + 2]):
            format_list_3_3 = []
            for g in sub_total_row_list_3_3:
                format_list_3_3.append(f)
                format_list_3_3.append(g)
            ws3_3.cell(row=song_total_row_no_3_3 + 2, column=col).value = grand_total_formula_3_3.format(
                *format_list_3_3)
            ws3_3.cell(row=song_total_row_no_3_3 + 2, column=col).style = 'lined_total_style'
        ws3_3.cell(row=song_total_row_no_3_3 + 2, column=1).value = 'Grand Total'
        ws3_3.cell(row=song_total_row_no_3_3 + 2, column=1).style = 'total_label_style'

    else:
    #Build song tables
        ws3_3.insert_cols(1, len(statement_period_half) + 3)
        for a in top_80_song_list:
            select_table_3_4_1_1 = '''SELECT Normalized_Income_Type_9LC,'''
            select_table_3_4_1_2 = ""
            for b in statement_period_half:
                select_table_3_4_1_2 += ('''sum( CASE WHEN Statement_Period_Half_9LC = "{}"
                                                THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(b, b))
            select_table_3_4_1_3 = '''sum( CASE WHEN Statement_Period_Half_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END)
                                              AS `TOTAL`
                                              FROM Master WHERE Song_Name_9LC = "{}" AND Statement_Period_Half_9LC <> ""
                                              GROUP BY Song_Name_9LC, Normalized_Income_Type_9LC 
                                              ORDER BY `Total` DESC'''.format(cut_off, a)
            select_table_3_4_1 = select_table_3_4_1_1 + select_table_3_4_1_2 + select_table_3_4_1_3
            mycursor.execute(select_table_3_4_1)
            table_3_4_1 = mycursor.fetchall()

            song_total_row_no_3_3 += len(table_3_4_1) + 3
            song_space_row_no_3_3 = song_total_row_no_3_3 + 1
            song_header_row_no_3_3 = song_total_row_no_3_3 - len(table_3_4_1)
            song_table_rows_3_3 = ws3_3[song_header_row_no_3_3:song_space_row_no_3_3]
            for row, c in zip(song_table_rows_3_3, table_3_4_1):
                for cell, d in zip(row, range(len(c))):
                    cell.value = c[d]
                    cell.style = 'number_style'
            for d in range(2, len(statement_period_half) + 3):
                ws3_3.cell(row=song_total_row_no_3_3, column=d).value = '=SUM({}{}:{}{})'.format(column_letters[d - 1],
                                                                                                song_header_row_no_3_3,
                                                                                                column_letters[d - 1],
                                                                                            song_total_row_no_3_3 - 1)
                ws3_3.cell(row=song_total_row_no_3_3, column=d).style = 'lined_total_style'
            for e in range(song_header_row_no_3_3, song_total_row_no_3_3):
                ws3_3.cell(row=e, column=1).style = 'name_style'
                ws3_3.cell(row=e, column=len(statement_period_half) + 2).style = 'total_style'
            ws3_3.cell(row=song_header_row_no_3_3-1, column=1).value = "{}".format(a)
            ws3_3.cell(row=song_header_row_no_3_3-1, column=1).style = 'sub_header_style'
            ws3_3.cell(row=song_total_row_no_3_3, column=1).value = 'Subtotal'
            ws3_3.cell(row=song_total_row_no_3_3, column=1).style = 'total_label_style'
            sub_total_row_list_3_3.append(song_total_row_no_3_3)

        #Column name list
        column_names_3_3 = ['Income Type']
        for c in statement_period_half:
            column_names_3_3.append(c)
        column_names_3_3.append('Total')

        #Add title row to worksheet
        ws3_3.merge_cells('B1:{}1'.format(column_letters[len(statement_period_half) + 1]))
        ws3_3.cell(row=1, column=1).value = '{}'.format(database[:-25])
        ws3_3.cell(row=1, column=1).style = 'title_style'
        ws3_3.cell(row=1, column=2).value = 'By Income Type'
        ws3_3.cell(row=1, column=2).style = 'title_style'

        #Add column names to worksheet
        column_name_row_3_3 = ws3_3[2]
        for cell, e in zip(column_name_row_3_3, range(len(column_names_3_3))):
            cell.value = column_names_3_3[e]
            cell.style = 'header_style'

        #Third row
            for col in range(1, len(column_names_3_3) + 1):
                ws3_3.cell(row=3, column=col).style = 'header_style'

        #Add grand total row
        grand_total_formula_initial_3_3 = '={}{}'
        grand_total_formula_3_3 = grand_total_formula_initial_3_3 + '+{}{}' * (len(sub_total_row_list_3_3) - 1)
        for col, f in zip(range(2, len(statement_period_half) + 4),column_letters[1:len(statement_period_half) + 2]):
            format_list_3_3 = []
            for g in sub_total_row_list_3_3:
                format_list_3_3.append(f)
                format_list_3_3.append(g)
            ws3_3.cell(row=song_total_row_no_3_3 + 2, column=col).value = grand_total_formula_3_3.format(*format_list_3_3)
            ws3_3.cell(row=song_total_row_no_3_3 + 2, column=col).style = 'lined_total_style'
        ws3_3.cell(row=song_total_row_no_3_3 + 2, column=1).value = 'Grand Total'
        ws3_3.cell(row=song_total_row_no_3_3 + 2, column=1).style = 'total_label_style'


#4. By Source By Year
    #Find total for each publisher
    total_publisher_list = []
    for a in publisher_list:
        find_total_publisher = '''SELECT Third_Party_9LC, sum( CASE WHEN Statement_Period_Half_9LC >= '{}'
                                  THEN Adjusted_Royalty_SB ELSE NULL END) AS `Total`
                                  FROM Master WHERE Third_Party_9LC = "{}" 
                                  GROUP BY Third_Party_9LC
                                  ORDER BY `Total` DESC'''.format(cut_off, a)
        mycursor.execute(find_total_publisher)
        total_publisher = mycursor.fetchall()
        total_publisher_list.append(total_publisher)
    publisher_royalty_eighty = []
    for b in range(len(publisher_list)):
        publisher_royalty_eighty.append(0.8*total_publisher_list[b][0][1])

    #Build publisher tables
    publisher_total_row_no_4 = 0
    sub_total_row_list_4 = []
    ws4.insert_cols(1, len(year_list) + 2)
    for a,c in zip(publisher_list, publisher_royalty_eighty):
        select_table_4_1_1_1 = '''SELECT Normalized_Source_9LC,'''
        select_table_4_1_1_2 = ""
        for b in year_list:
            select_table_4_1_1_2 += ('''sum( CASE WHEN Statement_Period_Half_9LC LIKE "{} H%" 
                                       AND Normalized_Source_9LC <> "Pool Revenue"
                                       THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(b, b))
        select_table_4_1_1_3 = '''sum( CASE WHEN Statement_Period_Half_9LC >= "{}" 
                                  AND Normalized_Source_9LC <> "Pool Revenue" THEN Adjusted_Royalty_SB ELSE "" END)
                                  AS `TOTAL`
                                  FROM Master WHERE Third_Party_9LC = "{}" AND Statement_Period_Half_9LC <> ""
                                  AND Normalized_Source_9LC <> "Pool Revenue"
                                  GROUP BY Third_Party_9LC, Normalized_Source_9LC 
                                  ORDER BY `Total` DESC'''.format(cut_off, a)
        select_table_4_1_1 = select_table_4_1_1_1 + select_table_4_1_1_2 + select_table_4_1_1_3
        mycursor.execute(select_table_4_1_1)
        table_4_1_1 = mycursor.fetchall()

        #Find eighty percent cutoff
        cumulative_total_publisher = 0
        source_cut_off_no_publisher = 0
        for i in table_4_1_1:
            if cumulative_total_publisher <= c:
                cumulative_total_publisher += i[len(year_list)+1]
                source_cut_off_publisher = i[0]
                source_cut_off_no_publisher += 1

        #Find pool revenue
        pool_rev_4_1_1_1 = '''SELECT Normalized_Source_9LC,'''
        pool_rev_4_1_1_2 = ""
        for d in year_list:
            pool_rev_4_1_1_2 += ('''sum( CASE WHEN Statement_Period_Half_9LC LIKE "{} H%" 
                                AND Normalized_Source_9LC = "Pool Revenue"
                                THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(d,d))
        pool_rev_4_1_1_3 = '''sum( CASE WHEN Normalized_Source_9LC = "Pool Revenue" AND Statement_Period_Half_9LC >= "{}"
                          THEN Adjusted_Royalty_SB ELSE "" END) AS `Total`
                          FROM Master WHERE Normalized_Source_9LC = "Pool Revenue" 
                          AND Statement_Period_Half_9LC <> ""
                          AND Third_Party_9LC = "{}"
                          GROUP BY Normalized_Source_9LC ORDER BY `Total` DESC'''.format(cut_off, a)
        pool_rev_4_1_1 = pool_rev_4_1_1_1 + pool_rev_4_1_1_2 + pool_rev_4_1_1_3
        mycursor.execute(pool_rev_4_1_1)
        initial_pool_revenue_4_1_1 = (mycursor.fetchall())
        pool_revenue_4_1_1 = []
        for i in initial_pool_revenue_4_1_1:
            for j in range(len(i)):
                pool_revenue_4_1_1.append(i[j])
        if len(pool_revenue_4_1_1) == 0:
            pool_revenue_4_1_1.append('Pool Revenue')
            pool_revenue_4_1_1_zeros = [0]*(len(year_list)+1)
            pool_revenue_4_1_1.extend(pool_revenue_4_1_1_zeros)

        #Add top eighty percent
        top_80_publisher_sources_4 = table_4_1_1[0:source_cut_off_no_publisher]
        publisher_total_row_no_4 += len(top_80_publisher_sources_4) + 6
        publisher_space_row_no_4 = publisher_total_row_no_4 + 1
        publisher_header_row_no_4 = publisher_total_row_no_4 - len(top_80_publisher_sources_4)
        publisher_table_rows_4 = ws4[publisher_header_row_no_4:publisher_space_row_no_4]
        for row, c in zip(publisher_table_rows_4, top_80_publisher_sources_4):
            for cell, d in zip(row, range(len(c))):
                cell.value = c[d]
                cell.style = 'number_style'
        ws4.cell(row=publisher_header_row_no_4 - 1, column=1).value = '{}'.format(a)
        ws4.cell(row=publisher_header_row_no_4 - 1, column=1).style = 'sub_header_style'
        for d in range(2, len(year_list) + 3):
            ws4.cell(row=publisher_total_row_no_4, column=d).value = '=SUM({}{}:{}{})'.format(column_letters[d - 1],
                                                                                            publisher_header_row_no_4,
                                                                                            column_letters[d - 1],
                                                                                        publisher_total_row_no_4 - 1)
            ws4.cell(row=publisher_total_row_no_4, column=d).style = 'lined_total_style'
        for e in range(publisher_header_row_no_4, publisher_total_row_no_4):
            ws4.cell(row=e, column=1).style = 'name_style'
            ws4.cell(row=e, column=len(year_list) + 2).style = 'total_style'
        ws4.cell(row=publisher_total_row_no_4, column=1).value = 'Subtotal (Top 80%)'
        ws4.cell(row=publisher_total_row_no_4, column=1).style = 'total_label_style'
        sub_total_row_list_4.append(publisher_total_row_no_4 + 3)

        #Add pool revenue line
        pool_rev_4_1_1_row = ws4[publisher_total_row_no_4+1]
        for cell, f in zip(pool_rev_4_1_1_row, range(len(pool_revenue_4_1_1))):
            cell.value = pool_revenue_4_1_1[f]
        for col in range(2,len(pool_revenue_4_1_1)+1):
            ws4.cell(row=publisher_total_row_no_4+1, column=col).style = 'number_style'
        ws4.cell(row=publisher_total_row_no_4+1, column=len(year_list)+2).style = 'total_style'

        #Add other sources row
        other_totals_4_1_1 = ['Other Songs']
        other_sources_4_1_1 = table_4_1_1[source_cut_off_no_publisher:]
        other_sources_row_4_1_1 = ws4[publisher_total_row_no_4+2]
        year_total_4_1 = 0
        for g in range(1, len(year_list)+2):
            for h in other_sources_4_1_1:
                year_total_4_1 += h[g]
            other_totals_4_1_1.append(year_total_4_1)
            year_total_4_1 = 0
        for cell, i in zip(other_sources_row_4_1_1, range(len(other_totals_4_1_1))):
            cell.value = other_totals_4_1_1[i]
            cell.style = 'number_style'
        ws4.cell(row=publisher_total_row_no_4+2, column=1).style = 'name_style'
        ws4.cell(row=publisher_total_row_no_4+2, column=len(year_list)+2).style = 'total_style'

        #Add total row
        for j, k in zip(range(2,len(year_list)+3), column_letters[1:]):
            ws4.cell(row=publisher_total_row_no_4+3, column=j).value = "={}{}+{}{}+{}{}".format(k,
                                                                                        publisher_total_row_no_4,k,
                                                                                        publisher_total_row_no_4+1,k,
                                                                                        publisher_total_row_no_4+2)
            ws4.cell(row=publisher_total_row_no_4+3, column=j).style = 'lined_total_style'
        ws4.cell(row=publisher_total_row_no_4+3, column=1).value = 'Total'
        ws4.cell(row=publisher_total_row_no_4+3, column=1).style = 'total_label_style'

    #Find total for each PRO
    total_PRO_list = []
    for a in PRO_list:
        find_total_PRO = '''SELECT Third_Party_9LC, sum( CASE WHEN Statement_Period_Half_9LC >= '{}'
                                  THEN Adjusted_Royalty_SB ELSE NULL END) AS `Total`
                                  FROM Master WHERE Third_Party_9LC = "{}" 
                                  GROUP BY Third_Party_9LC'''.format(cut_off, a)
        mycursor.execute(find_total_PRO)
        total_PRO = mycursor.fetchall()
        total_PRO_list.append(total_PRO)
        PRO_royalty_eighty = []
    for b in range(len(PRO_list)):
        PRO_royalty_eighty.append(0.8 * total_PRO_list[b][0][1])

    #Build PRO tables
    PRO_total_row_no_4 = publisher_total_row_no_4 + 1
    for a, c in zip(PRO_list, PRO_royalty_eighty):
        select_table_4_1_2_1 = '''SELECT Normalized_Source_9LC,'''
        select_table_4_1_2_2 = ""
        for b in year_list:
            select_table_4_1_2_2 += ('''sum( CASE WHEN Statement_Period_Half_9LC LIKE "{} H%"
                                          AND Normalized_Source_9LC <> "Pool Revenue"
                                          THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(b, b))
        select_table_4_1_2_3 = '''sum( CASE WHEN Statement_Period_Half_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END)
                                    AS `TOTAL`
                                    FROM Master WHERE Third_Party_9LC = "{}" AND Statement_Period_Half_9LC <> ""
                                    AND Normalized_Source_9LC <> "Pool Revenue"
                                    GROUP BY Third_Party_9LC, Normalized_Source_9LC 
                                    ORDER BY `Total` DESC'''.format(cut_off, a)
        select_table_4_1_2 = select_table_4_1_2_1 + select_table_4_1_2_2 + select_table_4_1_2_3
        mycursor.execute(select_table_4_1_2)
        table_4_1_2 = mycursor.fetchall()

        #Find eighty percent cutoff
        cumulative_total_PRO = 0
        source_cut_off_no_PRO = 0
        for i in table_4_1_2:
            if cumulative_total_PRO <= c:
                cumulative_total_PRO += i[len(year_list) + 1]
                source_cut_off_PRO = i[0]
                source_cut_off_no_PRO += 1

        #Find pool revenue line
        pool_rev_4_1_2_1 = '''SELECT Normalized_Source_9LC,'''
        pool_rev_4_1_2_2 = ""
        for d in year_list:
            pool_rev_4_1_2_2 += ('''sum( CASE WHEN Statement_Period_Half_9LC LIKE "{} H%" 
                                        AND Normalized_Source_9LC = "Pool Revenue"
                                        THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(d, d))
        pool_rev_4_1_2_3 = '''sum( CASE WHEN Normalized_Source_9LC = "Pool Revenue" AND Statement_Period_Half_9LC >= "{}"
                                  THEN Adjusted_Royalty_SB ELSE "" END) AS `Total`
                                  FROM Master WHERE Normalized_Source_9LC = "Pool Revenue" 
                                  AND Statement_Period_Half_9LC <> ""
                                  AND Third_Party_9LC = "{}"
                                  GROUP BY Normalized_Source_9LC ORDER BY `Total` DESC'''.format(cut_off, a)
        pool_rev_4_1_2 = pool_rev_4_1_2_1 + pool_rev_4_1_2_2 + pool_rev_4_1_2_3
        mycursor.execute(pool_rev_4_1_2)
        initial_pool_revenue_4_1_2 = (mycursor.fetchall())
        pool_revenue_4_1_2 = []
        for i in initial_pool_revenue_4_1_2:
            for j in range(len(i)):
                pool_revenue_4_1_2.append(i[j])
        if len(pool_revenue_4_1_2) == 0:
            pool_revenue_4_1_2.append('Pool Revenue')
            pool_revenue_4_1_2_zeros = [0] * (len(year_list) + 1)
            pool_revenue_4_1_2.extend(pool_revenue_4_1_2_zeros)

        #Add top eighty percent
        top_80_PRO_sources_4 = table_4_1_2[0:source_cut_off_no_PRO]
        PRO_total_row_no_4 += len(top_80_PRO_sources_4) + 6
        PRO_space_row_no_4 = PRO_total_row_no_4 + 1
        PRO_header_row_no_4 = PRO_total_row_no_4 - len(top_80_PRO_sources_4)
        PRO_table_rows_4 = ws4[PRO_header_row_no_4:PRO_space_row_no_4]
        for row, c in zip(PRO_table_rows_4, top_80_PRO_sources_4):
            for cell, d in zip(row, range(len(c))):
                cell.value = c[d]
                cell.style = 'number_style'
        for x in range(PRO_header_row_no_4, PRO_total_row_no_4):
            ws4.cell(row=x, column=1).style = 'name_style'
            ws4.cell(row=x, column=len(year_list)+2).style = 'total_style'
        ws4.cell(row=PRO_header_row_no_4 - 1, column=1).value = '{}'.format(a)
        ws4.cell(row=PRO_header_row_no_4 - 1, column=1).style = 'sub_header_style'
        for d in range(2, len(year_list) + 3):
            ws4.cell(row=PRO_total_row_no_4, column=d).value = '=SUM({}{}:{}{})'.format(column_letters[d - 1],
                                                                                              PRO_header_row_no_4,
                                                                                              column_letters[d - 1],
                                                                                              PRO_total_row_no_4 - 1)
            ws4.cell(row=PRO_total_row_no_4, column=d).style = 'lined_total_style'
        for e in range(PRO_header_row_no_4, publisher_total_row_no_4):
            ws4.cell(row=e, column=1).style = 'name_style'
            ws4.cell(row=e, column=len(year_list) + 2).style = 'total_style'
        ws4.cell(row=PRO_total_row_no_4, column=1).value = 'Subtotal (Top 80%)'
        ws4.cell(row=PRO_total_row_no_4, column=1).style = 'total_label_style'
        sub_total_row_list_4.append(PRO_total_row_no_4+3)

        #Add pool revenue line
        pool_rev_4_1_2_row = ws4[PRO_total_row_no_4 + 1]
        for cell, f in zip(pool_rev_4_1_2_row, range(len(pool_revenue_4_1_2))):
            cell.value = pool_revenue_4_1_2[f]
        for col in range(2, len(pool_revenue_4_1_2)+1):
            ws4.cell(row=PRO_total_row_no_4 + 1, column=col).style = 'number_style'
        ws4.cell(row=PRO_total_row_no_4 + 1, column=len(year_list) + 2).style = 'total_style'

        #Add other songs row
        other_totals_4_1_2 = ['Other Songs']
        other_songs_4_1_2 = table_4_1_2[source_cut_off_no_PRO:]
        other_songs_row_4_1_2 = ws4[PRO_total_row_no_4 + 2]
        year_total_4_2 = 0
        for g in range(1, len(year_list) + 2):
            for h in other_songs_4_1_2:
                year_total_4_2 += h[g]
            other_totals_4_1_2.append(year_total_4_2)
            year_total_4_2 = 0
        for cell, i in zip(other_songs_row_4_1_2, range(len(other_totals_4_1_2))):
            cell.value = other_totals_4_1_2[i]
            cell.style = 'number_style'
        ws4.cell(row=PRO_total_row_no_4 + 2, column=1).style = 'name_style'
        ws4.cell(row=PRO_total_row_no_4 + 2, column=len(year_list) + 2).style = 'total_style'

        #Add total row
        for j, k in zip(range(2, len(year_list) + 3), column_letters[1:]):
            ws4.cell(row=PRO_total_row_no_4 + 3, column=j).value = "={}{}+{}{}+{}{}".format(k, PRO_total_row_no_4,
                                                                                            k, PRO_total_row_no_4 + 1,
                                                                                            k, PRO_total_row_no_4 + 2)
            ws4.cell(row=PRO_total_row_no_4 + 3, column=j).style = 'lined_total_style'
        ws4.cell(row=PRO_total_row_no_4 + 3, column=1).value = 'Total'
        ws4.cell(row=PRO_total_row_no_4 + 3, column=1).style = 'total_label_style'

    #Add grand total row
    grand_total_formula_initial_4 = '={}{}'
    grand_total_formula_4 = grand_total_formula_initial_4 + '+{}{}' * (len(sub_total_row_list_4) - 1)
    for col, f in zip(range(2, len(year_list) + 3), column_letters[1:len(year_list) + 2]):
        format_list_4 = []
        for g in sub_total_row_list_4:
            format_list_4.append(f)
            format_list_4.append(g)
        ws4.cell(row=PRO_total_row_no_4 + 5, column=col).value = grand_total_formula_4.format(*format_list_4)
        ws4.cell(row=PRO_total_row_no_4 + 5, column=col).style = 'lined_total_style'
    ws4.cell(row=PRO_total_row_no_4 + 5, column=1).value = 'Grand Total'
    ws4.cell(row=PRO_total_row_no_4 + 5, column=1).style = 'total_label_style'

    #Publisher and PRO labels
    ws4.merge_cells('A4:{}4'.format(column_letters[len(year_list) + 1]))
    ws4.cell(row=4, column=1).value = 'Publishers'
    for w in range(1, len(year_list) + 3):
        ws4.cell(row=4, column=w).style = 'publisher_label_style'

    ws4.merge_cells('A{}:{}{}'.format(publisher_total_row_no_4 + 5, column_letters[len(year_list) + 1],
                                      publisher_total_row_no_4 + 5))
    ws4.cell(row=publisher_total_row_no_4 + 5, column=1).value = 'PROs'
    for x in range(1, len(year_list) + 3):
        ws4.cell(row=publisher_total_row_no_4 + 5, column=x).style = 'publisher_label_style'

    #Column name list
    column_names_4 = ['Source']
    for c in year_list:
        column_names_4.append(c)
    column_names_4.append('Total')

    #Add title row to worksheet
    ws4.merge_cells('B1:{}1'.format(column_letters[len(year_list) + 1]))
    ws4.cell(row=1, column=1).value = '{}'.format(database[:-25])
    ws4.cell(row=1, column=1).style = 'title_style'
    ws4.cell(row=1, column=2).value = 'By Source'
    ws4.cell(row=1, column=2).style = 'title_style'

    #Add column names to worksheet
    column_name_row_4 = ws4[2]
    for cell, e in zip(column_name_row_4, range(len(column_names_4))):
        cell.value = column_names_4[e]
        cell.style = 'header_style'

    #Specify H1/H2 for most recent year
    ws4.cell(row=3, column=len(year_list) + 1).value = final_half
    for col in range(1, len(column_names_4) + 1):
        ws4.cell(row=3, column=col).style = 'header_style'

#4.1. By Source By Half
    #Build publisher tables
    publisher_total_row_no_4_1 = 0
    sub_total_row_list_4_1 = []
    ws4_1.insert_cols(1, len(statement_period_half) + 2)
    for a, c in zip(publisher_list, publisher_royalty_eighty):
        select_table_4_2_1_1 = '''SELECT Normalized_Source_9LC,'''
        select_table_4_2_1_2 = ""
        for b in statement_period_half:
            select_table_4_2_1_2 += ('''sum( CASE WHEN Statement_Period_Half_9LC = "{}" 
                                           AND Normalized_Source_9LC <> "Pool Revenue"
                                           THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(b, b))
        select_table_4_2_1_3 = '''sum( CASE WHEN Statement_Period_Half_9LC >= "{}" 
                                      AND Normalized_Source_9LC <> "Pool Revenue" THEN Adjusted_Royalty_SB ELSE "" END)
                                      AS `TOTAL`
                                      FROM Master WHERE Third_Party_9LC = "{}" AND Statement_Period_Half_9LC <> ""
                                      AND Normalized_Source_9LC <> "Pool Revenue"
                                      GROUP BY Third_Party_9LC, Normalized_Source_9LC 
                                      ORDER BY `Total` DESC'''.format(cut_off, a)
        select_table_4_2_1 = select_table_4_2_1_1 + select_table_4_2_1_2 + select_table_4_2_1_3
        mycursor.execute(select_table_4_2_1)
        table_4_2_1 = mycursor.fetchall()

        #Find eighty percent cutoff
        cumulative_total_publisher = 0
        source_cut_off_no_publisher = 0
        for i in table_4_2_1:
            if cumulative_total_publisher <= c:
                cumulative_total_publisher += i[len(statement_period_half) + 1]
                source_cut_off_publisher = i[0]
                source_cut_off_no_publisher += 1

        #Find pool revenue
        pool_rev_4_2_1_1 = '''SELECT Normalized_Source_9LC,'''
        pool_rev_4_2_1_2 = ""
        for d in statement_period_half:
            pool_rev_4_2_1_2 += ('''sum( CASE WHEN Statement_Period_Half_9LC = "{}" 
                                    AND Normalized_Source_9LC = "Pool Revenue"
                                    THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(d, d))
        pool_rev_4_2_1_3 = '''sum( CASE WHEN Normalized_Source_9LC = "Pool Revenue" 
                              AND Statement_Period_Half_9LC >= "{}"
                              THEN Adjusted_Royalty_SB ELSE "" END) AS `Total`
                              FROM Master WHERE Normalized_Source_9LC = "Pool Revenue" 
                              AND Statement_Period_Half_9LC <> ""
                              AND Third_Party_9LC = "{}"
                              GROUP BY Normalized_Source_9LC ORDER BY `Total` DESC'''.format(cut_off, a)
        pool_rev_4_2_1 = pool_rev_4_2_1_1 + pool_rev_4_2_1_2 + pool_rev_4_2_1_3
        mycursor.execute(pool_rev_4_2_1)
        initial_pool_revenue_4_2_1 = (mycursor.fetchall())
        pool_revenue_4_2_1 = []
        for i in initial_pool_revenue_4_2_1:
            for j in range(len(i)):
                pool_revenue_4_2_1.append(i[j])
        if len(pool_revenue_4_2_1) == 0:
            pool_revenue_4_2_1.append('Pool Revenue')
            pool_revenue_4_2_1_zeros = [0] * (len(statement_period_half) + 1)
            pool_revenue_4_2_1.extend(pool_revenue_4_2_1_zeros)

        #Add top eighty percent
        top_80_publisher_sources_4_1 = table_4_2_1[0:source_cut_off_no_publisher]
        publisher_total_row_no_4_1 += len(top_80_publisher_sources_4_1) + 6
        publisher_space_row_no_4_1 = publisher_total_row_no_4_1 + 1
        publisher_header_row_no_4_1 = publisher_total_row_no_4_1 - len(top_80_publisher_sources_4_1)
        publisher_table_rows_4_1 = ws4_1[publisher_header_row_no_4_1:publisher_space_row_no_4_1]
        for row, c in zip(publisher_table_rows_4_1, top_80_publisher_sources_4_1):
            for cell, d in zip(row, range(len(c))):
                cell.value = c[d]
                cell.style = 'number_style'
        ws4_1.cell(row=publisher_header_row_no_4_1 - 1, column=1).value = '{}'.format(a)
        ws4_1.cell(row=publisher_header_row_no_4_1 - 1, column=1).style = 'sub_header_style'
        for d in range(2, len(statement_period_half) + 3):
            ws4_1.cell(row=publisher_total_row_no_4_1, column=d).value = '=SUM({}{}:{}{})'.format(column_letters[d - 1],
                                                                                        publisher_header_row_no_4_1,
                                                                                        column_letters[d - 1],
                                                                                        publisher_total_row_no_4_1 - 1)
            ws4_1.cell(row=publisher_total_row_no_4_1, column=d).style = 'lined_total_style'
        for e in range(publisher_header_row_no_4_1, publisher_total_row_no_4_1):
            ws4_1.cell(row=e, column=1).style = 'name_style'
            ws4_1.cell(row=e, column=len(statement_period_half) + 2).style = 'total_style'
        ws4_1.cell(row=publisher_total_row_no_4_1, column=1).value = 'Subtotal (Top 80%)'
        ws4_1.cell(row=publisher_total_row_no_4_1, column=1).style = 'total_label_style'
        sub_total_row_list_4_1.append(publisher_total_row_no_4_1 + 3)

        #Add pool revenue line
        pool_rev_4_2_1_row = ws4_1[publisher_total_row_no_4_1 + 1]
        for cell, f in zip(pool_rev_4_2_1_row, range(len(pool_revenue_4_2_1))):
            cell.value = pool_revenue_4_2_1[f]
        for col in range(2, len(pool_revenue_4_2_1) + 1):
            ws4_1.cell(row=publisher_total_row_no_4_1 + 1, column=col).style = 'number_style'
        ws4_1.cell(row=publisher_total_row_no_4_1 + 1, column=len(statement_period_half) + 2).style = 'total_style'

        #Add other sources row
        other_totals_4_2_1 = ['Other Songs']
        other_sources_4_2_1 = table_4_2_1[source_cut_off_no_publisher:]
        other_sources_row_4_2_1 = ws4_1[publisher_total_row_no_4_1 + 2]
        half_total_4_1 = 0
        for g in range(1, len(statement_period_half) + 2):
            for h in other_sources_4_2_1:
                half_total_4_1 += h[g]
            other_totals_4_2_1.append(half_total_4_1)
            half_total_4_1 = 0
        for cell, i in zip(other_sources_row_4_2_1, range(len(other_totals_4_2_1))):
            cell.value = other_totals_4_2_1[i]
            cell.style = 'number_style'
        ws4_1.cell(row=publisher_total_row_no_4_1 + 2, column=1).style = 'name_style'
        ws4_1.cell(row=publisher_total_row_no_4_1 + 2, column=len(statement_period_half) + 2).style = 'total_style'

        #Add total row
        for j, k in zip(range(2, len(statement_period_half) + 3), column_letters[1:]):
            ws4_1.cell(row=publisher_total_row_no_4_1 + 3, column=j).value = "={}{}+{}{}+{}{}".format(k,
                                                                                    publisher_total_row_no_4_1, k,
                                                                                    publisher_total_row_no_4_1 + 1, k,
                                                                                    publisher_total_row_no_4_1 + 2)
            ws4_1.cell(row=publisher_total_row_no_4_1 + 3, column=j).style = 'lined_total_style'
            ws4_1.cell(row=publisher_total_row_no_4_1 + 3, column=1).value = 'Total'
            ws4_1.cell(row=publisher_total_row_no_4_1 + 3, column=1).style = 'total_label_style'

    #Build PRO tables
    PRO_total_row_no_4_1 = publisher_total_row_no_4_1 + 1
    for a, c in zip(PRO_list, PRO_royalty_eighty):
        select_table_4_2_2_1 = '''SELECT Normalized_Source_9LC,'''
        select_table_4_2_2_2 = ""
        for b in statement_period_half:
            select_table_4_2_2_2 += ('''sum( CASE WHEN Statement_Period_Half_9LC = "{}"
                                              AND Normalized_Source_9LC <> "Pool Revenue"
                                              THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(b, b))
        select_table_4_2_2_3 = '''sum( CASE WHEN Statement_Period_Half_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END)
                                        AS `TOTAL`
                                        FROM Master WHERE Third_Party_9LC = "{}" AND Statement_Period_Half_9LC <> ""
                                        AND Normalized_Source_9LC <> "Pool Revenue"
                                        GROUP BY Third_Party_9LC, Normalized_Source_9LC 
                                        ORDER BY `Total` DESC'''.format(cut_off, a)
        select_table_4_2_2 = select_table_4_2_2_1 + select_table_4_2_2_2 + select_table_4_2_2_3
        mycursor.execute(select_table_4_2_2)
        table_4_2_2 = mycursor.fetchall()

        #Find eighty percent cutoff
        cumulative_total_PRO = 0
        source_cut_off_no_PRO = 0
        for i in table_4_2_2:
            if cumulative_total_PRO <= c:
                cumulative_total_PRO += i[len(statement_period_half) + 1]
                source_cut_off_PRO = i[0]
                source_cut_off_no_PRO += 1

        #Find pool revenue line
        pool_rev_4_2_2_1 = '''SELECT Normalized_Source_9LC,'''
        pool_rev_4_2_2_2 = ""
        for d in statement_period_half:
            pool_rev_4_2_2_2 += ('''sum( CASE WHEN Statement_Period_Half_9LC = "{}" 
                                            AND Normalized_Source_9LC = "Pool Revenue"
                                            THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(d, d))
        pool_rev_4_2_2_3 = '''sum( CASE WHEN Normalized_Source_9LC = "Pool Revenue" 
                                      AND Statement_Period_Half_9LC >= "{}"
                                      THEN Adjusted_Royalty_SB ELSE "" END) AS `Total`
                                      FROM Master WHERE Normalized_Source_9LC = "Pool Revenue" 
                                      AND Statement_Period_Half_9LC <> ""
                                      AND Third_Party_9LC = "{}"
                                      GROUP BY Normalized_Source_9LC ORDER BY `Total` DESC'''.format(cut_off, a)
        pool_rev_4_2_2 = pool_rev_4_2_2_1 + pool_rev_4_2_2_2 + pool_rev_4_2_2_3
        mycursor.execute(pool_rev_4_2_2)
        initial_pool_revenue_4_2_2 = (mycursor.fetchall())
        pool_revenue_4_2_2 = []
        for i in initial_pool_revenue_4_2_2:
            for j in range(len(i)):
                pool_revenue_4_2_2.append(i[j])
        if len(pool_revenue_4_2_2) == 0:
            pool_revenue_4_2_2.append('Pool Revenue')
            pool_revenue_4_2_2_zeros = [0] * (len(statement_period_half) + 1)
            pool_revenue_4_2_2.extend(pool_revenue_4_2_2_zeros)

        #Add top eighty percent
        top_80_PRO_sources_4_1 = table_4_2_2[0:source_cut_off_no_PRO]
        PRO_total_row_no_4_1 += len(top_80_PRO_sources_4_1) + 6
        PRO_space_row_no_4_1 = PRO_total_row_no_4_1 + 1
        PRO_header_row_no_4_1 = PRO_total_row_no_4_1 - len(top_80_PRO_sources_4_1)
        PRO_table_rows_4_1 = ws4_1[PRO_header_row_no_4_1:PRO_space_row_no_4_1]
        for row, c in zip(PRO_table_rows_4_1, top_80_PRO_sources_4_1):
            for cell, d in zip(row, range(len(c))):
                cell.value = c[d]
                cell.style = 'number_style'
        for x in range(PRO_header_row_no_4_1, PRO_total_row_no_4_1):
            ws4_1.cell(row=x, column=1).style = 'name_style'
            ws4_1.cell(row=x, column=len(statement_period_half) + 2).style = 'total_style'
        ws4_1.cell(row=PRO_header_row_no_4_1 - 1, column=1).value = '{}'.format(a)
        ws4_1.cell(row=PRO_header_row_no_4_1 - 1, column=1).style = 'sub_header_style'
        for d in range(2, len(statement_period_half) + 3):
            ws4_1.cell(row=PRO_total_row_no_4_1, column=d).value = '=SUM({}{}:{}{})'.format(column_letters[d - 1],
                                                                                            PRO_header_row_no_4_1,
                                                                                            column_letters[d - 1],
                                                                                            PRO_total_row_no_4_1 - 1)
            ws4_1.cell(row=PRO_total_row_no_4_1, column=d).style = 'lined_total_style'
        for e in range(PRO_header_row_no_4_1, publisher_total_row_no_4_1):
            ws4_1.cell(row=e, column=1).style = 'name_style'
            ws4_1.cell(row=e, column=len(statement_period_half) + 2).style = 'total_style'
        ws4_1.cell(row=PRO_total_row_no_4_1, column=1).value = 'Subtotal (Top 80%)'
        ws4_1.cell(row=PRO_total_row_no_4_1, column=1).style = 'total_label_style'
        sub_total_row_list_4_1.append(PRO_total_row_no_4_1 + 3)

        #Add pool revenue line
        pool_rev_4_2_2_row = ws4_1[PRO_total_row_no_4_1 + 1]
        for cell, f in zip(pool_rev_4_2_2_row, range(len(pool_revenue_4_2_2))):
            cell.value = pool_revenue_4_2_2[f]
        for col in range(2, len(pool_revenue_4_2_2) + 1):
            ws4_1.cell(row=PRO_total_row_no_4_1 + 1, column=col).style = 'number_style'
        ws4_1.cell(row=PRO_total_row_no_4_1 + 1, column=len(statement_period_half) + 2).style = 'total_style'

        #Add other songs row
        other_totals_4_2_2 = ['Other Songs']
        other_songs_4_2_2 = table_4_2_2[source_cut_off_no_PRO:]
        other_songs_row_4_2_2 = ws4_1[PRO_total_row_no_4_1 + 2]
        half_total_4_2 = 0
        for g in range(1, len(statement_period_half) + 2):
            for h in other_songs_4_2_2:
                half_total_4_2 += h[g]
            other_totals_4_2_2.append(half_total_4_2)
            half_total_4_2 = 0
        for cell, i in zip(other_songs_row_4_2_2, range(len(other_totals_4_2_2))):
            cell.value = other_totals_4_2_2[i]
            cell.style = 'number_style'
        ws4_1.cell(row=PRO_total_row_no_4_1 + 2, column=1).style = 'name_style'
        ws4_1.cell(row=PRO_total_row_no_4_1 + 2, column=len(statement_period_half) + 2).style = 'total_style'

        #Add total row
        for j, k in zip(range(2, len(statement_period_half) + 3), column_letters[1:]):
            ws4_1.cell(row=PRO_total_row_no_4_1 + 3, column=j).value = "={}{}+{}{}+{}{}".format(k, PRO_total_row_no_4_1,
                                                                                            k,
                                                                                            PRO_total_row_no_4_1 + 1,
                                                                                            k,
                                                                                            PRO_total_row_no_4_1 + 2)
            ws4_1.cell(row=PRO_total_row_no_4_1 + 3, column=j).style = 'lined_total_style'
        ws4_1.cell(row=PRO_total_row_no_4_1 + 3, column=1).value = 'Total'
        ws4_1.cell(row=PRO_total_row_no_4_1 + 3, column=1).style = 'total_label_style'

    #Add grand total row
    grand_total_formula_initial_4_1 = '={}{}'
    grand_total_formula_4_1 = grand_total_formula_initial_4_1 + '+{}{}' * (len(sub_total_row_list_4_1) - 1)
    for col, f in zip(range(2, len(statement_period_half) + 3), column_letters[1:len(statement_period_half) + 2]):
        format_list_4_1 = []
        for g in sub_total_row_list_4_1:
            format_list_4_1.append(f)
            format_list_4_1.append(g)
        ws4_1.cell(row=PRO_total_row_no_4_1 + 5, column=col).value = grand_total_formula_4_1.format(*format_list_4_1)
        ws4_1.cell(row=PRO_total_row_no_4_1 + 5, column=col).style = 'lined_total_style'
    ws4_1.cell(row=PRO_total_row_no_4_1 + 5, column=1).value = 'Grand Total'
    ws4_1.cell(row=PRO_total_row_no_4_1 + 5, column=1).style = 'total_label_style'

    #Publisher and PRO labels
    ws4_1.merge_cells('A4:{}4'.format(column_letters[len(statement_period_half) + 1]))
    ws4_1.cell(row=4, column=1).value = 'Publishers'
    for w in range(1, len(statement_period_half) + 3):
        ws4_1.cell(row=4, column=w).style = 'publisher_label_style'

    ws4_1.merge_cells('A{}:{}{}'.format(publisher_total_row_no_4_1 + 5, column_letters[len(statement_period_half) + 1],
                                          publisher_total_row_no_4_1 + 5))
    ws4_1.cell(row=publisher_total_row_no_4_1 + 5, column=1).value = 'PROs'
    for x in range(1, len(statement_period_half) + 3):
        ws4_1.cell(row=publisher_total_row_no_4_1 + 5, column=x).style = 'publisher_label_style'

    #Column name list
    column_names_4_1 = ['Source']
    for c in statement_period_half:
        column_names_4_1.append(c)
    column_names_4_1.append('Total')

    #Add title row to worksheet
    ws4_1.merge_cells('B1:{}1'.format(column_letters[len(statement_period_half) + 1]))
    ws4_1.cell(row=1, column=1).value = '{}'.format(database[:-25])
    ws4_1.cell(row=1, column=1).style = 'title_style'
    ws4_1.cell(row=1, column=2).value = 'By Source'
    ws4_1.cell(row=1, column=2).style = 'title_style'

    #Add column names to worksheet
    column_name_row_4_1 = ws4_1[2]
    for cell, e in zip(column_name_row_4_1, range(len(column_names_4_1))):
        cell.value = column_names_4_1[e]
        cell.style = 'header_style'

    #Third row
    for col in range(1, len(column_names_4_1) + 1):
        ws4_1.cell(row=3, column=col).style = 'header_style'

#4.2. By Source By Quarter
    #Build publisher tables
    ws4_2.insert_cols(1, len(statement_period_quarter) + 2)
    if len(quarterly_publisher_list) == len(publisher_list):
        publisher_total_row_no_4_2 = 0
    else:
        publisher_total_row_no_4_2 = 1
        ws4_2.merge_cells('A5:{}5'.format(column_letters[len(statement_period_quarter)+1]))
        if len(not_quarterly_publisher_list) == 1:
            ws4_2.cell(row=5, column=1).value = 'Quarterly data not available for {}'.format(not_quarterly_publisher_list[0])
        if len(not_quarterly_publisher_list) == 2:
            ws4_2.cell(row=5, column=1).value = 'Quarterly data not available for {} and {}'.format(not_quarterly_publisher_list[0], not_quarterly_publisher_list[1])
        if len(not_quarterly_publisher_list) > 2:
            not_available_statement = 'Quarterly data not available for ' + '{}, '*(len(not_quarterly_publisher_list)-1) + 'and {}'
            ws4_2.cell(row=5, column=1).value = not_available_statement.format(*not_quarterly_publisher_list)
        ws4_2.cell(row=5, column=1).style = 'not_available_style'
    sub_total_row_list_4_2 = []
    for a, c in zip(quarterly_publisher_list, publisher_royalty_eighty):
        select_table_4_3_1_1 = '''SELECT Normalized_Source_9LC,'''
        select_table_4_3_1_2 = ""
        for b, d in zip(year_statement_list, quarter_statement_list):
            select_table_4_3_1_2 += ('''sum( CASE WHEN Year_Statement_9LC = "{}" AND Quarter_Statement_9LC = "{}"
                                               AND Normalized_Source_9LC <> "Pool Revenue"
                                               THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(b,d,b,d))
        select_table_4_3_1_3 = '''sum( CASE WHEN Statement_Period_Half_9LC >= "{}" 
                                          AND Normalized_Source_9LC <> "Pool Revenue" 
                                          THEN Adjusted_Royalty_SB ELSE "" END) AS `TOTAL`
                                          FROM Master WHERE Third_Party_9LC = "{}" AND Statement_Period_Half_9LC <> ""
                                          AND Normalized_Source_9LC <> "Pool Revenue"
                                          GROUP BY Third_Party_9LC, Normalized_Source_9LC 
                                          ORDER BY `Total` DESC'''.format(cut_off, a)
        select_table_4_3_1 = select_table_4_3_1_1 + select_table_4_3_1_2 + select_table_4_3_1_3
        mycursor.execute(select_table_4_3_1)
        table_4_3_1 = mycursor.fetchall()
        print(table_4_3_1)

        #Find eighty percent cutoff
        cumulative_total_publisher = 0
        source_cut_off_no_publisher = 0
        for i in table_4_3_1:
            if cumulative_total_publisher <= c:
                cumulative_total_publisher += i[len(statement_period_quarter) + 1]
                source_cut_off_publisher = i[0]
                source_cut_off_no_publisher += 1

        #Find pool revenue
        pool_rev_4_3_1_1 = '''SELECT Normalized_Source_9LC,'''
        pool_rev_4_3_1_2 = ""
        for d,e in zip(year_statement_list,quarter_statement_list):
            pool_rev_4_3_1_2 += ('''sum( CASE WHEN Year_Statement_9LC = "{}" AND Quarter_Statement_9LC = "{}"
                                        AND Normalized_Source_9LC = "Pool Revenue"
                                        THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(d,e,d,e))
        pool_rev_4_3_1_3 = '''sum( CASE WHEN Normalized_Source_9LC = "Pool Revenue" 
                                  AND Statement_Period_Half_9LC >= "{}"
                                  THEN Adjusted_Royalty_SB ELSE "" END) AS `Total`
                                  FROM Master WHERE Normalized_Source_9LC = "Pool Revenue" 
                                  AND Statement_Period_Half_9LC <> ""
                                  AND Third_Party_9LC = "{}"
                                  GROUP BY Normalized_Source_9LC ORDER BY `Total` DESC'''.format(cut_off, a)
        pool_rev_4_3_1 = pool_rev_4_3_1_1 + pool_rev_4_3_1_2 + pool_rev_4_3_1_3
        mycursor.execute(pool_rev_4_3_1)
        initial_pool_revenue_4_3_1 = (mycursor.fetchall())
        pool_revenue_4_3_1 = []
        for i in initial_pool_revenue_4_3_1:
            for j in range(len(i)):
                pool_revenue_4_3_1.append(i[j])
        if len(pool_revenue_4_3_1) == 0:
            pool_revenue_4_3_1.append('Pool Revenue')
            pool_revenue_4_3_1_zeros = [0] * (len(statement_period_quarter) + 1)
            pool_revenue_4_3_1.extend(pool_revenue_4_3_1_zeros)

        #Add top eighty percent
        top_80_publisher_sources_4_2 = table_4_3_1[0:source_cut_off_no_publisher]
        publisher_total_row_no_4_2 += len(top_80_publisher_sources_4_2) + 6
        publisher_space_row_no_4_2 = publisher_total_row_no_4_2 + 1
        publisher_header_row_no_4_2 = publisher_total_row_no_4_2 - len(top_80_publisher_sources_4_2)
        publisher_table_rows_4_2 = ws4_2[publisher_header_row_no_4_2:publisher_space_row_no_4_2]
        for row, c in zip(publisher_table_rows_4_2, top_80_publisher_sources_4_2):
            for cell, d in zip(row, range(len(c))):
                cell.value = c[d]
                cell.style = 'number_style'
        ws4_2.cell(row=publisher_header_row_no_4_2 - 1, column=1).value = '{}'.format(a)
        ws4_2.cell(row=publisher_header_row_no_4_2 - 1, column=1).style = 'sub_header_style'
        for d in range(2, len(statement_period_quarter) + 3):
            ws4_2.cell(row=publisher_total_row_no_4_2, column=d).value = '=SUM({}{}:{}{})'.format(
                column_letters[d - 1],
                publisher_header_row_no_4_2,
                column_letters[d - 1],
                publisher_total_row_no_4_2 - 1)
            ws4_2.cell(row=publisher_total_row_no_4_2, column=d).style = 'lined_total_style'
        for e in range(publisher_header_row_no_4_2, publisher_total_row_no_4_2):
            ws4_2.cell(row=e, column=1).style = 'name_style'
            ws4_2.cell(row=e, column=len(statement_period_quarter) + 2).style = 'total_style'
        ws4_2.cell(row=publisher_total_row_no_4_2, column=1).value = 'Subtotal (Top 80%)'
        ws4_2.cell(row=publisher_total_row_no_4_2, column=1).style = 'total_label_style'
        sub_total_row_list_4_2.append(publisher_total_row_no_4_2 + 3)

        #Add pool revenue line
        pool_rev_4_3_1_row = ws4_2[publisher_total_row_no_4_2 + 1]
        for cell, f in zip(pool_rev_4_3_1_row, range(len(pool_revenue_4_3_1))):
            cell.value = pool_revenue_4_3_1[f]
        for col in range(2, len(pool_revenue_4_3_1) + 1):
            ws4_2.cell(row=publisher_total_row_no_4_2 + 1, column=col).style = 'number_style'
        ws4_2.cell(row=publisher_total_row_no_4_2 + 1, column=len(statement_period_quarter) + 2).style = 'total_style'

        #Add other sources row
        other_totals_4_3_1 = ['Other Songs']
        other_sources_4_3_1 = table_4_3_1[source_cut_off_no_publisher:]
        other_sources_row_4_3_1 = ws4_2[publisher_total_row_no_4_2 + 2]
        quarter_total_4_1 = 0
        for g in range(1, len(statement_period_quarter) + 2):
            for h in other_sources_4_3_1:
                quarter_total_4_1 += h[g]
            other_totals_4_3_1.append(quarter_total_4_1)
            quarter_total_4_1 = 0
        for cell, i in zip(other_sources_row_4_3_1, range(len(other_totals_4_3_1))):
            cell.value = other_totals_4_3_1[i]
            cell.style = 'number_style'
        ws4_2.cell(row=publisher_total_row_no_4_2 + 2, column=1).style = 'name_style'
        ws4_2.cell(row=publisher_total_row_no_4_2 + 2, column=len(statement_period_quarter) + 2).style = 'total_style'

        #Add total row
        for j, k in zip(range(2, len(statement_period_quarter) + 3), column_letters[1:]):
            ws4_2.cell(row=publisher_total_row_no_4_2 + 3, column=j).value = "={}{}+{}{}+{}{}".format(k,
                                                                                        publisher_total_row_no_4_2, k,
                                                                                        publisher_total_row_no_4_2 + 1, k,
                                                                                        publisher_total_row_no_4_2 + 2)
            ws4_2.cell(row=publisher_total_row_no_4_2 + 3, column=j).style = 'lined_total_style'
            ws4_2.cell(row=publisher_total_row_no_4_2 + 3, column=1).value = 'Total'
            ws4_2.cell(row=publisher_total_row_no_4_2 + 3, column=1).style = 'total_label_style'

    #Build PRO tables
    if len(quarterly_PRO_list) == len(PRO_list):
        PRO_total_row_no_4_2 = publisher_total_row_no_4_2 + 1
    else:
        PRO_total_row_no_4_2 = publisher_total_row_no_4_2 + 2
        ws4_2.merge_cells('A{}:{}{}'.format(PRO_total_row_no_4_2 + 4,column_letters[len(statement_period_quarter)+1],PRO_total_row_no_4_2 + 4))
        if len(not_quarterly_PRO_list) == 1:
            ws4_2.cell(row=PRO_total_row_no_4_2 + 4, column=1).value = 'Quarterly data not available for {}'.format(not_quarterly_PRO_list[0])
        if len(not_quarterly_PRO_list) == 2:
            ws4_2.cell(row=PRO_total_row_no_4_2 + 4, column=1).value = 'Quarterly data not available for {} and {}'.format(not_quarterly_PRO_list[0], not_quarterly_PRO_list[1])
        if len(not_quarterly_PRO_list) > 2:
            not_available_statement = 'Quarterly data not available for ' + '{}, '*(len(not_quarterly_PRO_list)-1) + 'and {}'
            ws4_2.cell(row=PRO_total_row_no_4_2 + 4, column=1).value = not_available_statement.format(*not_quarterly_PRO_list)
        ws4_2.cell(row=PRO_total_row_no_4_2 + 4, column=1).style = 'not_available_style'
    for a, c in zip(quarterly_PRO_list, PRO_royalty_eighty):
        select_table_4_3_2_1 = '''SELECT Normalized_Source_9LC,'''
        select_table_4_3_2_2 = ""
        for b,d in zip(year_statement_list, quarter_statement_list):
            select_table_4_3_2_2 += ('''sum( CASE WHEN Year_Statement_9LC = "{}" AND Quarter_Statement_9LC = "{}"
                                                  AND Normalized_Source_9LC <> "Pool Revenue"
                                                  THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(b,d,b,d))
        select_table_4_3_2_3 = '''sum( CASE WHEN Statement_Period_Half_9LC >= "{}" 
                                            THEN Adjusted_Royalty_SB ELSE "" END) AS `TOTAL`
                                            FROM Master WHERE Third_Party_9LC = "{}" AND Statement_Period_Half_9LC <> ""
                                            AND Normalized_Source_9LC <> "Pool Revenue"
                                            GROUP BY Third_Party_9LC, Normalized_Source_9LC 
                                            ORDER BY `Total` DESC'''.format(cut_off, a)
        select_table_4_3_2 = select_table_4_3_2_1 + select_table_4_3_2_2 + select_table_4_3_2_3
        mycursor.execute(select_table_4_3_2)
        table_4_3_2 = mycursor.fetchall()

        #Find eighty percent cutoff
        cumulative_total_PRO = 0
        source_cut_off_no_PRO = 0
        for i in table_4_3_2:
            if cumulative_total_PRO <= c:
                cumulative_total_PRO += i[len(statement_period_quarter) + 1]
                source_cut_off_PRO = i[0]
                source_cut_off_no_PRO += 1

        #Find pool revenue line
        pool_rev_4_3_2_1 = '''SELECT Normalized_Source_9LC,'''
        pool_rev_4_3_2_2 = ""
        for d, e in zip(year_statement_list, quarter_statement_list):
            pool_rev_4_3_2_2 += ('''sum( CASE WHEN Year_Statement_9LC = "{}" AND Quarter_Statement_9LC = "{}" 
                                                AND Normalized_Source_9LC = "Pool Revenue"
                                                THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(d,e,d,e))
        pool_rev_4_3_2_3 = '''sum( CASE WHEN Normalized_Source_9LC = "Pool Revenue" 
                                          AND Statement_Period_Half_9LC >= "{}"
                                          THEN Adjusted_Royalty_SB ELSE "" END) AS `Total`
                                          FROM Master WHERE Normalized_Source_9LC = "Pool Revenue" 
                                          AND Statement_Period_Half_9LC <> ""
                                          AND Third_Party_9LC = "{}"
                                          GROUP BY Normalized_Source_9LC ORDER BY `Total` DESC'''.format(cut_off, a)
        pool_rev_4_3_2 = pool_rev_4_3_2_1 + pool_rev_4_3_2_2 + pool_rev_4_3_2_3
        mycursor.execute(pool_rev_4_3_2)
        initial_pool_revenue_4_3_2 = (mycursor.fetchall())
        pool_revenue_4_3_2 = []
        for i in initial_pool_revenue_4_3_2:
            for j in range(len(i)):
                pool_revenue_4_3_2.append(i[j])
        if len(pool_revenue_4_3_2) == 0:
            pool_revenue_4_3_2.append('Pool Revenue')
            pool_revenue_4_3_2_zeros = [0] * (len(statement_period_quarter) + 1)
            pool_revenue_4_3_2.extend(pool_revenue_4_3_2_zeros)

        #Add top eighty percent
        top_80_PRO_sources_4_2 = table_4_3_2[0:source_cut_off_no_PRO]
        PRO_total_row_no_4_2 += len(top_80_PRO_sources_4_2) + 6
        PRO_space_row_no_4_2 = PRO_total_row_no_4_2 + 1
        PRO_header_row_no_4_2 = PRO_total_row_no_4_2 - len(top_80_PRO_sources_4_2)
        PRO_table_rows_4_2 = ws4_2[PRO_header_row_no_4_2:PRO_space_row_no_4_2]
        for row, c in zip(PRO_table_rows_4_2, top_80_PRO_sources_4_2):
            for cell, d in zip(row, range(len(c))):
                cell.value = c[d]
                cell.style = 'number_style'
        for x in range(PRO_header_row_no_4_2, PRO_total_row_no_4_2):
            ws4_2.cell(row=x, column=1).style = 'name_style'
            ws4_2.cell(row=x, column=len(statement_period_quarter) + 2).style = 'total_style'
        ws4_2.cell(row=PRO_header_row_no_4_2 - 1, column=1).value = '{}'.format(a)
        ws4_2.cell(row=PRO_header_row_no_4_2 - 1, column=1).style = 'sub_header_style'
        for d in range(2, len(statement_period_quarter) + 3):
            ws4_2.cell(row=PRO_total_row_no_4_2, column=d).value = '=SUM({}{}:{}{})'.format(column_letters[d - 1],
                                                                                                PRO_header_row_no_4_2,
                                                                                                column_letters[d - 1],
                                                                                                PRO_total_row_no_4_2 - 1)
            ws4_2.cell(row=PRO_total_row_no_4_2, column=d).style = 'lined_total_style'
        for e in range(PRO_header_row_no_4_2, publisher_total_row_no_4_2):
            ws4_2.cell(row=e, column=1).style = 'name_style'
            ws4_2.cell(row=e, column=len(statement_period_quarter) + 2).style = 'total_style'
        ws4_2.cell(row=PRO_total_row_no_4_2, column=1).value = 'Subtotal (Top 80%)'
        ws4_2.cell(row=PRO_total_row_no_4_2, column=1).style = 'total_label_style'
        sub_total_row_list_4_2.append(PRO_total_row_no_4_2 + 3)

        #Add pool revenue line
        pool_rev_4_3_2_row = ws4_2[PRO_total_row_no_4_2 + 1]
        for cell, f in zip(pool_rev_4_3_2_row, range(len(pool_revenue_4_3_2))):
            cell.value = pool_revenue_4_3_2[f]
        for col in range(2, len(pool_revenue_4_3_2) + 1):
            ws4_2.cell(row=PRO_total_row_no_4_2 + 1, column=col).style = 'number_style'
        ws4_2.cell(row=PRO_total_row_no_4_2 + 1, column=len(statement_period_quarter) + 2).style = 'total_style'

        #Add other songs row
        other_totals_4_3_2 = ['Other Songs']
        other_songs_4_3_2 = table_4_3_2[source_cut_off_no_PRO:]
        other_songs_row_4_3_2 = ws4_2[PRO_total_row_no_4_2 + 2]
        quarter_total_4_2 = 0
        for g in range(1, len(statement_period_quarter) + 2):
            for h in other_songs_4_3_2:
                quarter_total_4_2 += h[g]
            other_totals_4_3_2.append(quarter_total_4_2)
            quarter_total_4_2 = 0
        for cell, i in zip(other_songs_row_4_3_2, range(len(other_totals_4_3_2))):
            cell.value = other_totals_4_3_2[i]
            cell.style = 'number_style'
        ws4_2.cell(row=PRO_total_row_no_4_2 + 2, column=1).style = 'name_style'
        ws4_2.cell(row=PRO_total_row_no_4_2 + 2, column=len(statement_period_quarter) + 2).style = 'total_style'

        #Add total row
        for j, k in zip(range(2, len(statement_period_quarter) + 3), column_letters[1:]):
            ws4_2.cell(row=PRO_total_row_no_4_2 + 3, column=j).value = "={}{}+{}{}+{}{}".format(k, PRO_total_row_no_4_2,
                                                                                                k, PRO_total_row_no_4_2 + 1,
                                                                                                k, PRO_total_row_no_4_2 + 2)
            ws4_2.cell(row=PRO_total_row_no_4_2 + 3, column=j).style = 'lined_total_style'
        ws4_2.cell(row=PRO_total_row_no_4_2 + 3, column=1).value = 'Total'
        ws4_2.cell(row=PRO_total_row_no_4_2 + 3, column=1).style = 'total_label_style'

    #Add grand total row
    grand_total_formula_initial_4_2 = '={}{}'
    grand_total_formula_4_2 = grand_total_formula_initial_4_2 + '+{}{}' * (len(sub_total_row_list_4_2) - 1)
    for col, f in zip(range(2, len(statement_period_quarter) + 3), column_letters[1:len(statement_period_quarter) + 2]):
        format_list_4_2 = []
        for g in sub_total_row_list_4_2:
            format_list_4_2.append(f)
            format_list_4_2.append(g)
        ws4_2.cell(row=PRO_total_row_no_4_2 + 5, column=col).value = grand_total_formula_4_2.format(*format_list_4_2)
        ws4_2.cell(row=PRO_total_row_no_4_2 + 5, column=col).style = 'lined_total_style'
    ws4_2.cell(row=PRO_total_row_no_4_2 + 5, column=1).value = 'Grand Total'
    ws4_2.cell(row=PRO_total_row_no_4_2 + 5, column=1).style = 'total_label_style'

    #Publisher and PRO labels
    ws4_2.merge_cells('A4:{}4'.format(column_letters[len(statement_period_quarter) + 1]))
    ws4_2.cell(row=4, column=1).value = 'Publishers'
    for w in range(1, len(statement_period_quarter) + 3):
        ws4_2.cell(row=4, column=w).style = 'publisher_label_style'

    ws4_2.merge_cells('A{}:{}{}'.format(publisher_total_row_no_4_2 + 5, column_letters[len(statement_period_quarter) + 1],
                                                publisher_total_row_no_4_2 + 5))
    ws4_2.cell(row=publisher_total_row_no_4_2 + 5, column=1).value = 'PROs'
    for x in range(1, len(statement_period_quarter) + 3):
        ws4_2.cell(row=publisher_total_row_no_4_2 + 5, column=x).style = 'publisher_label_style'

    #Column name list
    column_names_4_2 = ['Source']
    for c in statement_period_quarter:
        column_names_4_2.append(c)
    column_names_4_2.append('Total')

    #Add title row to worksheet
    ws4_2.merge_cells('B1:{}1'.format(column_letters[len(statement_period_quarter) + 1]))
    ws4_2.cell(row=1, column=1).value = '{}'.format(database[:-25])
    ws4_2.cell(row=1, column=1).style = 'title_style'
    ws4_2.cell(row=1, column=2).value = 'By Source'
    ws4_2.cell(row=1, column=2).style = 'title_style'

    #Add column names to worksheet
    column_name_row_4_2 = ws4_2[2]
    for cell, e in zip(column_name_row_4_2, range(len(column_names_4_2))):
        cell.value = column_names_4_2[e]
        cell.style = 'header_style'

    #Third row
    for col in range(1, len(column_names_4_2) + 1):
        ws4_2.cell(row=3, column=col).style = 'header_style'

#4.3. By Source - Top Earner Breakdown
    #Top eighty song list
    top_80_song_list = song_list[:song_cut_off_no]

    if len(quarterly_third_party) == len(third_party_list):
        #Build song tables
        song_total_row_no_4_3 = -1
        sub_total_row_list_4_3 = []
        ws4_3.insert_cols(1, len(statement_period_quarter) + 3)
        for a in top_80_song_list:
            select_table_4_4_1_1 = '''SELECT Normalized_Source_9LC,'''
            select_table_4_4_1_2 = ""
            for b, c in zip(year_statement_list,quarter_statement_list):
                select_table_4_4_1_2 += ('''sum( CASE WHEN Year_Statement_9LC = "{}" AND Quarter_Statement_9LC = "{}"
                                                        THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(b,c,b,c))
            select_table_4_4_1_3 = '''sum( CASE WHEN Statement_Period_Half_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END)
                                                      AS `TOTAL`
                                                      FROM Master WHERE Song_Name_9LC = "{}" AND Statement_Period_Half_9LC <> ""
                                                      AND Normalized_Source_9LC <> "Pool Revenue"
                                                      GROUP BY Song_Name_9LC, Normalized_Source_9LC 
                                                      ORDER BY `Total` DESC'''.format(cut_off, a)
            select_table_4_4_1 = select_table_4_4_1_1 + select_table_4_4_1_2 + select_table_4_4_1_3
            mycursor.execute(select_table_4_4_1)
            table_4_4_1 = mycursor.fetchall()

            pool_rev_4_4_1_1 = '''SELECT Normalized_Source_9LC,'''
            pool_rev_4_4_1_2 = ""
            for b, c in zip(year_statement_list, quarter_statement_list):
                pool_rev_4_4_1_2 += ('''sum( CASE WHEN Year_Statement_9LC = "{}" AND Quarter_Statement_9LC = "{}"
                                            THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(b,c,b,c))
            pool_rev_4_4_1_3 = '''sum( CASE WHEN Statement_Period_Half_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END)
                                      AS `TOTAL`
                                      FROM Master WHERE Song_Name_9LC = "{}" AND Statement_Period_Half_9LC <> ""
                                      AND Normalized_Source_9LC = "Pool Revenue"
                                      GROUP BY Song_Name_9LC, Normalized_Source_9LC
                                      ORDER BY `Total` DESC'''.format(cut_off, a)
            pool_rev_4_4_1 = pool_rev_4_4_1_1 + pool_rev_4_4_1_2 + pool_rev_4_4_1_3
            mycursor.execute(pool_rev_4_4_1)
            initial_pool_revenue_4_4_1 = (mycursor.fetchall())
            pool_revenue_4_4_1 = []
            for i in initial_pool_revenue_4_4_1:
                for j in range(len(i)):
                    pool_revenue_4_4_1.append(i[j])
            if len(pool_revenue_4_4_1) == 0:
                pool_revenue_4_4_1.append('Pool Revenue')
                pool_revenue_4_4_1_zeros = [0] * (len(statement_period_quarter) + 1)
                pool_revenue_4_4_1.extend(pool_revenue_4_4_1_zeros)

            #Total per song
            find_song_total = '''SELECT Song_Name_9LC, 
                                     sum( CASE WHEN Statement_Period_Half_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END) 
                                     AS `TOTAL`
                                     FROM Master WHERE Song_Name_9LC = "{}" AND Statement_Period_Half_9LC <> ""
                                     GROUP BY Song_Name_9LC'''.format(cut_off, a)
            mycursor.execute(find_song_total)
            song_total = [i[1] for i in mycursor.fetchall()]
            song_80 = 0.8 * song_total[0]

            #Find eighty percent cutoff
            cumulative_total_song = 0
            cut_off_song_no = 0
            for i in table_4_4_1:
                if cumulative_total_song <= song_80:
                    cumulative_total_song += i[len(statement_period_quarter) + 1]
                    cut_off_song = i[0]
                    cut_off_song_no += 1
            top_80_song_sources = table_4_4_1[:cut_off_song_no]

            #Find other sources total
            other_totals_4_4_1 = ['Other Sources']
            other_sources_4_4_1 = table_4_4_1[cut_off_song_no:]
            song_source_total = 0
            for g in range(1, len(statement_period_quarter) + 2):
                for h in other_sources_4_4_1:
                    song_source_total += h[g]
                other_totals_4_4_1.append(song_source_total)
                song_source_total = 0

            #Add tables to sheet
            song_total_row_no_4_3 += len(top_80_song_sources) + 6
            song_space_row_no_4_3 = song_total_row_no_4_3 + 1
            song_header_row_no_4_3 = song_total_row_no_4_3 - len(top_80_song_sources)
            song_table_rows_4_3 = ws4_3[song_header_row_no_4_3:song_space_row_no_4_3]
            for row, c in zip(song_table_rows_4_3, top_80_song_sources):
                for cell, d in zip(row, range(len(c))):
                    cell.value = c[d]
                    cell.style = 'number_style'
            for d in range(2, len(statement_period_quarter) + 3):
                ws4_3.cell(row=song_total_row_no_4_3, column=d).value = '=SUM({}{}:{}{})'.format(column_letters[d - 1],
                                                                                                 song_header_row_no_4_3,
                                                                                                 column_letters[d - 1],
                                                                                                 song_total_row_no_4_3 - 1)
                ws4_3.cell(row=song_total_row_no_4_3, column=d).style = 'lined_total_style'
            for e in range(song_header_row_no_4_3, song_total_row_no_4_3):
                ws4_3.cell(row=e, column=1).style = 'name_style'
                ws4_3.cell(row=e, column=len(statement_period_quarter) + 2).style = 'total_style'
            ws4_3.cell(row=song_header_row_no_4_3 - 1, column=1).value = "{}".format(a)
            ws4_3.cell(row=song_header_row_no_4_3 - 1, column=1).style = 'sub_header_style'
            ws4_3.cell(row=song_total_row_no_4_3, column=1).value = 'Subtotal (Top 80%)'
            ws4_3.cell(row=song_total_row_no_4_3, column=1).style = 'total_label_style'
            sub_total_row_list_4_3.append(song_total_row_no_4_3 + 3)
            for f, g in zip(range(1, len(statement_period_quarter) + 3), pool_revenue_4_4_1):
                ws4_3.cell(row=song_total_row_no_4_3 + 1, column=f).value = g
                ws4_3.cell(row=song_total_row_no_4_3 + 1, column=f).style = 'number_style'
            ws4_3.cell(row=song_total_row_no_4_3 + 1, column=len(statement_period_quarter) + 2).style = 'total_style'
            for h, i in zip(range(1, len(statement_period_quarter) + 3), other_totals_4_4_1):
                ws4_3.cell(row=song_total_row_no_4_3 + 2, column=h).value = i
                ws4_3.cell(row=song_total_row_no_4_3 + 2, column=h).style = 'number_style'
            ws4_3.cell(row=song_total_row_no_4_3 + 1, column=1).style = 'name_style'
            ws4_3.cell(row=song_total_row_no_4_3 + 2, column=1).style = 'name_style'
            ws4_3.cell(row=song_total_row_no_4_3 + 2, column=len(statement_period_quarter) + 2).style = 'total_style'
            for j in range(2, len(statement_period_quarter) + 3):
                ws4_3.cell(row=song_total_row_no_4_3 + 3, column=j).value = '=SUM({}{}:{}{})'.format(
                    column_letters[j - 1],
                    song_total_row_no_4_3,
                    column_letters[j - 1],
                    song_total_row_no_4_3 + 2)
                ws4_3.cell(row=song_total_row_no_4_3 + 3, column=j).style = 'lined_total_style'
            ws4_3.cell(row=song_total_row_no_4_3 + 3, column=1).value = 'Total'
            ws4_3.cell(row=song_total_row_no_4_3 + 3, column=1).style = 'total_label_style'

        #Column name list
        column_names_4_3 = ['Income Type']
        for c in statement_period_quarter:
            column_names_4_3.append(c)
        column_names_4_3.append('Total')

        #Add title row to worksheet
        ws4_3.merge_cells('B1:{}1'.format(column_letters[len(statement_period_quarter) + 1]))
        ws4_3.cell(row=1, column=1).value = '{}'.format(database[:-25])
        ws4_3.cell(row=1, column=1).style = 'title_style'
        ws4_3.cell(row=1, column=2).value = 'By Income Type'
        ws4_3.cell(row=1, column=2).style = 'title_style'

        #Add column names to worksheet
        column_name_row_4_3 = ws4_3[2]
        for cell, e in zip(column_name_row_4_3, range(len(column_names_4_3))):
            cell.value = column_names_4_3[e]
            cell.style = 'header_style'

        #Third row
        for col in range(1, len(column_names_4_3) + 1):
            ws4_3.cell(row=3, column=col).style = 'header_style'

        #Add grand total row
        grand_total_formula_initial_4_3 = '={}{}'
        grand_total_formula_4_3 = grand_total_formula_initial_4_3 + '+{}{}' * (len(sub_total_row_list_4_3) - 1)
        for col, f in zip(range(2, len(statement_period_quarter) + 4),
                          column_letters[1:len(statement_period_quarter) + 2]):
            format_list_4_3 = []
            for g in sub_total_row_list_4_3:
                format_list_4_3.append(f)
                format_list_4_3.append(g)
            ws4_3.cell(row=song_total_row_no_4_3 + 5, column=col).value = grand_total_formula_4_3.format(
                *format_list_4_3)
            ws4_3.cell(row=song_total_row_no_4_3 + 5, column=col).style = 'lined_total_style'
        ws4_3.cell(row=song_total_row_no_4_3 + 5, column=1).value = 'Grand Total'
        ws4_3.cell(row=song_total_row_no_4_3 + 5, column=1).style = 'total_label_style'


    else:
        #Build song tables
        song_total_row_no_4_3 = -1
        sub_total_row_list_4_3 = []
        ws4_3.insert_cols(1, len(statement_period_half) + 3)
        for a in top_80_song_list:
            select_table_4_4_1_1 = '''SELECT Normalized_Source_9LC,'''
            select_table_4_4_1_2 = ""
            for b in statement_period_half:
                select_table_4_4_1_2 += ('''sum( CASE WHEN Statement_Period_Half_9LC = "{}"
                                                    THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(b, b))
            select_table_4_4_1_3 = '''sum( CASE WHEN Statement_Period_Half_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END)
                                                  AS `TOTAL`
                                                  FROM Master WHERE Song_Name_9LC = "{}" AND Statement_Period_Half_9LC <> ""
                                                  AND Normalized_Source_9LC <> "Pool Revenue"
                                                  GROUP BY Song_Name_9LC, Normalized_Source_9LC 
                                                  ORDER BY `Total` DESC'''.format(cut_off, a)
            select_table_4_4_1 = select_table_4_4_1_1 + select_table_4_4_1_2 + select_table_4_4_1_3
            mycursor.execute(select_table_4_4_1)
            table_4_4_1 = mycursor.fetchall()

            pool_rev_4_4_1_1 = '''SELECT Normalized_Source_9LC,'''
            pool_rev_4_4_1_2 = ""
            for b in statement_period_half:
                pool_rev_4_4_1_2 += ('''sum( CASE WHEN Statement_Period_9LC = "{}"
                                        THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(b,b))
            pool_rev_4_4_1_3 = '''sum( CASE WHEN Statement_Period_Half_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END)
                                  AS `TOTAL`
                                  FROM Master WHERE Song_Name_9LC = "{}" AND Statement_Period_Half_9LC <> ""
                                  AND Normalized_Source_9LC = "Pool Revenue"
                                  GROUP BY Song_Name_9LC, Normalized_Source_9LC
                                  ORDER BY `Total` DESC'''.format(cut_off, a)
            pool_rev_4_4_1 = pool_rev_4_4_1_1 + pool_rev_4_4_1_2 + pool_rev_4_4_1_3
            mycursor.execute(pool_rev_4_4_1)
            initial_pool_revenue_4_4_1 = (mycursor.fetchall())
            pool_revenue_4_4_1 = []
            for i in initial_pool_revenue_4_4_1:
                for j in range(len(i)):
                    pool_revenue_4_4_1.append(i[j])
            if len(pool_revenue_4_4_1) == 0:
                pool_revenue_4_4_1.append('Pool Revenue')
                pool_revenue_4_4_1_zeros = [0] * (len(statement_period_half) + 1)
                pool_revenue_4_4_1.extend(pool_revenue_4_4_1_zeros)

            #Total per song
            find_song_total = '''SELECT Song_Name_9LC, 
                                 sum( CASE WHEN Statement_Period_Half_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END) 
                                 AS `TOTAL`
                                 FROM Master WHERE Song_Name_9LC = "{}" AND Statement_Period_Half_9LC <> ""
                                 GROUP BY Song_Name_9LC'''.format(cut_off, a)
            mycursor.execute(find_song_total)
            song_total = [i[1] for i in mycursor.fetchall()]
            song_80 = 0.8*song_total[0]

            #Find eighty percent cutoff
            cumulative_total_song = 0
            cut_off_song_no = 0
            for i in table_4_4_1:
                if cumulative_total_song <= song_80:
                    cumulative_total_song += i[len(statement_period_half) + 1]
                    cut_off_song = i[0]
                    cut_off_song_no += 1
            top_80_song_sources = table_4_4_1[:cut_off_song_no]

            #Find other sources total
            other_totals_4_4_1 = ['Other Sources']
            other_sources_4_4_1 = table_4_4_1[cut_off_song_no:]
            song_source_total = 0
            for g in range(1, len(statement_period_half) + 2):
                for h in other_sources_4_4_1:
                    song_source_total += h[g]
                other_totals_4_4_1.append(song_source_total)
                song_source_total = 0

            #Add tables to sheet
            song_total_row_no_4_3 += len(top_80_song_sources) + 6
            song_space_row_no_4_3 = song_total_row_no_4_3 + 1
            song_header_row_no_4_3 = song_total_row_no_4_3 - len(top_80_song_sources)
            song_table_rows_4_3 = ws4_3[song_header_row_no_4_3:song_space_row_no_4_3]
            for row, c in zip(song_table_rows_4_3, top_80_song_sources):
                for cell, d in zip(row, range(len(c))):
                    cell.value = c[d]
                    cell.style = 'number_style'
            for d in range(2, len(statement_period_half) + 3):
                ws4_3.cell(row=song_total_row_no_4_3, column=d).value = '=SUM({}{}:{}{})'.format(column_letters[d - 1],
                                                                                                song_header_row_no_4_3,
                                                                                                column_letters[d - 1],
                                                                                                song_total_row_no_4_3 - 1)
                ws4_3.cell(row=song_total_row_no_4_3, column=d).style = 'lined_total_style'
            for e in range(song_header_row_no_4_3, song_total_row_no_4_3):
                ws4_3.cell(row=e, column=1).style = 'name_style'
                ws4_3.cell(row=e, column=len(statement_period_half) + 2).style = 'total_style'
            ws4_3.cell(row=song_header_row_no_4_3 - 1, column=1).value = "{}".format(a)
            ws4_3.cell(row=song_header_row_no_4_3 - 1, column=1).style = 'sub_header_style'
            ws4_3.cell(row=song_total_row_no_4_3, column=1).value = 'Subtotal (Top 80%)'
            ws4_3.cell(row=song_total_row_no_4_3, column=1).style = 'total_label_style'
            sub_total_row_list_4_3.append(song_total_row_no_4_3+3)
            for f, g in zip(range(1,len(statement_period_half)+3), pool_revenue_4_4_1):
                ws4_3.cell(row=song_total_row_no_4_3+1, column=f).value = g
                ws4_3.cell(row=song_total_row_no_4_3+1, column=f).style = 'number_style'
            ws4_3.cell(row=song_total_row_no_4_3+1, column=len(statement_period_half)+2).style = 'total_style'
            for h, i in zip(range(1, len(statement_period_half)+3), other_totals_4_4_1):
                ws4_3.cell(row=song_total_row_no_4_3+2, column=h).value = i
                ws4_3.cell(row=song_total_row_no_4_3+2, column=h).style = 'number_style'
            ws4_3.cell(row=song_total_row_no_4_3+1, column=1).style = 'name_style'
            ws4_3.cell(row=song_total_row_no_4_3+2, column=1).style = 'name_style'
            ws4_3.cell(row=song_total_row_no_4_3+2, column=len(statement_period_half)+2).style = 'total_style'
            for j in range(2, len(statement_period_half) + 3):
                ws4_3.cell(row=song_total_row_no_4_3+3, column=j).value = '=SUM({}{}:{}{})'.format(column_letters[j-1],
                                                                                                   song_total_row_no_4_3,
                                                                                                   column_letters[j-1],
                                                                                                   song_total_row_no_4_3+2)
                ws4_3.cell(row=song_total_row_no_4_3+3, column=j).style = 'lined_total_style'
            ws4_3.cell(row=song_total_row_no_4_3+3, column=1).value = 'Total'
            ws4_3.cell(row=song_total_row_no_4_3+3, column=1).style = 'total_label_style'


        #Column name list
        column_names_4_3 = ['Income Type']
        for c in statement_period_half:
            column_names_4_3.append(c)
        column_names_4_3.append('Total')

        #Add title row to worksheet
        ws4_3.merge_cells('B1:{}1'.format(column_letters[len(statement_period_half) + 1]))
        ws4_3.cell(row=1, column=1).value = '{}'.format(database[:-25])
        ws4_3.cell(row=1, column=1).style = 'title_style'
        ws4_3.cell(row=1, column=2).value = 'By Income Type'
        ws4_3.cell(row=1, column=2).style = 'title_style'

        #Add column names to worksheet
        column_name_row_4_3 = ws4_3[2]
        for cell, e in zip(column_name_row_4_3, range(len(column_names_4_3))):
            cell.value = column_names_4_3[e]
            cell.style = 'header_style'

        #Third row
        for col in range(1, len(column_names_4_3) + 1):
            ws4_3.cell(row=3, column=col).style = 'header_style'

        #Add grand total row
        grand_total_formula_initial_4_3 = '={}{}'
        grand_total_formula_4_3 = grand_total_formula_initial_4_3 + '+{}{}' * (len(sub_total_row_list_4_3) - 1)
        for col, f in zip(range(2, len(statement_period_half) + 4),
                            column_letters[1:len(statement_period_half) + 2]):
            format_list_4_3 = []
            for g in sub_total_row_list_4_3:
                format_list_4_3.append(f)
                format_list_4_3.append(g)
            ws4_3.cell(row=song_total_row_no_4_3 + 5, column=col).value = grand_total_formula_4_3.format(*format_list_4_3)
            ws4_3.cell(row=song_total_row_no_4_3 + 5, column=col).style = 'lined_total_style'
        ws4_3.cell(row=song_total_row_no_4_3 + 5, column=1).value = 'Grand Total'
        ws4_3.cell(row=song_total_row_no_4_3 + 5, column=1).style = 'total_label_style'



#Save workbook
    return(wb.save(filename))

#summary('DJ Battlecat_616c91a05f278b92afbaa5ae')
#summary('Kuya_615e052aecd455245e6954f2')
#summary('Jacknife Lee_61493586aa6eadb6793c9903')
#summary('Arden Altino_6103281d742b855f5c9fb6a7')
#summary('Josh Miller_616703cf0b2bca63a2d130dc')
#summary('Jerry Reed_6117d18ab51209aec72204af')
'Bob Morrison New_618ae6a6ee70bed89b394b82'
