import flask
import mysql.connector
import pathlib
import re
from datetime import date
from flask import request, jsonify
from flask_cors import cross_origin
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, NamedStyle
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties
from openpyxl.chart.axis import DateAxis
from pandas_utils.pandas_cursor import pandas_cursor


def data_periods(database, df):
  mycursor = pandas_cursor(df=df)

  data_list = ['Half']

  #Publisher list
  find_publishers = '''SELECT Third_Party_9LC FROM Master WHERE Third_Party_9LC NOT IN ('ASCAP', 'BMI', 'PRS', 'MCPS', 
                           'SoundExchange', 'SOCAN', 'APRA', 'AMCOS', 'SESAC', 'AMRA', 'GEMA', 'SUISA', 'ZAiKS') AND Third_Party_9LC <> "Vydia" AND Third_Party_9LC <> "STEM" GROUP BY Third_Party_9LC'''
  mycursor.execute(find_publishers)
  publisher_list = [i[0] for i in mycursor.fetchall()]
  publisher_string = ', '.join('"{}"'.format(str(x)) for x in publisher_list)

  #PRO list
  find_PRO = '''SELECT Third_Party_9LC FROM Master WHERE Third_Party_9LC IN ('ASCAP', 'BMI', 'PRS', 'MCPS', 
                             'SoundExchange', 'SOCAN', 'APRA', 'AMCOS', 'SESAC', 'AMRA', 'GEMA', 'SUISA', 'ZAiKS') GROUP BY Third_Party_9LC'''
  mycursor.execute(find_PRO)
  PRO_list = [i[0] for i in mycursor.fetchall()]
  PRO_string = ', '.join('"{}"'.format(str(x)) for x in PRO_list)

  #Third party list
  third_party_list = publisher_list + PRO_list

  #Find where quarterly data exists
  if len(publisher_list) > 0:
    find_quarterly_data_publisher = '''SELECT DISTINCT Third_Party_9LC FROM Master
                                                 WHERE Quarter_Statement_9LC <> ""
                                                 AND Third_Party_9LC IN ({})
                                                 GROUP BY Third_Party_9LC'''.format(publisher_string)
    mycursor.execute(find_quarterly_data_publisher)
    quarterly_publisher_list = [i[0] for i in mycursor.fetchall()]
    if len(quarterly_publisher_list) != 0:
      quarterly_publisher_string = ', '.join('"{}"'.format(str(x)) for x in quarterly_publisher_list)
    else:
      quarterly_publisher_string = "''"
  else:
    quarterly_publisher_string = "''"
    quarterly_publisher_list = []

  if len(PRO_list) > 0:
    find_quarterly_data_PRO = '''SELECT DISTINCT Third_Party_9LC FROM Master 
                                         WHERE Quarter_Statement_9LC <> "" 
                                         AND Third_Party_9LC IN ({})
                                         GROUP BY Third_Party_9LC'''.format(PRO_string)
    mycursor.execute(find_quarterly_data_PRO)
    quarterly_PRO_list = [i[0] for i in mycursor.fetchall()]
    if len(quarterly_PRO_list) != 0:
      quarterly_PRO_string = ', '.join('"{}"'.format(str(x)) for x in quarterly_PRO_list)
    else:
      quarterly_PRO_string = "''"
  else:
    quarterly_PRO_string = "''"
    quarterly_PRO_list = []

  if quarterly_publisher_string == '' or quarterly_PRO_string == '':
    quarterly_third_party = quarterly_publisher_string + quarterly_PRO_string
  else:
    quarterly_third_party = quarterly_publisher_string + ',' + quarterly_PRO_string

  if len(quarterly_publisher_list) + len(quarterly_PRO_list) == len(third_party_list):
    quarterly_data = True
    data_list.append('Quarter')
  else:
    quarterly_data = False

  #Find where monthly data exists
  if len(publisher_list) > 0:
    find_monthly_data_publisher = '''SELECT DISTINCT Third_Party_9LC FROM Master
                                                   WHERE Month_Statement_9LC <> ""
                                                   AND Third_Party_9LC IN ({})
                                                   GROUP BY Third_Party_9LC'''.format(publisher_string)
    mycursor.execute(find_monthly_data_publisher)
    monthly_publisher_list = [i[0] for i in mycursor.fetchall()]
    if len(monthly_publisher_list) != 0:
      monthly_publisher_string = ', '.join('"{}"'.format(str(x)) for x in monthly_publisher_list)
    else:
      monthly_publisher_string = "''"
  else:
    monthly_publisher_string = "''"
    monthly_publisher_list = []

  if len(PRO_list) > 0:
    find_monthly_data_PRO = '''SELECT DISTINCT Third_Party_9LC FROM Master 
                                             WHERE Month_Statement_9LC <> "" 
                                             AND Third_Party_9LC IN ({})
                                             GROUP BY Third_Party_9LC'''.format(PRO_string)
    mycursor.execute(find_monthly_data_PRO)
    monthly_PRO_list = [i[0] for i in mycursor.fetchall()]
    if len(monthly_PRO_list) != 0:
      monthly_PRO_string = ', '.join('"{}"'.format(str(x)) for x in monthly_PRO_list)
    else:
      monthly_PRO_string = "''"
  else:
    monthly_PRO_string = "''"
    monthly_PRO_list = []

  if monthly_publisher_string == '' or monthly_PRO_string == '':
    monthly_third_party = monthly_publisher_string + monthly_PRO_string
  else:
    monthly_third_party = monthly_publisher_string + ',' + monthly_PRO_string

  if len(monthly_publisher_list) + len(monthly_PRO_list) == len(third_party_list):
    monthly_data = True
    data_list.append('Month')
  else:
    monthly_data = False

  return data_list



def songvest(database, filename, period):
  mydb = mysql.connector.connect(
    host="34.65.111.142",
    user="external",
    password="musicpass",
    database="{}".format(database)
  )
  mycursor = mydb.cursor(buffered=True)

#Create sheets
  wb = Workbook()
  ws_1 = wb.active
  ws_1.title = "(All) Song List"
  ws_1_1 = wb.create_sheet(title='(Top 10) Song List')
  ws_1_2 = wb.create_sheet(title='(Top 5) Song List')
  ws_2 = wb.create_sheet(title='(All) Source List')
  ws_2_1 = wb.create_sheet(title='(Top 10) Source List')
  ws_3 = wb.create_sheet(title='(All) Source List by Payor')
  ws_3_1 = wb.create_sheet(title='(Top 10) Source List by Payor')
  ws_4 = wb.create_sheet(title='Revenue Over Time')
  ws_4_1 = wb.create_sheet(title='Revenue Over Time By Payor')
  ws_4_2 = wb.create_sheet(title='Revenue Over Time Data')

#Current year
  todays_date = date.today()
  current_year = todays_date.year - 1

#Find complete years
  find_complete = '''SELECT Quarter_Statement_9LC, Year_Statement_9LC FROM Master WHERE Year_Statement_9LC < "2021" GROUP BY Quarter_Statement_9LC, Year_Statement_9LC'''
  mycursor.execute(find_complete)
  quarters = [i[0] for i in mycursor.fetchall()]
  no_complete_years = int((len(quarters)-len(quarters)%4)/4)
  summary_years = []
  if len(summary_years) == 0:
    find_complete = '''SELECT Half_Statement_9LC, Year_Statement_9LC FROM Master WHERE Year_Statement_9LC < "2021" GROUP BY Half_Statement_9LC, Year_Statement_9LC'''
    mycursor.execute(find_complete)
    halves = [i[0] for i in mycursor.fetchall()]
    no_complete_years = int((len(halves) - len(halves) % 2) / 2)
    summary_years = []
  print(no_complete_years)

  if no_complete_years == 1:
    detailed_year_2 = current_year - 1
  if no_complete_years == 2:
    detailed_year_1 = current_year - 2
    detailed_year_2 = current_year - 1
  if no_complete_years > 2:
    detailed_year_1 = current_year - 2
    detailed_year_2 = current_year - 1
    no_summary_years = no_complete_years - 2
    for c in range(no_summary_years):
      summary_years.append(detailed_year_1-no_summary_years+c)


#Find last three full years
  find_period = '''SELECT Statement_Period_Half_9LC FROM Master WHERE Statement_Period_Half_9LC <> ""
                       GROUP BY Statement_Period_Half_9LC ORDER BY Statement_Period_Half_9LC'''
  mycursor.execute(find_period)
  statement_period_half = [i[0] for i in mycursor.fetchall()]
  base_year_value = current_year - 1
  base_year = []
  for a in statement_period_half:
    match = re.search(r'{}\sH\d'.format(base_year_value), a)
    if match:
      base_year.append(match.group())
  previous_year_1 = base_year_value - 1
  previous_year_2 = base_year_value - 2

#Cut off
  if len(summary_years) != 0:
    cut_off = summary_years[0]
  if len(summary_years) == 0 and no_complete_years > 1:
    cut_off = detailed_year_1
  if len(summary_years) == 0 and no_complete_years == 1:
    cut_off = detailed_year_2
  if no_complete_years == 1:
    detailed_cut_off = detailed_year_2
  if no_complete_years == 2:
    detailed_cut_off = detailed_year_1
  if no_complete_years > 2:
    detailed_cut_off = detailed_year_1


#Get complete year list
  find_all_years = '''SELECT Year_Statement_9LC FROM Master WHERE Year_Statement_9LC >= "{}" GROUP BY Year_Statement_9LC'''.format(cut_off)
  mycursor.execute(find_all_years)
  all_years = [i[0] for i in mycursor.fetchall()]

#Get list of statement half periods
  find_half_period = '''SELECT Year_Statement_9LC, Half_Statement_9LC FROM Master 
                               WHERE Half_Statement_9LC <> "" AND Year_Statement_9LC >= "{}" GROUP BY Year_Statement_9LC, Half_Statement_9LC'''.format(
    detailed_cut_off)
  mycursor.execute(find_half_period)
  complete_list = [i for i in mycursor.fetchall()]
  statement_period_half_list = [i[1] + ' ' + i[0] for i in complete_list]
  year_statement_list_half = [i[0] for i in complete_list]
  half_statement_list = [i[1] for i in complete_list]
  year_list = list(dict.fromkeys(year_statement_list_half))

#Get list of statement quarter periods
  find_quarter_period = '''SELECT Year_Statement_9LC, Quarter_Statement_9LC FROM Master 
                             WHERE Quarter_Statement_9LC <> "" AND Year_Statement_9LC >= "{}" GROUP BY Year_Statement_9LC, Quarter_Statement_9LC'''.format(detailed_cut_off)
  mycursor.execute(find_quarter_period)
  complete_list = [i for i in mycursor.fetchall()]
  statement_period_quarter_list = [i[1] + ' ' + i[0] for i in complete_list]
  year_statement_list_quarter = [i[0] for i in complete_list]
  quarter_statement_list = [i[1] for i in complete_list]
  print(statement_period_quarter_list)
  print(year_statement_list_quarter)

#Get list of statement month periods
  find_month_period = '''SELECT Year_Statement_9LC, Month_Statement_9LC FROM Master 
                               WHERE Quarter_Statement_9LC <> "" AND Year_Statement_9LC >= "{}" GROUP BY Year_Statement_9LC, Month_Statement_9LC'''.format(
    detailed_cut_off)
  mycursor.execute(find_month_period)
  complete_list = [i for i in mycursor.fetchall()]
  statement_period_month_list = [i[1] + ' ' + i[0] for i in complete_list]
  year_statement_list_month = [i[0] for i in complete_list]
  month_statement_list = [i[1] for i in complete_list]

#Publisher list
  find_publishers = '''SELECT Third_Party_9LC FROM Master WHERE Third_Party_9LC NOT IN ('ASCAP', 'BMI', 'PRS', 'MCPS', 
                         'SoundExchange', 'SOCAN', 'APRA', 'AMCOS', 'SESAC', 'AMRA', 'GEMA', 'SUISA', 'ZAiKS') GROUP BY Third_Party_9LC'''
  mycursor.execute(find_publishers)
  publisher_list = [i[0] for i in mycursor.fetchall()]
  publisher_string = ', '.join('"{}"'.format(str(x)) for x in publisher_list)

#PRO list
  find_PRO = '''SELECT Third_Party_9LC FROM Master WHERE Third_Party_9LC IN ('ASCAP', 'BMI', 'PRS', 'MCPS', 
                           'SoundExchange', 'SOCAN', 'APRA', 'AMCOS', 'SESAC', 'AMRA', 'GEMA', 'SUISA', 'ZAiKS') GROUP BY Third_Party_9LC'''
  mycursor.execute(find_PRO)
  PRO_list = [i[0] for i in mycursor.fetchall()]
  PRO_string = ', '.join('"{}"'.format(str(x)) for x in PRO_list)

#Third party list
  third_party_list = publisher_list + PRO_list

#Find smallest period
  if period == 'Month':
    mySQL_column = 'Month_Statement_9LC'
    statement_list = month_statement_list
    year_statement_list = year_statement_list_month
    divisions = 12
    statement_period_list = statement_period_month_list
  if period == 'Quarter':
    mySQL_column = 'Quarter_Statement_9LC'
    statement_list = quarter_statement_list
    year_statement_list = year_statement_list_quarter
    divisions = 4
    statement_period_list = statement_period_quarter_list
  if period == 'Half':
    mySQL_column = 'Half_Statement_9LC'
    statement_list = half_statement_list
    year_statement_list = year_statement_list_half
    divisions = 2
    statement_period_list = statement_period_half_list

#Find number of periods in most recent year
  find_remainder = '''SELECT {} FROM Master WHERE Year_Statement_9LC = "{}" AND {} <> "" GROUP BY {}'''.format(
       mySQL_column,current_year,mySQL_column, mySQL_column)
  mycursor.execute(find_remainder)
  remainder = [i[0] for i in mycursor.fetchall()]
  if len(remainder) == 0:
    find_remainder = '''SELECT {} FROM Master WHERE Year_Statement_9LC = "{}" AND {} <> "" GROUP BY {}'''.format(mySQL_column,all_years[-1],mySQL_column, mySQL_column)
    mycursor.execute(find_remainder)
    remainder = [i[0] for i in mycursor.fetchall()]
  print(remainder)

#Find number of periods in detailed cut off year
  find_detailed_remainder = '''SELECT {} FROM Master Where Year_Statement_9LC = "{}" AND {} <> "" GROUP BY {}'''.format(mySQL_column, detailed_cut_off, mySQL_column, mySQL_column)
  mycursor.execute(find_detailed_remainder)
  detailed_remainder = [i[0] for i in mycursor.fetchall()]
  if len(detailed_remainder) != 0 and len(summary_years) == 0:
    first_year_complete = False
  else:
    first_year_complete = True

#Find number of songs
  find_songs = '''SELECT Song_Name_9LC
                    FROM Master GROUP BY Song_Name_9LC'''
  mycursor.execute(find_songs)
  songs = [i[0] for i in mycursor.fetchall()]
  song_count = len(songs)

#Column letters
  column_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S',
                    'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ',
                    'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ']

#Create styles
  #Main numbers
  number_style = NamedStyle(name="number_style")
  number_style.alignment = Alignment(horizontal="right", vertical="center")
  number_style.font = Font(name="Calibri", size="11")
  number_style.number_format = '#,##0.00'

  #Title cells
  title_style = NamedStyle(name="title_style")
  title_style.alignment = Alignment(horizontal="center", vertical="center")
  title_style.font = Font(name="Calibri", size="11", bold=True)
  title_style.fill = PatternFill("solid", fgColor="7496CC")

  #Header cells
  header_style = NamedStyle(name="header_style")
  header_style.alignment = Alignment(horizontal="center", vertical="center")
  header_style.font = Font(name="Calibri", size="11", bold=True)
  header_style.fill = PatternFill("solid", fgColor="C9CDCF")

  #Sub header cells
  sub_header_style = NamedStyle(name="sub_header_style")
  sub_header_style.alignment = Alignment(horizontal="left", vertical="center")
  sub_header_style.font = Font(name="Calibri", size="11", bold=True, italic=True)

  #Name cells
  name_style = NamedStyle(name="name_style")
  name_style.alignment = Alignment(horizontal="left", vertical="center")
  name_style.font = Font(name="Calibri", size="11")

  #Total cells
  total_style = NamedStyle(name="total_style")
  total_style.alignment = Alignment(horizontal="right", vertical="center")
  total_style.font = Font(name="Calibri", size="11", bold=True)
  total_style.number_format = '#,##0.00'

  #Total row style
  total_row_style = NamedStyle(name="total_row_style")
  total_row_style.alignment = Alignment(horizontal="right", vertical="center")
  total_row_style.font = Font(name="Calibri", size="11", bold=True)
  total_row_style.fill = PatternFill("solid", fgColor="88B054")
  total_row_style.number_format = '#,##0.00'

  #Total label style
  total_label_style = NamedStyle(name="total_label_style")
  total_label_style.alignment = Alignment(horizontal="center", vertical="center")
  total_label_style.font = Font(name="Calibri", size="11", bold=True)
  total_label_style.fill = PatternFill("solid", fgColor="88B054")

  #Publisher/PRO label style
  thick = Side(border_style="medium", color="000000")
  publisher_label_style = NamedStyle(name="publisher_label_style")
  publisher_label_style.alignment = Alignment(horizontal="center", vertical="center")
  publisher_label_style.font = Font(name="Calibri", size="11", bold=True, italic=True)
  publisher_label_style.border = Border(bottom=thick)

  #Not available style
  not_available_style = NamedStyle(name="not_available_style")
  not_available_style.alignment = Alignment(horizontal="center", vertical="center")
  not_available_style.font = Font(name="Calibri", size="11")

  #Bold name style
  bold_name_style = NamedStyle(name="bold_name_style")
  bold_name_style.alignment = Alignment(horizontal="left", vertical="center")
  bold_name_style.font = Font(name="Calibri", size="11", bold=True)

  #Total revenue style
  total_revenue_style = NamedStyle(name="total_revenue_style")
  total_revenue_style.alignment = Alignment(horizontal="right", vertical="center")
  total_revenue_style.font = Font(name="Calibri", size="11", bold=True)
  total_revenue_style.number_format = '#,##0.00'

  #Add styles to workbook
  wb.add_named_style(number_style)
  wb.add_named_style(header_style)
  wb.add_named_style(title_style)
  wb.add_named_style(name_style)
  wb.add_named_style(total_style)
  wb.add_named_style(sub_header_style)
  wb.add_named_style(total_row_style)
  wb.add_named_style(total_label_style)
  wb.add_named_style(publisher_label_style)
  wb.add_named_style(not_available_style)
  wb.add_named_style(bold_name_style)
  wb.add_named_style(total_revenue_style)

#Sheet 1: Song List
  #Build table
  select_table_1_1 = '''SELECT Song_Name_9LC,'''
  select_table_1_2 = ""
  for c in summary_years:
    select_table_1_2 += '''sum( CASE WHEN Year_Statement_9LC = "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(c,c)
  for a, b in zip(year_statement_list, statement_list):
    select_table_1_2 += '''sum( CASE WHEN Year_Statement_9LC = "{}" AND {} = "{}"
                            THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(a,mySQL_column,b,a,b)
  select_table_1_3 = '''sum( CASE WHEN Year_Statement_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END)
                              AS `Total` FROM Master GROUP BY Song_Name_9LC ORDER BY `Total` DESC'''.format(cut_off)
  select_table_1 = select_table_1_1 + select_table_1_2 + select_table_1_3
  print(select_table_1)
  mycursor.execute(select_table_1)
  table_1 = mycursor.fetchall()

  #Title row list
  title_names_1 = ['{}'.format(database[:-25])]
  for c in all_years:
    title_names_1.append(c)

  #Column names list
  column_names_1 = ['Compositions']
  for d in summary_years:
    column_names_1.append("")
  for e in statement_list:
    column_names_1.append(e)
  column_names_1.append('Grand Total')

  #Size of worksheet
  for column_no in range(2, len(column_names_1)+3):
    for row_no in range(2, len(table_1)+4):
      ws_1.cell(row=row_no, column=column_no)

  #Title row
  ws_1.cell(row=2, column=2).value = '{}'.format(title_names_1[0])
  ws_1.cell(row=2, column=2).style = 'title_style'
  if first_year_complete:
    for d, e in zip(range(len(summary_years)+1, (len(title_names_1)-3)*divisions+2, divisions), title_names_1[1+len(summary_years):]):
      ws_1.merge_cells('{}2:{}2'.format(column_letters[d+1], column_letters[d+divisions]))
      ws_1.cell(row=2, column=d+2).value = e
      ws_1.cell(row=2, column=d+2).style = 'title_style'
    ws_1.cell(row=2, column=(len(summary_years) + divisions * (len(all_years) - len(summary_years) - 1) + 3)).value = title_names_1[-1]
    ws_1.cell(row=2, column=(len(summary_years) + divisions * (len(all_years) - len(summary_years) - 1) + 3)).style = 'title_style'
    ws_1.cell(row=2, column=(len(summary_years) + divisions * (len(all_years) - len(summary_years) - 1) + 3 + len(remainder))).style = 'title_style'
    ws_1.merge_cells('{}2:{}2'.format(column_letters[len(summary_years) + divisions * (len(all_years) - len(summary_years) - 1) + 2],
                       column_letters[len(summary_years) + divisions * (len(all_years) - len(summary_years) - 1) + len(remainder)+1]))
    for y, z in zip(range(3, 3 + len(summary_years)), title_names_1[1:1 + len(summary_years)]):
      ws_1.cell(row=2, column=y).value = z
      ws_1.cell(row=2, column=y).style = 'title_style'

  else:
    ws_1.merge_cells('C2:{}2'.format(column_letters[len(detailed_remainder)+1]))
    ws_1.cell(row=2, column=3).value = title_names_1[1]
    ws_1.cell(row=2, column=3).style = 'title_style'
    for d, e in zip(range(len(detailed_remainder)+3, (len(title_names_1)-3)*divisions+4, divisions), title_names_1[2:]):
      ws_1.merge_cells('{}2:{}2'.format(column_letters[d+1], column_letters[d+divisions]))
      ws_1.cell(row=2, column=d+3).value = e
      ws_1.cell(row=2, column=d+3).style = 'title_style'
    ws_1.cell(row=2, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 3)).value = title_names_1[-1]
    ws_1.cell(row=2, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 3)).style = 'title_style'
    ws_1.cell(row=2, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 3 + len(remainder))).style = 'title_style'
    ws_1.merge_cells('{}2:{}2'.format(column_letters[len(detailed_remainder) + divisions * (len(all_years) - 2) + 2],
                       column_letters[len(detailed_remainder) + divisions * (len(all_years) - 2) + len(remainder)+1]))

  #Header row
  for e, f in zip(range(2, len(column_names_1)+3), column_names_1):
    ws_1.cell(row=3, column=e).value = f
    ws_1.cell(row=3, column=e).style = 'title_style'

  #Add table
  table_rows_1 = ws_1[4:len(table_1)+3]
  for row, g in zip(table_rows_1, table_1):
    for cell, h in zip(row[1:], g):
      cell.value = h
      cell.style = 'number_style'
  for row in range(4, song_count+4):
    ws_1.cell(row=row, column=2).style = 'name_style'
    ws_1.cell(row=row, column=len(column_names_1)+1).style = 'total_style'

  #Total row
  for i, j in zip(range(2, len(column_names_1)+2), column_letters[1:]):
    ws_1.cell(row=song_count+4, column=i).value = "=SUM({}{}:{}{})".format(j,4,j,song_count+3)
    ws_1.cell(row=song_count+4, column=i).style = 'total_row_style'
  ws_1.cell(row=song_count+4, column=2).value = 'Grand Total'
  ws_1.cell(row=song_count+4, column=2).style = 'total_label_style'

  #Outlines
  thin = Side(border_style="thin", color="000000")
  if first_year_complete:
    for k in range(len(summary_years)+1, (len(title_names_1)-2)*divisions+2, divisions):
      for l in range(2, song_count + 5):
        ws_1.cell(row=l, column=k + 2).border = Border(left=thin)

    for m in range(2, len(column_names_1) + 2):
      ws_1.cell(row=1, column=m).border = Border(bottom=thin)
      ws_1.cell(row=3, column=m).border = Border(bottom=thin)
      ws_1.cell(row=song_count + 3, column=m).border = Border(bottom=thin)
      ws_1.cell(row=song_count + 4, column=m).border = Border(bottom=thin)

    for n in range(len(summary_years)+1, (len(title_names_1)-2)*divisions+2, divisions):
      ws_1.cell(row=3, column=n + 2).border = Border(left=thin, bottom=thin)
      ws_1.cell(row=song_count + 3, column=n + 2).border = Border(left=thin, bottom=thin)
      ws_1.cell(row=song_count + 4, column=n + 2).border = Border(left=thin, bottom=thin)

    for r in range(2, song_count+5):
      ws_1.cell(row=r, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(left=thin)
      ws_1.cell(row=r, column=1).border = Border(right=thin)
    ws_1.cell(row=3, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(
      left=thin, bottom=thin)
    ws_1.cell(row=2, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(
      left=thin, top=thin)
    ws_1.cell(row=song_count + 4, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(
      left=thin, bottom=thin)
    ws_1.cell(row=song_count + 3, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(
      left=thin, bottom=thin)

    for o in range(2, song_count+5):
      ws_1.cell(row=o, column=len(column_names_1)+2).border = Border(left=thin)

    for p in range(2, len(summary_years)+2):
      for q in range(2, song_count+5):
        ws_1.cell(row=q, column=p+1).border = Border(left=thin)

    for s in range(2, len(summary_years) + 3):
      ws_1.cell(row=2, column=s).border = Border(left=thin, top=thin)
      ws_1.cell(row=3, column=s).border = Border(left=thin, bottom=thin)
      ws_1.cell(row=song_count+4, column=s).border = Border(left=thin, bottom=thin)
      ws_1.cell(row=song_count+3, column=s).border = Border(left=thin, bottom=thin)

  else:
    for k in range(len(detailed_remainder)+1, (len(title_names_1) - 2) * divisions + 2, divisions):
      for l in range(2, song_count + 5):
        ws_1.cell(row=l, column=k + 2).border = Border(left=thin)

    for m in range(2, len(column_names_1) + 2):
      ws_1.cell(row=1, column=m).border = Border(bottom=thin)
      ws_1.cell(row=3, column=m).border = Border(bottom=thin)
      ws_1.cell(row=song_count + 3, column=m).border = Border(bottom=thin)
      ws_1.cell(row=song_count + 4, column=m).border = Border(bottom=thin)

    for n in range(len(detailed_remainder)+1, (len(title_names_1) - 2) * divisions + 2, divisions):
      ws_1.cell(row=3, column=n + 2).border = Border(left=thin, bottom=thin)
      ws_1.cell(row=song_count + 3, column=n + 2).border = Border(left=thin, bottom=thin)
      ws_1.cell(row=song_count + 4, column=n + 2).border = Border(left=thin, bottom=thin)

    for r in range(2, song_count + 5):
      ws_1.cell(row=r, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 3 + len(remainder))).border = Border(left=thin)
      ws_1.cell(row=r, column=1).border = Border(right=thin)
    ws_1.cell(row=3, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 3 + len(remainder))).border = Border(left=thin, bottom=thin)
    ws_1.cell(row=2, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 3 + len(remainder))).border = Border(left=thin, top=thin)
    ws_1.cell(row=song_count + 4, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 3 + len(remainder))).border = Border(left=thin, bottom=thin)
    ws_1.cell(row=song_count + 3, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 3 + len(remainder))).border = Border(left=thin, bottom=thin)

    for o in range(2, song_count + 5):
      ws_1.cell(row=o, column=len(column_names_1) + 2).border = Border(left=thin)

    for q in range(2, song_count + 5):
      ws_1.cell(row=q, column=3).border = Border(left=thin)

    for s in range(2, 4):
      ws_1.cell(row=2, column=s).border = Border(left=thin, top=thin)
      ws_1.cell(row=3, column=s).border = Border(left=thin, bottom=thin)
      ws_1.cell(row=song_count + 4, column=s).border = Border(left=thin, bottom=thin)
      ws_1.cell(row=song_count + 3, column=s).border = Border(left=thin, bottom=thin)

#Sheet 1.1: (Top 10) Song List
  #Build table
  select_table_1_1_1 = '''SELECT Song_Name_9LC,'''
  select_table_1_1_2 = ""
  for c in summary_years:
    select_table_1_1_2 += '''sum( CASE WHEN Year_Statement_9LC = "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(
      c, c)
  for a, b in zip(year_statement_list, statement_list):
    select_table_1_1_2 += '''sum( CASE WHEN Year_Statement_9LC = "{}" AND {} = "{}"
                            THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(a, mySQL_column, b, a, b)
  select_table_1_1_3 = '''sum( CASE WHEN Year_Statement_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END)
                              AS `Total` FROM Master GROUP BY Song_Name_9LC ORDER BY `Total` DESC'''.format(cut_off)
  select_table_1_1 = select_table_1_1_1 + select_table_1_1_2 + select_table_1_1_3
  mycursor.execute(select_table_1_1)
  table_1_1 = mycursor.fetchall()

  #Title row list
  title_names_1_1 = ['{}'.format(database[:-25])]
  for c in all_years:
    title_names_1_1.append(c)

  #Column names list
  column_names_1_1 = ['Compositions']
  for d in summary_years:
    column_names_1_1.append("")
  for e in statement_list:
    column_names_1_1.append(e)
  column_names_1_1.append('Grand Total')

  #Size of worksheet
  for column_no in range(2, len(column_names_1_1) + 3):
    for row_no in range(2, len(table_1_1) + 4):
      ws_1_1.cell(row=row_no, column=column_no)

  #Title row
  ws_1_1.cell(row=2, column=2).value = '{}'.format(title_names_1_1[0])
  ws_1_1.cell(row=2, column=2).style = 'title_style'
  if first_year_complete:
    for d, e in zip(range(len(summary_years)+1, (len(title_names_1_1)-3)*divisions+2, divisions),
                    title_names_1_1[1 + len(summary_years):]):
      ws_1_1.merge_cells('{}2:{}2'.format(column_letters[d + 1], column_letters[d + divisions]))
      ws_1_1.cell(row=2, column=d + 2).value = e
      ws_1_1.cell(row=2, column=d + 2).style = 'title_style'
    ws_1_1.cell(row=2, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3)).value = title_names_1_1[-1]
    ws_1_1.cell(row=2, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3)).style = 'title_style'
    ws_1_1.cell(row=2, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).style = 'title_style'
    ws_1_1.merge_cells('{}2:{}2'.format(column_letters[len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+2],
                                                        column_letters[len(summary_years) + divisions*(len(all_years)-len(summary_years)-1) + len(remainder)+1]))
    for y, z in zip(range(3, 3 + len(summary_years)), title_names_1_1[1:1 + len(summary_years)]):
      ws_1_1.cell(row=2, column=y).value = z
      ws_1_1.cell(row=2, column=y).style = 'title_style'

  else:
    ws_1_1.merge_cells('C2:{}2'.format(column_letters[len(detailed_remainder)+1]))
    ws_1_1.cell(row=2, column=3).value = title_names_1_1[1]
    ws_1_1.cell(row=2, column=3).style = 'title_style'
    for d, e in zip(range(len(detailed_remainder)+3, (len(title_names_1_1)-3)*divisions+4, divisions),
                    title_names_1_1[2:]):
      ws_1_1.merge_cells('{}2:{}2'.format(column_letters[d + 1], column_letters[d + divisions]))
      ws_1_1.cell(row=2, column=d + 3).value = e
      ws_1_1.cell(row=2, column=d + 3).style = 'title_style'
    ws_1_1.cell(row=2, column=(len(detailed_remainder) + divisions*(len(all_years)-2)+3)).value = title_names_1_1[-1]
    ws_1_1.cell(row=2, column=(len(detailed_remainder) + divisions*(len(all_years)-2)+3)).style = 'title_style'
    ws_1_1.cell(row=2, column=(len(detailed_remainder) + divisions*(len(all_years)-2)+3+len(remainder))).style = 'title_style'
    ws_1_1.merge_cells('{}2:{}2'.format(column_letters[len(detailed_remainder) + divisions*(len(all_years)-2)+2],
                                                        column_letters[len(detailed_remainder) + divisions*(len(all_years)-2) + len(remainder)+1]))

  #Header row
  for e, f in zip(range(2, len(column_names_1_1) + 3), column_names_1_1):
    ws_1_1.cell(row=3, column=e).value = f
    ws_1_1.cell(row=3, column=e).style = 'title_style'

  #Top songs
  top_10_songs = table_1_1[:10]

  #Other songs
  all_columns = summary_years + statement_list
  other_songs = table_1_1[10:]
  other_songs_element = ['Other Songs']
  period_total = 0
  for z in range(1, len(all_columns) + 2):
    for w in other_songs:
      if w[z] == None:
        period_total += 0
      else:
        period_total += w[z]
    other_songs_element.append(period_total)
    period_total = 0
  #other_songs_element = (other_songs_element)

  #Add table
  top_10_table = top_10_songs + [other_songs_element]
  table_rows_1_1 = ws_1_1[4:len(top_10_table) + 3]
  for row, g in zip(table_rows_1_1, top_10_table):
    for cell, h in zip(row[1:], g):
      cell.value = h
      cell.style = 'number_style'
  for row in range(4, len(top_10_table) + 4):
    ws_1_1.cell(row=row, column=2).style = 'name_style'
    ws_1_1.cell(row=row, column=len(column_names_1_1)+1).style = 'total_style'

  #Total row
  for i, j in zip(range(2, len(column_names_1_1) + 2), column_letters[1:]):
    ws_1_1.cell(row=len(top_10_table) + 4, column=i).value = "=SUM({}{}:{}{})".format(j, 3, j, len(top_10_table) + 3)
    ws_1_1.cell(row=len(top_10_table) + 4, column=i).style = 'total_row_style'
  ws_1_1.cell(row=len(top_10_table) + 4, column=2).value = 'Grand Total'
  ws_1_1.cell(row=len(top_10_table) + 4, column=2).style = 'total_label_style'

  #Add % of revenue column
  total_column_letter = column_letters[len(column_names_1_1)-1]
  for q in range(4,len(top_10_table)+4):
    ws_1_1.cell(row=q, column=len(column_names_1_1)+2).value = "=({}{}/{}{})".format(total_column_letter, q, total_column_letter, len(top_10_table)+4)
    ws_1_1.cell(row=q, column=len(column_names_1_1)+2).style = 'Percent'
  ws_1_1.cell(row=2, column=len(column_names_1_1)+2).value = '% Of Revenue'
  ws_1_1.cell(row=3, column=len(column_names_1_1)+2).value = '(Most Recent Period)'
  ws_1_1.cell(row=2, column=len(column_names_1_1)+2).style = 'title_style'
  ws_1_1.cell(row=3, column=len(column_names_1_1)+2).style = 'title_style'
  ws_1_1.cell(row=len(top_10_table)+4, column=len(column_names_1_1)+2).style = 'total_row_style'

  #Find artists and release dates
  find_artist = '''SELECT Song_Name_9LC, Release_Artist_9LC, Release_Date_9LC, 
                    sum( CASE WHEN Year_Statement_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END)
                    AS `Total` FROM Master GROUP BY Song_Name_9LC, Release_Artist_9LC, Release_Date_9LC ORDER BY `Total` DESC'''.format(cut_off)
  mycursor.execute(find_artist)
  artist_list = mycursor.fetchall()
  print(artist_list)
  top_10_artists = artist_list[:10]

  #Add artists and release dates
  artist_rows_1_1 = ws_1_1[4:len(top_10_table) + 3]
  for f, g in zip(range(4,len(top_10_artists)+4),top_10_artists):
    for h, i in zip(range(len(column_names_1_1)+3,len(column_names_1_1)+5),range(1,len(g))):
      ws_1_1.cell(row=f, column=h).value = g[i]
      ws_1_1.cell(row=f, column=h).style = 'number_style'

  ws_1_1.cell(row=3, column=len(column_names_1_1)+3).value = 'Release Artist'
  ws_1_1.cell(row=2, column=len(column_names_1_1)+3).style = 'title_style'
  ws_1_1.cell(row=3, column=len(column_names_1_1)+3).style = 'title_style'
  ws_1_1.cell(row=3, column=len(column_names_1_1)+4).value = 'Release Date'
  ws_1_1.cell(row=2, column=len(column_names_1_1)+4).style = 'title_style'
  ws_1_1.cell(row=3, column=len(column_names_1_1)+4).style = 'title_style'
  ws_1_1.cell(row=len(top_10_artists)+5, column=len(column_names_1_1)+3).style = 'total_row_style'
  ws_1_1.cell(row=len(top_10_artists)+5, column=len(column_names_1_1)+4).style = 'total_row_style'

  #Outlines
  thin = Side(border_style="thin", color="000000")
  if first_year_complete:
    for k in range(len(summary_years)+1, (len(title_names_1)-2)*divisions+2, divisions):
      for l in range(2, len(top_10_table) + 5):
        ws_1_1.cell(row=l, column=k + 2).border = Border(left=thin)

    for m in range(2, len(column_names_1_1) + 2):
      ws_1_1.cell(row=1, column=m).border = Border(bottom=thin)
      ws_1_1.cell(row=3, column=m).border = Border(bottom=thin)
      ws_1_1.cell(row=len(top_10_table) + 3, column=m).border = Border(bottom=thin)
      ws_1_1.cell(row=len(top_10_table) + 4, column=m).border = Border(bottom=thin)

    for n in range(len(summary_years)+1, (len(title_names_1)-2)*divisions+2, divisions):
      ws_1_1.cell(row=3, column=n + 2).border = Border(left=thin, bottom=thin)
      ws_1_1.cell(row=len(top_10_table) + 3, column=n + 2).border = Border(left=thin, bottom=thin)
      ws_1_1.cell(row=len(top_10_table) + 4, column=n + 2).border = Border(left=thin, bottom=thin)

    for r in range(2, len(top_10_table) + 5):
      ws_1_1.cell(row=r, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(left=thin)
      ws_1_1.cell(row=r, column=1).border = Border(right=thin)
    ws_1_1.cell(row=3, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(
      left=thin, bottom=thin)
    ws_1_1.cell(row=2, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(
      left=thin, top=thin)
    ws_1_1.cell(row=len(top_10_table) + 4, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(left=thin, bottom=thin)
    ws_1_1.cell(row=len(top_10_table) + 3, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(
      left=thin, bottom=thin)

    for o in range(2, len(top_10_table) + 5):
      ws_1_1.cell(row=o, column=len(column_names_1_1) + 2).border = Border(left=thin)
      ws_1_1.cell(row=o, column=len(column_names_1_1) + 3).border = Border(left=thin)
      ws_1_1.cell(row=o, column=len(column_names_1_1) + 4).border = Border(left=thin)
      ws_1_1.cell(row=o, column=len(column_names_1_1) + 5).border = Border(left=thin)

    for p in range(2, len(summary_years) + 2):
      for q in range(2, len(top_10_table) + 5):
        ws_1_1.cell(row=q, column=p + 1).border = Border(left=thin)

    for s in range(2, len(summary_years) + 3):
      ws_1_1.cell(row=2, column=s).border = Border(left=thin, top=thin)
      ws_1_1.cell(row=3, column=s).border = Border(left=thin, bottom=thin)
      ws_1_1.cell(row=len(top_10_table) + 4, column=s).border = Border(left=thin, bottom=thin)
      ws_1_1.cell(row=len(top_10_table) + 3, column=s).border = Border(left=thin, bottom=thin)

    for t in range(len(column_names_1_1)+2, len(column_names_1_1)+5):
      ws_1_1.cell(row=3, column=t).border = Border(left=thin, bottom=thin)
      ws_1_1.cell(row=1, column=t).border = Border(bottom=thin)
      ws_1_1.cell(row=len(top_10_table)+4, column=t).border = Border(left=thin, bottom=thin)
      ws_1_1.cell(row=len(top_10_table)+3, column=t).border = Border(left=thin, bottom=thin)

  else:
    for k in range(len(detailed_remainder)+1, (len(title_names_1) - 2) * divisions + 2, divisions):
      for l in range(2, len(top_10_table) + 5):
        ws_1_1.cell(row=l, column=k + 2).border = Border(left=thin)

    for m in range(2, len(column_names_1_1) + 2):
      ws_1_1.cell(row=1, column=m).border = Border(bottom=thin)
      ws_1_1.cell(row=3, column=m).border = Border(bottom=thin)
      ws_1_1.cell(row=len(top_10_table) + 3, column=m).border = Border(bottom=thin)
      ws_1_1.cell(row=len(top_10_table) + 4, column=m).border = Border(bottom=thin)

    for n in range(len(detailed_remainder)+1, (len(title_names_1) - 2) * divisions + 2, divisions):
      ws_1_1.cell(row=3, column=n + 2).border = Border(left=thin, bottom=thin)
      ws_1_1.cell(row=len(top_10_table) + 3, column=n + 2).border = Border(left=thin, bottom=thin)
      ws_1_1.cell(row=len(top_10_table) + 4, column=n + 2).border = Border(left=thin, bottom=thin)

    for r in range(2, len(top_10_table) + 5):
      ws_1_1.cell(row=r, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 3 + len(remainder))).border = Border(left=thin)
      ws_1_1.cell(row=r, column=1).border = Border(right=thin)
    ws_1_1.cell(row=2, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 3 + len(remainder))).border = Border(left=thin, bottom=thin)
    ws_1_1.cell(row=2, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 3 + len(remainder))).border = Border(left=thin, top=thin)
    ws_1_1.cell(row=len(top_10_table) + 4, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 3 + len(remainder))).border = Border(left=thin, bottom=thin)
    ws_1_1.cell(row=len(top_10_table) + 3, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 3 + len(remainder))).border = Border(left=thin, bottom=thin)

    for o in range(2, len(top_10_table) + 5):
      ws_1_1.cell(row=o, column=len(column_names_1_1) + 2).border = Border(left=thin)
      ws_1_1.cell(row=o, column=len(column_names_1_1) + 3).border = Border(left=thin)
      ws_1_1.cell(row=o, column=len(column_names_1_1) + 4).border = Border(left=thin)
      ws_1_1.cell(row=o, column=len(column_names_1_1) + 5).border = Border(left=thin)


    for q in range(2, len(top_10_table) + 5):
      ws_1_1.cell(row=q, column=3).border = Border(left=thin)

    for s in range(2, 4):
      ws_1_1.cell(row=2, column=s).border = Border(left=thin, top=thin)
      ws_1_1.cell(row=3, column=s).border = Border(left=thin, bottom=thin)
      ws_1_1.cell(row=len(top_10_table) + 4, column=s).border = Border(left=thin, bottom=thin)
      ws_1_1.cell(row=len(top_10_table) + 3, column=s).border = Border(left=thin, bottom=thin)

    for t in range(len(column_names_1_1) + 2, len(column_names_1_1) + 5):
      ws_1_1.cell(row=3, column=t).border = Border(left=thin, bottom=thin)
      ws_1_1.cell(row=1, column=t).border = Border(left=thin, bottom=thin)
      ws_1_1.cell(row=len(top_10_table) + 4, column=t).border = Border(left=thin, bottom=thin)
      ws_1_1.cell(row=len(top_10_table) + 3, column=t).border = Border(left=thin, bottom=thin)

  #Sheet 1.2: (Top 5) Song List
  #Build table
  select_table_1_2_1 = '''SELECT Song_Name_9LC,'''
  select_table_1_2_2 = ""
  for c in summary_years:
    select_table_1_2_2 += '''sum( CASE WHEN Year_Statement_9LC = "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(
      c, c)
  for a, b in zip(year_statement_list, statement_list):
    select_table_1_2_2 += '''sum( CASE WHEN Year_Statement_9LC = "{}" AND {} = "{}"
                            THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(a, mySQL_column, b, a, b)
  select_table_1_2_3 = '''sum( CASE WHEN Year_Statement_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END)
                              AS `Total` FROM Master GROUP BY Song_Name_9LC ORDER BY `Total` DESC'''.format(cut_off)
  select_table_1_2 = select_table_1_2_1 + select_table_1_2_2 + select_table_1_2_3
  mycursor.execute(select_table_1_2)
  table_1_2 = mycursor.fetchall()

  #Title row list
  title_names_1_2 = ['{}'.format(database[:-25])]
  for c in all_years:
    title_names_1_2.append(c)

  #Column names list
  column_names_1_2 = ['Compositions']
  for d in summary_years:
    column_names_1_2.append("")
  for e in statement_list:
    column_names_1_2.append(e)
  column_names_1_2.append('Grand Total')

  # Size of worksheet
  for column_no in range(2, len(column_names_1_2) + 3):
    for row_no in range(2, len(table_1_2) + 4):
      ws_1_2.cell(row=row_no, column=column_no)

  #Title row
  ws_1_2.cell(row=2, column=2).value = '{}'.format(title_names_1_2[0])
  ws_1_2.cell(row=2, column=2).style = 'title_style'
  if first_year_complete:
    for d, e in zip(range(len(summary_years)+1, (len(title_names_1_2) - 3) * divisions + 2, divisions),
                    title_names_1_2[1 + len(summary_years):]):
      ws_1_2.merge_cells('{}2:{}2'.format(column_letters[d + 1], column_letters[d + divisions]))
      ws_1_2.cell(row=2, column=d + 2).value = e
      ws_1_2.cell(row=2, column=d + 2).style = 'title_style'
    ws_1_2.cell(row=2, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3)).value = title_names_1_2[-1]
    ws_1_2.cell(row=2, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3)).style = 'title_style'
    ws_1_2.cell(row=2, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3)+len(remainder)).style = 'title_style'
    ws_1_2.merge_cells('{}2:{}2'.format(column_letters[len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+2],
                                        column_letters[len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+len(remainder)+1]))
    for y, z in zip(range(3, 3 + len(summary_years)), title_names_1_2[1:1 + len(summary_years)]):
      ws_1_2.cell(row=2, column=y).value = z
      ws_1_2.cell(row=2, column=y).style = 'title_style'

  else:
    ws_1_2.merge_cells('C2:{}2'.format(column_letters[len(detailed_remainder)+1]))
    ws_1_2.cell(row=2, column=3).value = title_names_1_2[1]
    ws_1_2.cell(row=2, column=3).style = 'title_style'
    for d, e in zip(range(len(detailed_remainder)+1, (len(title_names_1_2) - 3) * divisions + 2, divisions), title_names_1_2[2:]):
      ws_1_2.merge_cells('{}2:{}2'.format(column_letters[d + 1], column_letters[d + divisions]))
      ws_1_2.cell(row=2, column=d + 3).value = e
      ws_1_2.cell(row=2, column=d + 3).style = 'title_style'
    ws_1_2.cell(row=2, column=(len(detailed_remainder) + divisions*(len(all_years)-2)+3)).value = title_names_1_2[-1]
    ws_1_2.cell(row=2, column=(len(detailed_remainder) + divisions*(len(all_years)-2)+3)).style = 'title_style'
    ws_1_2.cell(row=2, column=(len(detailed_remainder) + divisions*(len(all_years)-2)+3)+len(remainder)).style = 'title_style'
    ws_1_2.merge_cells('{}2:{}2'.format(column_letters[len(detailed_remainder) + divisions*(len(all_years)-2)+2],
                                        column_letters[len(detailed_remainder) + divisions*(len(all_years)-2)+len(remainder)+1]))



  #Header row
  for e, f in zip(range(2, len(column_names_1_2) + 3), column_names_1_2):
    ws_1_2.cell(row=3, column=e).value = f
    ws_1_2.cell(row=3, column=e).style = 'title_style'

  #Top songs
  top_5_songs = table_1_2[:5]

  #Other songs
  all_columns = summary_years + statement_list
  other_songs = table_1_2[5:]
  other_songs_element = ['Other Songs']
  period_total = 0
  for z in range(1, len(all_columns) + 2):
    for w in other_songs:
      if w[z] == None:
        period_total += 0
      else:
        period_total += w[z]
    other_songs_element.append(period_total)
    period_total = 0

  #Add table
  top_5_table = top_5_songs + [other_songs_element]
  table_rows_1_2 = ws_1_2[4:len(top_5_table) + 3]
  for row, g in zip(table_rows_1_2, top_5_table):
    for cell, h in zip(row[1:], g):
      cell.value = h
      cell.style = 'number_style'
  for row in range(4, len(top_5_table) + 4):
    ws_1_2.cell(row=row, column=2).style = 'name_style'
    ws_1_2.cell(row=row, column=len(column_names_1_2)+1).style = 'total_style'

  #Total row
  for i, j in zip(range(3, len(column_names_1_2) + 2), column_letters[2:]):
    ws_1_2.cell(row=len(top_5_table) + 4, column=i).value = "=SUM({}{}:{}{})".format(j, 4, j, len(top_5_table) + 3)
    ws_1_2.cell(row=len(top_5_table) + 4, column=i).style = 'total_row_style'
  ws_1_2.cell(row=len(top_5_table) + 4, column=2).value = 'Grand Total'
  ws_1_2.cell(row=len(top_5_table) + 4, column=2).style = 'total_label_style'

  #Add % of revenue column
  total_column_letter = column_letters[len(column_names_1_2)-1]
  for q in range(4, len(top_5_table) + 4):
    ws_1_2.cell(row=q, column=len(column_names_1_2) + 2).value = "=({}{}/{}{})".format(total_column_letter, q,
                                                                                       total_column_letter,
                                                                                       len(top_5_table) + 4)
    ws_1_2.cell(row=q, column=len(column_names_1_2) + 2).style = 'Percent'
  ws_1_2.cell(row=2, column=len(column_names_1_2) + 2).value = '% Of Revenue'
  ws_1_2.cell(row=3, column=len(column_names_1_2) + 2).value = '(Most Recent Period)'
  ws_1_2.cell(row=2, column=len(column_names_1_2) + 2).style = 'title_style'
  ws_1_2.cell(row=3, column=len(column_names_1_2) + 2).style = 'title_style'
  ws_1_2.cell(row=len(top_5_table) + 4, column=len(column_names_1_2) + 2).style = 'total_row_style'

  #Add artists and release dates
  top_5_artists = artist_list[:5]
  for f, g in zip(range(4, len(top_5_artists) + 3), top_5_artists):
    for h, i in zip(range(len(column_names_1_2) + 3, len(column_names_1_2) + 5), range(1, len(g))):
      ws_1_2.cell(row=f, column=h).value = g[i]
      ws_1_2.cell(row=f, column=h).style = 'number_style'

  ws_1_2.cell(row=3, column=len(column_names_1_2) + 3).value = 'Release Artist'
  ws_1_2.cell(row=2, column=len(column_names_1_2) + 3).style = 'title_style'
  ws_1_2.cell(row=3, column=len(column_names_1_2) + 3).style = 'title_style'
  ws_1_2.cell(row=3, column=len(column_names_1_2) + 4).value = 'Release Date'
  ws_1_2.cell(row=2, column=len(column_names_1_2) + 4).style = 'title_style'
  ws_1_2.cell(row=3, column=len(column_names_1_2) + 4).style = 'title_style'
  ws_1_2.cell(row=len(top_5_artists) + 5, column=len(column_names_1_2) + 3).style = 'total_row_style'
  ws_1_2.cell(row=len(top_5_artists) + 5, column=len(column_names_1_2) + 4).style = 'total_row_style'

  #Outlines
  thin = Side(border_style="thin", color="000000")
  if first_year_complete:
    for k in range(len(summary_years)+1, (len(title_names_1)-2)*divisions+2, divisions):
      for l in range(2, len(top_5_table) + 5):
        ws_1_2.cell(row=l, column=k + 2).border = Border(left=thin)

    for m in range(2, len(column_names_1_2) + 2):
      ws_1_2.cell(row=1, column=m).border = Border(bottom=thin)
      ws_1_2.cell(row=3, column=m).border = Border(bottom=thin)
      ws_1_2.cell(row=len(top_5_table) + 3, column=m).border = Border(bottom=thin)
      ws_1_2.cell(row=len(top_5_table) + 4, column=m).border = Border(bottom=thin)

    for n in range(len(summary_years)+1, (len(title_names_1)-2)*divisions+2, divisions):
      ws_1_2.cell(row=3, column=n + 2).border = Border(left=thin, bottom=thin)
      ws_1_2.cell(row=len(top_5_table) + 3, column=n + 2).border = Border(left=thin, bottom=thin)
      ws_1_2.cell(row=len(top_5_table) + 4, column=n + 2).border = Border(left=thin, bottom=thin)

    for r in range(2, len(top_5_table) + 5):
      ws_1_2.cell(row=r, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(left=thin)
      ws_1_2.cell(row=r, column=1).border = Border(right=thin)
    ws_1_2.cell(row=3, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(
      left=thin, bottom=thin)
    ws_1_2.cell(row=2, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(
      left=thin, top=thin)
    ws_1_2.cell(row=len(top_5_table)+3, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(
      left=thin, bottom=thin)
    ws_1_2.cell(row=len(top_5_table)+4, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(
      left=thin, bottom=thin)

    for o in range(2, len(top_5_table) + 5):
      ws_1_2.cell(row=o, column=len(column_names_1_2) + 2).border = Border(left=thin)
      ws_1_2.cell(row=o, column=len(column_names_1_2) + 3).border = Border(left=thin)
      ws_1_2.cell(row=o, column=len(column_names_1_2) + 4).border = Border(left=thin)
      ws_1_2.cell(row=o, column=len(column_names_1_2) + 5).border = Border(left=thin)

    for p in range(2, len(summary_years) + 2):
      for q in range(2, len(top_5_table) + 5):
        ws_1_2.cell(row=q, column=p + 1).border = Border(left=thin)

    for s in range(2, len(summary_years) + 3):
      ws_1_2.cell(row=2, column=s).border = Border(left=thin, top=thin)
      ws_1_2.cell(row=3, column=s).border = Border(left=thin, bottom=thin)
      ws_1_2.cell(row=len(top_5_table) + 4, column=s).border = Border(left=thin, bottom=thin)
      ws_1_2.cell(row=len(top_5_table) + 3, column=s).border = Border(left=thin, bottom=thin)

    for t in range(len(column_names_1_2) + 2, len(column_names_1_2) + 5):
      ws_1_2.cell(row=1, column=t).border = Border(left=thin, bottom=thin)
      ws_1_2.cell(row=3, column=t).border = Border(left=thin, bottom=thin)
      ws_1_2.cell(row=len(top_5_table) + 4, column=t).border = Border(left=thin, bottom=thin)
      ws_1_2.cell(row=len(top_5_table) + 3, column=t).border = Border(left=thin, bottom=thin)

  else:
    for k in range(len(detailed_remainder)+1, (len(title_names_1) - 2) * divisions + 2, divisions):
      for l in range(2, len(top_5_table) + 5):
        ws_1_2.cell(row=l, column=k + 2).border = Border(left=thin)

    for m in range(2, len(column_names_1_2) + 2):
      ws_1_2.cell(row=1, column=m).border = Border(bottom=thin)
      ws_1_2.cell(row=3, column=m).border = Border(bottom=thin)
      ws_1_2.cell(row=len(top_5_table) + 3, column=m).border = Border(bottom=thin)
      ws_1_2.cell(row=len(top_5_table) + 4, column=m).border = Border(bottom=thin)

    for n in range(len(detailed_remainder)+1, (len(title_names_1) - 2) * divisions + 2, divisions):
      ws_1_2.cell(row=3, column=n + 2).border = Border(left=thin, bottom=thin)
      ws_1_2.cell(row=len(top_5_table) + 3, column=n + 2).border = Border(left=thin, bottom=thin)
      ws_1_2.cell(row=len(top_5_table) + 4, column=n + 2).border = Border(left=thin, bottom=thin)

    for r in range(2, len(top_5_table) + 5):
      ws_1_2.cell(row=r, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 3 + len(remainder))).border = Border(left=thin)
      ws_1_2.cell(row=r, column=1).border = Border(
        right=thin)
    ws_1_2.cell(row=3, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 3 + len(remainder))).border = Border(left=thin, bottom=thin)
    ws_1_2.cell(row=2, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 3 + len(remainder))).border = Border(left=thin, top=thin)
    ws_1_2.cell(row=len(top_5_table) + 3, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 3 + len(remainder))).border = Border(left=thin, bottom=thin)
    ws_1_2.cell(row=len(top_5_table) + 4, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 3 + len(remainder))).border = Border(left=thin, bottom=thin)

    for o in range(2, len(top_5_table) + 5):
      ws_1_2.cell(row=o, column=len(column_names_1_2) + 2).border = Border(left=thin)
      ws_1_2.cell(row=o, column=len(column_names_1_2) + 3).border = Border(left=thin)
      ws_1_2.cell(row=o, column=len(column_names_1_2) + 4).border = Border(left=thin)
      ws_1_2.cell(row=o, column=len(column_names_1_2) + 5).border = Border(left=thin)

    for q in range(2, len(top_5_table) + 5):
      ws_1_2.cell(row=q, column=3).border = Border(left=thin)

    for s in range(2, 4):
      ws_1_2.cell(row=2, column=s).border = Border(left=thin, top=thin)
      ws_1_2.cell(row=3, column=s).border = Border(left=thin, bottom=thin)
      ws_1_2.cell(row=len(top_5_table) + 4, column=s).border = Border(left=thin, bottom=thin)
      ws_1_2.cell(row=len(top_5_table) + 3, column=s).border = Border(left=thin, bottom=thin)

    for t in range(len(column_names_1_2) + 2, len(column_names_1_2) + 5):
      ws_1_2.cell(row=1, column=t).border = Border(left=thin, bottom=thin)
      ws_1_2.cell(row=3, column=t).border = Border(left=thin, bottom=thin)
      ws_1_2.cell(row=len(top_5_table) + 4, column=t).border = Border(left=thin, bottom=thin)
      ws_1_2.cell(row=len(top_5_table) + 3, column=t).border = Border(left=thin, bottom=thin)

#Sheet 2: Source x Rev x Half
  #Build table
  select_table_2_1 = '''SELECT Normalized_Source_9LC,'''
  select_table_2_2 = ""
  for c in summary_years:
    select_table_2_2 += '''sum( CASE WHEN Year_Statement_9LC = "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(c,c)
  for a, b in zip(year_statement_list, statement_list):
    select_table_2_2 += '''sum( CASE WHEN Year_Statement_9LC = "{}" AND {} = "{}"
                            THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(a, mySQL_column, b, a, b)
  select_table_2_3 = '''sum( CASE WHEN Year_Statement_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END)
                              AS `Total` FROM Master GROUP BY Normalized_Source_9LC ORDER BY `Total` DESC'''.format(cut_off)
  select_table_2 = select_table_2_1 + select_table_2_2 + select_table_2_3
  mycursor.execute(select_table_2)
  table_2 = mycursor.fetchall()

  #Title row list
  title_names_2 = ['{}'.format(database[:-25])]
  for c in all_years:
    title_names_2.append(c)

  #Column names list
  column_names_2 = ['Sources']
  for d in summary_years:
    column_names_2.append("")
  for e in statement_list:
    column_names_2.append(e)
  column_names_2.append('Grand Total')

  #Size of worksheet
  for column_no in range(2, len(column_names_2) + 3):
    for row_no in range(2, len(table_2) + 4):
      ws_2.cell(row=row_no, column=column_no)

  #Title row
  ws_2.cell(row=2, column=2).value = '{}'.format(title_names_2[0])
  ws_2.cell(row=2, column=2).style = 'title_style'
  if first_year_complete:
    for d, e in zip(range(len(summary_years)+1, (len(title_names_2) - 3) * divisions + 2, divisions), title_names_2[1+len(summary_years):]):
      ws_2.merge_cells('{}2:{}2'.format(column_letters[d + 1], column_letters[d + divisions]))
      ws_2.cell(row=2, column=d + 2).value = e
      ws_2.cell(row=2, column=d + 2).style = 'title_style'
    ws_2.cell(row=2, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3)).value = title_names_1[-1]
    ws_2.cell(row=2, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3)).style = 'title_style'
    ws_2.cell(row=2, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3) + len(remainder)).style = 'title_style'
    ws_2.merge_cells('{}2:{}2'.format(column_letters[len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+ 2],
                                              column_letters[len(summary_years) + divisions*(len(all_years)-len(summary_years)-1) + len(remainder)+1]))
    for y, z in zip(range(3, 3 + len(summary_years)), title_names_1[1:1 + len(summary_years)]):
      ws_2.cell(row=2, column=y).value = z
      ws_2.cell(row=2, column=y).style = 'title_style'

  else:
    ws_2.merge_cells('C2:{}2'.format(column_letters[len(detailed_remainder)+1]))
    ws_2.cell(row=2, column=3).value = title_names_2[1]
    ws_2.cell(row=2, column=3).style = 'title_style'
    for d, e in zip(range(len(detailed_remainder)+3, (len(title_names_2) - 3) * divisions + 4, divisions), title_names_2[2:]):
      ws_2.merge_cells('{}2:{}2'.format(column_letters[d + 1], column_letters[d + divisions]))
      ws_2.cell(row=2, column=d + 3).value = e
      ws_2.cell(row=2, column=d + 3).style = 'title_style'
    ws_2.cell(row=2, column=(len(detailed_remainder) + divisions*(len(all_years)-2)+3)).value = title_names_1[-1]
    ws_2.cell(row=2, column=(len(detailed_remainder) + divisions*(len(all_years)-2)+3)).style = 'title_style'
    ws_2.cell(row=2, column=(len(detailed_remainder) + divisions*(len(all_years)-2)+3) + len(remainder)).style = 'title_style'
    ws_2.merge_cells('{}2:{}2'.format(column_letters[len(detailed_remainder) + divisions*(len(all_years)-2)+ 2],
                                              column_letters[len(detailed_remainder) + divisions*(len(all_years)-2) + len(remainder)+1]))

  #Header row
  for e, f in zip(range(2, len(column_names_2) + 3), column_names_2):
    ws_2.cell(row=3, column=e).value = f
    ws_2.cell(row=3, column=e).style = 'title_style'

  #Add table
  table_rows_2 = ws_2[4:len(table_2) + 3]
  for row, g in zip(table_rows_2, table_2):
    for cell, h in zip(row[1:], g):
      cell.value = h
      cell.style = 'number_style'
  for row in range(4, len(table_2) + 4):
    ws_2.cell(row=row, column=2).style = 'name_style'
    ws_2.cell(row=row, column=len(column_names_2)+1).style = 'total_style'

  #Total row
  for i, j in zip(range(3, len(column_names_2) + 2), column_letters[2:]):
    ws_2.cell(row=len(table_2) + 4, column=i).value = "=SUM({}{}:{}{})".format(j, 4, j, len(table_2) + 3)
    ws_2.cell(row=len(table_2) + 4, column=i).style = 'total_row_style'
  ws_2.cell(row=len(table_2) + 4, column=2).value = 'Grand Total'
  ws_2.cell(row=len(table_2) + 4, column=2).style = 'total_label_style'

  #Outlines
  thin = Side(border_style="thin", color="000000")
  if first_year_complete:
    for k in range(len(summary_years)+1, (len(title_names_2)-2)*divisions+2, divisions):
      for l in range(2, len(table_2) + 5):
        ws_2.cell(row=l, column=k + 2).border = Border(left=thin)

    for m in range(2, len(column_names_2) + 2):
      ws_2.cell(row=1, column=m).border = Border(bottom=thin)
      ws_2.cell(row=3, column=m).border = Border(bottom=thin)
      ws_2.cell(row=len(table_2) + 3, column=m).border = Border(bottom=thin)
      ws_2.cell(row=len(table_2) + 4, column=m).border = Border(bottom=thin)

    for n in range(len(summary_years)+1, (len(title_names_2)-2)*divisions+2, divisions):
      ws_2.cell(row=3, column=n + 2).border = Border(left=thin, bottom=thin)
      ws_2.cell(row=len(table_2) + 3, column=n + 2).border = Border(left=thin, bottom=thin)
      ws_2.cell(row=len(table_2) + 4, column=n + 2).border = Border(left=thin, bottom=thin)

    for r in range(2, len(table_2) + 5):
      ws_2.cell(row=r, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(left=thin)
      ws_2.cell(row=r, column=1).border = Border(right=thin)
    ws_2.cell(row=3, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(
      left=thin, bottom=thin)
    ws_2.cell(row=2, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(
      left=thin, top=thin)
    ws_2.cell(row=len(table_2) + 4, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(
      left=thin, bottom=thin)
    ws_2.cell(row=len(table_2) + 3, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(
      left=thin, bottom=thin)

    for o in range(2, len(table_2) + 5):
      ws_2.cell(row=o, column=len(column_names_2) + 2).border = Border(left=thin)

    for p in range(2, len(summary_years) + 2):
      for q in range(2, len(table_2) + 5):
        ws_2.cell(row=q, column=p + 1).border = Border(left=thin)

    for s in range(2, len(summary_years) + 3):
      ws_2.cell(row=2, column=s).border = Border(left=thin, top=thin)
      ws_2.cell(row=3, column=s).border = Border(left=thin, bottom=thin)
      ws_2.cell(row=len(table_2) + 4, column=s).border = Border(left=thin, bottom=thin)
      ws_2.cell(row=len(table_2) + 3, column=s).border = Border(left=thin, bottom=thin)

  else:
    for k in range(len(detailed_remainder)+1, (len(title_names_2)-2)*divisions+2, divisions):
      for l in range(2, len(table_2) + 5):
        ws_2.cell(row=l, column=k + 2).border = Border(left=thin)

    for m in range(2, len(column_names_2) + 2):
      ws_2.cell(row=1, column=m).border = Border(bottom=thin)
      ws_2.cell(row=3, column=m).border = Border(bottom=thin)
      ws_2.cell(row=len(table_2) + 3, column=m).border = Border(bottom=thin)
      ws_2.cell(row=len(table_2) + 4, column=m).border = Border(bottom=thin)

    for n in range(len(detailed_remainder)+1, (len(title_names_2)-2)*divisions+2, divisions):
      ws_2.cell(row=3, column=n + 2).border = Border(left=thin, bottom=thin)
      ws_2.cell(row=len(table_2) + 3, column=n + 2).border = Border(left=thin, bottom=thin)
      ws_2.cell(row=len(table_2) + 4, column=n + 2).border = Border(left=thin, bottom=thin)

    for r in range(2, len(table_2) + 5):
      ws_2.cell(row=r, column=(len(detailed_remainder) + divisions*(len(all_years)-2)+3+len(remainder))).border = Border(left=thin)
      ws_2.cell(row=r, column=1).border = Border(
        right=thin)
    ws_2.cell(row=3, column=(len(detailed_remainder) + divisions*(len(all_years)-2)+3+len(remainder))).border = Border(left=thin, bottom=thin)
    ws_2.cell(row=2, column=(len(detailed_remainder) + divisions*(len(all_years)-2)+3+len(remainder))).border = Border(left=thin, top=thin)
    ws_2.cell(row=len(table_2) + 4, column=(len(detailed_remainder) + divisions*(len(all_years)-2)+3+len(remainder))).border = Border(left=thin, bottom=thin)
    ws_2.cell(row=len(table_2) + 3, column=(len(detailed_remainder) + divisions*(len(all_years)-2)+3+len(remainder))).border = Border(left=thin, bottom=thin)

    for o in range(2, len(table_2) + 5):
      ws_2.cell(row=o, column=len(column_names_2) + 2).border = Border(left=thin)

    for q in range(2, len(table_2) + 5):
      ws_2.cell(row=q, column=3).border = Border(left=thin)

    for s in range(2, 4):
      ws_2.cell(row=2, column=s).border = Border(left=thin, top=thin)
      ws_2.cell(row=3, column=s).border = Border(left=thin, bottom=thin)
      ws_2.cell(row=len(table_2) + 4, column=s).border = Border(left=thin, bottom=thin)
      ws_2.cell(row=len(table_2) + 3, column=s).border = Border(left=thin, bottom=thin)

#Sheet 2.1: (Top 10) Sources
  #Build table
  select_table_2_1_1 = '''SELECT Normalized_Source_9LC,'''
  select_table_2_1_2 = ""
  for c in summary_years:
    select_table_2_1_2 += '''sum( CASE WHEN Year_Statement_9LC = "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(
      c, c)
  for a, b in zip(year_statement_list, statement_list):
    select_table_2_1_2 += '''sum( CASE WHEN Year_Statement_9LC = "{}" AND {} = "{}"
                            THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(a, mySQL_column, b, a, b)
  select_table_2_1_3 = '''sum( CASE WHEN Year_Statement_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END)
                              AS `Total` FROM Master GROUP BY Normalized_Source_9LC ORDER BY `Total` DESC'''.format(
    cut_off)
  select_table_2_1 = select_table_2_1_1 + select_table_2_1_2 + select_table_2_1_3
  mycursor.execute(select_table_2_1)
  table_2_1 = mycursor.fetchall()

  #Top sources
  top_sources = table_2_1[:10]

  #Other sources
  other_sources = table_2_1[10:]
  other_sources_element = ['Other Sources']
  period_total = 0
  for z in range(1, len(all_columns) + 2):
    for w in other_sources:
      if w[z] == None:
        period_total += 0
      else:
        period_total += w[z]
    other_sources_element.append(period_total)
    period_total = 0

  #Final table
  top_sources.append((other_sources_element))

  #Title row list
  title_names_2_1 = ['{}'.format(database[:-25])]
  for c in all_years:
    title_names_2_1.append(c)

  #Column names list
  column_names_2_1 = ['Sources']
  for d in summary_years:
    column_names_2_1.append("")
  for e in statement_list:
    column_names_2_1.append(e)
  column_names_2_1.append('Grand Total')

  #Size of worksheet
  for column_no in range(2, len(column_names_2_1) + 3):
    for row_no in range(2, len(top_sources) + 4):
      ws_2_1.cell(row=row_no, column=column_no)

  #Title row
  ws_2_1.cell(row=2, column=2).value = '{}'.format(title_names_2_1[0])
  ws_2_1.cell(row=2, column=2).style = 'title_style'
  if first_year_complete:
    for d, e in zip(range(len(summary_years)+1, (len(title_names_2_1) - 3) * divisions + 2, divisions),
                    title_names_2_1[1 + len(summary_years):]):
      ws_2_1.merge_cells('{}2:{}2'.format(column_letters[d + 1], column_letters[d + divisions]))
      ws_2_1.cell(row=2, column=d + 2).value = e
      ws_2_1.cell(row=2, column=d + 2).style = 'title_style'
    ws_2_1.cell(row=2, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3)).value = title_names_1[-1]
    ws_2_1.cell(row=2, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3)).style = 'title_style'
    ws_2_1.cell(row=2, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).style = 'title_style'
    ws_2_1.merge_cells('{}2:{}2'.format(column_letters[len(summary_years) + divisions*(len(all_years)-len(summary_years)-1) + 2],
                            column_letters[len(summary_years) + divisions*(len(all_years)-len(summary_years)-1) + len(remainder)+1]))
    for y, z in zip(range(3, 3 + len(summary_years)), title_names_1[1:1 + len(summary_years)]):
      ws_2_1.cell(row=2, column=y).value = z
      ws_2_1.cell(row=2, column=y).style = 'title_style'

  else:
    ws_2_1.merge_cells('C2:{}2'.format(column_letters[len(detailed_remainder)+1]))
    ws_2_1.cell(row=2, column=3).value = title_names_1[1]
    ws_2_1.cell(row=2, column=3).style = 'title_style'
    for d, e in zip(range(len(detailed_remainder)+1, (len(title_names_2_1) - 3) * divisions + 2, divisions),
                    title_names_2_1[2:]):
      ws_2_1.merge_cells('{}2:{}2'.format(column_letters[d + 1], column_letters[d + divisions]))
      ws_2_1.cell(row=2, column=d + 3).value = e
      ws_2_1.cell(row=2, column=d + 3).style = 'title_style'
    ws_2_1.cell(row=2, column=(len(detailed_remainder) + divisions*(len(all_years)-2)+3)).value = title_names_1[-1]
    ws_2_1.cell(row=2, column=(len(detailed_remainder) + divisions*(len(all_years)-2)+3)).style = 'title_style'
    ws_2_1.cell(row=2, column=(len(detailed_remainder) + divisions*(len(all_years)-2)+3+len(remainder))).style = 'title_style'
    ws_2_1.merge_cells('{}2:{}2'.format(column_letters[len(detailed_remainder) + divisions*(len(all_years)-2) + 2],
                            column_letters[len(detailed_remainder) + divisions*(len(all_years)-2) + len(remainder)+1]))

  #Header row
  for e, f in zip(range(2, len(column_names_2_1) + 3), column_names_2_1):
    ws_2_1.cell(row=3, column=e).value = f
    ws_2_1.cell(row=3, column=e).style = 'title_style'

  #Add table
  table_rows_2_1 = ws_2_1[4:len(top_sources) + 3]
  for row, g in zip(table_rows_2_1, top_sources):
    for cell, h in zip(row[1:], g):
      cell.value = h
      cell.style = 'number_style'
  for row in range(4, len(top_sources) + 4):
    ws_2_1.cell(row=row, column=2).style = 'name_style'
    ws_2_1.cell(row=row, column=len(column_names_2_1)+1).style = 'total_style'

  #Total row
  for i, j in zip(range(3, len(column_names_2_1) + 2), column_letters[2:]):
    ws_2_1.cell(row=len(top_sources) + 4, column=i).value = "=SUM({}{}:{}{})".format(j, 4, j, len(top_sources) + 3)
    ws_2_1.cell(row=len(top_sources) + 4, column=i).style = 'total_row_style'
  ws_2_1.cell(row=len(top_sources) + 4, column=2).value = 'Grand Total'
  ws_2_1.cell(row=len(top_sources) + 4, column=2).style = 'total_label_style'

  #Add % of revenue column
  total_column_letter = column_letters[len(column_names_2_1) - 1]
  for q in range(4, len(top_sources) + 4):
    ws_2_1.cell(row=q, column=len(column_names_2_1) + 2).value = "=({}{}/{}{})".format(total_column_letter, q,
                                                                                       total_column_letter,
                                                                                       len(top_sources) + 4)
    ws_2_1.cell(row=q, column=len(column_names_2_1) + 2).style = 'Percent'
  ws_2_1.cell(row=2, column=len(column_names_2_1) + 2).value = '% Of Revenue'
  ws_2_1.cell(row=3, column=len(column_names_2_1) + 2).value = '(Most Recent Period)'
  ws_2_1.cell(row=2, column=len(column_names_2_1) + 2).style = 'title_style'
  ws_2_1.cell(row=3, column=len(column_names_2_1) + 2).style = 'title_style'
  ws_2_1.cell(row=len(top_sources) + 4, column=len(column_names_2_1) + 2).style = 'total_row_style'

  #Outlines
  thin = Side(border_style="thin", color="000000")
  if first_year_complete:
    for k in range(len(summary_years)+1, (len(title_names_1)-2)*divisions+2, divisions):
      for l in range(2, len(top_sources) + 5):
        ws_2_1.cell(row=l, column=k + 2).border = Border(left=thin)

    for m in range(2, len(column_names_2_1) + 2):
      ws_2_1.cell(row=1, column=m).border = Border(bottom=thin)
      ws_2_1.cell(row=3, column=m).border = Border(bottom=thin)
      ws_2_1.cell(row=len(top_sources) + 3, column=m).border = Border(bottom=thin)
      ws_2_1.cell(row=len(top_sources) + 4, column=m).border = Border(bottom=thin)

    for n in range(len(summary_years)+1, (len(title_names_1)-2)*divisions+2, divisions):
      ws_2_1.cell(row=3, column=n + 2).border = Border(left=thin, bottom=thin)
      ws_2_1.cell(row=len(top_sources) + 3, column=n + 2).border = Border(left=thin, bottom=thin)
      ws_2_1.cell(row=len(top_sources) + 4, column=n + 2).border = Border(left=thin, bottom=thin)

    for r in range(2, len(top_sources) + 5):
      ws_2_1.cell(row=r, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(left=thin)
      ws_2_1.cell(row=r, column=1).border = Border(right=thin)
    ws_2_1.cell(row=3, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(
      left=thin, bottom=thin)
    ws_2_1.cell(row=2, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(
      left=thin, top=thin)
    ws_2_1.cell(row=len(top_sources) + 4, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(
      left=thin, bottom=thin)
    ws_2_1.cell(row=len(top_sources) + 3, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(
      left=thin, bottom=thin)

    for o in range(2, len(top_sources) + 5):
      ws_2_1.cell(row=o, column=len(column_names_2_1) + 2).border = Border(left=thin)
      ws_2_1.cell(row=o, column=len(column_names_2_1) + 3).border = Border(left=thin)

    for p in range(2, len(summary_years) + 2):
      for q in range(2, len(top_sources) + 5):
        ws_2_1.cell(row=q, column=p + 1).border = Border(left=thin)

    for s in range(2, len(summary_years) + 3):
      ws_2_1.cell(row=2, column=s).border = Border(left=thin, top=thin)
      ws_2_1.cell(row=3, column=s).border = Border(left=thin, bottom=thin)
      ws_2_1.cell(row=len(top_sources) + 4, column=s).border = Border(left=thin, bottom=thin)
      ws_2_1.cell(row=len(top_sources) + 3, column=s).border = Border(left=thin, bottom=thin)

    for t in range(len(column_names_2_1) + 2, len(column_names_2_1) + 3):
      ws_2_1.cell(row=1, column=t).border = Border(bottom=thin)
      ws_2_1.cell(row=3, column=t).border = Border(left=thin, bottom=thin)
      ws_2_1.cell(row=len(top_sources) + 4, column=t).border = Border(left=thin, bottom=thin)
      ws_2_1.cell(row=len(top_sources) + 3, column=t).border = Border(left=thin, bottom=thin)

  else:
    for k in range(len(detailed_remainder)+1, (len(title_names_1) - 2) * divisions + 2, divisions):
      for l in range(2, len(top_sources) + 5):
        ws_2_1.cell(row=l, column=k + 2).border = Border(left=thin)

    for m in range(2, len(column_names_2_1) + 2):
      ws_2_1.cell(row=1, column=m).border = Border(bottom=thin)
      ws_2_1.cell(row=3, column=m).border = Border(bottom=thin)
      ws_2_1.cell(row=len(top_sources) + 3, column=m).border = Border(bottom=thin)
      ws_2_1.cell(row=len(top_sources) + 4, column=m).border = Border(bottom=thin)

    for n in range(len(detailed_remainder)+1, (len(title_names_1) - 2) * divisions + 2, divisions):
      ws_2_1.cell(row=3, column=n + 2).border = Border(left=thin, bottom=thin)
      ws_2_1.cell(row=len(top_sources) + 3, column=n + 2).border = Border(left=thin, bottom=thin)
      ws_2_1.cell(row=len(top_sources) + 4, column=n + 2).border = Border(left=thin, bottom=thin)

    for r in range(2, len(top_sources) + 5):
      ws_2_1.cell(row=r, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 3 + len(remainder))).border = Border(left=thin)
      ws_2_1.cell(row=r, column=1).border = Border(
        right=thin)
    ws_2_1.cell(row=3, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 3 + len(remainder))).border = Border(left=thin, bottom=thin)
    ws_2_1.cell(row=2, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 2 + len(remainder))).border = Border(left=thin, top=thin)
    ws_2_1.cell(row=len(top_sources) + 4, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 3 + len(remainder))).border = Border(left=thin, bottom=thin)
    ws_2_1.cell(row=len(top_sources) + 3, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 3 + len(remainder))).border = Border(left=thin, bottom=thin)

    for o in range(2, len(top_sources) + 5):
      ws_2_1.cell(row=o, column=len(column_names_2_1) + 2).border = Border(left=thin)
      ws_2_1.cell(row=o, column=len(column_names_2_1) + 3).border = Border(left=thin)

    for q in range(2, len(top_sources) + 5):
      ws_2_1.cell(row=q, column=3).border = Border(left=thin)

    for s in range(2, 4):
      ws_2_1.cell(row=2, column=s).border = Border(left=thin, top=thin)
      ws_2_1.cell(row=3, column=s).border = Border(left=thin, bottom=thin)
      ws_2_1.cell(row=len(top_sources) + 4, column=s).border = Border(left=thin, bottom=thin)
      ws_2_1.cell(row=len(top_sources) + 3, column=s).border = Border(left=thin, bottom=thin)

    for t in range(len(column_names_2_1) + 2, len(column_names_2_1) + 3):
      ws_2_1.cell(row=1, column=t).border = Border(bottom=thin)
      ws_2_1.cell(row=3, column=t).border = Border(left=thin, bottom=thin)
      ws_2_1.cell(row=len(top_sources) + 4, column=t).border = Border(left=thin, bottom=thin)
      ws_2_1.cell(row=len(top_sources) + 3, column=t).border = Border(left=thin, bottom=thin)

#Sheet 3: (Top 10) Source List by Third Party
  #Title row list
  title_names_3 = ['{}'.format(database[:-25])]
  for c in all_years:
    title_names_3.append(c)

  #Column names
  column_names_3 = ['Sources']
  for d in summary_years:
    column_names_3.append("")
  for e in statement_list:
    column_names_3.append(e)
  column_names_3.append('Grand Total')

  #Build tables
  total_row_no = -3
  ws_3.insert_cols(1, len(column_names_3)+1)
  for s in third_party_list:
    select_table_3_1 = '''SELECT Normalized_Source_9LC,'''
    select_table_3_2 = ""
    for c in summary_years:
      select_table_3_2 += '''sum( CASE WHEN Year_Statement_9LC = "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(c,c)
    for j, k in zip(year_statement_list, statement_list):
      select_table_3_2 += (''' sum( CASE WHEN Year_Statement_9LC = "{}" AND {} = "{}"
                             THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(j,mySQL_column,k,j,k))
    select_table_3_3 = '''sum( CASE WHEN Year_Statement_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `Total`
                        FROM Master WHERE Third_Party_9LC = "{}" AND Statement_Period_Half_9lC <> ""
                        GROUP BY Third_Party_9LC, Normalized_Source_9LC ORDER BY `Total` DESC'''.format(cut_off,s)
    select_table_3 = select_table_3_1 + select_table_3_2 + select_table_3_3
    mycursor.execute(select_table_3)
    table_3 = mycursor.fetchall()

    #Add tables to sheet
    total_row_no += len(table_3) + 5
    space_row_no = total_row_no + 1
    header_row_no = total_row_no - len(table_3)
    total_column_no = len(column_names_3)
    third_party_table_rows = ws_3[header_row_no+3:space_row_no+1]
    if len(table_3) > 1:
      for row,n in zip(third_party_table_rows, table_3):
        for cell,o in zip(row[1:], range(len(n))):
          cell.value = n[o]
          cell.style = 'number_style'
      for row in range(header_row_no+2, total_row_no+3):
        ws_3.cell(row=row, column=2).style = 'name_style'
        ws_3.cell(row=row, column=len(column_names_3)+1).style = 'total_style'
    else:
      for n in table_3:
        for cell, o in zip(third_party_table_rows, range(len(n))):
          cell.value = n[o]
          cell.style = 'number_style'
      for row in range(header_row_no + 2, total_row_no + 3):
        ws_3.cell(row=row, column=2).style = 'name_style'
        ws_3.cell(row=row, column=len(column_names_3)+1).style = 'total_style'

    #Title row
    ws_3.cell(row=header_row_no, column=2).value = '{}'.format(title_names_3[0])
    ws_3.cell(row=header_row_no, column=2).style = 'title_style'
    if first_year_complete:
      for d, e in zip(range(len(summary_years)+1, (len(title_names_3) - 3) * divisions + 2, divisions), title_names_3[1 + len(summary_years):]):
        ws_3.merge_cells('{}{}:{}{}'.format(column_letters[d + 1], header_row_no, column_letters[d + divisions],header_row_no))
        ws_3.cell(row=header_row_no, column=d + 2).value = e
        ws_3.cell(row=header_row_no, column=d + 2).style = 'title_style'
      ws_3.cell(row=header_row_no, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3)).value = title_names_1[-1]
      ws_3.cell(row=header_row_no, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3)).style = 'title_style'
      ws_3.cell(row=header_row_no, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3 + len(remainder))).style = 'title_style'
      ws_3.merge_cells('{}{}:{}{}'.format(column_letters[len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+2], header_row_no,
                                  column_letters[len(summary_years) + divisions*(len(all_years)-len(summary_years)-1) + len(remainder)+1],header_row_no))
      for y, z in zip(range(3, 3 + len(summary_years)), title_names_3[1:1 + len(summary_years)]):
        ws_3.cell(row=header_row_no, column=y).value = z
        ws_3.cell(row=header_row_no, column=y).style = 'title_style'

    else:
      ws_3.merge_cells('C{}:{}{}'.format(header_row_no,column_letters[len(detailed_remainder)+1],header_row_no))
      ws_3.cell(row=header_row_no, column=3).value = title_names_3[1]
      ws_3.cell(row=header_row_no, column=3).style = 'title_style'
      for d, e in zip(range(len(detailed_remainder)+1, (len(title_names_3) - 3) * divisions + 2, divisions), title_names_3[2:]):
        ws_3.merge_cells('{}{}:{}{}'.format(column_letters[d + 1],header_row_no+1, column_letters[d + divisions],header_row_no))
        ws_3.cell(row=header_row_no, column=d + 2).value = e
        ws_3.cell(row=header_row_no, column=d + 2).style = 'title_style'
      ws_3.cell(row=header_row_no, column=(len(detailed_remainder) + divisions*(len(all_years)-2)+3)).value = title_names_1[-1]
      ws_3.cell(row=header_row_no, column=(len(detailed_remainder) + divisions*(len(all_years)-2)+3)).style = 'title_style'
      ws_3.cell(row=header_row_no, column=(len(detailed_remainder) + divisions*(len(all_years)-2)+3 + len(remainder))).style = 'title_style'
      ws_3.merge_cells('{}{}:{}{}'.format(column_letters[len(detailed_remainder) + divisions*(len(all_years)-2)+2], header_row_no,
                                  column_letters[len(detailed_remainder) + divisions*(len(all_years)-2) + len(remainder)+1],header_row_no))



    #Header row
    for e, f in zip(range(2, len(column_names_3) + 3), column_names_3):
      ws_3.cell(row=header_row_no+1, column=e).value = f
      ws_3.cell(row=header_row_no+1, column=e).style = 'title_style'

    #Sub header row
    ws_3.merge_cells('B{}:{}{}'.format(header_row_no+2,column_letters[len(column_names_3)],header_row_no+2))
    ws_3.cell(row=header_row_no+2, column=2).value = '{}'.format(s)
    ws_3.cell(row=header_row_no + 2, column=2).style = 'publisher_label_style'

    #Total row
    ws_3.cell(row=total_row_no+3, column=2).value = 'Grand Total'
    ws_3.cell(row=total_row_no + 3, column=2).style = 'total_label_style'

    for g in range(3, total_column_no+2):
      ws_3.cell(row=total_row_no+3, column=g).value = '=SUM({}{}:{}{})'.format(column_letters[g-1], header_row_no+3, column_letters[g-1], total_row_no+2)
      ws_3.cell(row=total_row_no + 3, column=g).style = 'total_row_style'

    #Outlines
    thin = Side(border_style="thin", color="000000")
    if first_year_complete:
      for k in range(len(summary_years)+1, (len(title_names_1)-2)*divisions+2, divisions):
        for l in range(header_row_no, total_row_no+3):
          ws_3.cell(row=l, column=k + 2).border = Border(left=thin)

      for m in range(2, len(column_names_3) + 2):
        ws_3.cell(row=header_row_no+1, column=m).border = Border(bottom=thin)
        ws_3.cell(row=total_row_no+2, column=m).border = Border(bottom=thin)
        ws_3.cell(row=total_row_no+3, column=m).border = Border(bottom=thin)
        ws_3.cell(row=header_row_no, column=m).border = Border(top=thin)

      for n in range(len(summary_years)+1, (len(title_names_1)-2)*divisions+2, divisions):
        ws_3.cell(row=header_row_no, column=n + 2).border = Border(left=thin, top=thin)
        ws_3.cell(row=header_row_no+1, column=n + 2).border = Border(left=thin, bottom=thin)
        ws_3.cell(row=total_row_no+3, column=n + 2).border = Border(left=thin, bottom=thin)
        ws_3.cell(row=total_row_no+2, column=n + 2).border = Border(left=thin, bottom=thin)

      for o in range(2, total_row_no+4):
        ws_3.cell(row=o, column=len(column_names_3) + 2).border = Border(left=thin)

      for p in range(2, len(summary_years) + 2):
        for q in range(header_row_no, total_row_no+4):
          ws_3.cell(row=q, column=p + 1).border = Border(left=thin)

      for s in range(2, len(summary_years)+3):
        ws_3.cell(row=header_row_no, column=s).border = Border(left=thin, top=thin)
        ws_3.cell(row=header_row_no+1, column=s).border = Border(left=thin, bottom=thin)
        ws_3.cell(row=total_row_no+2, column=s).border = Border(left=thin, bottom=thin)
        ws_3.cell(row=total_row_no+3, column=s).border = Border(left=thin, bottom=thin)

      for r in range(header_row_no, total_row_no+4):
        ws_3.cell(row=r, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(left=thin)
        ws_3.cell(row=r, column=1).border = Border(right=thin)
      for a in range(2, len(column_names_3)+2):
        ws_3.cell(row=header_row_no+2, column=a).border = Border(bottom=thick)
      ws_3.cell(row=header_row_no+1, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(left=thin, bottom=thin)
      ws_3.cell(row=header_row_no, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(
        left=thin, top=thin)
      ws_3.cell(row=total_row_no+2, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(
        left=thin, bottom=thin)
      ws_3.cell(row=total_row_no+3, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(
        left=thin, bottom=thin)

    else:
      for k in range(len(detailed_remainder)+1, (len(title_names_1) - 2) * divisions + 2, divisions):
        for l in range(header_row_no, total_row_no + 3):
          ws_3.cell(row=l, column=k + 2).border = Border(left=thin)

      for m in range(2, len(column_names_3) + 2):
        ws_3.cell(row=header_row_no+1, column=m).border = Border(bottom=thin)
        ws_3.cell(row=total_row_no + 2, column=m).border = Border(bottom=thin)
        ws_3.cell(row=total_row_no + 3, column=m).border = Border(bottom=thin)
        ws_3.cell(row=header_row_no, column=m).border = Border(top=thin)

      for n in range(len(detailed_remainder)+1, (len(title_names_1) - 2) * divisions + 2, divisions):
        ws_3.cell(row=header_row_no, column=n + 2).border = Border(left=thin, top=thin)
        ws_3.cell(row=header_row_no+1, column=n + 2).border = Border(left=thin, bottom=thin)
        ws_3.cell(row=total_row_no + 3, column=n + 2).border = Border(left=thin, bottom=thin)
        ws_3.cell(row=total_row_no + 2, column=n + 2).border = Border(left=thin, bottom=thin)

      for o in range(2, total_row_no + 4):
        ws_3.cell(row=o, column=len(column_names_3) + 2).border = Border(left=thin)

      for q in range(header_row_no, total_row_no + 4):
        ws_3.cell(row=q, column=3).border = Border(left=thin)

      for s in range(2, 4):
        ws_3.cell(row=header_row_no, column=s).border = Border(left=thin, top=thin)
        ws_3.cell(row=header_row_no + 1, column=s).border = Border(left=thin, bottom=thin)
        ws_3.cell(row=total_row_no + 2, column=s).border = Border(left=thin, bottom=thin)
        ws_3.cell(row=total_row_no + 3, column=s).border = Border(left=thin, bottom=thin)

      for r in range(header_row_no, total_row_no + 4):
        ws_3.cell(row=r, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 3 + len(remainder))).border = Border(left=thin)
        ws_3.cell(row=r, column=1).border = Border(
          right=thin)

      for a in range(2, len(column_names_3) + 2):
        ws_3.cell(row=header_row_no + 2, column=a).border = Border(bottom=thick)
      ws_3.cell(row=header_row_no + 1, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 3 + len(remainder))).border = Border(left=thin, bottom=thin)
      ws_3.cell(row=header_row_no, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 3 + len(remainder))).border = Border(left=thin, top=thin)
      ws_3.cell(row=total_row_no + 2, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 3 + len(remainder))).border = Border(left=thin, bottom=thin)
      ws_3.cell(row=total_row_no + 3, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 3 + len(remainder))).border = Border(left=thin, bottom=thin)

#Sheet 3.1: (Top 10) Source List by Third Party
  #Title row list
  title_names_3_1 = ['{}'.format(database[:-25])]
  for c in all_years:
    title_names_3_1.append(c)

  #Column names
  column_names_3_1 = ['Sources']
  for d in summary_years:
    column_names_3_1.append("")
  for e in statement_list:
    column_names_3_1.append(e)
  column_names_3_1.append('Grand Total')

  #Build tables
  total_row_no = -3
  ws_3_1.insert_cols(1, len(column_names_3_1)+1)
  for s in third_party_list:
    select_table_3_1_1 = '''SELECT Normalized_Source_9LC,'''
    select_table_3_1_2 = ""
    for c in summary_years:
      select_table_3_1_2 += '''sum( CASE WHEN Year_Statement_9LC = "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(
        c, c)
    for j, k in zip(year_statement_list, statement_list):
      select_table_3_1_2 += (''' sum( CASE WHEN Year_Statement_9LC = "{}" AND {} = "{}" 
                             THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(j, mySQL_column, k, j, k))
    select_table_3_1_3 = '''sum( CASE WHEN Year_Statement_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `Total`
                        FROM Master WHERE Third_Party_9LC = "{}" AND Statement_Period_Half_9lC <> ""
                        GROUP BY Third_Party_9LC, Normalized_Source_9LC ORDER BY `Total` DESC'''.format(cut_off, s)
    select_table_3_1 = select_table_3_1_1 + select_table_3_1_2 + select_table_3_1_3
    mycursor.execute(select_table_3_1)
    table_3_1 = mycursor.fetchall()

    #Top sources
    top_sources = table_3_1[:10]

    #Other sources
    other_sources = table_3_1[10:]
    other_sources_element = ['Other Sources']
    period_total = 0

    for z in range(1, len(all_columns)+2):
      for w in other_sources:
        if w[z] == None:
          period_total += 0
        else:
          period_total += w[z]
      other_sources_element.append(period_total)
      period_total = 0
    print(other_sources_element)

    #Final table
    top_sources.append((other_sources_element))

    #Add tables to sheet
    total_row_no += len(top_sources) + 5
    space_row_no = total_row_no + 1
    header_row_no = total_row_no - len(top_sources)
    total_column_no = len(column_names_3_1)
    third_party_table_rows = ws_3_1[header_row_no + 3:space_row_no+2]
    for row, n in zip(third_party_table_rows, top_sources):
      for cell, o in zip(row[1:], range(len(n))):
        cell.value = n[o]
        cell.style = 'number_style'
    for row in range(header_row_no + 2, total_row_no + 3):
      ws_3_1.cell(row=row, column=2).style = 'name_style'
      ws_3_1.cell(row=row, column=len(column_names_3_1)+1).style = 'total_style'

    #Title row
    ws_3_1.cell(row=header_row_no, column=2).value = '{}'.format(title_names_3_1[0])
    ws_3_1.cell(row=header_row_no, column=2).style = 'title_style'
    if first_year_complete:
      for d, e in zip(range(len(summary_years)+1, (len(title_names_3_1) - 3) * divisions + 2, divisions),
                      title_names_3_1[1 + len(summary_years):]):
        ws_3_1.merge_cells('{}{}:{}{}'.format(column_letters[d + 1], header_row_no, column_letters[d + divisions],
                                            header_row_no))
        ws_3_1.cell(row=header_row_no, column=d + 2).value = e
        ws_3_1.cell(row=header_row_no, column=d + 2).style = 'title_style'
      ws_3_1.cell(row=header_row_no, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3)).value = title_names_3_1[-1]
      ws_3_1.cell(row=header_row_no, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3)).style = 'title_style'
      ws_3_1.cell(row=header_row_no, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3) + len(remainder)).style = 'title_style'
      ws_3_1.merge_cells('{}{}:{}{}'.format(column_letters[len(summary_years) + divisions*(len(all_years)-len(summary_years)-1) + 2], header_row_no,
                            column_letters[len(summary_years) + divisions*(len(all_years)-len(summary_years)-1) + len(remainder)+1],header_row_no))
      for y, z in zip(range(3, 3 + len(summary_years)), title_names_3_1[1:1 + len(summary_years)]):
        ws_3_1.cell(row=header_row_no, column=y).value = z
        ws_3_1.cell(row=header_row_no, column=y).style = 'title_style'

    else:
      ws_3_1.merge_cells('C{}:{}{}'.format(header_row_no,column_letters[len(detailed_remainder)+1],header_row_no))
      ws_3_1.cell(row=2, column=3).value = title_names_1[1]
      ws_3_1.cell(row=2, column=3).style = 'title_style'
      for d, e in zip(range(len(detailed_remainder)+1, (len(title_names_3_1) - 3) * divisions + 2, divisions),
                      title_names_3_1[2:]):
        ws_3_1.merge_cells('{}{}:{}{}'.format(column_letters[d + 1], header_row_no, column_letters[d + divisions],
                                            header_row_no))
        ws_3_1.cell(row=header_row_no, column=d + 2).value = e
        ws_3_1.cell(row=header_row_no, column=d + 2).style = 'title_style'
      ws_3_1.cell(row=header_row_no, column=(len(detailed_remainder) + divisions*(len(all_years)-2)+3)).value = title_names_3_1[-1]
      ws_3_1.cell(row=header_row_no, column=(len(detailed_remainder) + divisions*(len(all_years)-2)+3)).style = 'title_style'
      ws_3_1.cell(row=header_row_no, column=(len(detailed_remainder) + divisions*(len(all_years)-2)+3) + len(remainder)).style = 'title_style'
      ws_3_1.merge_cells('{}{}:{}{}'.format(column_letters[len(detailed_remainder) + divisions*(len(all_years)-2) + 2], header_row_no,
                            column_letters[len(detailed_remainder) + divisions*(len(all_years)-2) + len(remainder)+1],header_row_no))




    #Header row
    for e, f in zip(range(2, len(column_names_3_1) + 3), column_names_3_1):
      ws_3_1.cell(row=header_row_no+1, column=e).value = f
      ws_3_1.cell(row=header_row_no+1, column=e).style = 'title_style'

    #Sub header row
    ws_3_1.merge_cells('B{}:{}{}'.format(header_row_no + 2, column_letters[len(column_names_3_1)+1], header_row_no + 2))
    ws_3_1.cell(row=header_row_no + 2, column=2).value = '{}'.format(s)
    ws_3_1.cell(row=header_row_no + 2, column=2).style = 'publisher_label_style'

    #Total row
    ws_3_1.cell(row=total_row_no + 3, column=2).value = 'Grand Total'
    ws_3_1.cell(row=total_row_no + 3, column=2).style = 'total_label_style'

    for g in range(3, total_column_no + 2):
      ws_3_1.cell(row=total_row_no + 3, column=g).value = '=SUM({}{}:{}{})'.format(column_letters[g-1],
                                                                                 header_row_no + 3,
                                                                                 column_letters[g-1],
                                                                                 total_row_no + 2)
      ws_3_1.cell(row=total_row_no + 3, column=g).style = 'total_row_style'

    #Add % of revenue column
    total_column_letter = column_letters[len(column_names_3_1) - 1]
    for q in range(header_row_no + 3, total_row_no+3):
      ws_3_1.cell(row=q, column=len(column_names_3_1) + 2).value = "=({}{}/{}{})".format(total_column_letter, q,
                                                                                         total_column_letter,
                                                                                         total_row_no+3)
      ws_3_1.cell(row=q, column=len(column_names_3_1) + 2).style = 'Percent'
    ws_3_1.cell(row=header_row_no, column=len(column_names_3_1) + 2).value = '% Of Revenue'
    ws_3_1.cell(row=header_row_no+1, column=len(column_names_3_1) + 2).value = '(Most Recent Period)'
    ws_3_1.cell(row=header_row_no, column=len(column_names_3_1) + 2).style = 'title_style'
    ws_3_1.cell(row=header_row_no+1, column=len(column_names_3_1) + 2).style = 'title_style'
    ws_3_1.cell(row=total_row_no+3, column=len(column_names_3_1) + 2).style = 'total_row_style'

    #Outlines
    thin = Side(border_style="thin", color="000000")
    if first_year_complete:
      for k in range(len(summary_years)+1, (len(title_names_1)-2)*divisions+2, divisions):
        for l in range(header_row_no, total_row_no + 3):
          ws_3_1.cell(row=l, column=k + 2).border = Border(left=thin)

      for m in range(2, len(column_names_3_1) + 2):
        ws_3_1.cell(row=header_row_no+1, column=m).border = Border(bottom=thin)
        ws_3_1.cell(row=total_row_no + 2, column=m).border = Border(bottom=thin)
        ws_3_1.cell(row=total_row_no + 3, column=m).border = Border(bottom=thin)
        ws_3_1.cell(row=header_row_no, column=m).border = Border(top=thin)

      for n in range(len(summary_years)+1, (len(title_names_1)-2)*divisions+2, divisions):
        ws_3_1.cell(row=header_row_no, column=n + 2).border = Border(left=thin, top=thin)
        ws_3_1.cell(row=header_row_no+1, column=n + 2).border = Border(left=thin, bottom=thin)
        ws_3_1.cell(row=total_row_no + 3, column=n + 2).border = Border(left=thin, bottom=thin)
        ws_3_1.cell(row=total_row_no + 2, column=n + 2).border = Border(left=thin, bottom=thin)

      for o in range(header_row_no, total_row_no + 4):
        ws_3_1.cell(row=o, column=len(column_names_3_1) + 2).border = Border(left=thin)
        ws_3_1.cell(row=o, column=len(column_names_3_1) + 3).border = Border(left=thin)

      for p in range(2, len(summary_years) + 2):
        for q in range(header_row_no, total_row_no + 4):
          ws_3_1.cell(row=q, column=p + 1).border = Border(left=thin)

      for s in range(2, len(summary_years) + 3):
        ws_3_1.cell(row=header_row_no, column=s).border = Border(left=thin, top=thin)
        ws_3_1.cell(row=header_row_no+1, column=s).border = Border(left=thin, bottom=thin)
        ws_3_1.cell(row=total_row_no + 2, column=s).border = Border(left=thin, bottom=thin)
        ws_3_1.cell(row=total_row_no + 3, column=s).border = Border(left=thin, bottom=thin)

      for r in range(header_row_no, total_row_no + 4):
        ws_3_1.cell(row=r, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(left=thin)
        ws_3_1.cell(row=r, column=1).border = Border(right=thin)
      for a in range(2, len(column_names_3_1) + 3):
        ws_3_1.cell(row=header_row_no + 2, column=a).border = Border(bottom=thick)
      ws_3_1.cell(row=header_row_no+1, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(
        left=thin, bottom=thin)
      ws_3_1.cell(row=header_row_no, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(
        left=thin, top=thin)
      ws_3_1.cell(row=total_row_no + 2, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(
        left=thin, bottom=thin)
      ws_3_1.cell(row=total_row_no + 3, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(
        left=thin, bottom=thin)

      for t in range(len(column_names_3_1) + 2, len(column_names_3_1) + 3):
        ws_3_1.cell(row=header_row_no+1, column=t).border = Border(left=thin, bottom=thin)
        ws_3_1.cell(row=total_row_no+2, column=t).border = Border(left=thin, bottom=thin)
        ws_3_1.cell(row=total_row_no+3, column=t).border = Border(left=thin, bottom=thin)
        ws_3_1.cell(row=header_row_no, column=t).border = Border(top=thin, left=thin)

    else:
      for k in range(len(detailed_remainder)+1, (len(title_names_1)-2)*divisions+2, divisions):
        for l in range(header_row_no, total_row_no + 3):
          ws_3_1.cell(row=l, column=k + 2).border = Border(left=thin)

      for m in range(2, len(column_names_3_1) + 2):
        ws_3_1.cell(row=header_row_no+1, column=m).border = Border(bottom=thin)
        ws_3_1.cell(row=total_row_no + 2, column=m).border = Border(bottom=thin)
        ws_3_1.cell(row=total_row_no + 3, column=m).border = Border(bottom=thin)
        ws_3_1.cell(row=header_row_no, column=m).border = Border(top=thin)

      for n in range(len(detailed_remainder)+1, (len(title_names_1)-2)*divisions+2, divisions):
        ws_3_1.cell(row=header_row_no, column=n + 2).border = Border(left=thin, top=thin)
        ws_3_1.cell(row=header_row_no+1, column=n + 2).border = Border(left=thin, bottom=thin)
        ws_3_1.cell(row=total_row_no + 3, column=n + 2).border = Border(left=thin, bottom=thin)
        ws_3_1.cell(row=total_row_no + 2, column=n + 2).border = Border(left=thin, bottom=thin)

      for o in range(header_row_no, total_row_no + 4):
        ws_3_1.cell(row=o, column=len(column_names_3_1) + 2).border = Border(left=thin)
        ws_3_1.cell(row=o, column=len(column_names_3_1) + 3).border = Border(left=thin)

      for q in range(header_row_no, total_row_no + 4):
        ws_3_1.cell(row=q, column=3).border = Border(left=thin)

      for s in range(2, 4):
        ws_3_1.cell(row=header_row_no, column=s).border = Border(left=thin, top=thin)
        ws_3_1.cell(row=header_row_no+1, column=s).border = Border(left=thin, bottom=thin)
        ws_3_1.cell(row=total_row_no + 2, column=s).border = Border(left=thin, bottom=thin)
        ws_3_1.cell(row=total_row_no + 3, column=s).border = Border(left=thin, bottom=thin)

      for r in range(header_row_no, total_row_no + 4):
        ws_3_1.cell(row=r, column=(len(detailed_remainder) + divisions*(len(all_years)-2)+3+len(remainder))).border = Border(left=thin)
        ws_3_1.cell(row=r, column=1).border = Border(
          right=thin)
      for a in range(2, len(column_names_3_1) + 3):
        ws_3_1.cell(row=header_row_no + 2, column=a).border = Border(bottom=thick)
      ws_3_1.cell(row=header_row_no+1, column=(len(detailed_remainder) + divisions*(len(all_years)-2)+3+len(remainder))).border = Border(
        left=thin, bottom=thin)
      ws_3_1.cell(row=header_row_no, column=(len(detailed_remainder) + divisions*(len(all_years)-2)+3+len(remainder))).border = Border(
        left=thin, top=thin)
      ws_3_1.cell(row=total_row_no + 2, column=(len(detailed_remainder) + divisions*(len(all_years)-2)+3+len(remainder))).border = Border(
        left=thin, bottom=thin)
      ws_3_1.cell(row=total_row_no + 3, column=(len(detailed_remainder) + divisions*(len(all_years)-2)+3+len(remainder))).border = Border(
        left=thin, bottom=thin)

      for t in range(len(column_names_3_1) + 2, len(column_names_3_1) + 3):
        ws_3_1.cell(row=header_row_no+1, column=t).border = Border(left=thin, bottom=thin)
        ws_3_1.cell(row=total_row_no+2, column=t).border = Border(left=thin, bottom=thin)
        ws_3_1.cell(row=total_row_no+3, column=t).border = Border(left=thin, bottom=thin)
        ws_3_1.cell(row=header_row_no, column=t).border = Border(top=thin, left=thin)


#Sheet 4: Revenue Over Time
  master_total_list = []
  #Build table
  select_table_4_1 = '''SELECT Normalized_Income_Type_9LC,'''
  select_table_4_2 = ""
  for c in summary_years:
    select_table_4_2 += '''sum( CASE WHEN Year_Statement_9LC = "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(
      c, c)
  for a, b in zip(year_statement_list, statement_list):
    select_table_4_2 += '''sum( CASE WHEN Year_Statement_9LC = "{}" AND {} = "{}"
                              THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(a, mySQL_column, b, a, b)
  select_table_4_3 = '''sum( CASE WHEN Year_Statement_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END)
                                AS `Total` FROM Master GROUP BY Normalized_Income_Type_9LC ORDER BY `Total` DESC'''.format(cut_off)
  select_table_4 = select_table_4_1 + select_table_4_2 + select_table_4_3
  mycursor.execute(select_table_4)
  table_4 = mycursor.fetchall()
  print(table_4)

  #Total per period
  total_list = ['Total']
  for i in range(1,len(summary_years)+len(statement_list)+2):
    period_total = 0
    for j in table_4:
      if j[i] == None:
        period_total += 0
      else:
        period_total += j[i]
    total_list.append(period_total)
  master_total_list.append(total_list)
  print(total_list)

  #Title row list
  title_names_4 = ['{}'.format(database[:-25])]
  for c in all_years:
    title_names_4.append(c)

  #Column names list
  column_names_4 = [""]
  for d in summary_years:
    column_names_4.append("")
  for e in statement_list:
    column_names_4.append(e)
  column_names_4.append('Grand Total')

  #Size of worksheet
  for column_no in range(1, len(column_names_4) + 3):
    for row_no in range(1, len(table_4) + 4):
      ws_4.cell(row=row_no, column=column_no)

  #Title row
  ws_4.cell(row=2, column=2).value = '{}'.format(title_names_4[0])
  ws_4.cell(row=2, column=2).style = 'title_style'
  if first_year_complete:
    for d, e in zip(range(len(summary_years)+1, (len(title_names_4) - 3) * divisions + 2, divisions),
                    title_names_4[1 + len(summary_years):]):
      ws_4.merge_cells('{}2:{}2'.format(column_letters[d + 1], column_letters[d + divisions]))
      ws_4.cell(row=2, column=d + 2).value = e
      ws_4.cell(row=2, column=d + 2).style = 'title_style'
    ws_4.cell(row=2, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3)).value = title_names_4[-1]
    ws_4.cell(row=2, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3)).style = 'title_style'
    ws_4.cell(row=2, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3) + len(remainder)).style = 'title_style'
    ws_4.merge_cells('{}2:{}2'.format(column_letters[len(summary_years) + divisions*(len(all_years)-len(summary_years)-1) + 2],
                     column_letters[len(summary_years) + divisions*(len(all_years)-len(summary_years)-1) + len(remainder)+1]))

  else:
    ws_4.merge_cells('C2:{}2'.format(column_letters[len(detailed_remainder)+1]))
    ws_4.cell(row=2, column=3).value = title_names_1[1]
    ws_4.cell(row=2, column=3).style = 'title_style'
    for d, e in zip(range(len(detailed_remainder)+1, (len(title_names_4) - 3) * divisions + 2, divisions),
                    title_names_4[2:]):
      ws_4.merge_cells('{}2:{}2'.format(column_letters[d + 1], column_letters[d + divisions]))
      ws_4.cell(row=2, column=d + 2).value = e
      ws_4.cell(row=2, column=d + 2).style = 'title_style'
    ws_4.cell(row=2, column=(len(detailed_remainder) + divisions*(len(all_years)-2)+3)).value = title_names_4[-1]
    ws_4.cell(row=2, column=(len(detailed_remainder) + divisions*(len(all_years)-2)+3)).style = 'title_style'
    ws_4.cell(row=2, column=(len(detailed_remainder) + divisions*(len(all_years)-2)+3) + len(remainder)).style = 'title_style'
    ws_4.merge_cells('{}2:{}2'.format(column_letters[len(detailed_remainder) + divisions*(len(all_years)-2) + 2],
                     column_letters[len(detailed_remainder) + divisions*(len(all_years)-2) + len(remainder)+1]))

  for y, z in zip(range(3, 3 + len(summary_years)), title_names_4[1:1 + len(summary_years)]):
    ws_4.cell(row=2, column=y).value = z
    ws_4.cell(row=2, column=y).style = 'title_style'

  #Header row
  for e, f in zip(range(2, len(column_names_4) + 3), column_names_4):
    ws_4.cell(row=3, column=e).value = f
    ws_4.cell(row=3, column=e).style = 'title_style'

  #Add table
  for g, h in zip(total_list, range(2,len(column_names_4)+2)):
    ws_4.cell(row=4, column=h).value = g
    ws_4.cell(row=4, column=h).style = 'total_revenue_style'
  ws_4.cell(row=4, column=2).style = 'bold_name_style'

  #Outlines
  if first_year_complete:
    for i in range(2, len(column_names_4)+2):
      ws_4.cell(row=1, column=i).border = Border(bottom=thin)
      ws_4.cell(row=3, column=i).border = Border(bottom=thin)
      ws_4.cell(row=4, column=i).border = Border(bottom=thin)

    for k in range(len(summary_years)+1, (len(title_names_4) - 2) * divisions + 2, divisions):
      ws_4.cell(row=3, column=k + 2).border = Border(left=thin, bottom=thin)
      ws_4.cell(row=4, column=k + 2).border = Border(left=thin, bottom=thin)

    for r in range(2, 5):
      ws_4.cell(row=r, column=1).border = Border(right=thin)
      ws_4.cell(row=r, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(left=thin)
    ws_4.cell(row=3, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(
      left=thin, bottom=thin)
    ws_4.cell(row=2, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(
      left=thin, top=thin)
    ws_4.cell(row=4, column=(len(summary_years) + divisions*(len(all_years)-len(summary_years)-1)+3+len(remainder))).border = Border(
      left=thin, bottom=thin)

    for o in range(2, 5):
      ws_4.cell(row=o, column=len(column_names_4) + 2).border = Border(left=thin)

    for p in range(2, len(summary_years) + 2):
      for q in range(2, 5):
        ws_4.cell(row=q, column=p + 1).border = Border(left=thin)

    for q in range(len(summary_years)+1, (len(title_names_4)-2)*divisions+2, divisions):
      ws_4.cell(row=2, column=q+2).border = Border(left=thin)

    for r in range(3, len(summary_years)+3):
      ws_4.cell(row=3, column=r).border = Border(left=thin, bottom=thin)
      ws_4.cell(row=4, column=r).border = Border(left=thin, bottom=thin)

  else:
    for i in range(2, len(column_names_4) + 1):
      ws_4.cell(row=1, column=i).border = Border(bottom=thin)
      ws_4.cell(row=3, column=i).border = Border(bottom=thin)
      ws_4.cell(row=4, column=i).border = Border(bottom=thin)

    for k in range(len(detailed_remainder)+1, (len(title_names_4) - 2) * divisions + 2, divisions):
      ws_4.cell(row=3, column=k + 2).border = Border(left=thin, bottom=thin)
      ws_4.cell(row=4, column=k + 2).border = Border(left=thin, bottom=thin)

    for r in range(2, 5):
      ws_4.cell(row=r, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 3 + len(remainder))).border = Border(left=thin)
      ws_4.cell(row=r, column=1).border = Border(
        right=thin)
    ws_4.cell(row=3, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 3 + len(remainder))).border = Border(left=thin, bottom=thin)
    ws_4.cell(row=2, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 3 + len(remainder))).border = Border(left=thin, top=thin)
    ws_4.cell(row=4, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 3 + len(remainder))).border = Border(left=thin, bottom=thin)

    for o in range(2, 5):
      ws_4.cell(row=o, column=len(column_names_4) + 2).border = Border(left=thin)

    for q in range(2, 5):
      ws_4.cell(row=q, column=3).border = Border(left=thin)

    for q in range(len(detailed_remainder)+1, (len(title_names_4) - 2) * divisions + 2, divisions):
      ws_4.cell(row=2, column=q + 2).border = Border(left=thin)

    for r in range(3, 4):
      ws_4.cell(row=3, column=r).border = Border(left=thin, bottom=thin)
      ws_4.cell(row=4, column=r).border = Border(left=thin, bottom=thin)

#Sheet 4.1: Revenue Over Time By Third Party
  #Title row list
  title_names_4_1 = ['{}'.format(database[:-25])]
  for c in all_years:
    title_names_4_1.append(c)

  #Column names list
  column_names_4_1 = [""]
  for d in summary_years:
    column_names_4_1.append("")
  for e in statement_list:
    column_names_4_1.append(e)
  column_names_4_1.append('Grand Total')

  #Build table
  total_row_no = -1
  ws_4_1.insert_cols(1, len(column_names_4_1))
  for s in third_party_list:
    select_table_4_1_1 = '''SELECT Normalized_Income_Type_9LC,'''
    select_table_4_1_2 = ""
    for c in summary_years:
      select_table_4_1_2 += '''sum( CASE WHEN Year_Statement_9LC = "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(
        c, c)
    for a, b in zip(year_statement_list, statement_list):
      select_table_4_1_2 += '''sum( CASE WHEN Year_Statement_9LC = "{}" AND {} = "{}"
                                THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(a, mySQL_column, b, a, b)
    select_table_4_1_3 = '''sum( CASE WHEN Year_Statement_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END)
                                  AS `Total` FROM Master WHERE Third_Party_9LC = "{}" GROUP BY Normalized_Income_Type_9LC ORDER BY `Total` DESC'''.format(
      cut_off, s)
    select_table_4_1 = select_table_4_1_1 + select_table_4_1_2 + select_table_4_1_3
    mycursor.execute(select_table_4_1)
    table_4_1 = mycursor.fetchall()

    #Total per period
    total_list = ['Total']
    for i in range(1, len(summary_years) + len(statement_list) + 2):
      period_total = 0
      for j in table_4_1:
        if j[i] == None:
          period_total += 0
        else:
          period_total += j[i]
      total_list.append(period_total)
    master_total_list.append(total_list)
    print(master_total_list)

    #Add tables to sheet
    total_row_no += 5
    space_row_no = total_row_no + 1
    header_row_no = total_row_no - 1
    total_column_no = len(column_names_4_1)
    third_party_table_rows = ws_4_1[header_row_no + 3: space_row_no+1]
    for g, h in zip(total_list, range(2, len(column_names_4_1) + 2)):
      ws_4_1.cell(row=header_row_no+2, column=h).value = g
      ws_4_1.cell(row=header_row_no+2, column=h).style = 'total_revenue_style'
    ws_4_1.cell(row=header_row_no+2, column=2).style = 'bold_name_style'

    #Third party labels
    ws_4_1.cell(row=header_row_no-1, column=2).value = '{}'.format(s)
    ws_4_1.cell(row=header_row_no-1, column=2).style = 'sub_header_style'

    #Title row
    ws_4_1.cell(row=header_row_no, column=2).value = '{}'.format(title_names_4_1[0])
    ws_4_1.cell(row=header_row_no, column=2).style = 'title_style'
    if first_year_complete:
      for d, e in zip(range(len(summary_years)+1, (len(title_names_4_1) - 3) * divisions + 2, divisions),
                      title_names_4_1[1 + len(summary_years):]):
        ws_4_1.merge_cells('{}{}:{}{}'.format(column_letters[d + 1],header_row_no, column_letters[d + divisions],header_row_no))
        ws_4_1.cell(row=header_row_no, column=d + 2).value = e
        ws_4_1.cell(row=header_row_no, column=d + 2).style = 'title_style'
      ws_4_1.cell(row=header_row_no, column=(len(summary_years) + divisions * (len(all_years) - len(summary_years) - 1) + 3)).value = title_names_4_1[-1]
      ws_4_1.cell(row=header_row_no, column=(len(summary_years) + divisions * (len(all_years) - len(summary_years) - 1) + 3)).style = 'title_style'
      ws_4_1.cell(row=header_row_no, column=(len(summary_years) + divisions * (len(all_years) - len(summary_years) - 1) + 3) + len(remainder)).style = 'title_style'
      ws_4_1.merge_cells('{}{}:{}{}'.format(column_letters[len(summary_years) + divisions * (len(all_years) - len(summary_years) - 1) + 2],header_row_no,
                                        column_letters[len(summary_years) + divisions * (len(all_years) - len(summary_years) - 1) + len(remainder)+1],header_row_no))
      for y, z in zip(range(3, 3 + len(summary_years)), title_names_4_1[1:1 + len(summary_years)]):
        ws_4_1.cell(row=header_row_no, column=y).value = z
        ws_4_1.cell(row=header_row_no, column=y).style = 'title_style'

    else:
      ws_4_1.merge_cells('C{}:{}{}'.format(header_row_no,column_letters[len(detailed_remainder)+1],header_row_no))
      ws_4_1.cell(row=header_row_no, column=3).value = title_names_4_1[1]
      ws_4_1.cell(row=header_row_no, column=3).style = 'title_style'
      for d, e in zip(range(len(detailed_remainder), (len(title_names_4_1) - 3) * divisions + 2, divisions),title_names_4_1[2:]):
        ws_4_1.merge_cells('{}{}:{}{}'.format(column_letters[d + 1],header_row_no, column_letters[d + divisions],header_row_no))
        ws_4_1.cell(row=header_row_no, column=d + 2).value = e
        ws_4_1.cell(row=header_row_no, column=d + 2).style = 'title_style'
      ws_4_1.cell(row=header_row_no, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 3)).value = title_names_4_1[-1]
      ws_4_1.cell(row=header_row_no, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 3)).style = 'title_style'
      ws_4_1.cell(row=header_row_no, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 3) + len(remainder)).style = 'title_style'
      ws_4_1.merge_cells('{}{}:{}{}'.format(column_letters[len(detailed_remainder) + divisions * (len(all_years) - 2) + 2],header_row_no,
                                        column_letters[len(detailed_remainder) + divisions * (len(all_years) - 2) + len(remainder)+1],header_row_no))



    #Header row
    for e, f in zip(range(2, len(column_names_4_1) + 3), column_names_4_1):
      ws_4_1.cell(row=header_row_no+1, column=e).value = f
      ws_4_1.cell(row=header_row_no+1, column=e).style = 'title_style'

    #Outlines
    if first_year_complete:
      for i in range(2, len(column_names_4_1) + 2):
        ws_4_1.cell(row=header_row_no+1, column=i).border = Border(bottom=thin)
        ws_4_1.cell(row=header_row_no+2, column=i).border = Border(bottom=thin)
        ws_4_1.cell(row=header_row_no, column=i).border = Border(top=thin)
        ws_4_1.cell(row=header_row_no-1, column=i).border = Border(bottom=thin)

      for k in range(len(summary_years)+1, (len(title_names_4_1) - 2) * divisions + 1, divisions):
        ws_4_1.cell(row=header_row_no+1, column=k + 2).border = Border(left=thin, bottom=thin)
        ws_4_1.cell(row=header_row_no, column=k + 2).border = Border(left=thin, top=thin)
        ws_4_1.cell(row=header_row_no+2, column=k + 2).border = Border(left=thin, bottom=thin)

      for r in range(header_row_no, header_row_no+3):
        ws_4_1.cell(row=r, column=(len(summary_years) + divisions * (len(all_years) - len(summary_years) - 1) + 3 + len(
          remainder))).border = Border(left=thin, bottom=thin)
        ws_4_1.cell(row=r, column=(len(summary_years) + divisions * (len(all_years) - len(summary_years) - 1) + 4 + len(
          remainder))).border = Border(left=thin)
        ws_4_1.cell(row=r, column=1).border = Border(right=thin)
      ws_4_1.cell(row=header_row_no+1, column=(len(summary_years) + divisions * (len(all_years) - len(summary_years) - 1) + 3 + len(
        remainder))).border = Border(
        left=thin, bottom=thin)
      ws_4_1.cell(row=header_row_no, column=(len(summary_years) + divisions * (len(all_years) - len(summary_years) - 1) + 3 + len(remainder))).border = Border(left=thin, top=thin)
      ws_4_1.cell(row=header_row_no+2, column=(len(summary_years) + divisions * (len(all_years) - len(summary_years) - 1) + 3 + len(remainder))).border = Border(left=thin, bottom=thin)

      for o in range(header_row_no, header_row_no+1):
        ws_4_1.cell(row=o, column=len(column_names_4_1) + 2).border = Border(left=thin)

      for p in range(2, len(summary_years) + 2):
        for q in range(header_row_no+1, header_row_no+3):
          ws_4_1.cell(row=q, column=p + 1).border = Border(left=thin)
        ws_4_1.cell(row=header_row_no, column=p+1).border = Border(left=thin, top=thin)

      for q in range(len(summary_years)+1, (len(title_names_4_1) - 2) * divisions + 2, divisions):
        ws_4_1.cell(row=header_row_no+1, column=q + 2).border = Border(left=thin, bottom=thin)

      for r in range(3, len(summary_years) + 3):
        ws_4_1.cell(row=header_row_no+1, column=r).border = Border(left=thin, bottom=thin)
        ws_4_1.cell(row=header_row_no+2, column=r).border = Border(left=thin, bottom=thin)

    else:
      for i in range(2, len(column_names_4_1) + 2):
        ws_4_1.cell(row=header_row_no+1, column=i).border = Border(bottom=thin)
        ws_4_1.cell(row=header_row_no+2, column=i).border = Border(bottom=thin)
        ws_4_1.cell(row=header_row_no-1, column=i).border = Border(bottom=thin)

      for k in range(len(detailed_remainder)+1, (len(title_names_4_1) - 2) * divisions + 2, divisions):
        ws_4_1.cell(row=header_row_no+1, column=k + 2).border = Border(left=thin, bottom=thin)
        ws_4_1.cell(row=header_row_no+2, column=k + 2).border = Border(left=thin, bottom=thin)

      for r in range(header_row_no, header_row_no+3):
        ws_4_1.cell(row=r, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 3 + len(remainder))).border = Border(left=thin)
        ws_4_1.cell(row=r, column=1).border = Border(
          right=thin)
      ws_4_1.cell(row=header_row_no+1, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 3 + len(remainder))).border = Border(left=thin, bottom=thin)
      ws_4_1.cell(row=header_row_no, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 3 + len(remainder))).border = Border(left=thin, top=thin)
      ws_4_1.cell(row=header_row_no+2, column=(len(detailed_remainder) + divisions * (len(all_years) - 2) + 3 + len(remainder))).border = Border(left=thin, bottom=thin)

      for o in range(header_row_no, header_row_no+3):
        ws_4_1.cell(row=o, column=len(column_names_4_1) + 2).border = Border(left=thin)

      for q in range(header_row_no+1, header_row_no+3):
        ws_4_1.cell(row=q, column=3).border = Border(left=thin)
      ws_4_1.cell(row=header_row_no, column=3).border = Border(left=thin, top=thin)

      for q in range(len(detailed_remainder)+1, (len(title_names_4_1) - 2) * divisions + 2, divisions):
        ws_4_1.cell(row=header_row_no+1, column=q + 2).border = Border(left=thin, bottom=thin)
        ws_4_1.cell(row=header_row_no , column=q + 2).border = Border(left=thin, top=thin)

      for r in range(3, 4):
        ws_4_1.cell(row=header_row_no+1, column=r).border = Border(left=thin, bottom=thin)
        ws_4_1.cell(row=header_row_no+2, column=r).border = Border(left=thin, bottom=thin)

#Sheet 4.2: Revenue Over Time Data
  if first_year_complete:
    for z,a in zip(range(1, len(third_party_list)*4, 3), master_total_list):
      print(z)
      for x, y in zip(range(1,len(statement_period_list)-len(summary_years)+2,divisions),list(dict.fromkeys(year_statement_list))):
        ws_4_2.cell(row=x, column=z).value=y
      for x, y in zip(range(1,len(statement_period_list)-len(summary_years)+2),statement_list):
        ws_4_2.cell(row=x, column=z+1).value=y
      for x, y in zip(range(1,len(statement_period_list)-len(summary_years)+2),a[len(summary_years)+1:]):
        ws_4_2.cell(row=x, column=z+2).value=y

  else:
    ws_4_2.cell(row=1, column=1).value = list(dict.fromkeys(year_statement_list))[0]
    for x, y in zip(range(len(detailed_remainder)+1,len(statement_period_list)-len(detailed_remainder)+4,divisions),list(dict.fromkeys(year_statement_list))[1:]):
      ws_4_2.cell(row=x, column=1).value=y
    for v, w in zip(range(1, len(detailed_remainder)+1),statement_list[:len(detailed_remainder)]):
      ws_4_2.cell(row=v, column=2).value = w
    for x, y in zip(range(len(detailed_remainder)+1,len(statement_period_list)-len(detailed_remainder)+4),statement_list[len(detailed_remainder):]):
      ws_4_2.cell(row=x, column=2).value=y
    for v, w in zip(range(1, len(detailed_remainder)+1),total_list[1:2+len(detailed_remainder)]):
      ws_4_2.cell(row=v, column=3).value = w
    for x, y in zip(range(len(detailed_remainder)+1,len(statement_period_list)-len(detailed_remainder)+4),total_list[1+len(detailed_remainder):]):
      ws_4_2.cell(row=x, column=3).value=y

  #Add chart to sheet 4
  if first_year_complete:
    chart = LineChart()
    categories = Reference(ws_4_2, min_col=1, max_col=2, min_row=1,
                           max_row=len(statement_period_list) - len(summary_years) + 1)
    data = Reference(ws_4_2, min_col=3, max_col=3, min_row=1,
                     max_row=len(statement_period_list) - len(summary_years) + 1)
  else:
    chart = LineChart()
    categories = Reference(ws_4_2, min_col=1, max_col=2, min_row=1,
                           max_row=len(statement_period_list) - len(detailed_remainder) + 3)
    data = Reference(ws_4_2, min_col=3, max_col=3, min_row=1,
                     max_row=len(statement_period_list) - len(detailed_remainder) + 3)
  chart.add_data(data)
  chart.set_categories(categories)
  chart.y_axis.numFmt = '"$"#,##0.00'
  chart.title = 'Revenue Over Time'
  title_font = Font(name='Calibri')
  cp = CharacterProperties(sz=1200)
  chart.title.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp),endParaRPr=cp)])
  chart.legend = None
  ws_4.add_chart(chart, "{}2".format(column_letters[len(column_names_4) + 2]))

  #Add charts to sheet 4.1
  graph_no = 2
  for a,b in zip(third_party_list, range(4, len(third_party_list)*4, 3)):
    chart = LineChart()
    categories = Reference(ws_4_2, min_col=b, max_col=b+1, min_row=1, max_row=len(statement_period_list) - len(summary_years) + 1)
    data = Reference(ws_4_2, min_col=b+2, max_col=b+2, min_row=1, max_row=len(statement_period_list) - len(summary_years) + 1)
    chart.add_data(data)
    chart.set_categories(categories)
    chart.y_axis.numFmt = '"$"#,##0.00'
    chart.title = 'Revenue Over Time - {}'.format(a)
    chart.legend = None
    ws_4_1.add_chart(chart, "{}{}".format(column_letters[len(column_names_4) + 2], graph_no))
    graph_no += 15



  return(wb.save(filename))





#songvest('DJ Battlecat_616c91a05f278b92afbaa5ae')
#songvest('Songvest - Gianni_619e4e9eb2421ec5dfa7ee2e')
#songvest('Bizket_61ad4e64c5c7a7f539eb2c1e')
#songvest('Floyd Bentley_61ad59e3c5c7a7f539eb2c1f')
#songvest('Butta_61ad5e8bc5c7a7f539eb2c20')
#songvest('John Crawford_61ad6d17c5c7a7f539eb2c21')
#songvest('Steven Russell_619d6628b2421ec5dfa7ee2d')
#songvest('Santana TuneCore_61b0e77f5bf3453fdd9973b3')
#songvest('Todd Moore_61b0f0265bf3453fdd9973b4')
#songvest('Brandon Parrott_61b237110b966f46251d305b')
#Adman Khan_61b775f7c94b68a289900e81