import flask
import mysql.connector
import pathlib
import re
from datetime import date
from flask import request, jsonify
from flask_cors import cross_origin
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, NamedStyle
from pandas_utils.pandas_cursor import pandas_cursor
from pandas_utils.database_to_df import sql2pandas
import pandas as pd
import time


def databook(database, df, filename):
  mycursor = pandas_cursor(df=df)

  # Create sheets
  wb = Workbook()
  ws = wb.active
  ws.title = "Deal Key Metrics"

  # Column letters
  column_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
                    'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK',
                    'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ']

  # Create styles
  #Title style
  title_style = NamedStyle(name="title_style")
  title_style.alignment = Alignment(horizontal="left", vertical="center")
  title_style.font = Font(name="Calibri", size="12", color="002147", bold=True)

  #Dark blue header
  dark_blue_style = NamedStyle(name="dark_blue_style")
  dark_blue_style.alignment = Alignment(horizontal="left", vertical="center")
  dark_blue_style.font = Font(name="Calibri", size="11", color="ffffff", bold=True, italic=False)
  dark_blue_style.fill = PatternFill("solid", fgColor="002147")

  #Dark blue sub header
  dark_blue_sub_style = NamedStyle(name="dark_blue_sub_style")
  dark_blue_sub_style.alignment = Alignment(horizontal="left", vertical="center")
  dark_blue_sub_style.font = Font(name="Calibri", size="11", color="ffffff", italic=True)
  dark_blue_sub_style.fill = PatternFill("solid", fgColor="002147")

  #Number style
  number_style = NamedStyle(name="number_style")
  number_style.alignment = Alignment(horizontal="right", vertical="center")
  number_style.font = Font(name="Calibri", size="11")
  number_style.number_format = '#,##0'

  # Name style
  name_style = NamedStyle(name="name_style")
  name_style.alignment = Alignment(horizontal="left", vertical="center")
  name_style.font = Font(name="Calibri", size="11")

  # Total label style
  total_label_style = NamedStyle(name="total_label_style")
  total_label_style.alignment = Alignment(horizontal="left", vertical="center")
  total_label_style.font = Font(name="Calibri", size="11", bold=True)

  # Source label style
  source_label_style = NamedStyle(name="source_label_style")
  source_label_style.alignment = Alignment(horizontal="left", vertical="center")
  source_label_style.font = Font(name="Calibri", size="11", italic=True)

  # Bold number style
  bold_number_style = NamedStyle(name="bold_number_style")
  bold_number_style.alignment = Alignment(horizontal="right", vertical="center")
  bold_number_style.font = Font(name="Calibri", size="11", bold=True)
  bold_number_style.number_format = '#,##0'

  #Outline styles
  thick = Side(border_style="medium", color="000000")
  thin = Side(border_style="thin", color="000000")

  #Add styles to workbook
  wb.add_named_style(title_style)
  wb.add_named_style(dark_blue_style)
  wb.add_named_style(dark_blue_sub_style)
  wb.add_named_style(number_style)
  wb.add_named_style(name_style)
  wb.add_named_style(total_label_style)
  wb.add_named_style(source_label_style)
  wb.add_named_style(bold_number_style)

#Current year
  todays_date = date.today()
  current_year = todays_date.year

#Find complete list of statement half periods
  find_period = '''SELECT DISTINCT Statement_Period_Half_9LC FROM Master WHERE Statement_Period_Half_9LC <> ""
                       ORDER BY Statement_Period_Half_9LC'''
  mycursor.execute(find_period)
  statement_period_half_list = [i[0] for i in mycursor.fetchall()]
  if len(statement_period_half_list) > 12:
    statement_period_half = statement_period_half_list[-12:]
    cut_off = statement_period_half_list[-12]
  else:
    statement_period_half = statement_period_half_list
    cut_off = statement_period_half_list[0]
  year_list = []
  for period in statement_period_half:
    year_list.append(period[0:4])
  complete_year_list = list(dict.fromkeys(year_list))
  if complete_year_list[-1] == current_year:
    year_list = complete_year_list[-5:]
  else:
    year_list = complete_year_list[-4:]
  print(cut_off)

#Find total royalty amount in period
  find_total = '''SELECT Normalized_Income_Type_9LC, sum( CASE WHEN Statement_Period_Half_9LC >= '{}'
                      THEN Adjusted_Royalty_SB ELSE NULL END) AS `Total`
                      FROM Master GROUP BY Normalized_Income_Type_9LC'''.format(year_list[0])
  mycursor.execute(find_total)
  total_list = [i[1] for i in mycursor.fetchall()]
  royalty_total = 0
  for a in total_list:
    if a is None:
      royalty_total += 0
    else:
      royalty_total += a
  eighty_percent = royalty_total * 0.8
  print(year_list)
  print(eighty_percent)

#Year column headers
  year_headers = ['']
  for a in year_list:
    year_headers.append(int(a))
  for b in range(1,6):
    year_headers.append(int(year_list[-1])+b)

#Actual/forecast column headers
  desc_headers = ['$']
  for z in year_list:
    desc_headers.append('Actual')
  for y in range(1, 6):
    desc_headers.append('Forecast')

#YTD column
  ytd_col_no = 1
  for x in year_list:
    ytd_col_no += 1

#Title
  ws.cell(row=1, column=1).value = database[:-25]
  ws.cell(row=1, column=1).style = 'title_style'

#Initial title row number
  title_row_no = 3

#Summary table
  #Find data
  find_total_1 = '''SELECT Third_Party_9LC,'''
  find_total_2 = ''
  for c in year_list[:-1]:
    find_total_2 += '''sum( CASE WHEN Year_Statement_9LC = "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(c,c)
  find_total_3 = '''sum( CASE WHEN Year_Statement_9LC = "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `{}` FROM Master GROUP BY Third_Party_9LC'''.format(year_list[-1], year_list[-1])
  find_total = find_total_1 + find_total_2 + find_total_3
  mycursor.execute(find_total)
  total_table = mycursor.fetchall()
  total_list = ['Gross Income']
  year_total = 0
  for d in range(1, len(year_list)+1):
    for e in total_table:
      year_total += e[d]
    total_list.append(year_total)
    year_total = 0

  #Add headers
  ws.cell(row=title_row_no, column=1).value = 'NPS Summary'
  for f in range(1, len(year_list)+8):
    ws.cell(row=title_row_no, column=f).style = 'dark_blue_style'
  for g, h, i in zip(range(1, len(year_list)+7), year_headers, desc_headers):
    ws.cell(row=title_row_no+2, column=g).value = h
    ws.cell(row=title_row_no+2, column=g).style = 'dark_blue_style'
    ws.cell(row=title_row_no+3, column=ytd_col_no).value = 'YTD'
    ws.cell(row=title_row_no+3, column=g).style = 'dark_blue_style'
    ws.cell(row=title_row_no+4, column=g).value = i
    ws.cell(row=title_row_no+4, column=g).style = 'dark_blue_sub_style'
    ws.cell(row=title_row_no+4, column=1).style = 'dark_blue_style'

  #Add table
  for j, k in zip(range(1,len(total_list)+1), total_list):
    ws.cell(row=title_row_no+5, column=j).value = k
    ws.cell(row=title_row_no+5, column=j).style = 'number_style'
  ws.cell(row=title_row_no+5, column=1).style = 'name_style'
  ws.cell(row=title_row_no+6, column=1).value = '''Writer's share'''

  #Total row
  for l,m in zip(range(2,len(total_list)+6), column_letters[1:]):
    ws.cell(row=title_row_no+7, column=l).value = '=SUM({}8:{}9)'.format(m,m)
    ws.cell(row=title_row_no+7, column=l).style = 'bold_number_style'
  ws.cell(row=title_row_no+7, column=1).value = 'NPS'
  ws.cell(row=title_row_no+7, column=1).style = 'total_label_style'

  #Source label
  ws.cell(row=title_row_no+9, column=1).value = 'Source:'
  ws.cell(row=title_row_no+9, column=1).style = 'source_label_style'

  #Add to title row number
  title_row_no += 12

#Third Party Table
  #Find data
  find_third_party_table_1 = '''SELECT Third_Party_9LC,'''
  find_third_party_table_2 = ''
  for a in year_list[:-1]:
    find_third_party_table_2 += '''sum( CASE WHEN Year_Statement_9LC = "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(a,a)
  find_third_party_table_3 = '''sum( CASE WHEN Year_Statement_9LC = "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `{}` FROM Master GROUP BY Third_Party_9LC'''.format(year_list[-1], year_list[-1])
  find_third_party_table = find_third_party_table_1 + find_third_party_table_2 + find_third_party_table_3
  mycursor.execute(find_third_party_table)
  third_party_table = mycursor.fetchall()

  #Add headers
  ws.cell(row=title_row_no, column=1).value = 'NPS By Third Party'
  for f in range(1, len(year_list) + 8):
    ws.cell(row=title_row_no, column=f).style = 'dark_blue_style'
  for g, h, i in zip(range(1, len(year_list) + 7), year_headers, desc_headers):
    ws.cell(row=title_row_no + 2, column=g).value = h
    ws.cell(row=title_row_no + 2, column=g).style = 'dark_blue_style'
    ws.cell(row=title_row_no + 3, column=ytd_col_no).value = 'YTD'
    ws.cell(row=title_row_no + 3, column=g).style = 'dark_blue_style'
    ws.cell(row=title_row_no + 4, column=g).value = i
    ws.cell(row=title_row_no + 4, column=g).style = 'dark_blue_sub_style'
    ws.cell(row=title_row_no + 4, column=1).style = 'dark_blue_style'

  #Add table
  for j, k in zip(range(title_row_no+5, len(third_party_table) + title_row_no+6), third_party_table):
    for q,r in zip(range(1,len(k)+1), k):
      ws.cell(row=j, column=q).value = r
      ws.cell(row=j, column=q).style = 'number_style'
      ws.cell(row=j, column=1).style = 'name_style'

  #Total row
  for l, m in zip(range(2, len(year_list) + 7), column_letters[1:]):
    ws.cell(row=len(third_party_table) + title_row_no + 5, column=l).value = '=SUM({}{}:{}{})'.format(m,title_row_no+5, m,len(third_party_table) + title_row_no+4)
    ws.cell(row=len(third_party_table) + title_row_no + 5, column=l).style = 'bold_number_style'
  ws.cell(row=len(third_party_table) + title_row_no + 5, column=1).value = 'Total'
  ws.cell(row=len(third_party_table) + title_row_no + 5, column=1).style = 'total_label_style'

  #Source label
  ws.cell(row=len(third_party_table) + title_row_no + 7, column=1).value = 'Source:'
  ws.cell(row=len(third_party_table) + title_row_no + 7, column=1).style = 'source_label_style'

  #Check row
  ws.cell(row=len(third_party_table) + title_row_no + 8, column=1).value = 'Check'
  ws.cell(row=len(third_party_table) + title_row_no + 8, column=1).style = 'source_label_style'
  for n,o in zip(range(2,len(year_list)+7),column_letters[1:]):
    ws.cell(row=len(third_party_table) + title_row_no + 8, column=n).value = '={}10-{}{}'.format(o,o,len(third_party_table) + title_row_no + 5)
    ws.cell(row=len(third_party_table) + title_row_no + 8, column=n).style = 'number_style'

  #Add to title row number
  title_row_no += len(third_party_table) + 10

#NPS by Income Types
  #Create tables
  income_type_table = [['Publishing'], ['Publishing (non-admin)'], [''''Writers' Share'''], ['Masters'], ['Artist Royalties']]

  #Add headers
  ws.cell(row=title_row_no, column=1).value = 'NPS By Income Types'
  for f in range(1, len(year_list) + 8):
    ws.cell(row=title_row_no, column=f).style = 'dark_blue_style'
  for g, h, i in zip(range(1, len(year_list) + 7), year_headers, desc_headers):
    ws.cell(row=title_row_no + 2, column=g).value = h
    ws.cell(row=title_row_no + 2, column=g).style = 'dark_blue_style'
    ws.cell(row=title_row_no + 3, column=ytd_col_no).value = 'YTD'
    ws.cell(row=title_row_no + 3, column=g).style = 'dark_blue_style'
    ws.cell(row=title_row_no + 4, column=g).value = i
    ws.cell(row=title_row_no + 4, column=g).style = 'dark_blue_sub_style'
    ws.cell(row=title_row_no + 4, column=1).style = 'dark_blue_style'

  #Add table
  for j, k in zip(range(title_row_no + 5, len(income_type_table) + title_row_no + 6), income_type_table):
    for q, r in zip(range(1, len(k) + 1), k):
      ws.cell(row=j, column=q).value = r
      ws.cell(row=j, column=q).style = 'number_style'
      ws.cell(row=j, column=1).style = 'name_style'

  #Total row
  for l, m in zip(range(2, len(year_list) + 7), column_letters[1:]):
    ws.cell(row=len(income_type_table) + title_row_no + 5, column=l).value = '=SUM({}{}:{}{})'.format(m,
                                                                                                      title_row_no + 5,
                                                                                                      m, len(income_type_table) + title_row_no + 4)
    ws.cell(row=len(income_type_table) + title_row_no + 5, column=l).style = 'bold_number_style'
  ws.cell(row=len(income_type_table) + title_row_no + 5, column=1).value = 'Total'
  ws.cell(row=len(income_type_table) + title_row_no + 5, column=1).style = 'total_label_style'

  #Source label
  ws.cell(row=len(income_type_table) + title_row_no + 7, column=1).value = 'Source:'
  ws.cell(row=len(income_type_table) + title_row_no + 7, column=1).style = 'source_label_style'

  #Check row
  ws.cell(row=len(income_type_table) + title_row_no + 8, column=1).value = 'Check'
  ws.cell(row=len(income_type_table) + title_row_no + 8, column=1).style = 'source_label_style'
  for n, o in zip(range(2, len(year_list) + 7), column_letters[1:]):
    ws.cell(row=len(income_type_table) + title_row_no + 8, column=n).value = '={}10-{}{}'.format(o, o, len(income_type_table) + title_row_no + 5)
    ws.cell(row=len(income_type_table) + title_row_no + 8, column=n).style = 'number_style'

  #Add to title row number
  title_row_no += len(income_type_table) + 10

#NPS by Source
  #Create tables
  find_source_table_1 = '''SELECT Normalized_Source_9LC,'''
  find_source_table_2 = ''
  for a in year_list:
    find_source_table_2 += '''sum( CASE WHEN Year_Statement_9LC = "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(
      a, a)
  find_source_table_3 = '''sum( CASE WHEN Year_Statement_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `Total` FROM Master GROUP BY Normalized_Source_9LC ORDER BY `Total` DESC'''.format(
    year_list[0])
  find_source_table = find_source_table_1 + find_source_table_2 + find_source_table_3
  mycursor.execute(find_source_table)
  source_table = mycursor.fetchall()

  #Find eighty percent cut off
  cumulative_total = 0
  source_cut_off_no = 0
  for i in source_table:
    if cumulative_total <= eighty_percent:
      for j in range(1,len(year_list)+1):
        cumulative_total += i[j]
        source_cut_off = i[0]
      source_cut_off_no += 1
  top_80_sources = source_table[:source_cut_off_no]
  print(top_80_sources)

  #Add other sources row
  other_totals = ['Remaining revenue by source']
  other_sources = source_table[source_cut_off_no:]
  year_total = 0
  for k in range(1, len(year_list)+1):
    for l in other_sources:
      year_total += l[k]
    other_totals.append(year_total)
    year_total = 0
  final_table = top_80_sources
  print(final_table)

  #Add headers
  ws.cell(row=title_row_no, column=1).value = 'NPS By Source'
  for f in range(1, len(year_list) + 8):
    ws.cell(row=title_row_no, column=f).style = 'dark_blue_style'
  for g, h, i in zip(range(1, len(year_list) + 7), year_headers, desc_headers):
    ws.cell(row=title_row_no + 2, column=g).value = h
    ws.cell(row=title_row_no + 2, column=g).style = 'dark_blue_style'
    ws.cell(row=title_row_no + 3, column=ytd_col_no).value = 'YTD'
    ws.cell(row=title_row_no + 3, column=g).style = 'dark_blue_style'
    ws.cell(row=title_row_no + 4, column=g).value = i
    ws.cell(row=title_row_no + 4, column=g).style = 'dark_blue_sub_style'
    ws.cell(row=title_row_no + 4, column=1).style = 'dark_blue_style'

  #Add table
  for j, k in zip(range(title_row_no + 5, len(final_table) + title_row_no + 6), final_table):
    for q, r in zip(range(1, len(year_list) + 2), k):
      ws.cell(row=j, column=q).value = r
      ws.cell(row=j, column=q).style = 'number_style'
      ws.cell(row=j, column=1).style = 'name_style'

  #80% total row
  for l, m in zip(range(2, len(year_list) + 7), column_letters[1:]):
    ws.cell(row=len(final_table) + title_row_no + 5, column=l).value = '=SUM({}{}:{}{})'.format(m,
                                                                                                      title_row_no + 5,
                                                                                                      m, len(final_table) + title_row_no + 4)
    ws.cell(row=len(final_table) + title_row_no + 5, column=l).style = 'number_style'
  ws.cell(row=len(final_table) + title_row_no + 5, column=1).value = 'Top 80% of revenue by source'
  ws.cell(row=len(final_table) + title_row_no + 5, column=1).style = 'total_label_style'

  #Other row
  for n, o in zip(range(1, len(year_list)+2), other_totals):
    ws.cell(row=len(final_table) + title_row_no + 6, column=n).value = o
    ws.cell(row=len(final_table) + title_row_no + 6, column=n).style = 'number_style'
    ws.cell(row=len(final_table) + title_row_no + 6, column=1).style = 'name_style'

  #Grand total row
  for p, q in zip(range(2, len(year_list)+7), column_letters[1:]):
    ws.cell(row=len(final_table) + title_row_no + 7, column=p).value = '={}{}+{}{}'.format(q,len(final_table) + title_row_no + 5,q,len(final_table) + title_row_no + 6)
    ws.cell(row=len(final_table) + title_row_no + 7, column=p).style = 'bold_number_style'
  ws.cell(row=len(final_table) + title_row_no + 7, column=1).value = 'Total'
  ws.cell(row=len(final_table) + title_row_no + 7, column=1).style = 'total_label_style'

  #Source label
  ws.cell(row=len(final_table) + title_row_no + 9, column=1).value = 'Source:'
  ws.cell(row=len(final_table) + title_row_no + 9, column=1).style = 'source_label_style'

  #Bottom line
  for m in range(1, len(total_list) + 6):
    ws.cell(row=len(final_table) + title_row_no + 5, column=m).border = Border(top=thin)

  #Check row
  ws.cell(row=len(final_table) + title_row_no + 10, column=1).value = 'Check'
  ws.cell(row=len(final_table) + title_row_no + 10, column=1).style = 'source_label_style'
  for n, o in zip(range(2, len(year_list) + 7), column_letters[1:]):
    ws.cell(row=len(final_table) + title_row_no + 10, column=n).value = '={}10-{}{}'.format(o, o, len(
      final_table) + title_row_no + 7)
    ws.cell(row=len(final_table) + title_row_no + 10, column=n).style = 'number_style'

  #Add to title row number
  title_row_no += len(final_table) + 12

#NPS by Income Types
  #Create tables
  find_income_table_1 = '''SELECT Normalized_Income_Type_9LC,'''
  find_income_table_2 = ''
  for a in year_list:
    find_income_table_2 += '''sum( CASE WHEN Year_Statement_9LC = "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(
      a, a)
  find_income_table_3 = '''sum( CASE WHEN Year_Statement_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `Total` FROM Master GROUP BY Normalized_Income_Type_9LC ORDER BY `Total` DESC'''.format(
    year_list[0])
  find_income_table = find_income_table_1 + find_income_table_2 + find_income_table_3
  mycursor.execute(find_income_table)
  income_table = mycursor.fetchall()

  #Add headers
  ws.cell(row=title_row_no, column=1).value = 'NPS By Income Types'
  for f in range(1, len(year_list) + 8):
    ws.cell(row=title_row_no, column=f).style = 'dark_blue_style'
  for g, h, i in zip(range(1, len(year_list) + 7), year_headers, desc_headers):
    ws.cell(row=title_row_no + 2, column=g).value = h
    ws.cell(row=title_row_no + 2, column=g).style = 'dark_blue_style'
    ws.cell(row=title_row_no + 3, column=ytd_col_no).value = 'YTD'
    ws.cell(row=title_row_no + 3, column=g).style = 'dark_blue_style'
    ws.cell(row=title_row_no + 4, column=g).value = i
    ws.cell(row=title_row_no + 4, column=g).style = 'dark_blue_sub_style'
    ws.cell(row=title_row_no + 4, column=1).style = 'dark_blue_style'

  #Add table
  for j, k in zip(range(title_row_no + 5, len(income_table) + title_row_no + 6), income_table):
    for q, r in zip(range(1, len(k)), k):
      ws.cell(row=j, column=q).value = r
      ws.cell(row=j, column=q).style = 'number_style'
      ws.cell(row=j, column=1).style = 'name_style'

  #Total row
  for l, m in zip(range(2, len(year_list) + 7), column_letters[1:]):
    ws.cell(row=len(income_table) + title_row_no + 5, column=l).value = '=SUM({}{}:{}{})'.format(m,
                                                                                                      title_row_no + 5,
                                                                                                      m, len(
        income_table) + title_row_no + 4)
    ws.cell(row=len(income_table) + title_row_no + 5, column=l).style = 'bold_number_style'
  ws.cell(row=len(income_table) + title_row_no + 5, column=1).value = 'Total'
  ws.cell(row=len(income_table) + title_row_no + 5, column=1).style = 'total_label_style'

  #Source label
  ws.cell(row=len(income_table) + title_row_no + 7, column=1).value = 'Source:'
  ws.cell(row=len(income_table) + title_row_no + 7, column=1).style = 'source_label_style'

  #Check row
  ws.cell(row=len(income_table) + title_row_no + 8, column=1).value = 'Check'
  ws.cell(row=len(income_table) + title_row_no + 8, column=1).style = 'source_label_style'
  for n, o in zip(range(2, len(year_list) + 7), column_letters[1:]):
    ws.cell(row=len(income_table) + title_row_no + 8, column=n).value = '={}10-{}{}'.format(o, o, len(
      income_table) + title_row_no + 5)
    ws.cell(row=len(final_table) + title_row_no + 8, column=n).style = 'number_style'

  #Add to title row number
  title_row_no += len(income_table) + 10

#NPS by Song
  #Create tables
  find_song_table_1 = '''SELECT Song_Name_9LC,'''
  find_song_table_2 = ''
  for a in year_list:
    find_song_table_2 += '''sum( CASE WHEN Year_Statement_9LC = "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(
      a, a)
  find_song_table_3 = '''sum( CASE WHEN Year_Statement_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `Total` FROM Master GROUP BY Song_Name_9LC ORDER BY `Total` DESC'''.format(
    year_list[0])
  find_song_table = find_song_table_1 + find_song_table_2 + find_song_table_3
  mycursor.execute(find_song_table)
  song_table = mycursor.fetchall()

  #Find eighty percent cut off
  cumulative_total = 0
  song_cut_off_no = 0
  for i in song_table:
    if cumulative_total <= eighty_percent:
      for j in range(1, len(year_list)+1):
        cumulative_total += i[j]
        song_cut_off = i[0]
      song_cut_off_no += 1
  top_80_songs = song_table[:song_cut_off_no]
  print(top_80_songs)

  #Add other songs row
  other_totals = ['Remaining revenue by song']
  other_songs = song_table[song_cut_off_no:]
  year_total = 0
  for k in range(1, len(year_list) + 1):
    for l in other_songs:
      year_total += l[k]
    other_totals.append(year_total)
    year_total = 0
  final_table = top_80_songs
  print(final_table)

  #Add headers
  ws.cell(row=title_row_no, column=1).value = 'NPS By Song'
  for f in range(1, len(year_list) + 8):
    ws.cell(row=title_row_no, column=f).style = 'dark_blue_style'
  for g, h, i in zip(range(1, len(year_list) + 7), year_headers, desc_headers):
    ws.cell(row=title_row_no + 2, column=g).value = h
    ws.cell(row=title_row_no + 2, column=g).style = 'dark_blue_style'
    ws.cell(row=title_row_no + 3, column=ytd_col_no).value = 'YTD'
    ws.cell(row=title_row_no + 3, column=g).style = 'dark_blue_style'
    ws.cell(row=title_row_no + 4, column=g).value = i
    ws.cell(row=title_row_no + 4, column=g).style = 'dark_blue_sub_style'
    ws.cell(row=title_row_no + 4, column=1).style = 'dark_blue_style'

  #Add table
  for j, k in zip(range(title_row_no + 5, len(final_table) + title_row_no + 6), final_table):
    for q, r in zip(range(1, len(year_list) + 2), k):
      ws.cell(row=j, column=q).value = r
      ws.cell(row=j, column=q).style = 'number_style'
      ws.cell(row=j, column=1).style = 'name_style'

  #80% total row
  for l, m in zip(range(2, len(year_list) + 7), column_letters[1:]):
    ws.cell(row=len(final_table) + title_row_no + 5, column=l).value = '=SUM({}{}:{}{})'.format(m,
                                                                                                title_row_no + 5,
                                                                                                m, len(
        final_table) + title_row_no + 4)
    ws.cell(row=len(final_table) + title_row_no + 5, column=l).style = 'number_style'
  ws.cell(row=len(final_table) + title_row_no + 5, column=1).value = 'Top 80% of revenue by song'
  ws.cell(row=len(final_table) + title_row_no + 5, column=1).style = 'total_label_style'

  #Other row
  for n, o in zip(range(1, len(year_list) + 2), other_totals):
    ws.cell(row=len(final_table) + title_row_no + 6, column=n).value = o
    ws.cell(row=len(final_table) + title_row_no + 6, column=n).style = 'number_style'
    ws.cell(row=len(final_table) + title_row_no + 6, column=1).style = 'name_style'

  #Grand total row
  for p, q in zip(range(2, len(year_list) + 7), column_letters[1:]):
    ws.cell(row=len(final_table) + title_row_no + 7, column=p).value = '={}{}+{}{}'.format(q, len(
      final_table) + title_row_no + 5, q, len(final_table) + title_row_no + 6)
    ws.cell(row=len(final_table) + title_row_no + 7, column=p).style = 'bold_number_style'
  ws.cell(row=len(final_table) + title_row_no + 7, column=1).value = 'Total'
  ws.cell(row=len(final_table) + title_row_no + 7, column=1).style = 'total_label_style'

  #Source label
  ws.cell(row=len(final_table) + title_row_no + 9, column=1).value = 'Source:'
  ws.cell(row=len(final_table) + title_row_no + 9, column=1).style = 'source_label_style'

  #Bottom line
  for m in range(1, len(total_list) + 6):
    ws.cell(row=len(final_table) + title_row_no + 5, column=m).border = Border(top=thin)

  #Check row
  ws.cell(row=len(final_table) + title_row_no + 10, column=1).value = 'Check'
  ws.cell(row=len(final_table) + title_row_no + 10, column=1).style = 'source_label_style'
  for n, o in zip(range(2, len(year_list) + 7), column_letters[1:]):
    ws.cell(row=len(final_table) + title_row_no + 10, column=n).value = '={}10-{}{}'.format(o, o, len(
      final_table) + title_row_no + 7)
    ws.cell(row=len(final_table) + title_row_no + 10, column=n).style = 'number_style'

  #Add to title row number
  title_row_no += len(final_table) + 12

#NPS by Artist
  #Create tables
  find_artist_table_1 = '''SELECT Release_Artist_9LC,'''
  find_artist_table_2 = ''
  for a in year_list:
    find_artist_table_2 += '''sum( CASE WHEN Year_Statement_9LC = "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(
      a, a)
  find_artist_table_3 = '''sum( CASE WHEN Year_Statement_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `Total` FROM Master WHERE Release_Artist_9LC <> "" GROUP BY Release_Artist_9LC ORDER BY `Total` DESC'''.format(
    year_list[0])
  find_artist_table = find_artist_table_1 + find_artist_table_2 + find_artist_table_3
  mycursor.execute(find_artist_table)
  initial_artist_table = mycursor.fetchall()

  find_blank_artist_table_1 = '''SELECT Release_Artist_9LC,'''
  find_blank_artist_table_2 = ''
  for a in year_list:
    find_blank_artist_table_2 += '''sum( CASE WHEN Year_Statement_9LC = "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(
      a, a)
  find_blank_artist_table_3 = '''sum( CASE WHEN Year_Statement_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `Total` FROM Master WHERE Release_Artist_9LC = "" GROUP BY Release_Artist_9LC ORDER BY `Total` DESC'''.format(
    year_list[0])
  find_blank_artist_table = find_blank_artist_table_1 + find_blank_artist_table_2 + find_blank_artist_table_3
  mycursor.execute(find_blank_artist_table)
  blank_artist_table = mycursor.fetchall()
  artist_table = initial_artist_table + blank_artist_table

  #Find eighty percent cut off
  cumulative_total = 0
  artist_cut_off_no = 0
  for i in initial_artist_table:
    if cumulative_total <= eighty_percent:
      for j in range(1, len(year_list)+1):
        cumulative_total += i[j]
        artist_cut_off = i[0]
      artist_cut_off_no += 1
  print(artist_cut_off_no)
  top_80_artists = initial_artist_table[:artist_cut_off_no]
  print(top_80_artists)

  #Add other artists row
  other_totals = ['Remaining revenue by artist']
  other_artists = artist_table[artist_cut_off_no:]
  year_total = 0
  for k in range(1, len(year_list) + 1):
    for l in other_artists:
      year_total += l[k]
    other_totals.append(year_total)
    year_total = 0
  final_table = top_80_artists
  print(final_table)
  print(other_totals)
  #Add headers
  ws.cell(row=title_row_no, column=1).value = 'NPS By Artist'
  for f in range(1, len(year_list) + 8):
    ws.cell(row=title_row_no, column=f).style = 'dark_blue_style'
  for g, h, i in zip(range(1, len(year_list) + 7), year_headers, desc_headers):
    ws.cell(row=title_row_no + 2, column=g).value = h
    ws.cell(row=title_row_no + 2, column=g).style = 'dark_blue_style'
    ws.cell(row=title_row_no + 3, column=ytd_col_no).value = 'YTD'
    ws.cell(row=title_row_no + 3, column=g).style = 'dark_blue_style'
    ws.cell(row=title_row_no + 4, column=g).value = i
    ws.cell(row=title_row_no + 4, column=g).style = 'dark_blue_sub_style'
    ws.cell(row=title_row_no + 4, column=1).style = 'dark_blue_style'

  #Add table
  for j, k in zip(range(title_row_no + 5, len(final_table) + title_row_no + 6), final_table):
    for q, r in zip(range(1, len(year_list) + 2), k[:-1]):
      ws.cell(row=j, column=q).value = r
      ws.cell(row=j, column=q).style = 'number_style'
      ws.cell(row=j, column=1).style = 'name_style'

  #80% total row
  for l, m in zip(range(2, len(year_list) + 7), column_letters[1:]):
    ws.cell(row=len(final_table) + title_row_no + 5, column=l).value = '=SUM({}{}:{}{})'.format(m,
                                                                                                title_row_no + 5,
                                                                                                m, len(
        final_table) + title_row_no + 4)
    ws.cell(row=len(final_table) + title_row_no + 5, column=l).style = 'number_style'
  ws.cell(row=len(final_table) + title_row_no + 5, column=1).value = 'Top 80% of revenue by artist'
  ws.cell(row=len(final_table) + title_row_no + 5, column=1).style = 'total_label_style'

  #Other row
  for n, o in zip(range(1, len(year_list) + 2), other_totals):
    ws.cell(row=len(final_table) + title_row_no + 6, column=n).value = o
    ws.cell(row=len(final_table) + title_row_no + 6, column=n).style = 'number_style'
    ws.cell(row=len(final_table) + title_row_no + 6, column=1).style = 'name_style'

  #Grand total row
  for p, q in zip(range(2, len(year_list) + 7), column_letters[1:]):
    ws.cell(row=len(final_table) + title_row_no + 7, column=p).value = '={}{}+{}{}'.format(q, len(
      final_table) + title_row_no + 5, q, len(final_table) + title_row_no + 6)
    ws.cell(row=len(final_table) + title_row_no + 7, column=p).style = 'bold_number_style'
  ws.cell(row=len(final_table) + title_row_no + 7, column=1).value = 'Total'
  ws.cell(row=len(final_table) + title_row_no + 7, column=1).style = 'total_label_style'

  #Source label
  ws.cell(row=len(final_table) + title_row_no + 9, column=1).value = 'Source:'
  ws.cell(row=len(final_table) + title_row_no + 9, column=1).style = 'source_label_style'

  #Bottom line
  for m in range(1, len(total_list) + 6):
    ws.cell(row=len(final_table) + title_row_no + 5, column=m).border = Border(top=thin)

  #Check row
  ws.cell(row=len(final_table) + title_row_no + 10, column=1).value = 'Check'
  ws.cell(row=len(final_table) + title_row_no + 10, column=1).style = 'source_label_style'
  for n, o in zip(range(2, len(year_list) + 7), column_letters[1:]):
    ws.cell(row=len(final_table) + title_row_no + 10, column=n).value = '={}10-{}{}'.format(o, o, len(
      final_table) + title_row_no + 7)
    ws.cell(row=len(final_table) + title_row_no + 10, column=n).style = 'number_style'

  #Add to title row number
  title_row_no += len(final_table) + 12

#NPS by Geography
  #Create tables
  find_us_table_1 = '''SELECT '''
  find_us_table_2 = ""
  for b in year_list:
    find_us_table_2 += '''sum( CASE WHEN Year_Statement_9LC = "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(b,b)
  find_us_table_3 = '''sum( CASE WHEN Year_Statement_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `Total`
                        FROM Master WHERE Country_SB IN ("United States", "United States of America", "US", "USA", "USA AND DOMINIONS")
                        ORDER BY `Total` DESC'''.format(year_list[0])
  find_us_table = find_us_table_1 + find_us_table_2 +  find_us_table_3
  mycursor.execute(find_us_table)
  us_table = mycursor.fetchall()

  find_non_us_table_1 = '''SELECT '''
  find_non_us_table_2 = ""
  for b in year_list:
    find_non_us_table_2 += '''sum( CASE WHEN Year_Statement_9LC = "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(
      b, b)
  find_non_us_table_3 = '''sum( CASE WHEN Year_Statement_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `Total`
                          FROM Master WHERE Country_SB NOT IN ("United States", "United States of America", "US", "USA", "USA AND DOMINIONS")
                          ORDER BY `Total` DESC'''.format(year_list[0])
  find_non_us_table = find_non_us_table_1 + find_non_us_table_2 + find_non_us_table_3
  mycursor.execute(find_non_us_table)
  non_us_table = mycursor.fetchall()
  territory_table = us_table + non_us_table
  print(territory_table)

  #Add headers
  ws.cell(row=title_row_no, column=1).value = 'NPS By Geography'
  for f in range(1, len(year_list) + 8):
    ws.cell(row=title_row_no, column=f).style = 'dark_blue_style'
  for g, h, i in zip(range(1, len(year_list) + 7), year_headers, desc_headers):
    ws.cell(row=title_row_no + 2, column=g).value = h
    ws.cell(row=title_row_no + 2, column=g).style = 'dark_blue_style'
    ws.cell(row=title_row_no + 3, column=ytd_col_no).value = 'YTD'
    ws.cell(row=title_row_no + 3, column=g).style = 'dark_blue_style'
    ws.cell(row=title_row_no + 4, column=g).value = i
    ws.cell(row=title_row_no + 4, column=g).style = 'dark_blue_sub_style'
    ws.cell(row=title_row_no + 4, column=1).style = 'dark_blue_style'

  #Add table
  for j, k in zip(range(title_row_no + 5, len(territory_table) + title_row_no + 6), territory_table):
    for q, r in zip(range(2, len(k) + 2), k[:-1]):
      ws.cell(row=j, column=q).value = r
      ws.cell(row=j, column=q).style = 'number_style'
      ws.cell(row=j, column=1).style = 'name_style'
  ws.cell(row=title_row_no+5, column=1).value = 'US'
  ws.cell(row=title_row_no+6, column=1).value = 'Non-US'

  #Total row
  for l, m in zip(range(2, len(year_list) + 7), column_letters[1:]):
    ws.cell(row=len(territory_table) + title_row_no + 5, column=l).value = '=SUM({}{}:{}{})'.format(m,
                                                                                                      title_row_no + 5,
                                                                                                      m, len(
        territory_table) + title_row_no + 4)
    ws.cell(row=len(territory_table) + title_row_no + 5, column=l).style = 'bold_number_style'
  ws.cell(row=len(territory_table) + title_row_no + 5, column=1).value = 'Total'
  ws.cell(row=len(territory_table) + title_row_no + 5, column=1).style = 'total_label_style'

  #Source label
  ws.cell(row=len(territory_table) + title_row_no + 7, column=1).value = 'Source:'
  ws.cell(row=len(territory_table) + title_row_no + 7, column=1).style = 'source_label_style'

  #Check row
  ws.cell(row=len(territory_table) + title_row_no + 8, column=1).value = 'Check'
  ws.cell(row=len(territory_table) + title_row_no + 8, column=1).style = 'source_label_style'
  for n, o in zip(range(2, len(year_list) + 7), column_letters[1:]):
    ws.cell(row=len(territory_table) + title_row_no + 8, column=n).value = '={}10-{}{}'.format(o, o, len(
      territory_table) + title_row_no + 5)
    ws.cell(row=len(territory_table) + title_row_no + 8, column=n).style = 'number_style'

  #Add to title row number
  title_row_no += len(territory_table) + 10

  print(total_list)

  wb.save(filename)

  return pd.read_excel(filename)


if __name__ == "__main__":
  start =  time.time()
  dbname = 'Rick James - Hipgnosis - Final_61df0acaf5592bbb47ad9a86'
  df = sql2pandas(dbname)
  end = time.time()

  print(f"Translation Time: {end - start}")

  print("started ----")

  start = time.time()
  print(databook(dbname, df, 'output_files/sample_output.xlsx'))
  end = time.time()
  print(f"Creation Time: {end - start}")
  