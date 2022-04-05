from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, NamedStyle
from datetime import date
from pandas_utils.pandas_cursor import pandas_cursor


def findthirdparties(database, df):
    mycursor = pandas_cursor(df=df)

    find_third_parties = '''SELECT Third_Party_9LC FROM Master GROUP BY Third_Party_9LC'''
    mycursor.execute(find_third_parties)
    third_parties = [i[0] for i in mycursor.fetchall()]
    return third_parties

def thirdpartystandard(database, filename, thirdparty):
    mydb = mysql.connector.connect(
    host="34.65.111.142",
    user="external",
    password="musicpass",
    database="{}".format(database)
    )

    mycursor = mydb.cursor(buffered=True)

    #Find types of third party
    find_type = '''SELECT Rights_Type_9LC, Third_Party_9LC FROM Master GROUP BY Rights_Type_9LC, Third_Party_9LC'''
    mycursor.execute(find_type)
    type_list = list(dict.fromkeys([i[0] for i in mycursor.fetchall()]))
    test = [i[1] for i in mycursor.fetchall()]
    type_string = ', '.join('"{}"'.format(str(x)) for x in type_list)
    print(type_list)
    print(test)

    #Find third parties for each type
    third_party_type_list = []
    for a in type_list:
        third_party_sub_list = [a]
        find_third_party_type = '''SELECT Third_Party_9LC FROM Master WHERE Rights_Type_9LC = "{}" GROUP BY Third_Party_9LC'''.format(
          a)
        mycursor.execute(find_third_party_type)
        test_list = [i[0] for i in mycursor.fetchall()]
        for b in test_list:
            third_party_sub_list.append(b)
        third_party_type_list.append(third_party_sub_list)
        third_party_sub_list = []
    print(third_party_type_list)

    #Publisher list
    find_publishers = '''SELECT Third_Party_9LC FROM Master WHERE Third_Party_9LC NOT IN ('ASCAP', 'BMI', 'PRS', 'MCPS', 
                           'SoundExchange', 'SOCAN', 'APRA', 'AMCOS', 'SESAC', 'AMRA', 'GEMA', 'SUISA', 'ZAiKS') GROUP BY Third_Party_9LC'''
    mycursor.execute(find_publishers)
    publisher_list = [i[0] for i in mycursor.fetchall()]
    publisher_string = ', '.join('"{}"'.format(str(x)) for x in publisher_list)
    print(publisher_list)

    #PRO list
    find_PRO = '''SELECT Third_Party_9LC FROM Master WHERE Third_Party_9LC IN ('ASCAP', 'BMI', 'PRS', 'MCPS', 
                             'SoundExchange', 'SOCAN', 'APRA', 'AMCOS', 'SESAC', 'AMRA', 'GEMA', 'SUISA', 'ZAiKS') GROUP BY Third_Party_9LC'''
    mycursor.execute(find_PRO)
    PRO_list = [i[0] for i in mycursor.fetchall()]
    PRO_string = ', '.join('"{}"'.format(str(x)) for x in PRO_list)
    print(PRO_list)

    #Third party list
    third_party_list = publisher_list + PRO_list

    #Column letters
    column_letters = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T',
                    'U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO',
                    'AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY','AZ']

    #Create styles
    #Main numbers
    number_style = NamedStyle(name="number_style")
    number_style.alignment = Alignment(horizontal="right", vertical="center")
    number_style.font = Font(name="Calibri", size="11")
    number_style.number_format = '#,##0.00'

    #Title cells
    title_style = NamedStyle(name="title_style")
    title_style.alignment = Alignment(horizontal="center", vertical="center")
    title_style.font = Font(name="Calibri", size="11", bold=True)
    title_style.fill = PatternFill("solid", fgColor="7496CC")

    #Header cells
    header_style = NamedStyle(name="header_style")
    header_style.alignment = Alignment(horizontal="center", vertical="center")
    header_style.font = Font(name="Calibri", size="11", bold=True)
    header_style.fill = PatternFill("solid", fgColor="C9CDCF")

    #Sub header cells
    sub_header_style = NamedStyle(name="sub_header_style")
    sub_header_style.alignment = Alignment(horizontal="left", vertical="center")
    sub_header_style.font = Font(name="Calibri", size="11", bold=True, italic=True)

    #Name cells
    name_style = NamedStyle(name="name_style")
    name_style.alignment = Alignment(horizontal="left", vertical="center")
    name_style.font = Font(name="Calibri", size="11")

    #Total cells
    total_style = NamedStyle(name="total_style")
    total_style.alignment = Alignment(horizontal="right", vertical="center")
    total_style.font = Font(name="Calibri", size="11", bold=True)
    total_style.number_format = '#,##0.00'

    #Total row style
    total_row_style = NamedStyle(name="total_row_style")
    total_row_style.alignment = Alignment(horizontal="right", vertical="center")
    total_row_style.font = Font(name="Calibri", size="11", bold=True)
    total_row_style.fill = PatternFill("solid", fgColor="88B054")
    total_row_style.number_format = '#,##0.00'

    #Total label style
    total_label_style = NamedStyle(name="total_label_style")
    total_label_style.alignment = Alignment(horizontal="center", vertical="center")
    total_label_style.font = Font(name="Calibri", size="11", bold=True)
    total_label_style.fill = PatternFill("solid", fgColor="88B054")

    #Publisher/PRO label style
    thick = Side(border_style="medium", color="000000")
    publisher_label_style = NamedStyle(name="publisher_label_style")
    publisher_label_style.alignment = Alignment(horizontal="center", vertical="center")
    publisher_label_style.font = Font(name="Calibri", size="11", bold=True, italic=True)
    publisher_label_style.border = Border(bottom=thick)

    #Not available style
    not_available_style = NamedStyle(name="not_available_style")
    not_available_style.alignment = Alignment(horizontal="center", vertical="center")
    not_available_style.font = Font(name="Calibri", size="11")

    #Bold name style
    bold_name_style = NamedStyle(name="bold_name_style")
    bold_name_style.alignment = Alignment(horizontal="left", vertical="center")
    bold_name_style.font = Font(name="Calibri", size="11", bold=True)

    #Total revenue style
    total_revenue_style = NamedStyle(name="total_revenue_style")
    total_revenue_style.alignment = Alignment(horizontal="right", vertical="center")
    total_revenue_style.font = Font(name="Calibri", size="11", bold=True)
    total_revenue_style.number_format = '#,##0.00'

#Build tables
  #for e in third_party_list:
#Get list of statement half periods
    # Current year
    todays_date = date.today()
    current_year = todays_date.year

    # Find most recent year in data
    find_recent_year = '''SELECT Year_Statement_9LC FROM Master GROUP BY Year_Statement_9LC'''
    mycursor.execute(find_recent_year)
    recent_years = [i[0] for i in mycursor.fetchall()]
    recent_year = recent_years[-1]
    print(recent_year)
    print(current_year)
    if recent_year == current_year:
        find_cut_off_year = current_year - 1
    else:
        find_cut_off_year = int(recent_year) - 1

    # Find cut off
    find_period = 'SELECT DISTINCT Statement_Period_Half_9LC FROM Master WHERE Year_Statement_9LC <= "{}" ORDER BY Statement_Period_Half_9LC'.format(
        find_cut_off_year)
    mycursor.execute(find_period)
    statement_period_half_blank = [i[0] for i in mycursor.fetchall()]

    def check_blank(period):
        if period == '':
            return False
        else:
            return True

    remove_blank = filter(check_blank, statement_period_half_blank)
    statement_period_minus_blank = list(remove_blank)
    if len(statement_period_minus_blank) > 10:
        statement_period_half = statement_period_minus_blank[-10:]
        cut_off = statement_period_minus_blank[-10]
    else:
        statement_period_half = statement_period_minus_blank
        cut_off = statement_period_minus_blank[0]
    print(statement_period_half)
    print(cut_off)

#Get list of statement half periods
    find_half_period = '''SELECT Year_Statement_9LC, Half_Statement_9LC FROM Master 
                                         WHERE Half_Statement_9LC <> "" AND Statement_Period_Half_9LC >= "{}" AND Third_Party_9LC = "{}" GROUP BY Year_Statement_9LC, Half_Statement_9LC'''.format(
      cut_off, thirdparty)
    mycursor.execute(find_half_period)
    complete_list = [i for i in mycursor.fetchall()]
    statement_period_half_list = [i[0] + ' ' + i[1] for i in complete_list]
    year_statement_list_half = [i[0] for i in complete_list]
    half_statement_list = [i[1] for i in complete_list]
    year_list = list(dict.fromkeys(year_statement_list_half))
    print(year_list)

#Get list of statement quarter periods
    find_quarter_period = '''SELECT Year_Statement_9LC, Quarter_Statement_9LC FROM Master 
                                       WHERE Quarter_Statement_9LC <> "" AND Statement_Period_Half_9LC >= "{}" AND Third_Party_9LC = "{}" GROUP BY Year_Statement_9LC, Quarter_Statement_9LC'''.format(
      cut_off, thirdparty)
    mycursor.execute(find_quarter_period)
    complete_list = [i for i in mycursor.fetchall()]
    statement_period_quarter_list = [i[0] + ' ' + i[1] for i in complete_list]
    year_statement_list_quarter = [i[0] for i in complete_list]
    quarter_statement_list = [i[1] for i in complete_list]
    print(quarter_statement_list)

#Get list of statement month periods
    find_month_period = '''SELECT Year_Statement_9LC, Month_Statement_9LC FROM Master 
                                     WHERE Quarter_Statement_9LC <> "" AND Year_Statement_9LC >= "{}" AND Third_Party_9LC ="{}" GROUP BY Year_Statement_9LC, Month_Statement_9LC'''.format(
      cut_off, thirdparty)
    mycursor.execute(find_month_period)
    complete_list = [i for i in mycursor.fetchall()]
    statement_period_month_list = [i[1] + ' ' + i[0] for i in complete_list]
    year_statement_list_month = [i[0] for i in complete_list]
    month_statement_list = [i[1] for i in complete_list]

    find_quarterly_period = '''SELECT Third_Party_9LC FROM Master WHERE Quarter_Statement_9LC <> "" AND Third_Party_9LC = "{}" GROUP BY Third_Party_9LC'''.format(
      thirdparty)
    mycursor.execute(find_quarterly_period)
    quarterly = [i[0] for i in mycursor.fetchall()]

    #find_monthly_period = '''SELECT Third_Party_9LC FROM Master WHERE Month_Statement_9LC <> "" AND Third_Party_9LC = "{}" GROUP BY Third_Party_9LC'''.format(
      #thirdparty)
    #mycursor.execute(find_monthly_period)
    #monthly = [i[0] for i in mycursor.fetchall()]
    #print(monthly)

#Create list of column names
    #Song x Rev x Half Columns
    half_column_names = ['Song Title']
    for k in statement_period_half:
      half_column_names.append(k)
    half_column_names.append('Total')
    half_column_names.append('% Of Revenue')
    half_column_names.append('Cumulative %')
    half_column_names_final = [(half_column_names)]

    quarter_column_names = ['Song Title']
    for k in statement_period_quarter_list:
      quarter_column_names.append(k)
    quarter_column_names.append('Total')
    quarter_column_names.append('% Of Revenue')
    quarter_column_names.append('Cumulative %')
    quarter_column_names_final = [(quarter_column_names)]

    month_column_names = ['Song Title']
    for k in statement_period_month_list:
      month_column_names.append(k)
    month_column_names.append('Total')
    month_column_names.append('% Of Revenue')
    month_column_names.append('Cumulative %')
    month_column_names_final = [(month_column_names)]

    #Income x Rev x Half Columns
    half_column_names_1 = ['Income Type']
    for k in statement_period_half:
      half_column_names_1.append(k)
    half_column_names_1.append('Total')
    half_column_names_1.append('% Of Revenue')
    half_column_names_1.append('Cumulative %')
    half_column_names_final_1 = [(half_column_names_1)]

    quarter_column_names_1 = ['Income Type']
    for k in statement_period_quarter_list:
      quarter_column_names_1.append(k)
    quarter_column_names_1.append('Total')
    quarter_column_names_1.append('% Of Revenue')
    quarter_column_names_1.append('Cumulative %')
    quarter_column_names_final_1 = [(quarter_column_names_1)]

    month_column_names_1 = ['Income Type']
    for k in statement_period_month_list:
      month_column_names_1.append(k)
    month_column_names_1.append('Total')
    month_column_names_1.append('% Of Revenue')
    month_column_names_1.append('Cumulative %')
    month_column_names_final_1 = [(month_column_names_1)]

    #Source x Rev x Half Columns
    half_column_names_2 = ['Source']
    for k in statement_period_half:
      half_column_names_2.append(k)
    half_column_names_2.append('Total')
    half_column_names_2.append('% Of Revenue')
    half_column_names_2.append('Cumulative %')
    half_column_names_final_2 = [(half_column_names_2)]

    quarter_column_names_2 = ['Source']
    for k in statement_period_quarter_list:
      quarter_column_names_2.append(k)
    quarter_column_names_2.append('Total')
    quarter_column_names_2.append('% Of Revenue')
    quarter_column_names_2.append('Cumulative %')
    quarter_column_names_final_2 = [(quarter_column_names_2)]

    month_column_names_2 = ['Source']
    for k in statement_period_month_list:
      month_column_names_2.append(k)
    month_column_names_2.append('Total')
    month_column_names_2.append('% Of Revenue')
    month_column_names_2.append('Cumulative %')
    month_column_names_final_2 = [(month_column_names_2)]

    #Song x Income x Rev x Half Columns
    half_column_names_3 = ['Song Title', 'Income Type']
    for k in statement_period_half:
      half_column_names_3.append(k)
    half_column_names_3.append('Total')
    half_column_names_3.append('% Of Revenue')
    half_column_names_3.append('Cumulative %')
    half_column_names_final_3 = [(half_column_names_3)]

    quarter_column_names_3 = ['Song Title', 'Income Type']
    for k in statement_period_quarter_list:
      quarter_column_names_3.append(k)
    quarter_column_names_3.append('Total')
    quarter_column_names_3.append('% Of Revenue')
    quarter_column_names_3.append('Cumulative %')
    quarter_column_names_final_3 = [(quarter_column_names_3)]

    month_column_names_3 = ['Song Title', 'Income Type']
    for k in statement_period_month_list:
      month_column_names_3.append(k)
    month_column_names_3.append('Total')
    month_column_names_3.append('% Of Revenue')
    month_column_names_3.append('Cumulative %')
    month_column_names_final_3 = [(month_column_names_3)]

    #Artist x Rev x Half Columns
    half_column_names_4 = ['Release Artist']
    for k in statement_period_half:
      half_column_names_4.append(k)
    half_column_names_4.append('Total')
    half_column_names_4.append('% Of Revenue')
    half_column_names_4.append('Cumulative %')
    half_column_names_final_4 = [(half_column_names_4)]

    quarter_column_names_4 = ['Release Artist']
    for k in statement_period_quarter_list:
      quarter_column_names_4.append(k)
    quarter_column_names_4.append('Total')
    quarter_column_names_4.append('% Of Revenue')
    quarter_column_names_4.append('Cumulative %')
    quarter_column_names_final_4 = [(quarter_column_names_4)]

    month_column_names_4 = ['Release Artist']
    for k in statement_period_month_list:
      month_column_names_4.append(k)
    month_column_names_4.append('Total')
    month_column_names_4.append('% Of Revenue')
    month_column_names_4.append('Cumulative %')
    month_column_names_final_4 = [(month_column_names_4)]

    #if len(monthly) == 1:
      #smallest_period = 'Month'
      #mySQL_column = 'Month_Statement_9LC'
      #statement_list = month_statement_list
      #year_statement_list = year_statement_list_month
      #column_names_final = month_column_names_final
      #column_names = month_column_names
      #column_names_final_1 = month_column_names_final_1
      #column_names_1 = month_column_names_1
      #column_names_final_2 = month_column_names_final_2
      #column_names_2 = month_column_names_2
      #column_names_final_3 = month_column_names_final_3
      #column_names_3 = month_column_names_3
      #column_names_final_4 = month_column_names_final_4
      #column_names_4 = month_column_names_4
    if len(quarterly) == 1:
      smallest_period = 'Quarter'
      mySQL_column = 'Quarter_Statement_9LC'
      statement_list = quarter_statement_list
      year_statement_list = year_statement_list_quarter
      column_names_final = quarter_column_names_final
      column_names = quarter_column_names
      column_names_final_1 = quarter_column_names_final_1
      column_names_1 = quarter_column_names_1
      column_names_final_2 = quarter_column_names_final_2
      column_names_2 = quarter_column_names_2
      column_names_final_3 = quarter_column_names_final_3
      column_names_3 = quarter_column_names_3
      column_names_final_4 = quarter_column_names_final_4
      column_names_4 = quarter_column_names_4
    else:
      smallest_period = 'Half'
      mySQL_column = 'Half_Statement_9LC'
      statement_list = half_statement_list
      year_statement_list = year_statement_list_half
      column_names_final = half_column_names_final
      column_names = half_column_names
      column_names_final_1 = half_column_names_final_1
      column_names_1 = half_column_names_1
      column_names_final_2 = half_column_names_final_2
      column_names_2 = half_column_names_2
      column_names_final_3 = half_column_names_final_3
      column_names_3 = half_column_names_3
      column_names_final_4 = half_column_names_final_4
      column_names_4 = half_column_names_4


    wb = Workbook()
    ws = wb.active
    ws.title = 'Song x Rev x {}'.format(smallest_period)
    ws_1 = wb.create_sheet(title='Income x Rev x {}'.format(smallest_period))
    ws_2 = wb.create_sheet(title='Source x Rev x {}'.format(smallest_period))
    ws_3 = wb.create_sheet(title='Song x Income x Rev x {}'.format(smallest_period))
    ws_4 = wb.create_sheet(title='Artist x Rev x {}'.format(smallest_period))

  #Add styles to workbook
    wb.add_named_style(number_style)
    wb.add_named_style(header_style)
    wb.add_named_style(title_style)
    wb.add_named_style(name_style)
    wb.add_named_style(total_style)
    wb.add_named_style(sub_header_style)
    wb.add_named_style(total_row_style)
    wb.add_named_style(total_label_style)
    wb.add_named_style(publisher_label_style)
    wb.add_named_style(not_available_style)
    wb.add_named_style(bold_name_style)
    wb.add_named_style(total_revenue_style)

  #Sheet 1: Song x Rev x Half
    #Song list
    find_songs = '''SELECT Song_Name_9LC
                        FROM Master WHERE Third_Party_9LC = "{}" GROUP BY Song_Name_9LC'''.format(thirdparty)
    mycursor.execute(find_songs)
    songs = [i[0] for i in mycursor.fetchall()]
    song_count = len(songs)

    #Build main pivot table
    select_table_1_1 = '''SELECT Song_Name_9LC,'''
    select_table_1_2 = ""
    for j, k in zip(year_statement_list, statement_list):
      select_table_1_2 += (''' sum( CASE WHEN Year_Statement_9LC = "{}" AND {} = "{}" AND Song_Name_9LC <> "Pool Revenue" 
                             THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(j,mySQL_column,k,j,k))
    select_table_1_3 = '''sum( CASE WHEN Song_Name_9LC <> "Pool Revenue" AND Statement_Period_Half_9LC >= "{}"
                     THEN Adjusted_Royalty_SB ELSE "" END) AS `Total`
                     FROM Master WHERE Song_Name_9LC <> "Pool Revenue" AND Statement_Period_Half_9LC <> ""
                     AND Third_Party_9LC = "{}"
                     GROUP BY Song_Name_9LC ORDER BY `Total` DESC'''.format(cut_off,thirdparty)
    select_table_1 = select_table_1_1 + select_table_1_2 + select_table_1_3
    mycursor.execute(select_table_1)
    table_1 = mycursor.fetchall()

    #Add pool revenue line at bottom
    pool_rev_1_1 = '''SELECT Song_Name_9LC,'''
    pool_rev_1_2 = ""
    for l,m in zip(year_statement_list, statement_list):
      pool_rev_1_2 += ('''sum( CASE WHEN Year_Statement_9LC = "{}" AND {} = "{}" AND Song_Name_9LC = "Pool Revenue"
                        THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(l,mySQL_column,m,l,m))
    pool_rev_1_3 = '''sum( CASE WHEN Song_Name_9LC = "Pool Revenue" AND Statement_Period_Half_9LC >= "{}"
                  THEN Adjusted_Royalty_SB ELSE "" END) AS `Total` 
                  FROM Master WHERE Song_Name_9LC = "Pool Revenue" AND Statement_Period_Half_9LC <> ""
                  AND Third_Party_9LC = "{}"
                  GROUP BY Song_Name_9LC ORDER BY `Total` DESC'''.format(cut_off, thirdparty)
    pool_rev_1 = pool_rev_1_1 + pool_rev_1_2 + pool_rev_1_3
    mycursor.execute(pool_rev_1)
    pool_revenue_1 = mycursor.fetchall()
    final_table_1 = table_1 + pool_revenue_1

    #Size of worksheet
    for column_no in range(1, len(column_names) + 1):
      for row_no in range(1, song_count + 1):
        ws.cell(row=row_no, column=column_no)

    #Add column names to worksheet
    header_row = ws[1]
    for (header_cell, i) in zip(header_row, range(len(column_names))):
      header_cell.value = column_names[i]
    for cell in header_row:
      cell.font = Font(bold=True)
      cell.fill = PatternFill("solid", fgColor="A6ACAF")
      cell.alignment = Alignment(horizontal="center", vertical="center")

    #Add table to worksheet
    table_rows = ws[2:song_count + 1]
    for (row, j) in zip(table_rows, final_table_1):
      for (cell, k) in zip(row, range(len(j))):
        cell.value = j[k]
        cell.style = 'Comma'
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.font = Font(name="Calibri", size="11")
    for song_no in range(2, song_count + 2):
      ws.cell(row=song_no, column=1).alignment = Alignment(horizontal="left", vertical="center")
      ws.cell(row=song_no, column=(len(column_names) - 2)).font = Font(bold=True)

    #Add total row
    column_letters_2 = column_letters[1:]
    for (l, m) in zip(range(2, len(column_names) - 1), column_letters_2):
      ws.cell(row=song_count + 2, column=l).value = "=SUM({}{}:{}{})".format(m, 2, m, song_count + 1)
      ws.cell(row=song_count + 2, column=l).style = 'Comma'
      ws.cell(row=song_count + 2, column=l).alignment = Alignment(horizontal="right", vertical="center")
      ws.cell(row=song_count + 2, column=l).font = Font(bold=True)
    ws.cell(row=song_count + 2, column=1).value = 'Total'
    ws.cell(row=song_count + 2, column=1).font = Font(bold='True')
    ws.cell(row=song_count + 2, column=1).fill = PatternFill("solid", fgColor="A6ACAF")
    ws.cell(row=song_count + 2, column=1).alignment = Alignment(horizontal="center", vertical="center")

    #Add % of revenue column
    total_column_letter = column_letters[len(column_names) - 3]
    for n in range(2, song_count + 2):
      ws.cell(row=n, column=len(column_names) - 1).value = "=({}{}/{}{})".format(total_column_letter, n,
                                                                                     total_column_letter,
                                                                                     song_count + 2)
      ws.cell(row=n, column=len(column_names) - 1).style = 'Percent'

    #Add cumulative % column
    percent_rev_column_letter = column_letters[len(column_names) - 2]
    cumulative_rev_column_letter = column_letters[len(column_names) - 1]
    ws.cell(row=2, column=len(column_names)).value = "=({}2)".format(percent_rev_column_letter)
    ws.cell(row=2, column=len(column_names)).style = 'Percent'
    for o in range(3, song_count + 2):
      ws.cell(row=o, column=len(column_names)).value = "=({}{}+{}{})".format(cumulative_rev_column_letter, o - 1,
                                                                                 percent_rev_column_letter, o)
      ws.cell(row=o, column=len(column_names)).style = 'Percent'

#Sheet 2: Income x Rev x Half
    #Find number of income types
    find_income_type = '''SELECT Normalized_Income_Type_9LC
                          FROM Master WHERE Third_Party_9LC = "{}" GROUP BY Normalized_Income_Type_9LC'''.format(thirdparty)
    mycursor.execute(find_income_type)
    income_type = [i[0] for i in mycursor.fetchall()]
    income_type_count = len(income_type)

    #Build main pivot table
    select_table_1_1 = '''SELECT Normalized_Income_Type_9LC,'''
    select_table_1_2 = ""
    for j, k in zip(year_statement_list, statement_list):
      select_table_1_2 += (''' sum( CASE WHEN Year_Statement_9LC = "{}" AND {} = "{}"
                             THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(j,mySQL_column,k, j,k))
    select_table_1_3 = '''sum( CASE WHEN Statement_Period_Half_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END) 
                     AS `Total`
                     FROM Master WHERE Statement_Period_Half_9LC <> ""
                     AND Third_Party_9LC = "{}"
                     GROUP BY Normalized_Income_Type_9LC ORDER BY `Total` DESC'''.format(cut_off, thirdparty)
    select_table_1 = select_table_1_1 + select_table_1_2 + select_table_1_3
    mycursor.execute(select_table_1)
    table_1 = mycursor.fetchall()

    #Size of worksheet
    for column_no in range(1, len(column_names_1) + 1):
      for row_no in range(1, income_type_count + 1):
        ws_1.cell(row=row_no, column=column_no)

    #Add column names to worksheet
    header_row = ws_1[1]
    for (header_cell, i) in zip(header_row, range(len(column_names_1))):
      header_cell.value = column_names_1[i]
    for cell in header_row:
      cell.font = Font(bold=True)
      cell.fill = PatternFill("solid", fgColor="A6ACAF")
      cell.alignment = Alignment(horizontal="center", vertical="center")

    #Add table to worksheet
    table_rows = ws_1[2:income_type_count + 1]
    for j, k in zip(range(2, income_type_count + 2), table_1):
      for l, m in zip(range(1, len(column_names_1)), range(len(k))):
        ws_1.cell(row=j, column=l).value = k[m]
        ws_1.cell(row=j, column=l).style = 'number_style'

    for income_no in range(2, income_type_count + 2):
      ws_1.cell(row=income_no, column=1).alignment = Alignment(horizontal="left", vertical="center")
      ws_1.cell(row=income_no, column=(len(column_names_1) - 2)).style = 'total_style'

    #Add total row
    column_letters_2 = column_letters[1:]
    for (l, m) in zip(range(2, len(column_names_1) - 1), column_letters_2):
      ws_1.cell(row=income_type_count + 2, column=l).value = "=SUM({}{}:{}{})".format(m, 2, m, income_type_count + 1)
      ws_1.cell(row=income_type_count + 2, column=l).style = 'total_style'
    ws_1.cell(row=income_type_count + 2, column=1).value = 'Total'
    ws_1.cell(row=income_type_count + 2, column=1).font = Font(bold='True')
    ws_1.cell(row=income_type_count + 2, column=1).fill = PatternFill("solid", fgColor="A6ACAF")
    ws_1.cell(row=income_type_count + 2, column=1).alignment = Alignment(horizontal="center", vertical="center")

    #Add % of revenue column
    total_column_letter = column_letters[len(column_names_1) - 3]
    for n in range(2, income_type_count + 2):
      ws_1.cell(row=n, column=len(column_names_1) - 1).value = "=({}{}/{}{})".format(total_column_letter, n,
                                                                                     total_column_letter,
                                                                                     income_type_count + 2)
      ws_1.cell(row=n, column=len(column_names_1) - 1).style = 'Percent'

    #Add cumulative % column
    percent_rev_column_letter = column_letters[len(column_names_1) - 2]
    cumulative_rev_column_letter = column_letters[len(column_names_1) - 1]
    ws_1.cell(row=2, column=len(column_names_1)).value = "=({}2)".format(percent_rev_column_letter)
    ws_1.cell(row=2, column=len(column_names_1)).style = 'Percent'
    for o in range(3, income_type_count + 2):
      ws_1.cell(row=o, column=len(column_names_1)).value = "=({}{}+{}{})".format(cumulative_rev_column_letter, o - 1,
                                                                                 percent_rev_column_letter, o)
      ws_1.cell(row=o, column=len(column_names_1)).style = 'Percent'

  #Sheet 3: Source x Rev x Half
    #Find number of sources
    find_source = '''SELECT Normalized_Source_9LC
                      FROM Master WHERE Third_Party_9LC = "{}" GROUP BY Normalized_Source_9LC'''.format(thirdparty)
    mycursor.execute(find_source)
    source_type = [i[0] for i in mycursor.fetchall()]
    source_type_count = len(source_type)

    #Build main pivot table
    select_table_2_1 = '''SELECT Normalized_Source_9LC,'''
    select_table_2_2 = ""
    for j,k in zip(year_statement_list, statement_list):
      select_table_2_2 += (''' sum( CASE WHEN Year_Statement_9LC = "{}" AND {} = "{}" 
                             AND Normalized_Source_9LC <> "Pool Revenue"
                             THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(j,mySQL_column,k, j,k))
    select_table_2_3 = '''sum( CASE WHEN Normalized_Source_9LC <> "Pool Revenue" AND Statement_Period_Half_9LC >= "{}" 
                     THEN Adjusted_Royalty_SB ELSE "" END) 
                     AS `Total`
                     FROM Master WHERE Normalized_Source_9LC <> "Pool Revenue" AND Statement_Period_Half_9LC <> ""
                     AND Third_Party_9LC = "{}"
                     GROUP BY Normalized_Source_9LC ORDER BY `Total` DESC'''.format(cut_off, thirdparty)
    select_table_2 = select_table_2_1 + select_table_2_2 + select_table_2_3
    mycursor.execute(select_table_2)
    table_2 = mycursor.fetchall()

    #Add pool revenue line at bottom
    pool_rev_2_1 = '''SELECT Normalized_Source_9LC,'''
    pool_rev_2_2 = ""
    for l,m in zip(year_statement_list, statement_list):
      pool_rev_2_2 += ('''sum( CASE WHEN Year_Statement_9LC = "{}" AND {} = "{}" AND Normalized_Source_9LC = "Pool Revenue"
                          THEN Adjusted_Royalty_SB ELSE NULL END) AS `{} {}`,'''.format(l,mySQL_column,m, l,m))
    pool_rev_2_3 = '''sum( CASE WHEN Normalized_Source_9LC = "Pool Revenue" AND Statement_Period_Half_9LC >= "{}" 
                    THEN Adjusted_Royalty_SB ELSE NULL END) AS `Total` 
                    FROM Master WHERE Normalized_Source_9LC = "Pool Revenue" AND Statement_Period_Half_9LC <> ""
                    AND Third_Party_9LC = "{}"
                    GROUP BY Normalized_Source_9LC ORDER BY `Total` DESC'''.format(cut_off, thirdparty)
    pool_rev_2 = pool_rev_2_1 + pool_rev_2_2 + pool_rev_2_3
    mycursor.execute(pool_rev_2)
    pool_revenue_2 = mycursor.fetchall()
    final_table_2 = table_2 + pool_revenue_2

    #Size of worksheet
    for column_no in range(1, len(column_names_2) + 1):
      for row_no in range(1, source_type_count + 1):
        ws_2.cell(row=row_no, column=column_no)

    #Add column names to worksheet
    header_row = ws_2[1]
    for (header_cell, i) in zip(header_row, range(len(column_names_2))):
      header_cell.value = column_names_2[i]
    for cell in header_row:
      cell.font = Font(bold=True)
      cell.fill = PatternFill("solid", fgColor="A6ACAF")
      cell.alignment = Alignment(horizontal="center", vertical="center")

    #Add table to worksheet
    table_rows = ws_2[2:source_type_count + 1]
    for (row, j) in zip(table_rows, final_table_2):
      for (cell, k) in zip(row, range(len(j))):
        cell.value = j[k]
        cell.style = 'Comma'
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.font = Font(name="Calibri", size="11")
    for source_no in range(2, source_type_count + 2):
      ws_2.cell(row=source_no, column=1).alignment = Alignment(horizontal="left", vertical="center")
      ws_2.cell(row=source_no, column=(len(column_names_2) - 2)).font = Font(bold=True)

    #Add total row
    column_letters_2 = column_letters[1:]
    for (l, m) in zip(range(2, len(column_names_2) - 1), column_letters_2):
      ws_2.cell(row=source_type_count + 2, column=l).value = "=SUM({}{}:{}{})".format(m, 2, m, source_type_count + 1)
      ws_2.cell(row=source_type_count + 2, column=l).style = 'Comma'
      ws_2.cell(row=source_type_count + 2, column=l).alignment = Alignment(horizontal="right", vertical="center")
      ws_2.cell(row=source_type_count + 2, column=l).font = Font(bold=True)
    ws_2.cell(row=source_type_count + 2, column=1).value = 'Total'
    ws_2.cell(row=source_type_count + 2, column=1).font = Font(bold='True')
    ws_2.cell(row=source_type_count + 2, column=1).fill = PatternFill("solid", fgColor="A6ACAF")
    ws_2.cell(row=source_type_count + 2, column=1).alignment = Alignment(horizontal="center", vertical="center")

    #Add % of revenue column
    total_column_letter = column_letters[len(column_names_2) - 3]
    for n in range(2, source_type_count + 2):
      ws_2.cell(row=n, column=len(column_names_2) - 1).value = "=({}{}/{}{})".format(total_column_letter, n,
                                                                                     total_column_letter,
                                                                                     source_type_count + 2)
      ws_2.cell(row=n, column=len(column_names_2) - 1).style = 'Percent'

    #Add cumulative % column
    percent_rev_column_letter = column_letters[len(column_names_2) - 2]
    cumulative_rev_column_letter = column_letters[len(column_names_2) - 1]
    ws_2.cell(row=2, column=len(column_names_2)).value = "=({}2)".format(percent_rev_column_letter)
    ws_2.cell(row=2, column=len(column_names_2)).style = 'Percent'
    for o in range(3, source_type_count + 2):
      ws_2.cell(row=o, column=len(column_names_2)).value = "=({}{}+{}{})".format(cumulative_rev_column_letter, o - 1,
                                                                                 percent_rev_column_letter, o)
      ws_2.cell(row=o, column=len(column_names_2)).style = 'Percent'

  #Sheet 4: Song x Income x Rev x Half
    #Find number of songs
    find_songs = '''SELECT Song_Name_9LC, sum( Adjusted_Royalty_SB) 
                    FROM Master WHERE Song_Name_9LC <> "Pool Revenue" AND Statement_Period_Half_9LC >= "{}"
                    AND Third_Party_9LC = "{}"
                    GROUP BY Song_Name_9LC ORDER BY sum( Adjusted_Royalty_SB) DESC'''.format(cut_off, thirdparty)
    mycursor.execute(find_songs)
    songs = [i[0] for i in mycursor.fetchall()]
    find_pool_rev = '''SELECT Song_Name_9LC, sum( Adjusted_Royalty_SB)
                       FROM Master WHERE Song_Name_9LC = "Pool Revenue" AND Statement_Period_Half_9LC >= "{}"
                       AND Third_Party_9LC = "{}"
                       GROUP BY Song_Name_9LC'''.format(cut_off, thirdparty)
    mycursor.execute(find_pool_rev)
    pool_rev = [i[0] for i in mycursor.fetchall()]
    all_songs = songs + pool_rev

    #Build tables
    total_row_no = 0
    ws_3.insert_cols(1, len(column_names_3))
    for s in all_songs:
      s_string = mydb.converter.escape(s)
      select_table_3_1 = '''SELECT Song_Name_9LC,
                              Normalized_Income_Type_9LC,'''
      select_table_3_2 = ""
      for j, k in zip(year_statement_list, statement_list):
        select_table_3_2 += (''' sum( CASE WHEN Year_Statement_9LC = "{}" AND {} = "{}"
                                   THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(j,mySQL_column,k, j,k))
      select_table_3_3 = '''sum( CASE WHEN Statement_Period_Half_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `Total`
                              FROM Master WHERE Song_Name_9LC = "{}" AND Statement_Period_Half_9lC <> ""
                              AND Third_Party_9LC = "{}"
                              GROUP BY Song_Name_9LC, Normalized_Income_Type_9LC ORDER BY `Total` DESC'''.format(
        cut_off, s_string, thirdparty)
      select_table_3 = select_table_3_1 + select_table_3_2 + select_table_3_3
      mycursor.execute(select_table_3)
      table_3 = mycursor.fetchall()
      song_table_3 = column_names_final_3 + table_3

      #Add tables to sheet
      total_row_no += len(song_table_3) + 1
      space_row_no = total_row_no + 1
      header_row_no = total_row_no - len(song_table_3)
      total_column_no = len(column_names_3) - 1
      ws_3.append(range(len(song_table_3) + 2))
      song_table_rows = ws_3[header_row_no:space_row_no]
      for (row, l) in zip(song_table_rows, song_table_3):
        for (cell, m) in zip(row, range(len(l))):
          cell.value = l[m]

      #Add total row
      ws_3.cell(row=total_row_no, column=2).value = 'Total'
      for n in range(3, total_column_no):
        ws_3.cell(row=total_row_no, column=n).value = '=SUM({}{}:{}{})'.format(column_letters[n - 1], header_row_no + 1,
                                                                               column_letters[n - 1], total_row_no - 1)

      #Add percentage and cumulative percentage columns
      total_column_letter = column_letters[len(column_names_3) - 3]
      percentage_column_letter = column_letters[len(column_names_3) - 2]
      cumulative_column_letter = column_letters[len(column_names_3) - 1]
      for o in range(header_row_no + 1, total_row_no):
        ws_3.cell(row=o, column=total_column_no).value = '={}{}/{}{}'.format(total_column_letter, o,
                                                                             total_column_letter, total_row_no)
        ws_3.cell(row=o, column=total_column_no).style = 'Percent'
        ws_3.cell(row=o, column=total_column_no).font = Font(name="Calibri", size="11")
      ws_3.cell(row=header_row_no + 1, column=total_column_no + 1).value = '={}{}'.format(percentage_column_letter,
                                                                                          header_row_no + 1)
      ws_3.cell(row=header_row_no + 1, column=total_column_no + 1).style = 'Percent'
      ws_3.cell(row=header_row_no + 1, column=total_column_no + 1).font = Font(name="Calibri", size=11)
      for p in range(header_row_no + 2, total_row_no):
        ws_3.cell(row=p, column=total_column_no + 1).value = '={}{}+{}{}'.format(cumulative_column_letter, p - 1,
                                                                                 percentage_column_letter, p)
        ws_3.cell(row=p, column=total_column_no + 1).style = 'Percent'
        ws_3.cell(row=p, column=total_column_no + 1).font = Font(name="Calibri", size=11)

      #Format header row
      for b in range(1, len(column_names_3) + 1):
        ws_3.cell(row=header_row_no, column=b).fill = PatternFill("solid", fgColor="A6ACAF")
        ws_3.cell(row=header_row_no, column=b).alignment = Alignment(horizontal="center", vertical="center")
        ws_3.cell(row=header_row_no, column=b).font = Font(bold=True)

      #Format numbers
      for c in range(3, total_column_no):
        for d in range(header_row_no + 1, total_row_no + 1):
          ws_3.cell(row=d, column=c).style = 'Comma'
          ws_3.cell(row=d, column=c).font = Font(name="Calibri", size="11")

      #Format total row header
      ws_3.cell(row=total_row_no, column=2).fill = PatternFill("solid", fgColor="A6ACAF")
      ws_3.cell(row=total_row_no, column=2).font = Font(bold=True)
      ws_3.cell(row=total_row_no, column=2).alignment = Alignment(horizontal="center", vertical="center")

      #Format total row and column
      for g in range(header_row_no + 1, total_row_no):
        ws_3.cell(row=g, column=total_column_no - 1).font = Font(bold=True)
      for f in range(3, total_column_no):
        ws_3.cell(row=total_row_no, column=f).font = Font(bold=True)

      #Insert row between tables
      ws_3.insert_rows(total_row_no + 1)
      total_row_no += 1


    #Sheet 5: Artist x Rev x Half
    # Find number of artists
    find_songs = '''SELECT Release_Artist_9LC
                  FROM Master WHERE Third_Party_9LC = "{}" GROUP BY Release_Artist_9LC'''.format(thirdparty)
    mycursor.execute(find_songs)
    artists = [i[0] for i in mycursor.fetchall()]
    artist_count = len(artists)

    # Build main pivot table
    select_table_4_1 = '''SELECT Release_Artist_9LC,'''
    select_table_4_2 = ""
    for j,k in zip(year_statement_list, statement_list):
      select_table_4_2 += (''' sum( CASE WHEN Year_Statement_9LC = "{}" AND {} = "{}" AND Release_Artist_9LC <> "" 
                           THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(j,mySQL_column,k, j,k))
    select_table_4_3 = '''sum( CASE WHEN Statement_Period_Half_9LC >= "{}" AND Release_Artist_9LC <> ""
                   THEN Adjusted_Royalty_SB ELSE "" END) AS `Total`
                   FROM Master WHERE Statement_Period_Half_9LC <> "" AND Release_Artist_9LC <> ""
                   AND Third_Party_9LC = "{}"
                   GROUP BY Release_Artist_9LC ORDER BY `Total` DESC'''.format(cut_off, thirdparty)
    select_table_4 = select_table_4_1 + select_table_4_2 + select_table_4_3
    mycursor.execute(select_table_4)
    table_4 = mycursor.fetchall()

      # Add line for unknown artists
    unknown_artists_4_1 = '''SELECT Release_Artist_9LC,'''
    unknown_artists_4_2 = ""
    for l,m in zip(year_statement_list, statement_list):
      unknown_artists_4_2 += ('''sum( CASE WHEN Year_Statement_9LC = "{}" AND {} = "{}" AND Release_Artist_9LC = ""
                             THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(l,mySQL_column,m, l,m))
    unknown_artists_4_3 = '''sum( CASE WHEN Release_Artist_9LC = "" AND Statement_Period_Half_9LC >= "{}"
                         THEN Adjusted_Royalty_SB ELSE "" END) AS `Total`
                         FROM Master WHERE Statement_Period_Half_9LC <> "" AND Release_Artist_9LC = ""
                         AND Third_Party_9LC = "{}"
                         GROUP BY Release_Artist_9LC ORDER BY `Total` DESC'''.format(cut_off, thirdparty)
    unknown_artists_4 = unknown_artists_4_1 + unknown_artists_4_2 + unknown_artists_4_3
    mycursor.execute(unknown_artists_4)
    unknown_artist_line_4 = mycursor.fetchall()
    final_table_4 = table_4 + unknown_artist_line_4


    # Size of worksheet
    for column_no in range(1, len(column_names_4) + 1):
      for row_no in range(1, artist_count + 1):
          ws_4.cell(row=row_no, column=column_no)

    # Add column names to worksheet
    header_row = ws_4[1]
    for (header_cell, i) in zip(header_row, range(len(column_names_4))):
      header_cell.value = column_names_4[i]
    for cell in header_row:
      cell.font = Font(bold=True)
      cell.fill = PatternFill("solid", fgColor="A6ACAF")
      cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add table to worksheet
    table_rows = ws_4[2:artist_count + 1]
    for (row, j) in zip(table_rows, final_table_4):
      for (cell, k) in zip(row, range(len(j))):
          cell.value = j[k]
          cell.style = 'Comma'
          cell.alignment = Alignment(horizontal="right", vertical="center")
          cell.font = Font(name="Calibri", size="11")
    for song_no in range(2, artist_count + 2):
      ws_4.cell(row=song_no, column=1).alignment = Alignment(horizontal="left", vertical="center")
      ws_4.cell(row=song_no, column=(len(column_names_4) - 2)).font = Font(bold=True)
    ws_4.cell(row=artist_count + 1, column=1).value = 'Unknown Artists'

    # Add total row
    column_letters_2 = column_letters[1:]
    for (l, m) in zip(range(2, len(column_names_4) - 1), column_letters_2):
      ws_4.cell(row=artist_count + 2, column=l).value = "=SUM({}{}:{}{})".format(m, 2, m, artist_count + 1)
      ws_4.cell(row=artist_count + 2, column=l).style = 'Comma'
      ws_4.cell(row=artist_count + 2, column=l).alignment = Alignment(horizontal="right", vertical="center")
      ws_4.cell(row=artist_count + 2, column=l).font = Font(bold=True)
    ws_4.cell(row=artist_count + 2, column=1).value = 'Total'
    ws_4.cell(row=artist_count + 2, column=1).font = Font(bold='True')
    ws_4.cell(row=artist_count + 2, column=1).fill = PatternFill("solid", fgColor="A6ACAF")
    ws_4.cell(row=artist_count + 2, column=1).alignment = Alignment(horizontal="center", vertical="center")

    # Add % of revenue column
    total_column_letter = column_letters[len(column_names_4) - 3]
    for n in range(2, artist_count + 2):
      ws_4.cell(row=n, column=len(column_names_4) - 1).value = "=({}{}/{}{})".format(total_column_letter, n,
                                                                                     total_column_letter,
                                                                                     artist_count + 2)
      ws_4.cell(row=n, column=len(column_names_4) - 1).style = 'Percent'

    # Add cumulative % column
    percent_rev_column_letter = column_letters[len(column_names_4) - 2]
    cumulative_rev_column_letter = column_letters[len(column_names_4) - 1]
    ws_4.cell(row=2, column=len(column_names_4)).value = "=({}2)".format(percent_rev_column_letter)
    ws_4.cell(row=2, column=len(column_names_4)).style = 'Percent'
    for o in range(3, artist_count + 2):
      ws_4.cell(row=o, column=len(column_names_4)).value = "=({}{}+{}{})".format(cumulative_rev_column_letter,
                                                                                 o - 1,
                                                                                 percent_rev_column_letter, o)
      ws_4.cell(row=o, column=len(column_names_4)).style = 'Percent'

    wb.save(filename)

#thirdpartybooks('Neil Ormandy New_61b3b26dc94b68a289900e7e')