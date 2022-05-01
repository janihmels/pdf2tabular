from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, NamedStyle
import pymysql
from pymysql.converters import escape_string
from datetime import date
from pandas_utils.pandas_cursor import pandas_cursor


def songxUSAxrevxhalf(database, df, filename):
  mycursor = pandas_cursor(df=df)

  # Third party list
  find_third_parties = '''SELECT Third_Party_9LC FROM Master GROUP BY Third_Party_9LC'''
  mycursor.execute(find_third_parties)
  third_party_list = [i[0] for i in mycursor.fetchall()]
  third_party_string = ', '.join('"{}"'.format(str(x)) for x in third_party_list)

  #Find whether rights type filled in
  find_rights_type = '''SELECT Rights_Type_9LC FROM Master WHERE Rights_Type_9LC <> "" GROUP BY Rights_Type_9LC'''
  mycursor.execute(find_rights_type)
  rights_type_list = [i[0] for i in mycursor.fetchall()]
  print(len(rights_type_list))
  if len(rights_type_list) != 0:
    mojo = True
  else:
    mojo = False

#Create sheets
  wb = Workbook()
  ws = wb.active
  ws_1 = wb.create_sheet(title = 'S Song x Territory x Rev x Half', index=0)

#Current year
  todays_date = date.today()
  current_year = todays_date.year

#Find most recent year in data
  find_recent_year = '''SELECT Year_Statement_9LC FROM Master GROUP BY Year_Statement_9LC'''
  mycursor.execute(find_recent_year)
  recent_years = [i[0] for i in mycursor.fetchall()]
  recent_year = recent_years[-1]
  print(recent_year)
  print(current_year)
  if recent_year == current_year:
    find_cut_off_year = current_year - 1
  else:
    find_cut_off_year = int(recent_year) - 1

#Find cut off
  find_period = 'SELECT DISTINCT Statement_Period_Half_9LC FROM Master WHERE Year_Statement_9LC <= "{}" ORDER BY Statement_Period_Half_9LC'.format(find_cut_off_year)
  mycursor.execute(find_period)
  statement_period_half_blank = [i[0] for i in mycursor.fetchall()]

  def check_blank(period):
    if period == '':
      return False
    else:
      return True

  remove_blank = filter(check_blank, statement_period_half_blank)
  statement_period_minus_blank = list(remove_blank)
  if len(statement_period_minus_blank) > 10:
    statement_period_half = statement_period_minus_blank[-10:]
    standard_cut_off = statement_period_minus_blank[-10]
  else:
    statement_period_half = statement_period_minus_blank
    standard_cut_off = statement_period_minus_blank[0]
  print(statement_period_half)
  print(standard_cut_off)
  year_cut_off = standard_cut_off[0:4]

#Get list of statement year periods
  find_year_period = '''SELECT Year_Statement_9LC FROM Master 
                                          WHERE Year_Statement_9LC >= "{}" GROUP BY Year_Statement_9LC'''.format(
    year_cut_off)
  mycursor.execute(find_year_period)
  complete_list = [i for i in mycursor.fetchall()]
  statement_period_year_list = [i[0] for i in complete_list]
  year_statement_list_year = [i[0] for i in complete_list]
  yearly_statement_list = [i[0] for i in complete_list]
  year_list = list(dict.fromkeys(year_statement_list_year))

#Get list of statement half periods
  find_half_period = '''SELECT Year_Statement_9LC, Half_Statement_9LC FROM Master 
                                       WHERE Half_Statement_9LC <> "" AND Statement_Period_Half_9LC >= "{}" GROUP BY Year_Statement_9LC, Half_Statement_9LC'''.format(
    standard_cut_off)
  mycursor.execute(find_half_period)
  complete_list = [i for i in mycursor.fetchall()]
  statement_period_half_list = [i[0] + ' ' + i[1] for i in complete_list]
  year_statement_list_half = [i[0] for i in complete_list]
  half_statement_list = [i[1] for i in complete_list]

#Get list of statement quarter periods
  find_quarter_period = '''SELECT Year_Statement_9LC, Quarter_Statement_9LC FROM Master 
                                     WHERE Quarter_Statement_9LC <> "" AND Statement_Period_Half_9LC >= "{}" GROUP BY Year_Statement_9LC, Quarter_Statement_9LC'''.format(
    standard_cut_off)
  mycursor.execute(find_quarter_period)
  complete_list = [i for i in mycursor.fetchall()]
  statement_period_quarter_list = [i[0] + ' ' + i[1] for i in complete_list]
  year_statement_list_quarter = [i[0] for i in complete_list]
  quarter_statement_list = [i[1] for i in complete_list]

#Most recent year
  test_year = int(year_statement_list_half[-1]) - 1
  test_third_parties = []
  for b in third_party_list:
    find_complete_year = '''SELECT Statement_Period_Half_9LC FROM Master WHERE Year_Statement_9LC = "{}" GROUP BY Statement_Period_Half_9LC'''.format(
      test_year)
    mycursor.execute(find_complete_year)
    complete_year = [i[0] for i in mycursor.fetchall()]
    if len(complete_year) / 2 == 1:
      test_third_parties.append(b)
  if len(test_third_parties) != len(third_party_list):
    base_year_value = int(test_year) - 2
  else:
    base_year_value = test_year

#Find types of third party
  find_type = '''SELECT Rights_Type_9LC, Third_Party_9LC FROM Master GROUP BY Rights_Type_9LC, Third_Party_9LC'''
  mycursor.execute(find_type)
  type_list = list(dict.fromkeys([i[0] for i in mycursor.fetchall()]))
  test = [i[1] for i in mycursor.fetchall()]
  type_string = ', '.join('"{}"'.format(str(x)) for x in type_list)
  print(type_list)
  print(test)

#Find third parties for each type
  third_party_type_list = []
  for a in type_list:
    third_party_sub_list = [a]
    find_third_party_type = '''SELECT Third_Party_9LC FROM Master WHERE Rights_Type_9LC = "{}" GROUP BY Third_Party_9LC'''.format(
      a)
    mycursor.execute(find_third_party_type)
    test_list = [i[0] for i in mycursor.fetchall()]
    for b in test_list:
      third_party_sub_list.append(b)
    third_party_type_list.append(third_party_sub_list)
    third_party_sub_list = []

#Find where quarterly data exists
  quarterly_types = []
  for b in type_list:
    find_third_parties = '''SELECT Third_Party_9LC FROM Master WHERE Rights_Type_9LC = "{}"'''.format(b)
    mycursor.execute(find_third_parties)
    third_parties = [i[0] for i in mycursor.fetchall()]
    find_quarterly_third_parties = '''SELECT Third_Party_9LC FROM Master WHERE Rights_Type_9LC = "{}" AND Quarter_Statement_9LC <> ""'''.format(
      b)
    mycursor.execute(find_quarterly_third_parties)
    quarterly_third_parties = [i[0] for i in mycursor.fetchall()]
    if len(third_parties) == len(quarterly_third_parties):
      quarterly_types.append(b)

#Find where half data exists
  half_types = []
  for b in type_list:
    find_third_parties = '''SELECT Third_Party_9LC FROM Master WHERE Rights_Type_9LC = "{}"'''.format(b)
    mycursor.execute(find_third_parties)
    third_parties = [i[0] for i in mycursor.fetchall()]
    find_half_third_parties = '''SELECT Third_Party_9LC FROM Master WHERE Rights_Type_9LC = "{}" AND Half_Statement_9LC <> ""'''.format(
      b)
    mycursor.execute(find_half_third_parties)
    half_third_parties = [i[0] for i in mycursor.fetchall()]
    if len(third_parties) == len(half_third_parties):
      half_types.append(b)

#Find number of songs
  find_songs = '''SELECT Song_Name_9LC, sum( Adjusted_Royalty_SB) 
                  FROM Master WHERE Song_Name_9LC <> "Pool Revenue" AND Statement_Period_Half_9LC >= "{}" 
                  GROUP BY Song_Name_9LC ORDER BY sum( Adjusted_Royalty_SB) DESC'''.format(standard_cut_off)
  mycursor.execute(find_songs)
  songs = [i[0] for i in mycursor.fetchall()]
  find_pool_rev = '''SELECT Song_Name_9LC, sum( Adjusted_Royalty_SB)
                     FROM Master WHERE Song_Name_9LC = "Pool Revenue" AND Statement_Period_Half_9LC >= "{}" 
                     GROUP BY Song_Name_9LC'''.format(standard_cut_off)
  mycursor.execute(find_pool_rev)
  pool_rev = [i[0] for i in mycursor.fetchall()]
  all_songs_combined = songs + pool_rev
  print(all_songs_combined)

#Create styles
  #Main numbers
  number_style = NamedStyle(name="number_style")
  number_style.alignment = Alignment(horizontal="right", vertical="center")
  number_style.font = Font(name="Calibri", size="11")
  number_style.number_format = '#,##0.00'

  #Title cells
  title_style = NamedStyle(name="title_style")
  title_style.alignment = Alignment(horizontal="center", vertical="center")
  title_style.font = Font(name="Calibri", size="11", bold=True)
  title_style.fill = PatternFill("solid", fgColor="A6ACAF")

  #Header cells
  header_style = NamedStyle(name="header_style")
  header_style.alignment = Alignment(horizontal="center", vertical="center")
  header_style.font = Font(name="Calibri", size="11", bold=True)
  header_style.fill = PatternFill("solid", fgColor="C9CDCF")

  #Sub header cells
  sub_header_style = NamedStyle(name="sub_header_style")
  sub_header_style.alignment = Alignment(horizontal="left", vertical="center")
  sub_header_style.font = Font(name="Calibri", size="11", bold=True, italic=True)

  #Name cells
  name_style = NamedStyle(name="name_style")
  name_style.alignment = Alignment(horizontal="left", vertical="center")
  name_style.font = Font(name="Calibri", size="11")

  #Total cells
  total_style = NamedStyle(name="total_style")
  total_style.alignment = Alignment(horizontal="right", vertical="center")
  total_style.font = Font(name="Calibri", size="11", bold=True)
  total_style.number_format = '#,##0.00'

  #Lined total row
  thin = Side(border_style="thin", color="000000")
  lined_total_style = NamedStyle(name="lined_total_style")
  lined_total_style.alignment = Alignment(horizontal="right", vertical="center")
  lined_total_style.font = Font(name="Calibri", size="11", bold=True)
  lined_total_style.number_format = '#,##0.00'
  lined_total_style.border = Border(top=thin, left=None, right=None, bottom=thin)

  #Total label style
  total_label_style = NamedStyle(name="total_label_style")
  total_label_style.alignment = Alignment(horizontal="center", vertical="center")
  total_label_style.font = Font(name="Calibri", size="11", bold=True)
  total_label_style.fill = PatternFill("solid", fgColor="C9CDCF")
  total_label_style.border = Border(top=thin, left=None, right=None, bottom=thin)

  #Publisher/PRO label style
  thick = Side(border_style="medium", color="000000")
  publisher_label_style = NamedStyle(name="publisher_label_style")
  publisher_label_style.alignment = Alignment(horizontal="center", vertical="center")
  publisher_label_style.font = Font(name="Calibri", size="11", bold=True, italic=True)
  publisher_label_style.border = Border(bottom=thick)

  #Not available style
  not_available_style = NamedStyle(name="not_available_style")
  not_available_style.alignment = Alignment(horizontal="center", vertical="center")
  not_available_style.font = Font(name="Calibri", size="11")

  #Add styles to workbook
  wb.add_named_style(number_style)
  wb.add_named_style(header_style)
  wb.add_named_style(title_style)
  wb.add_named_style(name_style)
  wb.add_named_style(total_style)
  wb.add_named_style(sub_header_style)
  wb.add_named_style(lined_total_style)
  wb.add_named_style(total_label_style)
  wb.add_named_style(publisher_label_style)
  wb.add_named_style(not_available_style)

#Create list of column names
  year_column_names = ['Song Title', 'Territory']
  for k in statement_period_year_list:
    year_column_names.append(k)
  year_column_names.append('Total')
  year_column_names.append('% Of Revenue')
  year_column_names.append('Cumulative %')
  year_column_names_final = [(year_column_names)]

  half_column_names = ['Song Title', 'Territory']
  for k in statement_period_half_list:
    half_column_names.append(k)
  half_column_names.append('Total')
  half_column_names.append('% Of Revenue')
  half_column_names.append('Cumulative %')
  half_column_names_final = [(half_column_names)]

  quarter_column_names = ['Song Title', 'Territory']
  for k in statement_period_quarter_list:
    quarter_column_names.append(k)
  quarter_column_names.append('Total')
  quarter_column_names.append('% Of Revenue')
  quarter_column_names.append('Cumulative %')
  quarter_column_names_final = [(quarter_column_names)]

#Column letters
  column_letters = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T',
                    'U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO',
                    'AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY','AZ']

#Build tables
#Build first table
  #ws = wb.active
  if third_party_type_list[0][0] == "":
    ws.title = "M Song x Territory x Rev x Half"
  else:
    ws.title = "M ({}) Song x Territory x Rev x Half".format(third_party_type_list[0][0])

  # Find smallest period
  if third_party_type_list[0][0] in quarterly_types:
    smallest_period = 'Quarter'
    mySQL_column = 'Quarter_Statement_9LC'
    statement_list = quarter_statement_list
    year_statement_list = year_statement_list_quarter
    column_names_final = quarter_column_names_final
    column_names = quarter_column_names
    cut_off_field = 'Statement_Period_Half_9LC'
    cut_off = standard_cut_off
  elif third_party_type_list[0][0] in half_types:
    smallest_period = 'Half'
    mySQL_column = 'Half_Statement_9LC'
    statement_list = half_statement_list
    year_statement_list = year_statement_list_half
    column_names_final = half_column_names_final
    column_names = half_column_names
    cut_off_field = 'Statement_Period_Half_9LC'
    cut_off = standard_cut_off
  else:
    smallest_period = 'Year'
    mySQL_column = 'Year_Statement_9LC'
    statement_list = yearly_statement_list
    year_statement_list = year_statement_list_year
    column_names_final = year_column_names_final
    column_names = year_column_names
    cut_off_field = 'Year_Statement_9LC'
    cut_off = year_cut_off



  #Song list per type
  find_songs = '''SELECT Song_Name_9LC, sum( Adjusted_Royalty_SB) 
                    FROM Master WHERE Song_Name_9LC <> "Pool Revenue" AND {} >= "{}" 
                    AND Rights_Type_9LC = "{}" 
                    GROUP BY Song_Name_9LC ORDER BY sum( Adjusted_Royalty_SB) DESC'''.format(cut_off_field,cut_off, third_party_type_list[0][0])
  mycursor.execute(find_songs)
  songs = [i[0] for i in mycursor.fetchall()]
  find_pool_rev = '''SELECT Song_Name_9LC, sum( Adjusted_Royalty_SB)
                       FROM Master WHERE Song_Name_9LC = "Pool Revenue" AND {} >= "{}"
                       AND Rights_Type_9LC = "{}" 
                       GROUP BY Song_Name_9LC'''.format(cut_off_field,cut_off, third_party_type_list[0][0])
  mycursor.execute(find_pool_rev)
  pool_rev = [i[0] for i in mycursor.fetchall()]
  all_songs = songs + pool_rev



  #Build song tables
  total_row_no = 0
  for column_no in range(1, len(column_names) + 1):
    ws.cell(row=1, column=column_no)
  #ws.insert_cols(1,len(column_names))
  for s in all_songs:
    print(s)
    s_string = escape_string(s)
    select_table_1_1 = '''SELECT Song_Name_9LC,
                        Territory_Status_DD,'''
    select_table_1_2 = ""
    for j,k in zip(year_statement_list, statement_list):
      select_table_1_2 += (''' sum( CASE WHEN Year_Statement_9LC = "{}" AND {} = "{}"
                             THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(j,mySQL_column,k,j,k))
    select_table_1_3 = '''sum( CASE WHEN {} >= "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `Total`
                        FROM Master WHERE Song_Name_9LC = "{}" AND Rights_Type_9LC = "{}" AND Territory_Status_DD = 1
                        GROUP BY Song_Name_9LC, Territory_Status_DD ORDER BY `Total` DESC'''.format(cut_off_field,cut_off,s_string, third_party_type_list[0][0])
    select_table_1 = select_table_1_1 + select_table_1_2 + select_table_1_3
    mycursor.execute(select_table_1)
    table_1 = mycursor.fetchall()

    select_table_2_1 = '''SELECT Song_Name_9LC,
                            Territory_Status_DD,'''
    select_table_2_2 = ""
    for j, k in zip(year_statement_list, statement_list):
      select_table_2_2 += (''' sum( CASE WHEN Year_Statement_9LC = "{}" AND {} = "{}"
                                 THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(j, mySQL_column, k, j, k))
    select_table_2_3 = '''sum( CASE WHEN {} >= "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `Total`
                            FROM Master WHERE Song_Name_9LC = "{}" AND Rights_Type_9LC = "{}" AND Territory_Status_DD = 0
                            GROUP BY Song_Name_9LC, Territory_Status_DD ORDER BY `Total` DESC'''.format(
      cut_off_field, cut_off, s_string, third_party_type_list[0][0])
    select_table_2 = select_table_2_1 + select_table_2_2 + select_table_2_3
    mycursor.execute(select_table_2)
    table_2 = mycursor.fetchall()

    song_table = column_names_final + table_1 + table_2

  #Add tables to sheet
    total_row_no += len(song_table)+1
    space_row_no = total_row_no + 1
    header_row_no = total_row_no - len(song_table)
    total_column_no = len(column_names)-1
    ws.append(range(len(song_table)+2))
    song_table_rows = ws[header_row_no:space_row_no]
    for (row,l) in zip(song_table_rows, song_table):
      for (cell, m) in zip(row, range(len(l))):
        cell.value = l[m]
    ws.cell(row=header_row_no, column=2).value = 'USA'
    ws.cell(row=header_row_no+1, column=2).value = 'Non-USA'

  #Add total row
    ws.cell(row=total_row_no, column=2).value = 'Total'
    for n in range(3,total_column_no):
      ws.cell(row=total_row_no, column=n).value = '=SUM({}{}:{}{})'.format(column_letters[n-1],header_row_no+1,
                                                                           column_letters[n-1],total_row_no-1)

  #Add percentage and cumulative percentage columns
    total_column_letter = column_letters[len(column_names)-3]
    percentage_column_letter = column_letters[len(column_names)-2]
    cumulative_column_letter = column_letters[len(column_names)-1]
    for o in range(header_row_no+1, total_row_no):
      ws.cell(row=o, column=total_column_no).value = '={}{}/{}{}'.format(total_column_letter,o,
                                                                           total_column_letter,total_row_no)
      ws.cell(row=o, column=total_column_no).number_format ='0.00%'
      ws.cell(row=o, column=total_column_no).font = Font(name="Calibri", size="11")
    ws.cell(row=header_row_no+1, column=total_column_no+1).value = '={}{}'.format(percentage_column_letter,
                                                                                 header_row_no+1)
    ws.cell(row=header_row_no+1, column=total_column_no+1).number_format ='0.00%'
    ws.cell(row=header_row_no+1, column=total_column_no+1).font = Font(name="Calibri", size=11)
    for p in range(header_row_no+2, total_row_no):
      ws.cell(row=p, column=total_column_no+1).value = '={}{}+{}{}'.format(cumulative_column_letter,p-1,
                                                                           percentage_column_letter,p)
      ws.cell(row=p, column=total_column_no+1).number_format ='0.00%'
      ws.cell(row=p, column=total_column_no+1).font = Font(name="Calibri", size=11)

  #Format header row
    for b in range(1, len(column_names)+1):
      ws.cell(row=header_row_no, column=b).fill = PatternFill("solid", fgColor="A6ACAF")
      ws.cell(row=header_row_no, column=b).alignment = Alignment(horizontal="center", vertical="center")
      ws.cell(row=header_row_no, column=b).font = Font(bold=True)

  #Format numbers
    for c in range(3,total_column_no):
      for d in range(header_row_no+1, total_row_no+1):
        ws.cell(row=d, column=c).style = 'Comma'
        ws.cell(row=d, column=c).font = Font(name="Calibri", size="11")

  #Format total row header
    ws.cell(row=total_row_no, column=2).fill = PatternFill("solid", fgColor="A6ACAF")
    ws.cell(row=total_row_no, column=2).font = Font(bold=True)
    ws.cell(row=total_row_no, column=2).alignment = Alignment(horizontal="center", vertical="center")

  #Format total row and column
    for e in range(header_row_no+1, total_row_no):
      ws.cell(row=e, column=total_column_no-1).font = Font(bold=True)
    for f in range(3, total_column_no):
      ws.cell(row=total_row_no, column=f).font = Font(bold=True)

  #Insert row between tables
    ws.insert_rows(total_row_no+1)
    total_row_no += 1

#Build rest of tables
  for i in third_party_type_list[1:]:
    ws_new = wb.create_sheet(title="({}) Song x Territory x Rev x Half".format(i[0]))
    ws = ws_new

  #Find smallest period
    if i[0] in quarterly_types:
      smallest_period = 'Quarter'
      mySQL_column = 'Quarter_Statement_9LC'
      statement_list = quarter_statement_list
      year_statement_list = year_statement_list_quarter
      column_names_final = quarter_column_names_final
      column_names = quarter_column_names
      cut_off_field = 'Statement_Period_Half_9LC'
      cut_off = standard_cut_off
    elif i[0] in half_types:
      smallest_period = 'Half'
      mySQL_column = 'Half_Statement_9LC'
      statement_list = half_statement_list
      year_statement_list = year_statement_list_half
      column_names_final = half_column_names_final
      column_names = half_column_names
      cut_off_field = 'Statement_Period_Half_9LC'
      cut_off = standard_cut_off
    else:
      smallest_period = 'Year'
      mySQL_column = 'Year_Statement_9LC'
      statement_list = yearly_statement_list
      year_statement_list = year_statement_list_year
      column_names_final = year_column_names_final
      column_names = year_column_names
      cut_off_field = 'Year_Statement_9LC'
      cut_off = year_cut_off

    smallest_period = 'Half'
    mySQL_column = 'Half_Statement_9LC'
    statement_list = half_statement_list
    year_statement_list = year_statement_list_half
    column_names_final = half_column_names_final
    column_names = half_column_names

  #Song list per type
    find_songs = '''SELECT Song_Name_9LC, sum( Adjusted_Royalty_SB) 
                        FROM Master WHERE Song_Name_9LC <> "Pool Revenue" AND {} >= "{}" 
                        AND Rights_Type_9LC = "{}"
                        GROUP BY Song_Name_9LC ORDER BY sum( Adjusted_Royalty_SB) DESC'''.format(cut_off_field,cut_off, i[0])
    mycursor.execute(find_songs)
    songs = [i[0] for i in mycursor.fetchall()]
    find_pool_rev = '''SELECT Song_Name_9LC, sum( Adjusted_Royalty_SB)
                           FROM Master WHERE Song_Name_9LC = "Pool Revenue" AND {} >= "{}" 
                           AND Rights_Type_9LC = "{}"
                           GROUP BY Song_Name_9LC'''.format(cut_off_field,cut_off, i[0])
    mycursor.execute(find_pool_rev)
    pool_rev = [i[0] for i in mycursor.fetchall()]
    all_songs = songs + pool_rev



  #Build song tables
    total_row_no = 0
    ws.insert_cols(1, len(column_names))
    for s in all_songs:
      print(s)
      s_string = escape_string(s)
      select_table_1_1 = '''SELECT Song_Name_9LC,
                            Territory_Status_DD,'''
      select_table_1_2 = ""
      for j, k in zip(year_statement_list, statement_list):
        select_table_1_2 += (''' sum( CASE WHEN Year_Statement_9LC = "{}" AND {} = "{}"
                                 THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(j, mySQL_column, k, j, k))
      select_table_1_3 = '''sum( CASE WHEN {} >= "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `Total`
                            FROM Master WHERE Song_Name_9LC = "{}" AND Rights_Type_9LC = "{}" AND Territory_Status_DD = 1
                            GROUP BY Song_Name_9LC, Territory_Status_DD ORDER BY `Total` DESC'''.format(cut_off_field,cut_off,
                                                                                                               s_string,
                                                                                                               i[0])
      select_table_1 = select_table_1_1 + select_table_1_2 + select_table_1_3
      mycursor.execute(select_table_1)
      table_1 = mycursor.fetchall()

      select_table_2_1 = '''SELECT Song_Name_9LC,
                                  Territory_Status_DD,'''
      select_table_2_2 = ""
      for j, k in zip(year_statement_list, statement_list):
        select_table_2_2 += (''' sum( CASE WHEN Year_Statement_9LC = "{}" AND {} = "{}"
                                       THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(j, mySQL_column, k,
                                                                                                   j, k))
      select_table_2_3 = '''sum( CASE WHEN {} >= "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `Total`
                                  FROM Master WHERE Song_Name_9LC = "{}" AND Rights_Type_9LC = "{}" AND Territory_Status_DD = 0
                                  GROUP BY Song_Name_9LC, Territory_Status_DD ORDER BY `Total` DESC'''.format(
        cut_off_field, cut_off,
        s_string,
        i[0])
      select_table_2 = select_table_2_1 + select_table_2_2 + select_table_2_3
      mycursor.execute(select_table_2)
      table_2 = mycursor.fetchall()
      song_table = column_names_final + table_1 + table_2

  #Add tables to sheet
      total_row_no += len(song_table) + 1
      space_row_no = total_row_no + 1
      header_row_no = total_row_no - len(song_table)
      total_column_no = len(column_names) - 1
      ws.append(range(len(song_table) + 2))
      song_table_rows = ws[header_row_no:space_row_no]
      for (row, l) in zip(song_table_rows, song_table):
        for (cell, m) in zip(row, range(len(l))):
          cell.value = l[m]
      ws.cell(row=header_row_no, column=2).value = 'USA'
      ws.cell(row=header_row_no+1, column=2).value = 'Non-USA'

  #Add total row
      ws.cell(row=total_row_no, column=2).value = 'Total'
      for n in range(3, total_column_no):
        ws.cell(row=total_row_no, column=n).value = '=SUM({}{}:{}{})'.format(column_letters[n - 1], header_row_no + 1,
                                                                             column_letters[n - 1], total_row_no - 1)

  #Add percentage and cumulative percentage columns
      total_column_letter = column_letters[len(column_names) - 3]
      percentage_column_letter = column_letters[len(column_names) - 2]
      cumulative_column_letter = column_letters[len(column_names) - 1]
      for o in range(header_row_no + 1, total_row_no):
        ws.cell(row=o, column=total_column_no).value = '={}{}/{}{}'.format(total_column_letter, o,
                                                                           total_column_letter, total_row_no)
        ws.cell(row=o, column=total_column_no).style = 'Percent'
        ws.cell(row=o, column=total_column_no).font = Font(name="Calibri", size="11")
      ws.cell(row=header_row_no + 1, column=total_column_no + 1).value = '={}{}'.format(percentage_column_letter,
                                                                                        header_row_no + 1)
      ws.cell(row=header_row_no + 1, column=total_column_no + 1).style = 'Percent'
      ws.cell(row=header_row_no + 1, column=total_column_no + 1).font = Font(name="Calibri", size=11)
      for p in range(header_row_no + 2, total_row_no):
        ws.cell(row=p, column=total_column_no + 1).value = '={}{}+{}{}'.format(cumulative_column_letter, p - 1,
                                                                               percentage_column_letter, p)
        ws.cell(row=p, column=total_column_no + 1).style = 'Percent'
        ws.cell(row=p, column=total_column_no + 1).font = Font(name="Calibri", size=11)

  #Format header row
      for b in range(1, len(column_names) + 1):
        ws.cell(row=header_row_no, column=b).fill = PatternFill("solid", fgColor="A6ACAF")
        ws.cell(row=header_row_no, column=b).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=header_row_no, column=b).font = Font(bold=True)

  #Format numbers
      for c in range(3, total_column_no):
        for d in range(header_row_no + 1, total_row_no + 1):
          ws.cell(row=d, column=c).style = 'Comma'
          ws.cell(row=d, column=c).font = Font(name="Calibri", size="11")

  #Format total row header
      ws.cell(row=total_row_no, column=2).fill = PatternFill("solid", fgColor="A6ACAF")
      ws.cell(row=total_row_no, column=2).font = Font(bold=True)
      ws.cell(row=total_row_no, column=2).alignment = Alignment(horizontal="center", vertical="center")

  #Format total row and column
      for e in range(header_row_no + 1, total_row_no):
        ws.cell(row=e, column=total_column_no - 1).font = Font(bold=True)
      for f in range(3, total_column_no):
        ws.cell(row=total_row_no, column=f).font = Font(bold=True)

  #Insert row between tables
      ws.insert_rows(total_row_no + 1)
      total_row_no += 1

#Create combined sheet
  monthly_third_parties = []
  quarterly_third_parties = []
  half_third_parties = []
  for a in third_party_list:
    find_statement_number_months = '''SELECT Statement_Period_9LC, Month_Statement_9LC FROM Master WHERE Third_Party_9LC = "{}" GROUP BY Statement_Period_9LC, Month_Statement_9LC'''.format(
      a)
    mycursor.execute(find_statement_number_months)
    full_list_months = mycursor.fetchall()
    statement_number_blanks = [i[0] for i in full_list_months]
    statement_number_months_blanks = [i[1] for i in full_list_months]
    statement_number = len(list(filter(check_blank, statement_number_blanks)))
    statement_number_months = len(list(filter(check_blank, statement_number_months_blanks)))
    find_months_test = '''SELECT Month_Statement_9LC FROM Master WHERE Month_Statement_9LC <> "" AND Third_Party_9LC = "{}" GROUP BY Month_Statement_9LC'''.format(
      base_year_value, a)
    mycursor.execute(find_months_test)
    months_test = [i[0] for i in mycursor.fetchall()]
    if statement_number == statement_number_months:
      monthly = True
    else:
      monthly = False

    find_statement_number_quarters = '''SELECT Statement_Period_9LC, Quarter_Statement_9LC FROM Master WHERE Third_Party_9LC = "{}" GROUP BY Statement_Period_9LC, Quarter_Statement_9LC'''.format(
      a)
    mycursor.execute(find_statement_number_quarters)
    full_list_quarters = mycursor.fetchall()
    statement_number_blanks = [i[0] for i in full_list_quarters]
    statement_number_quarters_blanks = [i[1] for i in full_list_quarters]
    statement_number = len(list(filter(check_blank, statement_number_blanks)))
    statement_number_quarters = len(list(filter(check_blank, statement_number_quarters_blanks)))
    print(statement_number)
    print(statement_number_quarters)
    find_quarters_test = '''SELECT Quarter_Statement_9LC FROM Master WHERE Quarter_Statement_9LC <> "" AND Year_Statement_9LC = "{}" AND Third_Party_9LC = "{}" GROUP BY Quarter_Statement_9LC'''.format(
      base_year_value, a)
    mycursor.execute(find_quarters_test)
    quarters_test = [i[0] for i in mycursor.fetchall()]
    print(quarters_test)
    if statement_number == statement_number_quarters:
      quarterly = True
    else:
      quarterly = False

    if monthly == True:
      monthly_third_parties.append(a)
    if monthly == False and quarterly == True:
      quarterly_third_parties.append(a)
    if monthly == False and quarterly == False:
      half_third_parties.append(a)

  monthly_data = False
  quarterly_data = False
  half_data = False

  if len(monthly_third_parties) == len(third_party_list):
    monthly_data = True
  if len(quarterly_third_parties) == len(third_party_list):
    quarterly_data = True
  if len(half_third_parties) == len(third_party_list):
    half_data = True

  print(monthly_third_parties)
  print(quarterly_third_parties)
  print(half_third_parties)

  # Smallest period criteria
  if quarterly_data:
    smallest_period = 'quarterly'
    period_name = 'Quarter'
    mySQL_column = 'Quarter_Statement_9LC'
    statement_list = quarter_statement_list
    year_statement_list = year_statement_list_quarter
    statement_period_list = statement_period_quarter_list
    cut_off_field = 'Statement_Period_Half_9LC'
    cut_off = standard_cut_off
  else:
    smallest_period = 'half'
    period_name = 'Half'
    mySQL_column = 'Half_Statement_9LC'
    statement_list = half_statement_list
    year_statement_list = year_statement_list_half
    statement_period_list = statement_period_half_list
    cut_off_field = 'Statement_Period_Half_9LC'
    cut_off = year_cut_off


  #Create list of column names
  column_names_1 = ['Song Title', 'Territory']
  for k in statement_period_list:
    column_names_1.append(k)
  column_names_1.append('Total')
  column_names_1.append('% Of Revenue')
  column_names_1.append('Cumulative %')
  column_names_final_1 = [(column_names_1)]

  #Build tables
  total_row_no = 0
  ws_1.insert_cols(1, len(column_names_1))
  for s in all_songs_combined:
    print(s)
    s_string = escape_string(s)
    select_table_1_1 = '''SELECT Song_Name_9LC,
                          Territory_Status_DD,'''
    select_table_1_2 = ""
    for j,k in zip(year_statement_list, statement_list):
      select_table_1_2 += (''' sum( CASE WHEN Year_Statement_9LC = "{}" AND {} = "{}"
                               THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(j,mySQL_column,k, j,k))
    select_table_1_3 = '''sum( CASE WHEN {} >= "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `Total`
                          FROM Master WHERE Song_Name_9LC = "{}" AND Territory_Status_DD = 1
                          GROUP BY Song_Name_9LC, Territory_Status_DD ORDER BY `Total` DESC'''.format(cut_off_field,cut_off, s_string)
    select_table_1 = select_table_1_1 + select_table_1_2 + select_table_1_3
    mycursor.execute(select_table_1)
    table_1 = mycursor.fetchall()

    select_table_2_1 = '''SELECT Song_Name_9LC,
                              Territory_Status_DD,'''
    select_table_2_2 = ""
    for j, k in zip(year_statement_list, statement_list):
      select_table_2_2 += (''' sum( CASE WHEN Year_Statement_9LC = "{}" AND {} = "{}"
                                   THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(j, mySQL_column, k, j,
                                                                                               k))
    select_table_2_3 = '''sum( CASE WHEN {} >= "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `Total`
                              FROM Master WHERE Song_Name_9LC = "{}" AND Territory_Status_DD = 0
                              GROUP BY Song_Name_9LC, Territory_Status_DD ORDER BY `Total` DESC'''.format(cut_off_field,
                                                                                                          cut_off,
                                                                                                          s_string)
    select_table_2 = select_table_2_1 + select_table_2_2 + select_table_2_3
    mycursor.execute(select_table_2)
    table_2 = mycursor.fetchall()
    song_table_1 = column_names_final_1 + table_1 + table_2

    #Add tables to sheet
    total_row_no += len(song_table_1) + 1
    space_row_no = total_row_no + 1
    header_row_no = total_row_no - len(song_table_1)
    total_column_no = len(column_names_1) - 1
    ws_1.append(range(len(song_table_1) + 2))
    song_table_rows = ws_1[header_row_no:space_row_no]
    for (row, l) in zip(song_table_rows, song_table_1):
      for (cell, m) in zip(row, range(len(l))):
        cell.value = l[m]
    ws_1.cell(row=header_row_no+1, column=2).value = 'USA'
    ws_1.cell(row=header_row_no+2, column=2).value = 'Non-USA'

    #Add total row
    ws_1.cell(row=total_row_no, column=2).value = 'Total'
    for n in range(3, total_column_no):
      ws_1.cell(row=total_row_no, column=n).value = '=SUM({}{}:{}{})'.format(column_letters[n - 1], header_row_no + 1,
                                                                           column_letters[n - 1], total_row_no - 1)

    #Add percentage and cumulative percentage columns
    total_column_letter = column_letters[len(column_names_1) - 3]
    percentage_column_letter = column_letters[len(column_names_1) - 2]
    cumulative_column_letter = column_letters[len(column_names_1) - 1]
    for o in range(header_row_no + 1, total_row_no):
      ws_1.cell(row=o, column=total_column_no).value = '={}{}/{}{}'.format(total_column_letter, o,
                                                                         total_column_letter, total_row_no)
      ws_1.cell(row=o, column=total_column_no).number_format = '0.00%'
      ws_1.cell(row=o, column=total_column_no).font = Font(name="Calibri", size="11")
    ws_1.cell(row=header_row_no + 1, column=total_column_no + 1).value = '={}{}'.format(percentage_column_letter,
                                                                                      header_row_no + 1)
    ws_1.cell(row=header_row_no + 1, column=total_column_no + 1).number_format = '0.00%'
    ws_1.cell(row=header_row_no + 1, column=total_column_no + 1).font = Font(name="Calibri", size=11)
    for p in range(header_row_no + 2, total_row_no):
      ws_1.cell(row=p, column=total_column_no + 1).value = '={}{}+{}{}'.format(cumulative_column_letter, p - 1,
                                                                             percentage_column_letter, p)
      ws_1.cell(row=p, column=total_column_no + 1).number_format = '0.00%'
      ws_1.cell(row=p, column=total_column_no + 1).font = Font(name="Calibri", size=11)

    #Format header row
    for b in range(1, len(column_names_1) + 1):
      ws_1.cell(row=header_row_no, column=b).fill = PatternFill("solid", fgColor="A6ACAF")
      ws_1.cell(row=header_row_no, column=b).alignment = Alignment(horizontal="center", vertical="center")
      ws_1.cell(row=header_row_no, column=b).font = Font(bold=True)

    #Format numbers
    for c in range(3, total_column_no):
      for d in range(header_row_no + 1, total_row_no + 1):
        ws_1.cell(row=d, column=c).style = 'Comma'
        ws_1.cell(row=d, column=c).font = Font(name="Calibri", size="11")

    #Format total row header
    ws_1.cell(row=total_row_no, column=2).fill = PatternFill("solid", fgColor="A6ACAF")
    ws_1.cell(row=total_row_no, column=2).font = Font(bold=True)
    ws_1.cell(row=total_row_no, column=2).alignment = Alignment(horizontal="center", vertical="center")

    #Format total row and column
    for e in range(header_row_no + 1, total_row_no):
      ws_1.cell(row=e, column=total_column_no - 1).font = Font(bold=True)
    for f in range(3, total_column_no):
      ws_1.cell(row=total_row_no, column=f).font = Font(bold=True)

    #Insert row between tables
    ws_1.insert_rows(total_row_no + 1)
    total_row_no += 1




  return wb.save(filename)

#songxincomexrevxhalf('Havoc_61c0ae68c94b68a289900e85')