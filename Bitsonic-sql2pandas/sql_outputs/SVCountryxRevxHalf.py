from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, NamedStyle
from pymysql.converters import escape_string

import mysql.connector

def countryxrevxhalf(database,filename):
  mydb = mysql.connector.connect(
    host="34.65.111.142",
    user="external",
    password="musicpass",
    database="{}".format(database)
  )
  mycursor = mydb.cursor(buffered=True)

#Third party list
  find_third_parties = '''SELECT Third_Party_9LC FROM Master GROUP BY Third_Party_9LC'''
  mycursor.execute(find_third_parties)
  third_party_list = [i[0] for i in mycursor.fetchall()]
  third_party_string = ', '.join('"{}"'.format(str(x)) for x in third_party_list)

#Find whether rights type filled in
  find_rights_type = '''SELECT Rights_Type_9LC FROM Master WHERE Rights_Type_9LC <> "" GROUP BY Rights_Type_9LC'''
  mycursor.execute(find_rights_type)
  rights_type_list = [i[0] for i in mycursor.fetchall()]
  if len(rights_type_list) != 0:
    mojo = True
  else:
    mojo = False

#Create sheets
  wb = Workbook()
  ws = wb.active
  #ws.title = 'M Song x Rev x Half'


#Song list


#Song string
  selected_song_string = ("(I'm Not) A Candle In The Wind","Afterthoughts","All Roads Lead To The Geart",
                          "An American Tail: The Treasure Of Manhattan Island","Anywhere","Anywhere In Your Dreams",
                          "Before","Blinded","Business As Usual","Champion","Cheater's Road","Dark In Here","Dealer",
                          "Deja Vu","Different Times","Dynamite","Enough About Me","Every Heart","Everybody's Walking",
                          "Feeling Is Believing","Finally You","Fire To Fire","Flying Tour","For Tonight","Freedom","Get In Out Of The Rain",
                          "Good Morning Goodbye","Heart Of Hearts","Here Comes The Rain","Here I Am","Here You Come Raining On Me","I Could'A Had You",
                          "I Don't Want To See The Light","I Hurt All Over","I Used To Know Her","I Want To Forget You (Just The Way You Are)",
                          "If I Ever See You Again","If I'd Only","I'm Going Out Of Your Mind","I'm Not That Lonely Yet",
                          "In The Corner Of My Heart","It Ain't All Over Over Here","It Ain't All That Bad","It Ain't Nothin'",
                          "It Goes Without Saying","It Is What It Is","It's All Coming Back To Me Now","It's Me","It's Not Too Late",
                          "I've Been Wrong Before","I've Never Lost You Before","Just A Thought","Just This Once (One More Time)",
                          "Ladies Choice","Leaving Underground #1","Let Love Come Between Us","Liar's Roses","Lily","Listen To Your Heart",
                          "Lonely Too Long","Long Way Back","Looking Forward To Looking Back","Love Don't Be A Stranger","Love Til It Hurts","Love Won't Listen",
                          "Main Attraction","Man Behind The Curtain","Maybe","Maybe We Can Fly","Montana","More Than A Memory",
                          "My Heroes Have Always Been Cowboys","My Only Claim To Fame Is You","Never Ever And Forever","Nobody Gets Hurt",
                          "Nothing But Love Songs (For You)","Oh Moon","Old Boyfriend","One","One Dream","One Step Closer","Other Side Of Goodbye",
                          "Out Of My Bones","Pacific Blue","Powerful Thing","Pretty Little Baby","Pretty Little Baby Child",
                          "Promise","Reach Out (I'll Rescue You)","Right On Time","Road Behind Me","Road To Jackson County",
                          "Rock Me Back","Ruby Shoes","Scuttlebutt Steals","She's Got A Lot On Her Heart","She's In A Better Place",
                          "She's Only A Cowboy When It Rains","Shine On","Shot For The Moon","So Close","So Far So Good","Somebody Stop Me","Soul Searchin'",
                          "Stand In The Middle Of Texas","Tear Me Out Of The Picture","That Boy Sure Can Jump A Train","That's How He Rides",
                         "The Last Time","There Ain't No Future In This","There Is No End","There's Nothing I Can Do About It","Til A Tear Becomes A Rose",
                          "Til I Said It To You","Till Your Memory's Gone","Time In Between","Too Close To Home","Trip Around The Sun","True Lies","Warm Place In The Snow",
                          "Weight Of The World","What Are You Doing Here With Me","What Do We Do With the Rest of the Night","What Do You Want From Me",
                          "What Happens To Me","When A Man Can't Keep A Woman Off His Mind","When The Moon Walks On Water","Who's Gonna Tell Her Goodbye",
                          "Y'all Come Back Saloon","You Can't Love Too Much","You Didn't Miss A Thing","You Really Better Love Me After This")
  #escaped_string = mydb.converter.escape(selected_song_string)
  #print(escaped_string)

#Find cut off
  find_period = "SELECT DISTINCT Statement_Period_Half_9LC FROM Master ORDER BY Statement_Period_Half_9LC"
  mycursor.execute(find_period)
  statement_period_half_blank = [i[0] for i in mycursor.fetchall()]
  def check_blank(period):
    if period == '':
      return False
    else:
      return True
  remove_blank = filter(check_blank,statement_period_half_blank)
  statement_period_minus_blank = list(remove_blank)
  if len(statement_period_minus_blank) > 12:
    statement_period_half = statement_period_minus_blank[-12:]
    cut_off = statement_period_minus_blank[-12]
  else:
    statement_period_half = statement_period_minus_blank
    cut_off=statement_period_minus_blank[0]

#Get list of statement half periods
  find_half_period = '''SELECT Year_Statement_9LC, Half_Statement_9LC FROM Master 
                                     WHERE Half_Statement_9LC <> "" AND Statement_Period_Half_9LC >= "{}" GROUP BY Year_Statement_9LC, Half_Statement_9LC'''.format(
    cut_off)
  mycursor.execute(find_half_period)
  complete_list = [i for i in mycursor.fetchall()]
  statement_period_half_list = [i[0] + ' ' + i[1] for i in complete_list]
  year_statement_list_half = [i[0] for i in complete_list]
  half_statement_list = [i[1] for i in complete_list]
  year_list = list(dict.fromkeys(year_statement_list_half))

  print(half_statement_list)

#Get list of statement quarter periods
  find_quarter_period = '''SELECT Year_Statement_9LC, Quarter_Statement_9LC FROM Master 
                                   WHERE Quarter_Statement_9LC <> "" AND Statement_Period_Half_9LC >= "{}" GROUP BY Year_Statement_9LC, Quarter_Statement_9LC'''.format(
    cut_off)
  mycursor.execute(find_quarter_period)
  complete_list = [i for i in mycursor.fetchall()]
  statement_period_quarter_list = [i[0] + ' ' + i[1] for i in complete_list]
  year_statement_list_quarter = [i[0] for i in complete_list]
  quarter_statement_list = [i[1] for i in complete_list]

#Get list of statement month periods
  find_month_period = '''SELECT Year_Statement_9LC, Month_Statement_9LC FROM Master 
                                         WHERE Month_Statement_9LC <> "" AND Statement_Period_Half_9LC >= "{}" GROUP BY Year_Statement_9LC, Month_Statement_9LC'''.format(
    cut_off)
  mycursor.execute(find_month_period)
  complete_list = [i for i in mycursor.fetchall()]
  statement_period_month_list = [i[0] + ' ' + i[1] for i in complete_list]
  year_statement_list_month = [i[0] for i in complete_list]
  month_statement_list = [i[1] for i in complete_list]

#Most recent year
  test_year = int(year_statement_list_half[-1]) - 1
  test_third_parties = []
  for b in third_party_list:
    find_complete_year = '''SELECT Statement_Period_Half_9LC FROM Master WHERE Year_Statement_9LC = "{}" GROUP BY Statement_Period_Half_9LC'''.format(
      test_year)
    mycursor.execute(find_complete_year)
    complete_year = [i[0] for i in mycursor.fetchall()]
    if len(complete_year) / 2 == 1:
      test_third_parties.append(b)
  if len(test_third_parties) != len(third_party_list):
    base_year_value = int(test_year) - 2
  else:
    base_year_value = test_year

#Find types of third party
  find_type = '''SELECT Rights_Type_9LC, Third_Party_9LC FROM Master GROUP BY Rights_Type_9LC, Third_Party_9LC'''
  mycursor.execute(find_type)
  type_list = list(dict.fromkeys([i[0] for i in mycursor.fetchall()]))
  test = [i[1] for i in mycursor.fetchall()]
  type_string = ', '.join('"{}"'.format(str(x)) for x in type_list)


#Find third parties for each type
  third_party_type_list = []
  for a in type_list:
    third_party_sub_list = [a]
    find_third_party_type = '''SELECT Third_Party_9LC FROM Master WHERE Rights_Type_9LC = "{}" GROUP BY Third_Party_9LC'''.format(
      a)
    mycursor.execute(find_third_party_type)
    test_list = [i[0] for i in mycursor.fetchall()]
    for b in test_list:
      third_party_sub_list.append(b)
    third_party_type_list.append(third_party_sub_list)
    third_party_sub_list = []

#Find where quarterly data exists
  quarterly_types = []
  for b in type_list:
    find_third_parties = '''SELECT Third_Party_9LC FROM Master WHERE Rights_Type_9LC = "{}"'''.format(b)
    mycursor.execute(find_third_parties)
    third_parties = [i[0] for i in mycursor.fetchall()]
    find_quarterly_third_parties = '''SELECT Third_Party_9LC FROM Master WHERE Rights_Type_9LC = "{}" AND Quarter_Statement_9LC <> ""'''.format(b)
    mycursor.execute(find_quarterly_third_parties)
    quarterly_third_parties = [i[0] for i in mycursor.fetchall()]
    if len(third_parties) == len(quarterly_third_parties):
      quarterly_types.append(b)

#Find number of songs
  find_songs = '''SELECT Normalized_Country_SB
                  FROM Master WHERE Statement_Period_Half_9LC <> "" AND Third_Party_9LC = "ASCAP" AND Song_Name_9LC IN {} GROUP BY Normalized_Country_SB'''.format(selected_song_string)
  mycursor.execute(find_songs)
  songs = [i[0] for i in mycursor.fetchall()]
  song_count = len(songs)

  #Column letters
  column_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
                    'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK',
                    'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ']

#Create styles
#Main numbers
  number_style = NamedStyle(name="number_style")
  number_style.alignment = Alignment(horizontal="right", vertical="center")
  number_style.font = Font(name="Calibri", size="11")
  number_style.number_format = '#,##0.00'

#Title cells
  title_style = NamedStyle(name="title_style")
  title_style.alignment = Alignment(horizontal="center", vertical="center")
  title_style.font = Font(name="Calibri", size="11", bold=True)
  title_style.fill = PatternFill("solid", fgColor="A6ACAF")

#Header cells
  header_style = NamedStyle(name="header_style")
  header_style.alignment = Alignment(horizontal="center", vertical="center")
  header_style.font = Font(name="Calibri", size="11", bold=True)
  header_style.fill = PatternFill("solid", fgColor="C9CDCF")

#Sub header cells
  sub_header_style = NamedStyle(name="sub_header_style")
  sub_header_style.alignment = Alignment(horizontal="left", vertical="center")
  sub_header_style.font = Font(name="Calibri", size="11", bold=True, italic=True)

#Name cells
  name_style = NamedStyle(name="name_style")
  name_style.alignment = Alignment(horizontal="left", vertical="center")
  name_style.font = Font(name="Calibri", size="11")

#Total cells
  total_style = NamedStyle(name="total_style")
  total_style.alignment = Alignment(horizontal="right", vertical="center")
  total_style.font = Font(name="Calibri", size="11", bold=True)
  total_style.number_format = '#,##0.00'

#Lined total row
  thin = Side(border_style="thin", color="000000")
  lined_total_style = NamedStyle(name="lined_total_style")
  lined_total_style.alignment = Alignment(horizontal="right", vertical="center")
  lined_total_style.font = Font(name="Calibri", size="11", bold=True)
  lined_total_style.number_format = '#,##0.00'
  lined_total_style.border = Border(top=thin, left=None, right=None, bottom=thin)

#Total label style
  total_label_style = NamedStyle(name="total_label_style")
  total_label_style.alignment = Alignment(horizontal="center", vertical="center")
  total_label_style.font = Font(name="Calibri", size="11", bold=True)
  total_label_style.fill = PatternFill("solid", fgColor="C9CDCF")
  total_label_style.border = Border(top=thin, left=None, right=None, bottom=thin)

#Publisher/PRO label style
  thick = Side(border_style="medium", color="000000")
  publisher_label_style = NamedStyle(name="publisher_label_style")
  publisher_label_style.alignment = Alignment(horizontal="center", vertical="center")
  publisher_label_style.font = Font(name="Calibri", size="11", bold=True, italic=True)
  publisher_label_style.border = Border(bottom=thick)

#Not available style
  not_available_style = NamedStyle(name="not_available_style")
  not_available_style.alignment = Alignment(horizontal="center", vertical="center")
  not_available_style.font = Font(name="Calibri", size="11")

#Add styles to workbook
  wb.add_named_style(number_style)
  wb.add_named_style(header_style)
  wb.add_named_style(title_style)
  wb.add_named_style(name_style)
  wb.add_named_style(total_style)
  wb.add_named_style(sub_header_style)
  wb.add_named_style(lined_total_style)
  wb.add_named_style(total_label_style)
  wb.add_named_style(publisher_label_style)
  wb.add_named_style(not_available_style)

#Create list of column names
  half_column_names = ['Song Title']
  for k in statement_period_half:
    half_column_names.append(k)
  half_column_names.append('Total')
  half_column_names.append('% Of Revenue')
  half_column_names.append('Cumulative %')
  half_column_names_final = [(half_column_names)]

  quarter_column_names = ['Song Title']
  for k in statement_period_quarter_list:
    quarter_column_names.append(k)
  quarter_column_names.append('Total')
  quarter_column_names.append('% Of Revenue')
  quarter_column_names.append('Cumulative %')
  quarter_column_names_final = [(quarter_column_names)]

#Sheet 1: Split sheet
#Build main pivot table
  ws.title = "M Country x Rev x Half"
  total_row_no = -2
  for s in third_party_type_list:
    if s[0] in quarterly_types:
      smallest_period = 'Quarter'
      mySQL_column = 'Quarter_Statement_9LC'
      statement_list = quarter_statement_list
      year_statement_list = year_statement_list_quarter
      column_names_final = quarter_column_names_final
      column_names = quarter_column_names
    else:
      smallest_period = 'Half'
      mySQL_column = 'Half_Statement_9LC'
      statement_list = half_statement_list
      year_statement_list = year_statement_list_half
      column_names_final = half_column_names_final
      column_names = half_column_names


    smallest_period = 'Quarter'
    mySQL_column = 'Quarter_Statement_9LC'
    statement_list = quarter_statement_list
    year_statement_list = year_statement_list_quarter
    column_names_final = quarter_column_names_final
    column_names = quarter_column_names

    print(column_names_final)

    select_table_1 = '''SELECT Normalized_Income_Type_9LC,'''
    select_table_2 = ""
    for j,k in zip(year_statement_list, statement_list):
      select_table_2 += (''' sum( CASE WHEN Year_Statement_9LC = "{}" AND {} = "{}" 
                             THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(j,mySQL_column,k,j,k))
    select_table_3 = '''sum( CASE WHEN Statement_Period_Half_9LC >= "{}"
                     THEN Adjusted_Royalty_SB ELSE "" END) AS `Total`
                     FROM Master WHERE Statement_Period_Half_9LC <> ""
                     AND Rights_Type_9LC = "{}" AND Song_Name_9LC IN {} AND Third_Party_9LC = "ASCAP"
                     GROUP BY Normalized_Income_Type_9LC ORDER BY `Total` DESC'''.format(cut_off, s[0],selected_song_string)
    select_table = select_table_1 + select_table_2 + select_table_3
    print(select_table)
    mycursor.execute(select_table)
    table = mycursor.fetchall()
    print(table)
    final_table = table
    print(final_table)


  #Size of worksheet
    for column_no in range(1, len(column_names)+1):
      for row_no in range(1, song_count+1):
        ws.cell(row=row_no, column=column_no)

  #Row numbers
    total_row_no += len(final_table) + 4
    space_row_no = total_row_no + 1
    header_row_no = total_row_no - len(final_table) - 1
    final_table_rows = ws[header_row_no+2:space_row_no+1]

  #Add column names to worksheet
    header_row = ws[header_row_no]
    for (header_cell, i) in zip(header_row, range(len(column_names))):
      header_cell.value = column_names[i]
    for cell in header_row[:len(column_names)]:
      cell.font = Font(bold=True)
      cell.fill = PatternFill("solid", fgColor="A6ACAF")
      cell.alignment = Alignment(horizontal="center", vertical="center")

  #Type label
    ws.merge_cells('A{}:{}{}'.format(header_row_no+1,column_letters[len(column_names)-1],header_row_no+1))
    if s[0] == "":
      type_label = '({}'.format(s[1])
    else:
      type_label = '{} - ({}'.format(s[0],s[1])
    if len(s) > 2:
      for c in s[2:]:
        type_label += ', {}'.format(c)
      type_label += ')'
    else:
      type_label += ')'
    ws.cell(row=header_row_no+1, column=1).value = '{}'.format(type_label)
    ws.cell(row=header_row_no+1, column=1).style = 'publisher_label_style'
    ws.cell(row=header_row_no + 1, column=1).alignment = Alignment(vertical="center", horizontal="center")
    for d in range(1, len(column_names)+1):
      ws.cell(row=header_row_no+1, column=d).border = Border(bottom=thick)

  #Add table to worksheet (song x rev x half)
    table_rows = ws[2:song_count+1]
    for (row, j) in zip(final_table_rows, final_table):
      for (cell, k) in zip(row,range(len(j))):
        cell.value = j[k]
        cell.style = 'Comma'
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.font = Font(name="Calibri", size="11")
    for k in range(header_row_no+2,total_row_no+1):
      ws.cell(row=k, column=1).alignment = Alignment(horizontal="left", vertical="center")
      ws.cell(row=k, column=(len(column_names) - 2)).font = Font(bold=True)

  #Add total row
    column_letters_2 = column_letters[1:]
    for (l,m) in zip(range(2,len(column_names)-1),column_letters_2):
      ws.cell(row=total_row_no+1, column=l).value = "=SUM({}{}:{}{})".format(m,header_row_no+1,m,total_row_no)
      ws.cell(row=total_row_no+1, column=l).style = 'Comma'
      ws.cell(row=total_row_no+1, column=l).alignment = Alignment(horizontal="right", vertical="center")
      ws.cell(row=total_row_no+1, column=l).font = Font(bold=True)
    ws.cell(row=total_row_no+1, column=1).value = 'Total'
    ws.cell(row=total_row_no+1, column=1).font = Font(bold='True')
    ws.cell(row=total_row_no+1, column=1).fill = PatternFill("solid", fgColor="A6ACAF")
    ws.cell(row=total_row_no+1, column=1).alignment = Alignment(horizontal="center", vertical="center")

  #Add % of revenue column
    total_column_letter = column_letters[len(column_names)-3]
    for n in range(header_row_no+2,total_row_no+1):
      ws.cell(row=n, column=len(column_names)-1).value = "=({}{}/{}{})".format(total_column_letter,n,
                                                                               total_column_letter,total_row_no+1)
      ws.cell(row=n, column=len(column_names)-1).number_format = '0.00%'

  #Add cumulative % column
    percent_rev_column_letter = column_letters[len(column_names)-2]
    cumulative_rev_column_letter = column_letters[len(column_names)-1]
    ws.cell(row=header_row_no+2, column=len(column_names)).value = "=({}{})".format(percent_rev_column_letter,header_row_no+2)
    ws.cell(row=header_row_no+2, column=len(column_names)).number_format = '0.00%'
    for o in range(header_row_no+3,total_row_no+1):
      ws.cell(row=o, column=len(column_names)).value = "=({}{}+{}{})".format(cumulative_rev_column_letter,o-1,
                                                                            percent_rev_column_letter,o)
      ws.cell(row=o, column=len(column_names)).number_format = '0.00%'


  #Save workbook
  return wb.save(filename)

#songxrevxhalf(61e5d39fa6da181907e64459')














