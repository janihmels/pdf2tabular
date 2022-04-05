from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
from pymysql.converters import escape_string
import mysql.connector

def combined_outputs(database, filename):
    mydb = mysql.connector.connect(
        host="34.65.111.142",
        user="external",
        password="musicpass",
        database="{}".format(database)
    )

    mycursor = mydb.cursor(buffered=True)

#Create sheets
    wb = Workbook()
    ws_1 = wb.active
    ws_1.title = "Song x Rev x Half"
    ws_2 = wb.create_sheet(title = "Income Type x Rev x Half")
    ws_3 = wb.create_sheet(title="Source x Rev x Half")
    ws_4 = wb.create_sheet(title="Artist x Rev x Half")
    ws_5 = wb.create_sheet(title="Song x Income x Rev x Half")
    ws_6 = wb.create_sheet(title="3rd Party x Song x Rev x Half")
    ws_7 = wb.create_sheet(title="3rd Party x Income x Rev x Half")
    ws_8 = wb.create_sheet(title="3rd Party x Source x Half")

#Get list of statement half periods
    find_period = 'SELECT DISTINCT Statement_Period_Half_9LC FROM Master WHERE Year_Statement_9LC <= "2021" ORDER BY Statement_Period_Half_9LC'
    mycursor.execute(find_period)
    statement_period_half_blank = [i[0] for i in mycursor.fetchall()]

    def check_blank(period):
        if period == '':
            return False
        else:
            return True

    remove_blank = filter(check_blank, statement_period_half_blank)
    statement_period_minus_blank = list(remove_blank)
    if len(statement_period_minus_blank) > 10:
        statement_period_half = statement_period_minus_blank[-10:]
        cut_off = statement_period_minus_blank[-10]
    else:
        statement_period_half = statement_period_minus_blank
        cut_off = statement_period_minus_blank[0]

#Get list of statement half periods
    find_half_period = '''SELECT Year_Statement_9LC, Half_Statement_9LC FROM Master 
                                     WHERE Half_Statement_9LC <> "" AND Statement_Period_Half_9LC >= "{}" GROUP BY Year_Statement_9LC, Half_Statement_9LC'''.format(
        cut_off)
    mycursor.execute(find_half_period)
    complete_list = [i for i in mycursor.fetchall()]
    statement_period_half_list = [i[0] + ' ' + i[1] for i in complete_list]
    year_statement_list_half = [i[0] for i in complete_list]
    half_statement_list = [i[1] for i in complete_list]
    year_list = list(dict.fromkeys(year_statement_list_half))

#Get list of statement quarter periods
    find_quarter_period = '''SELECT Year_Statement_9LC, Quarter_Statement_9LC FROM Master 
                                   WHERE Quarter_Statement_9LC <> "" AND Statement_Period_Half_9LC >= "{}" GROUP BY Year_Statement_9LC, Quarter_Statement_9LC'''.format(
        cut_off)
    mycursor.execute(find_quarter_period)
    complete_list = [i for i in mycursor.fetchall()]
    statement_period_quarter_list = [i[0] + ' ' + i[1] for i in complete_list]
    year_statement_list_quarter = [i[0] for i in complete_list]
    quarter_statement_list = [i[1] for i in complete_list]

#Column letters
    column_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S',
                      'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

#Find number of songs
    find_songs = '''SELECT Song_Name_9LC
                      FROM Master GROUP BY Song_Name_9LC'''
    mycursor.execute(find_songs)
    songs = [i[0] for i in mycursor.fetchall()]
    song_count = len(songs)

#Find ordered song list
    find_songs_2 = '''SELECT Song_Name_9LC, sum( Adjusted_Royalty_SB) 
                      FROM Master WHERE Song_Name_9LC <> "Pool Revenue" AND Statement_Period_Half_9LC >= "{}"
                      GROUP BY Song_Name_9LC ORDER BY sum( Adjusted_Royalty_SB) DESC'''.format(cut_off)
    mycursor.execute(find_songs_2)
    songs_2 = [i[0] for i in mycursor.fetchall()]
    find_pool_rev = '''SELECT Song_Name_9LC, sum( Adjusted_Royalty_SB)
                         FROM Master WHERE Song_Name_9LC = "Pool Revenue" AND Statement_Period_Half_9LC >= "{}"
                         GROUP BY Song_Name_9LC'''.format(cut_off)
    mycursor.execute(find_pool_rev)
    pool_rev = [i[0] for i in mycursor.fetchall()]
    all_songs = songs_2 + pool_rev

#Find number of income types
    find_income_type = '''SELECT Normalized_Income_Type_9LC
                            FROM Master GROUP BY Normalized_Income_Type_9LC'''
    mycursor.execute(find_income_type)
    income_type = [i[0] for i in mycursor.fetchall()]
    income_type_count = len(income_type)

#Find number of sources
    find_source = '''SELECT Normalized_Source_9LC
                        FROM Master GROUP BY Normalized_Source_9LC'''
    mycursor.execute(find_source)
    source_type = [i[0] for i in mycursor.fetchall()]
    source_type_count = len(source_type)

#Find number of artists
    find_songs = '''SELECT Release_Artist_9LC
                      FROM Master GROUP BY Release_Artist_9LC'''
    mycursor.execute(find_songs)
    artists = [i[0] for i in mycursor.fetchall()]
    artist_count = len(artists)

#Find number of third parties
    find_third_party = '''SELECT Third_Party_9LC
                            FROM Master
                            GROUP BY Third_Party_9LC'''
    mycursor.execute(find_third_party)
    third_party_list = [i[0] for i in mycursor.fetchall()]

#Worksheet 1: Song x Rev x Half
    #Build main table
    select_table_1_1 = '''SELECT Song_Name_9LC,'''
    select_table_1_2 = ""
    for j in statement_period_half:
        select_table_1_2 += (''' sum( CASE WHEN Statement_Period_Half_9LC = "{}" AND Song_Name_9LC <> "Pool Revenue" 
                               THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(j,j))
    select_table_1_3 = '''sum( CASE WHEN Song_Name_9LC <> "Pool Revenue" AND Statement_Period_Half_9LC >= "{}"
                       THEN Adjusted_Royalty_SB ELSE "" END) AS `Total`
                       FROM Master WHERE Song_Name_9LC <> "Pool Revenue" AND Statement_Period_Half_9LC <> "" 
                       GROUP BY Song_Name_9LC ORDER BY `Total` DESC'''.format(cut_off)
    select_table_1 = select_table_1_1 + select_table_1_2 + select_table_1_3
    mycursor.execute(select_table_1)
    table_1 = mycursor.fetchall()

    #Find pool revenue
    pool_rev_1_1 = '''SELECT Song_Name_9LC,'''
    pool_rev_1_2 = ""
    for l in statement_period_half:
        pool_rev_1_2 += ('''sum( CASE WHEN Statement_Period_Half_9LC = "{}" AND Song_Name_9LC = "Pool Revenue"
                          THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(l, l))
    pool_rev_1_3 = '''sum( CASE WHEN Song_Name_9LC = "Pool Revenue" AND Statement_Period_Half_9LC >= "{}"
                    THEN Adjusted_Royalty_SB ELSE "" END) AS `Total` 
                    FROM Master WHERE Song_Name_9LC = "Pool Revenue" AND Statement_Period_Half_9LC <> ""
                    GROUP BY Song_Name_9LC ORDER BY `Total` DESC'''.format(cut_off)
    pool_rev_1 = pool_rev_1_1 + pool_rev_1_2 + pool_rev_1_3
    mycursor.execute(pool_rev_1)
    pool_revenue_1 = mycursor.fetchall()
    final_table_1 = table_1 + pool_revenue_1

    #Create list of column names
    column_names_1 = ['Song Title']
    for k in statement_period_half:
        column_names_1.append(k)
    column_names_1.append('Total')
    column_names_1.append('% Of Revenue')
    column_names_1.append('Cumulative %')

    #Size of worksheet
    for column_no in range(1, len(column_names_1)+1):
        for row_no in range(1, song_count+1):
            ws_1.cell(row=row_no, column=column_no)

    #Add column names to worksheet (song x rev x half)
    header_row = ws_1[1]
    for (header_cell, i) in zip(header_row, range(len(column_names_1))):
        header_cell.value = column_names_1[i]
    for cell in header_row:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="A6ACAF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    #Add table to worksheet (song x rev x half)
    table_rows = ws_1[2:song_count+1]
    for (row, j) in zip(table_rows, final_table_1):
        for (cell, k) in zip(row,range(len(j))):
            cell.value = j[k]
            cell.style = 'Comma'
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.font = Font(name="Calibri", size="11")
    for song_no in range(2,song_count+2):
        ws_1.cell(row=song_no, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws_1.cell(row=song_no, column=(len(column_names_1) - 2)).font = Font(bold=True)

    #Add total row
    column_letters_2 = column_letters[1:]
    for (l,m) in zip(range(2,len(column_names_1)-1),column_letters_2):
        ws_1.cell(row=song_count+2, column=l).value = "=SUM({}{}:{}{})".format(m,2,m,song_count+1)
        ws_1.cell(row=song_count+2, column=l).style = 'Comma'
        ws_1.cell(row=song_count+2, column=l).alignment = Alignment(horizontal="right", vertical="center")
        ws_1.cell(row=song_count+2, column=l).font = Font(bold=True)
        ws_1.cell(row=song_count+2, column=1).value = 'Total'
        ws_1.cell(row=song_count+2, column=1).font = Font(bold='True')
        ws_1.cell(row=song_count+2, column=1).fill = PatternFill("solid", fgColor="A6ACAF")
        ws_1.cell(row=song_count+2, column=1).alignment = Alignment(horizontal="center", vertical="center")

    #Add % of revenue column
    total_column_letter = column_letters[len(column_names_1)-3]
    for n in range(2,song_count+2):
        ws_1.cell(row=n, column=len(column_names_1)-1).value = "=({}{}/{}{})".format(total_column_letter,n,
                                                                                     total_column_letter,song_count+2)
        ws_1.cell(row=n, column=len(column_names_1)-1).style = 'Percent'

    #Add cumulative % column
    percent_rev_column_letter = column_letters[len(column_names_1)-2]
    cumulative_rev_column_letter = column_letters[len(column_names_1)-1]
    ws_1.cell(row=2, column=len(column_names_1)).value = "=({}2)".format(percent_rev_column_letter)
    ws_1.cell(row=2, column=len(column_names_1)).style = 'Percent'
    for o in range(3,song_count+2):
        ws_1.cell(row=o, column=len(column_names_1)).value = "=({}{}+{}{})".format(cumulative_rev_column_letter,o-1,
                                                                              percent_rev_column_letter,o)
        ws_1.cell(row=o, column=len(column_names_1)).style = 'Percent'

#Worksheet 2: Income Type x Rev x Half
    #Build main table
    select_table_2_1 = '''SELECT Normalized_Income_Type_9LC,'''
    select_table_2_2 = ""
    for j in statement_period_half:
        select_table_2_2 += (''' sum( CASE WHEN Statement_Period_Half_9LC = "{}" 
                               THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(j, j))
    select_table_2_3 = '''sum( CASE WHEN Statement_Period_Half_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END) 
                       AS `Total`
                       FROM Master WHERE Statement_Period_Half_9LC <> "" 
                       GROUP BY Normalized_Income_Type_9LC ORDER BY `Total` DESC'''.format(cut_off)
    select_table_2 = select_table_2_1 + select_table_2_2 + select_table_2_3
    mycursor.execute(select_table_2)
    table_2 = mycursor.fetchall()

    #Create list of column names
    column_names_2 = ['Normalized Income Type']
    for k in statement_period_half:
        column_names_2.append(k)
    column_names_2.append('Total')
    column_names_2.append('% Of Revenue')
    column_names_2.append('Cumulative %')

    #Size of worksheet
    for column_no in range(1, len(column_names_2) + 1):
        for row_no in range(1, income_type_count + 1):
            ws_2.cell(row=row_no, column=column_no)

    #Add column names to worksheet
    header_row = ws_2[1]
    for (header_cell, i) in zip(header_row, range(len(column_names_2))):
        header_cell.value = column_names_2[i]
    for cell in header_row:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="A6ACAF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    #Add table to worksheet
    table_rows = ws_2[2:income_type_count + 1]
    for j,k in zip(range(2, income_type_count+2), table_2):
        for l,m in zip(range(1, len(column_names_2)), range(len(k))):
            ws_2.cell(row=j, column=l).value = k[m]
            ws_2.cell(row=j, column=l).style = 'Comma'
            ws_2.cell(row=j, column=l).alignment = Alignment(horizontal="right", vertical="center")
            ws_2.cell(row=j, column=l).font = Font(name="Calibri", size="11")
    for income_no in range(2, income_type_count + 2):
        ws_2.cell(row=income_no, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws_2.cell(row=income_no, column=(len(column_names_2) - 2)).font = Font(bold=True)

    #Add total row
    column_letters_2 = column_letters[1:]
    for (l, m) in zip(range(2, len(column_names_2) - 1), column_letters_2):
        ws_2.cell(row=income_type_count + 2, column=l).value = "=SUM({}{}:{}{})".format(m, 2, m, income_type_count + 1)
        ws_2.cell(row=income_type_count + 2, column=l).style = 'Comma'
        ws_2.cell(row=income_type_count + 2, column=l).alignment = Alignment(horizontal="right", vertical="center")
        ws_2.cell(row=income_type_count + 2, column=l).font = Font(bold=True)
    ws_2.cell(row=income_type_count + 2, column=1).value = 'Total'
    ws_2.cell(row=income_type_count + 2, column=1).font = Font(bold='True')
    ws_2.cell(row=income_type_count + 2, column=1).fill = PatternFill("solid", fgColor="A6ACAF")
    ws_2.cell(row=income_type_count + 2, column=1).alignment = Alignment(horizontal="center", vertical="center")

    #Add % of revenue column
    total_column_letter = column_letters[len(column_names_2) - 3]
    for n in range(2, income_type_count + 2):
        ws_2.cell(row=n, column=len(column_names_2) - 1).value = "=({}{}/{}{})".format(total_column_letter, n,
                                                                                   total_column_letter,
                                                                                   income_type_count + 2)
        ws_2.cell(row=n, column=len(column_names_2) - 1).style = 'Percent'

    #Add cumulative % column
    percent_rev_column_letter = column_letters[len(column_names_2) - 2]
    cumulative_rev_column_letter = column_letters[len(column_names_2) - 1]
    ws_2.cell(row=2, column=len(column_names_2)).value = "=({}2)".format(percent_rev_column_letter)
    ws_2.cell(row=2, column=len(column_names_2)).style = 'Percent'
    for o in range(3, income_type_count + 2):
        ws_2.cell(row=o, column=len(column_names_2)).value = "=({}{}+{}{})".format(cumulative_rev_column_letter, o - 1,
                                                                               percent_rev_column_letter, o)
        ws_2.cell(row=o, column=len(column_names_2)).style = 'Percent'

#Worksheet 3: Source x Rev x Half
    #Build main table
    select_table_3_1 = '''SELECT Normalized_Source_9LC,'''
    select_table_3_2 = ""
    for j in statement_period_half:
        select_table_3_2 += (''' sum( CASE WHEN Statement_Period_Half_9LC = "{}" AND Normalized_Source_9LC <> "Pool Revenue"
                               THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(j, j))
    select_table_3_3 = '''sum( CASE WHEN Normalized_Source_9LC <> "Pool Revenue" AND Statement_Period_Half_9LC >= "{}" 
                       THEN Adjusted_Royalty_SB ELSE "" END) 
                       AS `Total`
                       FROM Master WHERE Normalized_Source_9LC <> "Pool Revenue" AND Statement_Period_Half_9LC <> "" 
                       GROUP BY Normalized_Source_9LC ORDER BY `Total` DESC'''.format(cut_off)
    select_table_3 = select_table_3_1 + select_table_3_2 + select_table_3_3
    mycursor.execute(select_table_3)
    table_3 = mycursor.fetchall()

    #Find pool revenue
    pool_rev_3_1 = '''SELECT Normalized_Source_9LC,'''
    pool_rev_3_2 = ""
    for l in statement_period_half:
        pool_rev_3_2 += ('''sum( CASE WHEN Statement_Period_Half_9LC = "{}" AND Normalized_Source_9LC = "Pool Revenue"
                            THEN Adjusted_Royalty_SB ELSE NULL END) AS `{}`,'''.format(l, l))
    pool_rev_3_3 = '''sum( CASE WHEN Normalized_Source_9LC = "Pool Revenue" AND Statement_Period_Half_9LC >= "{}" 
                      THEN Adjusted_Royalty_SB ELSE NULL END) AS `Total` 
                      FROM Master WHERE Normalized_Source_9LC = "Pool Revenue" AND Statement_Period_Half_9LC <> "" 
                      GROUP BY Normalized_Source_9LC ORDER BY `Total` DESC'''.format(cut_off)
    pool_rev_3 = pool_rev_3_1 + pool_rev_3_2 + pool_rev_3_3
    mycursor.execute(pool_rev_3)
    pool_revenue_3 = mycursor.fetchall()
    final_table_3 = table_3 + pool_revenue_3

    #Create list of column names
    column_names_3 = ['Normalized Source']
    for k in statement_period_half:
        column_names_3.append(k)
    column_names_3.append('Total')
    column_names_3.append('% Of Revenue')
    column_names_3.append('Cumulative %')

    #Size of worksheet
    for column_no in range(1, len(column_names_3) + 1):
        for row_no in range(1, source_type_count + 1):
            ws_3.cell(row=row_no, column=column_no)

    #Add column names to worksheet
    header_row = ws_3[1]
    for (header_cell, i) in zip(header_row, range(len(column_names_3))):
        header_cell.value = column_names_3[i]
    for cell in header_row:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="A6ACAF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    #Add table to worksheet
    table_rows = ws_3[2:source_type_count + 1]
    for (row, j) in zip(table_rows, final_table_3):
        for (cell, k) in zip(row, range(len(j))):
            cell.value = j[k]
            cell.style = 'Comma'
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.font = Font(name="Calibri", size="11")
    for source_no in range(2, source_type_count + 2):
        ws_3.cell(row=source_no, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws_3.cell(row=source_no, column=(len(column_names_3) - 2)).font = Font(bold=True)

    #Add total row
    column_letters_2 = column_letters[1:]
    for (l, m) in zip(range(2, len(column_names_3) - 1), column_letters_2):
        ws_3.cell(row=source_type_count + 2, column=l).value = "=SUM({}{}:{}{})".format(m, 2, m, source_type_count + 1)
        ws_3.cell(row=source_type_count + 2, column=l).style = 'Comma'
        ws_3.cell(row=source_type_count + 2, column=l).alignment = Alignment(horizontal="right", vertical="center")
        ws_3.cell(row=source_type_count + 2, column=l).font = Font(bold=True)
    ws_3.cell(row=source_type_count + 2, column=1).value = 'Total'
    ws_3.cell(row=source_type_count + 2, column=1).font = Font(bold='True')
    ws_3.cell(row=source_type_count + 2, column=1).fill = PatternFill("solid", fgColor="A6ACAF")
    ws_3.cell(row=source_type_count + 2, column=1).alignment = Alignment(horizontal="center", vertical="center")

    #Add % of revenue column
    total_column_letter = column_letters[len(column_names_3) - 3]
    for n in range(2, source_type_count + 2):
        ws_3.cell(row=n, column=len(column_names_3) - 1).value = "=({}{}/{}{})".format(total_column_letter, n,
                                                                                   total_column_letter,
                                                                                   source_type_count + 2)
        ws_3.cell(row=n, column=len(column_names_3) - 1).style = 'Percent'

    #Add cumulative % column
    percent_rev_column_letter = column_letters[len(column_names_3) - 2]
    cumulative_rev_column_letter = column_letters[len(column_names_3) - 1]
    ws_3.cell(row=2, column=len(column_names_3)).value = "=({}2)".format(percent_rev_column_letter)
    ws_3.cell(row=2, column=len(column_names_3)).style = 'Percent'
    for o in range(3, source_type_count + 2):
        ws_3.cell(row=o, column=len(column_names_3)).value = "=({}{}+{}{})".format(cumulative_rev_column_letter, o - 1,
                                                                               percent_rev_column_letter, o)
        ws_3.cell(row=o, column=len(column_names_3)).style = 'Percent'

#Worksheet 4: Artist x Rev x Half
    #Build main table
    select_table_4_1 = '''SELECT Release_Artist_9LC,'''
    select_table_4_2 = ""
    for j in statement_period_half:
        select_table_4_2 += (''' sum( CASE WHEN Statement_Period_Half_9LC = "{}" AND Release_Artist_9LC <> "" 
                               THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(j, j))
    select_table_4_3 = '''sum( CASE WHEN Statement_Period_Half_9LC >= "{}" AND Release_Artist_9LC <> ""
                       THEN Adjusted_Royalty_SB ELSE "" END) AS `Total`
                       FROM Master WHERE Statement_Period_Half_9LC <> "" AND Release_Artist_9LC <> ""
                       GROUP BY Release_Artist_9LC ORDER BY `Total` DESC'''.format(cut_off)
    select_table_4 = select_table_4_1 + select_table_4_2 + select_table_4_3
    mycursor.execute(select_table_4)
    table_4 = mycursor.fetchall()

    #Add line for unknown artists
    unknown_artists_4_1 = '''SELECT Release_Artist_9LC,'''
    unknown_artists_4_2 = ""
    for l in statement_period_half:
        unknown_artists_4_2 += ('''sum( CASE WHEN Statement_Period_Half_9LC = "{}" AND Release_Artist_9LC = ""
                                 THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(l, l))
    unknown_artists_4_3 = '''sum( CASE WHEN Release_Artist_9LC = "" AND Statement_Period_Half_9LC >= "{}"
                             THEN Adjusted_Royalty_SB ELSE "" END) AS `Total`
                             FROM Master WHERE Statement_Period_Half_9LC <> "" AND Release_Artist_9LC = ""
                             GROUP BY Release_Artist_9LC ORDER BY `Total` DESC'''.format(cut_off)
    unknown_artists_4 = unknown_artists_4_1 + unknown_artists_4_2 + unknown_artists_4_3
    mycursor.execute(unknown_artists_4)
    unknown_artist_line_4 = mycursor.fetchall()
    final_table_4 = table_4 + unknown_artist_line_4

    #Create list of column names
    column_names_4 = ['Release Artist']
    for k in statement_period_half:
        column_names_4.append(k)
    column_names_4.append('Total')
    column_names_4.append('% Of Revenue')
    column_names_4.append('Cumulative %')

    #Size of worksheet
    for column_no in range(1, len(column_names_4) + 1):
        for row_no in range(1, artist_count + 1):
            ws_4.cell(row=row_no, column=column_no)

    #Add column names to worksheet
    header_row = ws_4[1]
    for (header_cell, i) in zip(header_row, range(len(column_names_4))):
        header_cell.value = column_names_4[i]
    for cell in header_row:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="A6ACAF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    #Add table to worksheet
    table_rows = ws_4[2:artist_count + 1]
    for (row, j) in zip(table_rows, final_table_4):
        for (cell, k) in zip(row, range(len(j))):
            cell.value = j[k]
            cell.style = 'Comma'
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.font = Font(name="Calibri", size="11")
    for song_no in range(2, artist_count + 2):
        ws_4.cell(row=song_no, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws_4.cell(row=song_no, column=(len(column_names_4) - 2)).font = Font(bold=True)
    ws_4.cell(row=artist_count + 1, column=1).value = 'Unknown Artists'

    #Add total row
    column_letters_2 = column_letters[1:]
    for (l, m) in zip(range(2, len(column_names_4) - 1), column_letters_2):
        ws_4.cell(row=artist_count + 2, column=l).value = "=SUM({}{}:{}{})".format(m, 2, m, artist_count + 1)
        ws_4.cell(row=artist_count + 2, column=l).style = 'Comma'
        ws_4.cell(row=artist_count + 2, column=l).alignment = Alignment(horizontal="right", vertical="center")
        ws_4.cell(row=artist_count + 2, column=l).font = Font(bold=True)
    ws_4.cell(row=artist_count + 2, column=1).value = 'Total'
    ws_4.cell(row=artist_count + 2, column=1).font = Font(bold='True')
    ws_4.cell(row=artist_count + 2, column=1).fill = PatternFill("solid", fgColor="A6ACAF")
    ws_4.cell(row=artist_count + 2, column=1).alignment = Alignment(horizontal="center", vertical="center")

    #Add % of revenue column
    total_column_letter = column_letters[len(column_names_4) - 3]
    for n in range(2, artist_count + 2):
        ws_4.cell(row=n, column=len(column_names_4) - 1).value = "=({}{}/{}{})".format(total_column_letter, n,
                                                                                   total_column_letter,
                                                                                   artist_count + 2)
        ws_4.cell(row=n, column=len(column_names_4) - 1).style = 'Percent'

    #Add cumulative % column
    percent_rev_column_letter = column_letters[len(column_names_4) - 2]
    cumulative_rev_column_letter = column_letters[len(column_names_4) - 1]
    ws_4.cell(row=2, column=len(column_names_4)).value = "=({}2)".format(percent_rev_column_letter)
    ws_4.cell(row=2, column=len(column_names_4)).style = 'Percent'
    for o in range(3, artist_count + 2):
        ws_4.cell(row=o, column=len(column_names_4)).value = "=({}{}+{}{})".format(cumulative_rev_column_letter, o - 1,
                                                                               percent_rev_column_letter, o)
        ws_4.cell(row=o, column=len(column_names_4)).style = 'Percent'

#Worksheet 5: Song x Income Type x Rev x Half
    #Create list of column names
    column_names_5 = ['Song Title', 'Income Type']
    for k in statement_period_half:
        column_names_5.append(k)
    column_names_5.append('Total')
    column_names_5.append('% Of Revenue')
    column_names_5.append('Cumulative %')
    column_names_final_5 = [(column_names_5)]

    #Build tables
    total_row_no = 0
    ws_5.insert_cols(1, len(column_names_5))
    for s in all_songs:
        s_string = escape_string(s)
        select_table_5_1 = '''SELECT Song_Name_9LC,
                            Normalized_Income_Type_9LC,'''
        select_table_5_2 = ""
        for j in statement_period_half:
            select_table_5_2 += (''' sum( CASE WHEN Statement_Period_Half_9LC = "{}"
                                 THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(j, j))
        select_table_5_3 = '''sum( CASE WHEN Statement_Period_Half_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `Total`
                            FROM Master WHERE Song_Name_9LC = "{}" AND Statement_Period_Half_9lC <> "" 
                            GROUP BY Song_Name_9LC, Normalized_Income_Type_9LC ORDER BY `Total` DESC'''.format(cut_off,
                                                                                                               s_string)
        select_table_5 = select_table_5_1 + select_table_5_2 + select_table_5_3
        mycursor.execute(select_table_5)
        table_5 = mycursor.fetchall()
        song_table_5 = column_names_final_5 + table_5

    #Add tables to sheet
        total_row_no += len(song_table_5) + 1
        space_row_no = total_row_no + 1
        header_row_no = total_row_no - len(song_table_5)
        total_column_no = len(column_names_5) - 1
        ws_5.append(range(len(song_table_5) + 2))
        song_table_rows = ws_5[header_row_no:space_row_no]
        for (row, l) in zip(song_table_rows, song_table_5):
            for (cell, m) in zip(row, range(len(l))):
                cell.value = l[m]

    #Add total row
        ws_5.cell(row=total_row_no, column=2).value = 'Total'
        for n in range(3, total_column_no):
            ws_5.cell(row=total_row_no, column=n).value = '=SUM({}{}:{}{})'.format(column_letters[n - 1], header_row_no + 1,
                                                                                 column_letters[n - 1], total_row_no - 1)

    #Add % of revenue and cumulative % columns
        total_column_letter = column_letters[len(column_names_5) - 3]
        percentage_column_letter = column_letters[len(column_names_5) - 2]
        cumulative_column_letter = column_letters[len(column_names_5) - 1]
        for o in range(header_row_no + 1, total_row_no):
            ws_5.cell(row=o, column=total_column_no).value = '={}{}/{}{}'.format(total_column_letter, o,
                                                                               total_column_letter, total_row_no)
            ws_5.cell(row=o, column=total_column_no).style = 'Percent'
            ws_5.cell(row=o, column=total_column_no).font = Font(name="Calibri", size="11")
        ws_5.cell(row=header_row_no + 1, column=total_column_no + 1).value = '={}{}'.format(percentage_column_letter,
                                                                                          header_row_no + 1)
        ws_5.cell(row=header_row_no + 1, column=total_column_no + 1).style = 'Percent'
        ws_5.cell(row=header_row_no + 1, column=total_column_no + 1).font = Font(name="Calibri", size=11)
        for p in range(header_row_no + 2, total_row_no):
            ws_5.cell(row=p, column=total_column_no + 1).value = '={}{}+{}{}'.format(cumulative_column_letter, p - 1,
                                                                                   percentage_column_letter, p)
            ws_5.cell(row=p, column=total_column_no + 1).style = 'Percent'
            ws_5.cell(row=p, column=total_column_no + 1).font = Font(name="Calibri", size=11)

    #Format header row
        for b in range(1, len(column_names_5)+1):
          ws_5.cell(row=header_row_no, column=b).fill = PatternFill("solid", fgColor="A6ACAF")
          ws_5.cell(row=header_row_no, column=b).alignment = Alignment(horizontal="center", vertical="center")
          ws_5.cell(row=header_row_no, column=b).font = Font(bold=True)

    #Format numbers
        for c in range(3,total_column_no):
          for d in range(header_row_no+1, total_row_no+1):
            ws_5.cell(row=d, column=c).style = 'Comma'
            ws_5.cell(row=d, column=c).font = Font(name="Calibri", size="11")

    #Format total row header
        ws_5.cell(row=total_row_no, column=2).fill = PatternFill("solid", fgColor="A6ACAF")
        ws_5.cell(row=total_row_no, column=2).font = Font(bold=True)
        ws_5.cell(row=total_row_no, column=2).alignment = Alignment(horizontal="center", vertical="center")

    #Format total row and column
        for e in range(header_row_no+1, total_row_no):
          ws_5.cell(row=e, column=total_column_no-1).font = Font(bold=True)
        for f in range(3, total_column_no):
          ws_5.cell(row=total_row_no, column=f).font = Font(bold=True)

    #Insert row between tables
        ws_5.insert_rows(total_row_no + 1)
        total_row_no += 1

#Worksheet 6: Third Party x Song x Rev x Half
    #Create list of column names
    half_column_names_6 = ['Payor', 'Song Title']
    for k in statement_period_half:
        half_column_names_6.append(k)
    half_column_names_6.append('Total')
    half_column_names_6.append('% Of Revenue')
    half_column_names_6.append('Cumulative %')
    half_column_names_final_6 = [(half_column_names_6)]

    quarter_column_names_6 = ['Payor', 'Song Title']
    for k in statement_period_quarter_list:
        quarter_column_names_6.append(k)
    quarter_column_names_6.append('Total')
    quarter_column_names_6.append('% Of Revenue')
    quarter_column_names_6.append('Cumulative %')
    quarter_column_names_final_6 = [(quarter_column_names_6)]

    #Build tables
    #Main table
    total_row_no = 0
    for y in range(1, len(quarter_column_names_6)):
        ws_6.cell(row=1, column=y)
    for s in third_party_list:
    #Find smallest period
        find_quarterly_period = '''SELECT Third_Party_9LC FROM Master WHERE Quarter_Statement_9LC <> "" AND Third_Party_9LC = "{}" GROUP BY Third_Party_9LC'''.format(
            s)
        mycursor.execute(find_quarterly_period)
        quarterly = [i[0] for i in mycursor.fetchall()]
        if len(quarterly) == 1:
            smallest_period = 'Quarter'
            mySQL_column = 'Quarter_Statement_9LC'
            statement_list = quarter_statement_list
            year_statement_list = year_statement_list_quarter
            column_names_final_6 = quarter_column_names_final_6
            column_names_6 = quarter_column_names_6
        else:
            smallest_period = 'Half'
            mySQL_column = 'Half_Statement_9LC'
            statement_list = half_statement_list
            year_statement_list = year_statement_list_half
            column_names_final_6 = half_column_names_final_6
            column_names_6 = half_column_names_6

        select_table_6_1 = '''SELECT Third_Party_9LC,
                            Song_Name_9LC,'''
        select_table_6_2 = ""
        for j, k in zip(year_statement_list, statement_list):
            select_table_6_2 += (''' sum( CASE WHEN Year_Statement_9LC = "{}" AND {} = "{}" AND Song_Name_9LC <> "Pool Revenue"
                                 THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(j,mySQL_column,k, j,k))
        select_table_6_3 = '''sum( CASE WHEN Song_Name_9LC <> "Pool Revenue" AND 
                            Statement_Period_Half_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `Total`
                            FROM Master WHERE Third_Party_9LC = "{}" AND Song_Name_9LC <> "Pool Revenue"
                            AND Statement_Period_Half_9lC <> "" 
                            GROUP BY Third_Party_9LC, Song_Name_9LC ORDER BY `Total` DESC'''.format(cut_off, s)
        select_table_6 = select_table_6_1 + select_table_6_2 + select_table_6_3
        mycursor.execute(select_table_6)
        table_6 = mycursor.fetchall()

    #Pool revenue
        pool_rev_6_1 = '''SELECT Third_Party_9LC,
                            Song_Name_9LC,'''
        pool_rev_6_2 = ""
        for l, m in zip(year_statement_list, statement_list):
            pool_rev_6_2 += ('''sum( CASE WHEN Year_Statement_9LC = "{}" AND {} = "{}" AND Song_Name_9LC = "Pool Revenue"
                                  THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(l,mySQL_column,m, l,m))
        pool_rev_6_3 = '''sum( CASE WHEN Song_Name_9LC = "Pool Revenue" AND Statement_Period_Half_9LC >= "{}"
                            THEN Adjusted_Royalty_SB ELSE "" END) AS `Total` 
                            FROM Master WHERE Third_Party_9LC = "{}" AND Song_Name_9LC = "Pool Revenue" 
                            AND Statement_Period_Half_9LC <> "" 
                            GROUP BY Third_Party_9LC, Song_Name_9LC ORDER BY `Total` DESC'''.format(cut_off, s)
        pool_rev_6 = pool_rev_6_1 + pool_rev_6_2 + pool_rev_6_3
        mycursor.execute(pool_rev_6)
        pool_revenue_6 = mycursor.fetchall()
        third_party_table_6 = column_names_final_6 + table_6 + pool_revenue_6

    #Add tables to sheet
        total_row_no += len(third_party_table_6) + 1
        space_row_no = total_row_no + 1
        header_row_no = total_row_no - len(third_party_table_6)
        total_column_no = len(column_names_6) - 1
        third_party_table_rows = ws_6[header_row_no:space_row_no]
        for (row, l) in zip(third_party_table_rows, third_party_table_6):
            for (cell, m) in zip(row, range(len(l))):
                cell.value = l[m]

    #Add total row
        ws_6.cell(row=total_row_no, column=2).value = 'Total'
        for n in range(3, total_column_no):
            ws_6.cell(row=total_row_no, column=n).value = '=SUM({}{}:{}{})'.format(column_letters[n - 1], header_row_no + 1,
                                                                                 column_letters[n - 1], total_row_no - 1)

    #Add % of revenue and cumulative % columns
        total_column_letter = column_letters[len(column_names_6) - 3]
        percentage_column_letter = column_letters[len(column_names_6) - 2]
        cumulative_column_letter = column_letters[len(column_names_6) - 1]
        for o in range(header_row_no + 1, total_row_no):
            ws_6.cell(row=o, column=total_column_no).value = '={}{}/{}{}'.format(total_column_letter, o,
                                                                               total_column_letter, total_row_no)
            ws_6.cell(row=o, column=total_column_no).style = 'Percent'
            ws_6.cell(row=o, column=total_column_no).font = Font(name="Calibri", size="11")
        ws_6.cell(row=header_row_no + 1, column=total_column_no + 1).value = '={}{}'.format(percentage_column_letter,
                                                                                          header_row_no + 1)
        ws_6.cell(row=header_row_no + 1, column=total_column_no + 1).style = 'Percent'
        ws_6.cell(row=header_row_no + 1, column=total_column_no + 1).font = Font(name="Calibri", size=11)
        for p in range(header_row_no + 2, total_row_no):
            ws_6.cell(row=p, column=total_column_no + 1).value = '={}{}+{}{}'.format(cumulative_column_letter, p - 1,
                                                                                   percentage_column_letter, p)
            ws_6.cell(row=p, column=total_column_no + 1).style = 'Percent'
            ws_6.cell(row=p, column=total_column_no + 1).font = Font(name="Calibri", size=11)
        ws_6.cell(row=header_row_no, column=total_column_no + 1).value = 'Cumulative %'

    #Format header row
        for b in range(1, len(column_names_6)+1):
          ws_6.cell(row=header_row_no, column=b).fill = PatternFill("solid", fgColor="A6ACAF")
          ws_6.cell(row=header_row_no, column=b).alignment = Alignment(horizontal="center", vertical="center")
          ws_6.cell(row=header_row_no, column=b).font = Font(bold=True)

    #Format numbers
        for c in range(3,total_column_no):
          for d in range(header_row_no+1, total_row_no+1):
            ws_6.cell(row=d, column=c).style = 'Comma'
            ws_6.cell(row=d, column=c).font = Font(name="Calibri", size="11")

    #Format total row header
        ws_6.cell(row=total_row_no, column=2).fill = PatternFill("solid", fgColor="A6ACAF")
        ws_6.cell(row=total_row_no, column=2).font = Font(bold=True)
        ws_6.cell(row=total_row_no, column=2).alignment = Alignment(horizontal="center", vertical="center")

    #Format total row and column
        for e in range(header_row_no+1, total_row_no):
          ws_6.cell(row=e, column=total_column_no-1).font = Font(bold=True)
        for f in range(3, total_column_no):
          ws_6.cell(row=total_row_no, column=f).font = Font(bold=True)

    #Insert row between tables
        ws_6.insert_rows(total_row_no + 1)
        total_row_no += 1

#Worksheet 7: Third Party x Income Type x Rev x Half
    #Create list of column names
    half_column_names_7 = ['Payor', 'Income Type']
    for k in statement_period_half:
        half_column_names_7.append(k)
    half_column_names_7.append('Total')
    half_column_names_7.append('% Of Revenue')
    half_column_names_7.append('Cumulative %')
    half_column_names_final_7 = [(half_column_names_7)]

    quarter_column_names_7 = ['Payor', 'Song Title']
    for k in statement_period_quarter_list:
        quarter_column_names_7.append(k)
    quarter_column_names_7.append('Total')
    quarter_column_names_7.append('% Of Revenue')
    quarter_column_names_7.append('Cumulative %')
    quarter_column_names_final_7 = [(quarter_column_names_7)]

    #Build tables
    total_row_no = 0
    ws_7.insert_cols(1, len(quarter_column_names_7))
    for s in third_party_list:
        #Find smallest period
        find_quarterly_period = '''SELECT Third_Party_9LC FROM Master WHERE Quarter_Statement_9LC <> "" AND Third_Party_9LC = "{}" GROUP BY Third_Party_9LC'''.format(
            s)
        mycursor.execute(find_quarterly_period)
        quarterly = [i[0] for i in mycursor.fetchall()]
        if len(quarterly) == 1:
            smallest_period = 'Quarter'
            mySQL_column = 'Quarter_Statement_9LC'
            statement_list = quarter_statement_list
            year_statement_list = year_statement_list_quarter
            column_names_final_7 = quarter_column_names_final_7
            column_names_7 = quarter_column_names_7
        else:
            smallest_period = 'Half'
            mySQL_column = 'Half_Statement_9LC'
            statement_list = half_statement_list
            year_statement_list = year_statement_list_half
            column_names_final_7 = half_column_names_final_7
            column_names_7 = half_column_names_7

        select_table_7_1 = '''SELECT Third_Party_9LC,
                            Normalized_Income_Type_9LC,'''
        select_table_7_2 = ""
        for j, k in zip(year_statement_list, statement_list):
            select_table_7_2 += (''' sum( CASE WHEN Year_Statement_9LC = "{}" AND {} = "{}"
                                 THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(j,mySQL_column,k, j,k))
        select_table_7_3 = '''sum( CASE WHEN Statement_Period_Half_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END) 
                            AS `Total`
                            FROM Master WHERE Third_Party_9LC = "{}" AND Statement_Period_Half_9lC <> "" 
                            GROUP BY Third_Party_9LC, Normalized_Income_Type_9LC ORDER BY `Total` DESC'''.format(
            cut_off, s)
        select_table_7 = select_table_7_1 + select_table_7_2 + select_table_7_3
        mycursor.execute(select_table_7)
        table_7 = mycursor.fetchall()
        third_party_table_7 = column_names_final_7 + table_7

        #Add tables to sheet
        total_row_no += len(third_party_table_7) + 1
        space_row_no = total_row_no + 1
        header_row_no = total_row_no - len(third_party_table_7)
        total_column_no = len(column_names_7) - 1
        third_party_table_rows = ws_7[header_row_no:space_row_no]
        for (row, l) in zip(third_party_table_rows, third_party_table_7):
            for (cell, m) in zip(row, range(len(l))):
                cell.value = l[m]

        #Add total row
        ws_7.cell(row=total_row_no, column=2).value = 'Total'
        for n in range(3, total_column_no):
            ws_7.cell(row=total_row_no, column=n).value = '=SUM({}{}:{}{})'.format(column_letters[n - 1],
                                                                                 header_row_no + 1,
                                                                                 column_letters[n - 1],
                                                                                 total_row_no - 1)

        #Add % of revenue and cumulative % columns
        total_column_letter = column_letters[len(column_names_7) - 3]
        percentage_column_letter = column_letters[len(column_names_7) - 2]
        cumulative_column_letter = column_letters[len(column_names_7) - 1]
        for o in range(header_row_no + 1, total_row_no):
            ws_7.cell(row=o, column=total_column_no).value = '={}{}/{}{}'.format(total_column_letter, o,
                                                                               total_column_letter, total_row_no)
            ws_7.cell(row=o, column=total_column_no).style = 'Percent'
            ws_7.cell(row=o, column=total_column_no).font = Font(name="Calibri", size="11")
        ws_7.cell(row=header_row_no + 1, column=total_column_no + 1).value = '={}{}'.format(percentage_column_letter,
                                                                                          header_row_no + 1)
        ws_7.cell(row=header_row_no + 1, column=total_column_no + 1).style = 'Percent'
        ws_7.cell(row=header_row_no + 1, column=total_column_no + 1).font = Font(name="Calibri", size=11)
        for p in range(header_row_no + 2, total_row_no):
            ws_7.cell(row=p, column=total_column_no + 1).value = '={}{}+{}{}'.format(cumulative_column_letter, p - 1,
                                                                                   percentage_column_letter, p)
            ws_7.cell(row=p, column=total_column_no + 1).style = 'Percent'
            ws_7.cell(row=p, column=total_column_no + 1).font = Font(name="Calibri", size=11)

        #Format header row
        for b in range(1, len(column_names_7) + 1):
            ws_7.cell(row=header_row_no, column=b).fill = PatternFill("solid", fgColor="A6ACAF")
            ws_7.cell(row=header_row_no, column=b).alignment = Alignment(horizontal="center", vertical="center")
            ws_7.cell(row=header_row_no, column=b).font = Font(bold=True)

        #Format numbers
        for c in range(3, total_column_no):
            for d in range(header_row_no + 1, total_row_no + 1):
                ws_7.cell(row=d, column=c).style = 'Comma'
                ws_7.cell(row=d, column=c).font = Font(name="Calibri", size="11")

        #Format total row header
        ws_7.cell(row=total_row_no, column=2).fill = PatternFill("solid", fgColor="A6ACAF")
        ws_7.cell(row=total_row_no, column=2).font = Font(bold=True)
        ws_7.cell(row=total_row_no, column=2).alignment = Alignment(horizontal="center", vertical="center")

        #Format total row and column
        for e in range(header_row_no + 1, total_row_no):
            ws_7.cell(row=e, column=total_column_no - 1).font = Font(bold=True)
        for f in range(3, total_column_no):
            ws_7.cell(row=total_row_no, column=f).font = Font(bold=True)

        #Insert row between tables
        ws_7.insert_rows(total_row_no + 1)
        total_row_no += 1

#Worksheet 8: Third Party x Source x Rev x Half
    #Create list of column names
    half_column_names_8 = ['Payor', 'Source']
    for k in statement_period_half:
        half_column_names_8.append(k)
    half_column_names_8.append('Total')
    half_column_names_8.append('% Of Revenue')
    half_column_names_8.append('Cumulative %')
    half_column_names_final_8 = [(half_column_names_8)]

    quarter_column_names_8 = ['Payor', 'Source']
    for k in statement_period_quarter_list:
        quarter_column_names_8.append(k)
    quarter_column_names_8.append('Total')
    quarter_column_names_8.append('% Of Revenue')
    quarter_column_names_8.append('Cumulative %')
    quarter_column_names_final_8 = [(quarter_column_names_8)]

    #Build tables
    #Main table
    total_row_no = 0
    ws_8.insert_cols(1, len(quarter_column_names_8))
    for s in third_party_list:
        find_quarterly_period = '''SELECT Third_Party_9LC FROM Master WHERE Quarter_Statement_9LC <> "" AND Third_Party_9LC = "{}" GROUP BY Third_Party_9LC'''.format(
            s)
        mycursor.execute(find_quarterly_period)
        quarterly = [i[0] for i in mycursor.fetchall()]
        if len(quarterly) == 1:
            smallest_period = 'Quarter'
            mySQL_column = 'Quarter_Statement_9LC'
            statement_list = quarter_statement_list
            year_statement_list = year_statement_list_quarter
            column_names_final_8 = quarter_column_names_final_8
            column_names_8 = quarter_column_names_8
        else:
            smallest_period = 'Half'
            mySQL_column = 'Half_Statement_9LC'
            statement_list = half_statement_list
            year_statement_list = year_statement_list_half
            column_names_final_8 = half_column_names_final_8
            column_names_8 = half_column_names_8
        select_table_8_1 = '''SELECT Third_Party_9LC,
                            Normalized_Source_9LC,'''
        select_table_8_2 = ""
        for j, k in zip(year_statement_list, statement_list):
            select_table_8_2 += (''' sum( CASE WHEN Year_Statement_9LC = "{}" AND {} = "{}" AND Normalized_Source_9LC <> "Pool Revenue"
                                 THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(j,mySQL_column,k, j,k))
        select_table_8_3 = '''sum( CASE WHEN Normalized_Source_9LC <> "Pool Revenue" AND 
                            Statement_Period_Half_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `Total`
                            FROM Master WHERE Normalized_Source_9LC <> "Pool Revenue" AND
                            Third_Party_9LC = "{}" AND Statement_Period_Half_9lC <> "" 
                            GROUP BY Third_Party_9LC, Normalized_Source_9LC ORDER BY `Total` DESC'''.format(cut_off, s)
        select_table_8 = select_table_8_1 + select_table_8_2 + select_table_8_3
        mycursor.execute(select_table_8)
        table_8 = mycursor.fetchall()

    #Pool revenue
        pool_rev_8_1 = '''SELECT Third_Party_9LC,
                            Normalized_Source_9LC,'''
        pool_rev_8_2 = ""
        for l, m in zip(year_statement_list, statement_list):
            pool_rev_8_2 += ('''sum( CASE WHEN Year_Statement_9LC = "{}" AND {} = "{}" AND Normalized_Source_9LC = "Pool Revenue"
                                  THEN Adjusted_Royalty_SB ELSE "" END) AS `{}`,'''.format(l,mySQL_column,m, l,m))
        pool_rev_8_3 = '''sum( CASE WHEN Normalized_Source_9LC = "Pool Revenue" AND Statement_Period_Half_9LC >= "{}"
                            THEN Adjusted_Royalty_SB ELSE "" END) AS `Total` 
                            FROM Master WHERE Third_Party_9LC = "{}" AND Normalized_Source_9LC = "Pool Revenue" 
                            AND Statement_Period_Half_9LC <> ""
                            GROUP BY Third_Party_9LC, Normalized_Source_9LC ORDER BY `Total` DESC'''.format(cut_off, s)
        pool_rev_8 = pool_rev_8_1 + pool_rev_8_2 + pool_rev_8_3
        mycursor.execute(pool_rev_8)
        pool_revenue_8 = mycursor.fetchall()
        third_party_table_8 = column_names_final_8 + table_8 + pool_revenue_8

    #Add tables to sheet
        total_row_no += len(third_party_table_8) + 1
        space_row_no = total_row_no + 1
        header_row_no = total_row_no - len(third_party_table_8)
        total_column_no = len(column_names_8) - 1
        third_party_table_rows = ws_8[header_row_no:space_row_no]
        for (row, l) in zip(third_party_table_rows, third_party_table_8):
            for (cell, m) in zip(row, range(len(l))):
                cell.value = l[m]

    #Add total row
        ws_8.cell(row=total_row_no, column=2).value = 'Total'
        for n in range(3, total_column_no):
            ws_8.cell(row=total_row_no, column=n).value = '=SUM({}{}:{}{})'.format(column_letters[n - 1], header_row_no + 1,
                                                                                 column_letters[n - 1], total_row_no - 1)

    #Add % of revenue and cumulative % columns
        total_column_letter = column_letters[len(column_names_8) - 3]
        percentage_column_letter = column_letters[len(column_names_8) - 2]
        cumulative_column_letter = column_letters[len(column_names_8) - 1]
        for o in range(header_row_no + 1, total_row_no):
            ws_8.cell(row=o, column=total_column_no).value = '={}{}/{}{}'.format(total_column_letter, o,
                                                                               total_column_letter, total_row_no)
            ws_8.cell(row=o, column=total_column_no).style = 'Percent'
            ws_8.cell(row=o, column=total_column_no).font = Font(name="Calibri", size="11")
        ws_8.cell(row=header_row_no + 1, column=total_column_no + 1).value = '={}{}'.format(percentage_column_letter,
                                                                                          header_row_no + 1)
        ws_8.cell(row=header_row_no + 1, column=total_column_no + 1).style = 'Percent'
        ws_8.cell(row=header_row_no + 1, column=total_column_no + 1).font = Font(name="Calibri", size=11)
        for p in range(header_row_no + 2, total_row_no):
            ws_8.cell(row=p, column=total_column_no + 1).value = '={}{}+{}{}'.format(cumulative_column_letter, p - 1,
                                                                                   percentage_column_letter, p)
            ws_8.cell(row=p, column=total_column_no + 1).style = 'Percent'
            ws_8.cell(row=p, column=total_column_no + 1).font = Font(name="Calibri", size=11)

    #Format header row
        for b in range(1, len(column_names_8)+1):
          ws_8.cell(row=header_row_no, column=b).fill = PatternFill("solid", fgColor="A6ACAF")
          ws_8.cell(row=header_row_no, column=b).alignment = Alignment(horizontal="center", vertical="center")
          ws_8.cell(row=header_row_no, column=b).font = Font(bold=True)

    #Format numbers
        for c in range(3,total_column_no):
          for d in range(header_row_no+1, total_row_no+1):
            ws_8.cell(row=d, column=c).style = 'Comma'
            ws_8.cell(row=d, column=c).font = Font(name="Calibri", size="11")

    #Format total row header
        ws_8.cell(row=total_row_no, column=2).fill = PatternFill("solid", fgColor="A6ACAF")
        ws_8.cell(row=total_row_no, column=2).font = Font(bold=True)
        ws_8.cell(row=total_row_no, column=2).alignment = Alignment(horizontal="center", vertical="center")

    #Format total row and column
        for e in range(header_row_no+1, total_row_no):
          ws_8.cell(row=e, column=total_column_no-1).font = Font(bold=True)
        for f in range(3, total_column_no):
          ws_8.cell(row=total_row_no, column=f).font = Font(bold=True)

    #Insert row between tables
        ws_8.insert_rows(total_row_no + 1)
        total_row_no += 1



    return(wb.save(filename))

#combined_outputs('DJ Battlecat_616c91a05f278b92afbaa5ae')


