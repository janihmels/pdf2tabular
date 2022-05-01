from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, NamedStyle
from datetime import date
import re
import mysql.connector
import scipy.optimize as optimize

def dcf(database, filename, growth_rates, incremental_sync, discount_rate, tv_multiple, tax_rate, initial_cost):
    mydb = mysql.connector.connect(
        host="34.65.111.142",
        user="external",
        password="musicpass",
        database="{}".format(database)
    )
    mycursor = mydb.cursor(buffered=True)

    discount_rate = float(discount_rate)
    tv_multiple = float(tv_multiple)
    tax_rate = float(tax_rate)
    initial_cost = float(initial_cost)

    def npv(discount_rate, cf, cost):
        net_present_value = 0
        forecast_values = cf[:5]
        terminal_value = cf[5]
        for i in forecast_values:
            net_present_value += i / (1 + discount_rate) ** (cf.index(i) + 0.5)
        net_present_value += terminal_value / (1 + discount_rate) ** 4.5
        net_present_value -= cost
        return net_present_value

    def irr(cf, initial_cost):
        f = lambda x: npv(x, cf, initial_cost)
        root = optimize.root(f, 0.1)
        return root.x[0]

#Find last three full years
    find_period = '''SELECT Statement_Period_Half_9LC FROM Master WHERE Statement_Period_Half_9LC <> ""
                     GROUP BY Statement_Period_Half_9LC ORDER BY Statement_Period_Half_9LC'''
    mycursor.execute(find_period)
    statement_period_half = [i[0] for i in mycursor.fetchall()]
    todays_date = date.today()
    current_year = todays_date.year
    base_year_value = current_year - 1
    base_year = []
    for a in statement_period_half:
        match = re.search(r'{}\sH\d'.format(base_year_value), a)
        if match:
            base_year.append(match.group())
    previous_year_1 = base_year_value - 1
    previous_year_2 = base_year_value - 2

#Create sheets
    wb = Workbook()
    ws = wb.active
    ws.title = "DCF"

#Create styles
    #Main numbers
    number_style = NamedStyle(name="number_style")
    number_style.alignment = Alignment(horizontal="right", vertical="center")
    number_style.font = Font(name="Calibri", size="11")
    number_style.number_format = '#,##0'

    #Title cells
    title_style = NamedStyle(name="title_style")
    title_style.alignment = Alignment(horizontal="center", vertical="center")
    title_style.font = Font(name="Calibri", size="11", bold=True)
    title_style.fill = PatternFill("solid", fgColor="A6ACAF")

    #Header cells
    header_style = NamedStyle(name="header_style")
    header_style.alignment = Alignment(horizontal="center", vertical="center")
    header_style.font = Font(name="Calibri", size="11", bold=True)
    header_style.fill = PatternFill("solid", fgColor="C9CDCF")

    #Sub header cells
    sub_header_style = NamedStyle(name="sub_header_style")
    sub_header_style.alignment = Alignment(horizontal="left", vertical="center")
    sub_header_style.font = Font(name="Calibri", size="11", italic=True)

    #Name cells
    name_style = NamedStyle(name="name_style")
    name_style.alignment = Alignment(horizontal="left", vertical="center")
    name_style.font = Font(name="Calibri", size="11")

    #Bold name cells
    bold_name_style = NamedStyle(name="bold_name_style")
    bold_name_style.alignment = Alignment(horizontal="left", vertical="center")
    bold_name_style.font = Font(name="Calibri", size="11", bold=True)

    #Total cells
    total_style = NamedStyle(name="total_style")
    total_style.alignment = Alignment(horizontal="right", vertical="center")
    total_style.font = Font(name="Calibri", size="11", bold=True)
    total_style.number_format = '#,##0'

    #Lined total row
    thin = Side(border_style="thin", color="000000")
    lined_total_style = NamedStyle(name="lined_total_style")
    lined_total_style.alignment = Alignment(horizontal="right", vertical="center")
    lined_total_style.font = Font(name="Calibri", size="11", bold=True)
    lined_total_style.number_format = '#,##0.00'
    lined_total_style.border = Border(top=thin, left=None, right=None, bottom=thin)

    #Total label style
    total_label_style = NamedStyle(name="total_label_style")
    total_label_style.alignment = Alignment(horizontal="center", vertical="center")
    total_label_style.font = Font(name="Calibri", size="11", bold=True)
    total_label_style.fill = PatternFill("solid", fgColor="C9CDCF")
    total_label_style.border = Border(top=thin, left=None, right=None, bottom=thin)

    #Publisher/PRO label style
    thick = Side(border_style="medium", color="000000")
    publisher_label_style = NamedStyle(name="publisher_label_style")
    publisher_label_style.alignment = Alignment(horizontal="center", vertical="center")
    publisher_label_style.font = Font(name="Calibri", size="11", bold=True, italic=True)
    publisher_label_style.border = Border(bottom=thick)

    #Percentage style
    percentage_style = NamedStyle(name="percentage_style")
    percentage_style.alignment = Alignment(horizontal="right", vertical="center")
    percentage_style.font = Font(name="Calibri", size="11")
    percentage_style.number_format = '0%'

    #Incremental style
    incremental_style = NamedStyle(name="incremental_style")
    incremental_style.alignment = Alignment(horizontal="right", vertical="center")
    incremental_style.font = Font(name="Calibri", size="11")
    incremental_style.fill = PatternFill("solid", fgColor="C9CDCF")
    incremental_style.number_format = '#,##0'

    #Incremental label style
    incremental_label_style = NamedStyle(name="incremental_label_style")
    incremental_label_style.alignment = Alignment(horizontal="left", vertical="center")
    incremental_label_style.font = Font(name="Calibri", size="11", bold=True)
    incremental_label_style.fill = PatternFill("solid", fgColor="C9CDCF")

    #Negative number style
    negative_number_style = NamedStyle(name="negative_number_style")
    negative_number_style.number_format = '#,##0_);[Red](#,##0)'

    #Implied multiple style
    implied_multiple_style = NamedStyle(name="implied_multiple_style")
    implied_multiple_style.alignment = Alignment(horizontal="right", vertical="center")
    implied_multiple_style.font = Font(name="Calibri", size="11", bold=True)
    implied_multiple_style.number_format = '#,##0.0'

    #Bold percentage style
    bold_percentage_style = NamedStyle(name="bold_percentage_style")
    bold_percentage_style.alignment = Alignment(horizontal="right", vertical="center")
    bold_percentage_style.font = Font(name="Calibri", size="11", bold=True)
    bold_percentage_style.number_format = '0.0%'

    #Add styles to workbook
    wb.add_named_style(number_style)
    wb.add_named_style(header_style)
    wb.add_named_style(title_style)
    wb.add_named_style(name_style)
    wb.add_named_style(bold_name_style)
    wb.add_named_style(total_style)
    wb.add_named_style(sub_header_style)
    wb.add_named_style(lined_total_style)
    wb.add_named_style(total_label_style)
    wb.add_named_style(publisher_label_style)
    wb.add_named_style(percentage_style)
    wb.add_named_style(incremental_style)
    wb.add_named_style(incremental_label_style)
    wb.add_named_style(negative_number_style)
    wb.add_named_style(implied_multiple_style)
    wb.add_named_style(bold_percentage_style)

#Column letters
    column_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S',
                      'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

#Find last three years by income type
    find_base = '''SELECT Normalized_Income_Type_9LC,
                     sum( CASE WHEN Statement_Period_Half_9LC LIKE "{} H%" 
                     THEN Adjusted_Royalty_SB ELSE NULL END) AS `{}`,
                     sum( CASE WHEN Statement_Period_Half_9LC LIKE "{} H%"
                     THEN Adjusted_Royalty_SB ELSE NULL END) AS `{}`,
                     sum( CASE WHEN Statement_Period_Half_9LC LIKE "{} H%"
                     THEN Adjusted_Royalty_SB ELSE NULL END) AS `{}`,
                     sum( CASE WHEN Statement_Period_Half_9LC >= "{}" AND Statement_Period_Half_9LC < "{}"
                     THEN Adjusted_Royalty_SB ELSE "" END) AS `Total`
                     FROM Master WHERE Normalized_Income_Type_9LC <> "" GROUP BY Normalized_Income_Type_9LC
                     ORDER BY Normalized_Income_Type_9LC ASC'''.format(previous_year_2, previous_year_2,
                                                                       previous_year_1, previous_year_1,
                                                                       base_year_value, base_year_value,
                                                                       previous_year_2, current_year)
    mycursor.execute(find_base)
    initial_base_table = mycursor.fetchall()

    base_table = []
    digital_base_income = []
    mechanical_base_income = []
    other_base_income = []
    performance_base_income = []
    print_base_income = []
    sync_base_income = []
    income_type_list = ['Digital', 'Mechanical', 'Other', 'Performance', 'Print', 'Sync']
    if len(initial_base_table) < 6:
        for j in range(len(initial_base_table)):
            if initial_base_table[j][0] == income_type_list[0]:
                digital_base_income.append(initial_base_table[j])
            if initial_base_table[j][0] == income_type_list[1]:
                mechanical_base_income.append(initial_base_table[j])
            if initial_base_table[j][0] == income_type_list[2]:
                other_base_income.append(initial_base_table[j])
            if initial_base_table[j][0] == income_type_list[3]:
                performance_base_income.append(initial_base_table[j])
            if initial_base_table[j][0] == income_type_list[4]:
                print_base_income.append(initial_base_table[j])
            if initial_base_table[j][0] == income_type_list[5]:
                sync_base_income.append(initial_base_table[j])
        if len(digital_base_income) == 0:
            digital_base_income.append(('Digital', 0.00, 0.00, 0.00, 0.00))
        if len(mechanical_base_income) == 0:
            mechanical_base_income.append(('Mechanical', 0.00, 0.00, 0.00, 0.00))
        if len(other_base_income) == 0:
            other_base_income.append(('Other', 0.00, 0.00, 0.00, 0.00))
        if len(performance_base_income) == 0:
            performance_base_income.append(('Performance', 0.00, 0.00, 0.00, 0.00))
        if len(print_base_income) == 0:
            print_base_income.append(('Print', 0.00, 0.00, 0.00, 0.00))
        if len(sync_base_income) == 0:
            sync_base_income.append(('Sync', 0.00, 0.00, 0.00, 0.00))

        base_table.append(digital_base_income[0])
        base_table.append(mechanical_base_income[0])
        base_table.append(other_base_income[0])
        base_table.append(performance_base_income[0])
        base_table.append(print_base_income[0])
        base_table.append(sync_base_income[0])

    else:
        base_table = initial_base_table

#List of three year average per income type
    three_year_average_list = []
    for a in base_table:
        first_year_income = a[1]
        second_year_income = a[2]
        third_year_income = a[3]
        if first_year_income is None:
            first_year_income = 0
        if second_year_income is None:
            second_year_income = 0
        if third_year_income is None:
            third_year_income = 0
        three_years_income = first_year_income + second_year_income + third_year_income
        average = (first_year_income + second_year_income + third_year_income) / 3
        three_year_average_list.append(average)

#Size of worksheet
    for column_no in range(1, 19):
        for row_no in range(1, 10 + len(base_table)):
            ws.cell(row=row_no, column=column_no)

#Growth rate table
    #Growth rate column names
    growth_column_names = ['Income Type', 'Base', 'Growth Rates']
    forecasted_years = []
    for c in range(0, 5):
        forecasted_years.append(current_year + c)

    #Build growth rate table
    ws.merge_cells('C3:G3')
    for d, e in zip(range(1, 4), growth_column_names):
        ws.cell(row=3, column=d).value = e
        ws.cell(row=3, column=d).style = 'title_style'

    for f, g in zip(range(3, 8), forecasted_years):
        ws.cell(row=4, column=f).value = g

    for h in range(1, 8):
        ws.cell(row=4, column=h).style = 'header_style'

    for i, j in zip(range(5, len(base_table) + 5), range(len(base_table))):
        ws.cell(row=i, column=1).value = base_table[j][0]
        ws.cell(row=i, column=1).style = 'bold_name_style'

    for k, l in zip(range(5, len(base_table) + 5), three_year_average_list):
        ws.cell(row=k, column=2).value = l
        ws.cell(row=k, column=2).style = 'number_style'

    for m, n in zip(range(5, len(base_table) + 5), growth_rates):
        for o, p in zip(range(3, 8), n):
            ws.cell(row=m, column=o).value = p
            ws.cell(row=m, column=o).style = 'percentage_style'
            ws.cell(row=m, column=o).font = Font(color='000066CC')

    for o, p in zip(range(3, 8), incremental_sync):
        ws.cell(row=len(base_table) + 6, column=o).value = p
        ws.cell(row=len(base_table) + 6, column=o).style = 'incremental_style'
    ws.cell(row=len(base_table) + 6, column=1).value = 'Incremental Sync'
    ws.cell(row=len(base_table) + 6, column=1).style = 'incremental_label_style'
    ws.cell(row=len(base_table) + 6, column=2).style = 'incremental_style'

    ws.cell(row=len(base_table) + 7, column=2).value = "=SUM({}{}:{}{})".format('B', 4, 'B', len(base_table) + 4)
    ws.cell(row=len(base_table) + 7, column=2).style = 'total_style'
    ws.cell(row=len(base_table) + 7, column=1).value = 'Total'
    ws.cell(row=len(base_table) + 7, column=1).style = 'title_style'

#Input table
    ws.merge_cells('I3:J3')
    ws.cell(row=3, column=9).value = 'Inputs'
    ws.cell(row=3, column=9).style = 'title_style'

    input_table_labels = ['Discount Rate', 'TV Multiple', 'Tax Rate', 'Initial Cost']
    for q, r in zip(range(4, 8), input_table_labels):
        ws.cell(row=q, column=9).value = r
        ws.cell(row=q, column=9).style = 'name_style'
    ws.cell(row=4, column=10).value = discount_rate
    ws.cell(row=4, column=10).style = 'percentage_style'
    ws.cell(row=4, column=10).font = Font(color='000066CC')
    ws.cell(row=5, column=10).value = tv_multiple
    ws.cell(row=5, column=10).style = 'number_style'
    ws.cell(row=5, column=10).font = Font(color='000066CC')
    ws.cell(row=6, column=10).value = tax_rate
    ws.cell(row=6, column=10).style = 'percentage_style'
    ws.cell(row=6, column=10).font = Font(color='000066CC')
    ws.cell(row=7, column=10).value = initial_cost
    ws.cell(row=7, column=10).style = 'number_style'
    ws.cell(row=7, column=10).font = Font(color='000066CC')


#Valuation table
    ws.merge_cells('I9:J9')
    #Valuation table headers
    ws.cell(row=9, column=9).value = 'Valuation'
    ws.cell(row=9, column=9).style = 'title_style'
    ws.cell(row=10, column=9).value = 'NPV'
    ws.cell(row=10, column=9).style = 'bold_name_style'
    ws.cell(row=11, column=9).value = 'Explicit Forecasts'
    ws.cell(row=11, column=9).style = 'name_style'
    ws.cell(row=12, column=9).value = 'Terminal Value'
    ws.cell(row=12, column=9).style = 'name_style'
    ws.cell(row=13, column=9).value = 'Total (Enterprise Value)'
    ws.cell(row=13, column=9).style = 'title_style'
    ws.cell(row=14, column=9).value = 'Base'
    ws.cell(row=14, column=9).style = 'bold_name_style'
    ws.cell(row=15, column=9).value = 'Implied Multiple'
    ws.cell(row=15, column=9).style = 'bold_name_style'

    #Valuation table values
    ws.cell(row=11, column=10).value = "=SUM(N17:R17)"
    ws.cell(row=11, column=10).style = 'number_style'
    ws.cell(row=12, column=10).value = "=S17"
    ws.cell(row=12, column=10).style = 'number_style'
    ws.cell(row=13, column=10).value = "=J11+J12"
    ws.cell(row=13, column=10).style = 'total_style'
    ws.cell(row=14, column=10).value = "=M13"
    ws.cell(row=14, column=10).style = 'total_style'
    ws.cell(row=15, column=10).value = "=J13/J14"
    ws.cell(row=15, column=10).style = 'implied_multiple_style'

#DCF table
    #Column names
    dcf_table_columns = ['Income Type']
    dcf_table_columns.append(base_year_value)
    for s in forecasted_years:
        dcf_table_columns.append(s)
    dcf_table_columns.append(forecasted_years[4])

    for t, u in zip(range(12, 20), dcf_table_columns):
        ws.cell(row=3, column=t).value = u
        ws.cell(row=3, column=t).style = 'title_style'

    #Year labels
    year_labels = ['Base', 'Forecast', 'Forecast', 'Forecast', 'Forecast', 'Forecast', 'TV']
    for v, w in zip(range(13, 20), year_labels):
        ws.cell(row=4, column=v).value = w
        ws.cell(row=4, column=v).style = 'header_style'
    ws.cell(row=4, column=12).style = 'header_style'

    #Year numbers
    year_numbers = [1, 2, 3, 4, 5, 5]
    for x, y in zip(range(14, 20), year_numbers):
        ws.cell(row=2, column=x).value = y
        ws.cell(row=2, column=x).alignment = Alignment(horizontal="center", vertical="center")

    #Income types
    for z, a in zip(range(5, len(base_table) + 5), base_table):
        ws.cell(row=z, column=12).value = a[0]
        ws.cell(row=z, column=12).style = 'bold_name_style'
    ws.cell(row=len(base_table) + 6, column=12).value = 'Incremental Sync'
    ws.cell(row=len(base_table) + 6, column=12).style = 'incremental_label_style'

    #Base income
    for b in range(5, len(base_table) + 5):
        ws.cell(row=b, column=13).value = "=B{}".format(b)
        ws.cell(row=b, column=13).style = 'number_style'

    #Forecast year 1
    for c in range(5, len(base_table) + 5):
        ws.cell(row=c, column=14).value = "=M{}*(1+C{})".format(c, c)
        ws.cell(row=c, column=14).style = 'number_style'

    #Forecast year 2
    for d in range(5, len(base_table) + 5):
        ws.cell(row=d, column=15).value = "=N{}*(1+D{})".format(d, d)
        ws.cell(row=d, column=15).style = 'number_style'

    #Forecast year 3
    for e in range(5, len(base_table) + 5):
        ws.cell(row=e, column=16).value = "=O{}*(1+E{})".format(e, e)
        ws.cell(row=e, column=16).style = 'number_style'

    #Forecast year 4
    for f in range(5, len(base_table) + 5):
        ws.cell(row=f, column=17).value = "=P{}*(1+F{})".format(f, f)
        ws.cell(row=f, column=17).style = 'number_style'

    #Forecast year 5
    for g in range(5, len(base_table) + 5):
        ws.cell(row=g, column=18).value = "=Q{}*(1+G{})".format(g, g)
        ws.cell(row=g, column=18).style = 'number_style'

    #Add incremental sync line
    for h, i in zip(range(14, 19), column_letters[2:7]):
        ws.cell(row=12, column=h).value = "={}12".format(i)
        ws.cell(row=12, column=h).style = 'incremental_style'
    ws.cell(row=12, column=13).style = 'incremental_style'

    #Total row
    for j, k in zip(range(13, 19), column_letters[12:18]):
        ws.cell(row=13, column=j).value = "=SUM({}{}:{}{})".format(k, 5, k, len(base_table) + 6)
        ws.cell(row=13, column=j).style = 'total_style'
    ws.cell(row=13, column=12).value = 'Total'
    ws.cell(row=13, column=12).style = 'title_style'

    #Terminal value column
    ws.cell(row=13, column=19).value = "={}{}*{}{}".format('R', 13, 'J', 5)
    ws.cell(row=13, column=19).style = 'total_style'

    #Number of quarters pre receipts
    current_month = todays_date.month
    quarter = 0
    if current_month >= 1 and current_month <= 3:
        quarter = 1
    if current_month >= 4 and current_month <= 6:
        quarter = 2
    if current_month >= 7 and current_month <= 9:
        quarter = 3
    if current_month >= 10 and current_month <= 12:
        quarter = 4

    ws.cell(row=14, column=12).value = 'No of Quarters Pre Receipts ({})'.format(str(current_year)[-2:])
    ws.cell(row=14, column=12).style = 'bold_name_style'

    ws.cell(row=14, column=13).value = quarter

    ws.cell(row=14, column=14).value = "=-(M14/4)*N13"
    ws.cell(row=14, column=14).style = 'negative_number_style'

    #Adjusted income row
    for l, m in zip(range(14, 19), column_letters[13:18]):
        ws.cell(row=15, column=l).value = "=({}13+{}14)*(1-J6)".format(m, m)
        ws.cell(row=15, column=l).style = 'total_style'
    ws.cell(row=15, column=12).value = 'Adjusted Income (post tax)'
    ws.cell(row=15, column=12).style = 'title_style'
    print(three_year_average_list)
    #Adjusted income list
    year_total_list = []
    year_1_total = 0
    year_2_total = 0
    year_3_total = 0
    year_4_total = 0
    year_5_total = 0
    year_1 = 0
    year_2 = 0
    year_3 = 0
    year_4 = 0
    year_5 = 0
    for i, j in zip(range(len(three_year_average_list)), three_year_average_list):
        year_1 = j*(1+growth_rates[i][0])
        year_2 = year_1*(1+growth_rates[i][1])
        year_3 = year_2*(1+growth_rates[i][2])
        year_4 = year_3*(1+growth_rates[i][3])
        year_5 = year_4*(1+growth_rates[i][4])
        year_1_total += year_1
        year_2_total += year_2
        year_3_total += year_3
        year_4_total += year_4
        year_5_total += year_5
        year_1 = 0
        year_2 = 0
        year_3 = 0
        year_4 = 0
        year_5 = 0
    year_1_total += incremental_sync[0]
    year_1_total = year_1_total*(1-(quarter/4))
    year_2_total += incremental_sync[1]
    year_3_total += incremental_sync[2]
    year_4_total += incremental_sync[3]
    year_5_total += incremental_sync[4]
    terminal_value_total = year_5_total*float(tv_multiple)
    year_total_list = [year_1_total] + [year_2_total] + [year_3_total] + [year_4_total] + [year_5_total] + [terminal_value_total]
    print(tv_multiple)
    print(year_5_total)

    #NPV line
    for n, o in zip(range(14, 19), column_letters[13:18]):
        ws.cell(row=17, column=n).value = "={}15/(1+J4)^({}2-0.5)".format(o, o)
        ws.cell(row=17, column=n).style = 'total_style'
    ws.cell(row=17, column=19).value = "={}13/(1+J4)^({}2-0.5)".format('S', 'S')
    ws.cell(row=17, column=19).style = 'total_style'
    ws.cell(row=17, column=13).value = 'NPV'
    ws.cell(row=17, column=13).style = 'title_style'

    #IRR Calculation
    net_present_value = npv(discount_rate, year_total_list, initial_cost)
    rate_return = irr(year_total_list, initial_cost)

    #IRR Table
    ws.merge_cells('I17:J17')
    ws.cell(row=17, column=9).value = 'IRR'
    ws.cell(row=17, column=9).style = 'title_style'
    ws.cell(row=18, column=9).value = 'IRR'
    ws.cell(row=18, column=9).style = 'bold_name_style'
    ws.cell(row=18, column=10).value = rate_return
    ws.cell(row=18, column=10).style = 'bold_percentage_style'

#Outlines
    for p in range(1, 8):
        ws.cell(row=2, column=p).border = Border(bottom=thin)
        ws.cell(row=12, column=p).border = Border(bottom=thin)
        ws.cell(row=13, column=p).border = Border(bottom=thin)

    for q in range(9, 11):
        ws.cell(row=2, column=q).border = Border(bottom=thin)
        ws.cell(row=7, column=q).border = Border(bottom=thin)
        ws.cell(row=8, column=q).border = Border(bottom=thin)
        ws.cell(row=15, column=q).border = Border(bottom=thin)
        ws.cell(row=16, column=q).border = Border(bottom=thin)
        ws.cell(row=18, column=q).border = Border(bottom=thin)

    for r in range(12, 20):
        ws.cell(row=2, column=r).border = Border(bottom=thin)
        ws.cell(row=12, column=r).border = Border(bottom=thin)
        ws.cell(row=14, column=r).border = Border(bottom=thin)
        ws.cell(row=15, column=r).border = Border(bottom=thin)

    for s in range(13, 20):
        ws.cell(row=16, column=s).border = Border(bottom=thin)
        ws.cell(row=17, column=s).border = Border(bottom=thin)

    for t in range(3, 14):
        ws.cell(row=t, column=1).border = Border(left=thin)
        ws.cell(row=t, column=2).border = Border(left=thin)
        ws.cell(row=t, column=8).border = Border(left=thin)

    for u in range(3, 8):
        ws.cell(row=u, column=9).border = Border(left=thin)
        ws.cell(row=u, column=11).border = Border(left=thin)

    for v in range(9, 16):
        ws.cell(row=v, column=9).border = Border(left=thin)
        ws.cell(row=v, column=11).border = Border(left=thin)

    for w in range(3, 16):
        ws.cell(row=w, column=12).border = Border(left=thin)
        ws.cell(row=w, column=13).border = Border(left=thin)
        ws.cell(row=w, column=14).border = Border(left=thin)
        ws.cell(row=w, column=19).border = Border(left=thin)
        ws.cell(row=w, column=20).border = Border(left=thin)

    for x in range(17, 19):
        ws.cell(row=x, column=9).border = Border(left=thin)
        ws.cell(row=x, column=11).border = Border(left=thin)

    ws.cell(row=12, column=1).border = Border(left=thin, bottom=thin)
    ws.cell(row=12, column=2).border = Border(left=thin, bottom=thin)
    ws.cell(row=13, column=1).border = Border(left=thin, bottom=thin)
    ws.cell(row=13, column=2).border = Border(left=thin, bottom=thin)
    ws.cell(row=7, column=9).border = Border(left=thin, bottom=thin)
    ws.cell(row=15, column=9).border = Border(left=thin, bottom=thin)
    ws.cell(row=17, column=13).border = Border(left=thin, bottom=thin)
    ws.cell(row=17, column=14).border = Border(left=thin, bottom=thin)
    ws.cell(row=17, column=19).border = Border(left=thin, bottom=thin)
    ws.cell(row=12, column=12).border = Border(left=thin, bottom=thin)
    ws.cell(row=12, column=13).border = Border(left=thin, bottom=thin)
    ws.cell(row=12, column=14).border = Border(left=thin, bottom=thin)
    ws.cell(row=12, column=19).border = Border(left=thin, bottom=thin)
    ws.cell(row=14, column=12).border = Border(left=thin, bottom=thin)
    ws.cell(row=14, column=13).border = Border(left=thin, bottom=thin)
    ws.cell(row=14, column=14).border = Border(left=thin, bottom=thin)
    ws.cell(row=14, column=19).border = Border(left=thin, bottom=thin)
    ws.cell(row=15, column=12).border = Border(left=thin, bottom=thin)
    ws.cell(row=15, column=13).border = Border(left=thin, bottom=thin)
    ws.cell(row=15, column=14).border = Border(left=thin, bottom=thin)
    ws.cell(row=15, column=19).border = Border(left=thin, bottom=thin)
    ws.cell(row=18, column=9).border = Border(left=thin, bottom=thin)
    ws.cell(row=17, column=20).border = Border(left=thin)

#Save workbook
    #wb.save(filename)
    return (wb.save(filename))


# Income type order: digital, mechanical, other, performance, print, sync
battlecat_growth_rates = [(0.08, 0.08, 0.08, 0.08, 0.08), (-0.08, -0.08, -0.08, -0.08, -0.08, -0.08),
                          (0.00, 0.00, 0.00, 0.00, 0.00, 0.00), (0.03, 0.03, 0.03, 0.03, 0.03, 0.03),
                          (0.00, 0.00, 0.00, 0.00, 0.00, 0.00), (0.03, 0.03, 0.03, 0.03, 0.03, 0.03, 0.03)]
battlecat_incremental_sync = [1000, 2000, 3000, 4000, 5000]
battlecat_discount_rate = 0.09
battlecat_tv_multiple = 12
battlecat_tax_rate = 0
battlecat_initial_cost = 190000
bush_initial_cost = 60000
#dcf('DJ Battlecat_616c91a05f278b92afbaa5ae', battlecat_growth_rates, battlecat_incremental_sync,
  #battlecat_discount_rate, battlecat_tv_multiple, battlecat_tax_rate, battlecat_initial_cost)
#dcf('Kristian Bush_615db5a2ecd455245e6954f1', battlecat_growth_rates, battlecat_incremental_sync,
    #battlecat_discount_rate, battlecat_tv_multiple, battlecat_tax_rate, bush_initial_cost)








