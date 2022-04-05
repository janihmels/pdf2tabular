from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from datetime import date
import mysql.connector

def thirdpartyxincomexrevxhalf(database, filename):
  mydb = mysql.connector.connect(
    host="34.65.111.142",
    user="external",
    password="musicpass",
    database="{}".format(database)
    )

  mycursor = mydb.cursor(buffered=True)

#Create sheets
  wb = Workbook()
  ws = wb.active
  ws.title = "Payor x Income x Rev x Half"

#Current year
  todays_date = date.today()
  current_year = todays_date.year

#Find most recent year in data
  find_recent_year = '''SELECT Year_Statement_9LC FROM Master GROUP BY Year_Statement_9LC'''
  mycursor.execute(find_recent_year)
  recent_years = [i[0] for i in mycursor.fetchall()]
  recent_year = recent_years[-1]
  print(recent_year)
  print(current_year)
  if recent_year == current_year:
    find_cut_off_year = current_year
  else:
    find_cut_off_year = current_year - 1

#Find cut off
  find_period = 'SELECT DISTINCT Statement_Period_Half_9LC FROM Master WHERE Year_Statement_9LC <= "{}" ORDER BY Statement_Period_Half_9LC'.format(find_cut_off_year)
  mycursor.execute(find_period)
  statement_period_half_blank = [i[0] for i in mycursor.fetchall()]

  def check_blank(period):
    if period == '':
      return False
    else:
      return True

  remove_blank = filter(check_blank, statement_period_half_blank)
  statement_period_minus_blank = list(remove_blank)
  if len(statement_period_minus_blank) > 10:
    statement_period_half = statement_period_minus_blank[-10:]
    standard_cut_off = statement_period_minus_blank[-10]
  else:
    statement_period_half = statement_period_minus_blank
    standard_cut_off = statement_period_minus_blank[0]
  print(statement_period_half)
  print(standard_cut_off)
  year_cut_off = standard_cut_off[0:4]

#Get list of statement year periods
  find_year_period = '''SELECT Year_Statement_9LC FROM Master 
                                            WHERE Year_Statement_9LC >= "{}" GROUP BY Year_Statement_9LC'''.format(
    year_cut_off)
  mycursor.execute(find_year_period)
  complete_list = [i for i in mycursor.fetchall()]
  statement_period_year_list = [i[0] for i in complete_list]
  year_statement_list_year = [i[0] for i in complete_list]
  yearly_statement_list = [i[0] for i in complete_list]
  year_list = list(dict.fromkeys(year_statement_list_year))

#Get list of statement half periods
  find_half_period = '''SELECT Year_Statement_9LC, Half_Statement_9LC FROM Master 
                                 WHERE Half_Statement_9LC <> "" AND Statement_Period_Half_9LC >= "{}" GROUP BY Year_Statement_9LC, Half_Statement_9LC'''.format(
    standard_cut_off)
  mycursor.execute(find_half_period)
  complete_list = [i for i in mycursor.fetchall()]
  statement_period_half_list = [i[0] + ' ' + i[1] for i in complete_list]
  year_statement_list_half = [i[0] for i in complete_list]
  half_statement_list = [i[1] for i in complete_list]

#Get list of statement quarter periods
  find_quarter_period = '''SELECT Year_Statement_9LC, Quarter_Statement_9LC FROM Master 
                               WHERE Quarter_Statement_9LC <> "" AND Statement_Period_Half_9LC >= "{}" GROUP BY Year_Statement_9LC, Quarter_Statement_9LC'''.format(
    standard_cut_off)
  mycursor.execute(find_quarter_period)
  complete_list = [i for i in mycursor.fetchall()]
  statement_period_quarter_list = [i[0] + ' ' + i[1] for i in complete_list]
  year_statement_list_quarter = [i[0] for i in complete_list]
  quarter_statement_list = [i[1] for i in complete_list]
  print(quarter_statement_list)

#Find types of third party
  find_type = '''SELECT Rights_Type_9LC, Third_Party_9LC FROM Master GROUP BY Rights_Type_9LC, Third_Party_9LC'''
  mycursor.execute(find_type)
  type_list = list(dict.fromkeys([i[0] for i in mycursor.fetchall()]))
  test = [i[1] for i in mycursor.fetchall()]
  type_string = ', '.join('"{}"'.format(str(x)) for x in type_list)
  print(type_list)
  print(test)

#Find third parties for each type
  third_party_type_list = []
  for a in type_list:
    third_party_sub_list = [a]
    find_third_party_type = '''SELECT Third_Party_9LC FROM Master WHERE Rights_Type_9LC = "{}" GROUP BY Third_Party_9LC'''.format(
      a)
    mycursor.execute(find_third_party_type)
    test_list = [i[0] for i in mycursor.fetchall()]
    for b in test_list:
      third_party_sub_list.append(b)
    third_party_type_list.append(third_party_sub_list)
    third_party_sub_list = []

#Publisher list
  find_publishers = '''SELECT Third_Party_9LC FROM Master WHERE Third_Party_9LC NOT IN ('ASCAP', 'BMI', 'PRS', 'MCPS', 
                           'SoundExchange', 'SOCAN', 'APRA', 'AMCOS', 'SESAC', 'AMRA', 'GEMA', 'SUISA', 'ZAiKS') GROUP BY Third_Party_9LC'''
  mycursor.execute(find_publishers)
  publisher_list = [i[0] for i in mycursor.fetchall()]
  publisher_string = ', '.join('"{}"'.format(str(x)) for x in publisher_list)
  print(publisher_list)

#PRO list
  find_PRO = '''SELECT Third_Party_9LC FROM Master WHERE Third_Party_9LC IN ('ASCAP', 'BMI', 'PRS', 'MCPS', 
                             'SoundExchange', 'SOCAN', 'APRA', 'AMCOS', 'SESAC', 'AMRA', 'GEMA', 'SUISA', 'ZAiKS') GROUP BY Third_Party_9LC'''
  mycursor.execute(find_PRO)
  PRO_list = [i[0] for i in mycursor.fetchall()]
  PRO_string = ', '.join('"{}"'.format(str(x)) for x in PRO_list)
  print(PRO_list)

#Third party list
  third_party_list = publisher_list + PRO_list

#Create list of column names
  half_column_names = ['Payor', 'Income Type']
  for k in statement_period_half_list:
    half_column_names.append(k)
  half_column_names.append('Total')
  half_column_names.append('% Of Revenue')
  half_column_names.append('Cumulative %')
  half_column_names_final = [(half_column_names)]

  quarter_column_names = ['Payor', 'Income Type']
  for k in statement_period_quarter_list:
    quarter_column_names.append(k)
  quarter_column_names.append('Total')
  quarter_column_names.append('% Of Revenue')
  quarter_column_names.append('Cumulative %')
  quarter_column_names_final = [(quarter_column_names)]
  print(quarter_column_names_final)

  year_column_names = ['Payor', 'Income Type']
  for k in statement_period_year_list:
    year_column_names.append(k)
  year_column_names.append('Total')
  year_column_names.append('% Of Revenue')
  year_column_names.append('Cumulative %')
  year_column_names_final = [(year_column_names)]

#Column letters
  column_letters = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T',
                    'U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO',
                    'AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY','AZ']

#Build publisher tables
  total_row_no = 0
  ws.insert_cols(1,len(quarter_column_names))
  for r in third_party_type_list:
    for s in r[1:]:
      find_statement_number_months = '''SELECT Statement_Period_9LC, Month_Statement_9LC FROM Master WHERE Third_Party_9LC = "{}" GROUP BY Statement_Period_9LC, Month_Statement_9LC'''.format(
        s)
      mycursor.execute(find_statement_number_months)
      full_list_months = mycursor.fetchall()
      statement_number_blanks = [i[0] for i in full_list_months]
      statement_number_months_blanks = [i[1] for i in full_list_months]
      statement_number = len(list(filter(check_blank, statement_number_blanks)))
      statement_number_months = len(list(filter(check_blank, statement_number_months_blanks)))
      if statement_number == statement_number_months:
        monthly = True
      else:
        monthly = False

      find_statement_number_quarters = '''SELECT Statement_Period_9LC, Quarter_Statement_9LC FROM Master WHERE Third_Party_9LC = "{}" GROUP BY Statement_Period_9LC, Quarter_Statement_9LC'''.format(
        s)
      mycursor.execute(find_statement_number_quarters)
      full_list_quarters = mycursor.fetchall()
      statement_number_blanks = [i[0] for i in full_list_quarters]
      statement_number_quarters_blanks = [i[1] for i in full_list_quarters]
      statement_number = len(list(filter(check_blank, statement_number_blanks)))
      statement_number_quarters = len(list(filter(check_blank, statement_number_quarters_blanks)))
      print(statement_number)
      print(statement_number_quarters)
      if statement_number == statement_number_quarters:
        quarterly = True
      else:
        quarterly = False

      find_statement_number_halves = '''SELECT Statement_Period_9LC, Half_Statement_9LC FROM Master WHERE Third_Party_9LC = "{}" GROUP BY Statement_Period_9LC, Half_Statement_9LC'''.format(
        s)
      mycursor.execute(find_statement_number_halves)
      full_list_halves = mycursor.fetchall()
      statement_number_blanks = [i[0] for i in full_list_halves]
      statement_number_halves_blanks = [i[1] for i in full_list_halves]
      statement_number = len(list(filter(check_blank, statement_number_blanks)))
      statement_number_halves = len(list(filter(check_blank, statement_number_halves_blanks)))
      print(statement_number)
      print(statement_number_halves)
      if statement_number == statement_number_halves:
        half = True
      else:
        half = False

      if quarterly:
        smallest_period = 'Quarter'
        mySQL_column = 'Quarter_Statement_9LC'
        statement_list = quarter_statement_list
        year_statement_list = year_statement_list_quarter
        column_names_final = quarter_column_names_final
        column_names = quarter_column_names
        cut_off_field = 'Statement_Period_Half_9LC'
        cut_off = standard_cut_off
      elif half:
        smallest_period = 'Half'
        mySQL_column = 'Half_Statement_9LC'
        statement_list = half_statement_list
        year_statement_list = year_statement_list_half
        column_names_final = half_column_names_final
        column_names = half_column_names
        cut_off_field = 'Statement_Period_Half_9LC'
        cut_off = standard_cut_off
      else:
        smallest_period = 'Year'
        mySQL_column = 'Year_Statement_9LC'
        statement_list = yearly_statement_list
        year_statement_list = year_statement_list_year
        column_names_final = year_column_names_final
        column_names = year_column_names
        cut_off_field = 'Year_Statement_9LC'
        cut_off = year_cut_off

      select_table_1 = '''SELECT Third_Party_9LC,
                          Normalized_Income_Type_9LC,'''
      select_table_2 = ""
      for j, k in zip(year_statement_list, statement_list):
        select_table_2 += (''' sum( CASE WHEN Year_Statement_9LC = "{}" AND {} = "{}" 
                               THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(j,mySQL_column,k,j,k))
      select_table_3 = '''sum( CASE WHEN {} >= "{}" THEN Adjusted_Royalty_SB ELSE "" END) 
                          AS `Total`
                          FROM Master WHERE Third_Party_9LC = "{}"
                          GROUP BY Third_Party_9LC, Normalized_Income_Type_9LC ORDER BY `Total` DESC'''.format(cut_off_field,cut_off,s)
      select_table = select_table_1 + select_table_2 + select_table_3
      mycursor.execute(select_table)
      print(select_table)
      table = mycursor.fetchall()
      third_party_table = column_names_final + table


  #Add tables to sheet
      total_row_no += len(third_party_table)+1
      space_row_no = total_row_no + 1
      header_row_no = total_row_no - len(third_party_table)
      total_column_no = len(column_names)-1
      third_party_table_rows = ws[header_row_no:space_row_no]
      for (row,l) in zip(third_party_table_rows, third_party_table):
        for (cell, m) in zip(row, range(len(l))):
          cell.value = l[m]

  #Add total row
      ws.cell(row=total_row_no, column=2).value = 'Total'
      for n in range(3,total_column_no):
        ws.cell(row=total_row_no, column=n).value = '=SUM({}{}:{}{})'.format(column_letters[n-1],header_row_no+1,
                                                                             column_letters[n-1],total_row_no-1)

  #Add percentage and cumulative percentage columns
      total_column_letter = column_letters[len(column_names)-3]
      percentage_column_letter = column_letters[len(column_names)-2]
      cumulative_column_letter = column_letters[len(column_names)-1]
      for o in range(header_row_no+1, total_row_no):
        ws.cell(row=o, column=total_column_no).value = '={}{}/{}{}'.format(total_column_letter,o,
                                                                             total_column_letter,total_row_no)
        ws.cell(row=o, column=total_column_no).number_format = '0.00%'
        ws.cell(row=o, column=total_column_no).font = Font(name="Calibri", size="11")
      ws.cell(row=header_row_no+1, column=total_column_no+1).value = '={}{}'.format(percentage_column_letter,
                                                                                   header_row_no+1)
      ws.cell(row=header_row_no+1, column=total_column_no+1).number_format = '0.00%'
      ws.cell(row=header_row_no+1, column=total_column_no+1).font = Font(name="Calibri", size=11)
      for p in range(header_row_no+2, total_row_no):
        ws.cell(row=p, column=total_column_no+1).value = '={}{}+{}{}'.format(cumulative_column_letter,p-1,
                                                                             percentage_column_letter,p)
        ws.cell(row=p, column=total_column_no+1).number_format = '0.00%'
        ws.cell(row=p, column=total_column_no+1).font = Font(name="Calibri", size=11)

  #Format header row
      for b in range(1, len(column_names)+1):
        ws.cell(row=header_row_no, column=b).fill = PatternFill("solid", fgColor="A6ACAF")
        ws.cell(row=header_row_no, column=b).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=header_row_no, column=b).font = Font(bold=True)

  #Format numbers
      for c in range(3,total_column_no):
        for d in range(header_row_no+1, total_row_no+1):
          ws.cell(row=d, column=c).style = 'Comma'
          ws.cell(row=d, column=c).font = Font(name="Calibri", size="11")

  #Format total row header
      ws.cell(row=total_row_no, column=2).fill = PatternFill("solid", fgColor="A6ACAF")
      ws.cell(row=total_row_no, column=2).font = Font(bold=True)
      ws.cell(row=total_row_no, column=2).alignment = Alignment(horizontal="center", vertical="center")

  #Format total row and column
      for e in range(header_row_no+1, total_row_no):
        ws.cell(row=e, column=total_column_no-1).font = Font(bold=True)
      for f in range(3, total_column_no):
        ws.cell(row=total_row_no, column=f).font = Font(bold=True)

  #Insert row between tables
      ws.insert_rows(total_row_no+1)
      total_row_no += 1


    print(total_row_no)
    print(header_row_no)
    print(third_party_table)
    print(len(third_party_table))


  return wb.save(filename)

#thirdpartyxincomexrevxhalf('DJ Battlecat_616c91a05f278b92afbaa5ae')