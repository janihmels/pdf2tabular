from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, NamedStyle
import pymysql
from pymysql.converters import escape_string
from datetime import date
import mysql.connector

def songxincomexplaysxrevenue(database, filename):
  mydb = mysql.connector.connect(
    host="34.65.111.142",
    user="external",
    password="musicpass",
    database="{}".format(database)
    )

  mycursor = mydb.cursor(buffered=True)

#Create sheets
  wb = Workbook()
  ws = wb.active

# Current year
  todays_date = date.today()
  current_year = todays_date.year

  # Find most recent year in data
  find_recent_year = '''SELECT Year_Statement_9LC FROM Master GROUP BY Year_Statement_9LC'''
  mycursor.execute(find_recent_year)
  recent_years = [i[0] for i in mycursor.fetchall()]
  recent_year = recent_years[-1]
  print(recent_year)
  print(current_year)
  if recent_year == current_year:
    find_cut_off_year = current_year - 1
  else:
    find_cut_off_year = int(recent_year) - 1

  # Find cut off
  find_period = 'SELECT DISTINCT Statement_Period_Half_9LC FROM Master WHERE Year_Statement_9LC <= "{}" ORDER BY Statement_Period_Half_9LC'.format(
    find_cut_off_year)
  mycursor.execute(find_period)
  statement_period_half_blank = [i[0] for i in mycursor.fetchall()]

  def check_blank(period):
    if period == '':
      return False
    else:
      return True

  remove_blank = filter(check_blank, statement_period_half_blank)
  statement_period_minus_blank = list(remove_blank)
  if len(statement_period_minus_blank) > 10:
    statement_period_half = statement_period_minus_blank[-10:]
    cut_off = statement_period_minus_blank[-10]
  else:
    statement_period_half = statement_period_minus_blank
    cut_off = statement_period_minus_blank[0]
  print(statement_period_half)
  print(cut_off)

#Get list of statement half periods
  find_half_period = '''SELECT Year_Statement_9LC, Half_Statement_9LC FROM Master 
                                       WHERE Half_Statement_9LC <> "" AND Statement_Period_Half_9LC >= "{}" GROUP BY Year_Statement_9LC, Half_Statement_9LC'''.format(
    cut_off)
  mycursor.execute(find_half_period)
  complete_list = [i for i in mycursor.fetchall()]
  statement_period_half_list = [i[0] + ' ' + i[1] for i in complete_list]
  year_statement_list_half = [i[0] for i in complete_list]
  half_statement_list = [i[1] for i in complete_list]
  year_list = list(dict.fromkeys(year_statement_list_half))

#Get list of statement quarter periods
  find_quarter_period = '''SELECT Year_Statement_9LC, Quarter_Statement_9LC FROM Master 
                                     WHERE Quarter_Statement_9LC <> "" AND Statement_Period_Half_9LC >= "{}" GROUP BY Year_Statement_9LC, Quarter_Statement_9LC'''.format(
    cut_off)
  mycursor.execute(find_quarter_period)
  complete_list = [i for i in mycursor.fetchall()]
  statement_period_quarter_list = [i[0] + ' ' + i[1] for i in complete_list]
  year_statement_list_quarter = [i[0] for i in complete_list]
  quarter_statement_list = [i[1] for i in complete_list]

#Find number of songs
  find_songs = '''SELECT Song_Name_9LC, sum( Adjusted_Royalty_SB) 
                  FROM Master WHERE Song_Name_9LC <> "Pool Revenue" AND Statement_Period_Half_9LC >= "{}" 
                  GROUP BY Song_Name_9LC ORDER BY sum( Adjusted_Royalty_SB) DESC'''.format(cut_off)
  mycursor.execute(find_songs)
  songs = [i[0] for i in mycursor.fetchall()]
  find_pool_rev = '''SELECT Song_Name_9LC, sum( Adjusted_Royalty_SB)
                     FROM Master WHERE Song_Name_9LC = "Pool Revenue" AND Statement_Period_Half_9LC >= "{}" 
                     GROUP BY Song_Name_9LC'''.format(cut_off)
  mycursor.execute(find_pool_rev)
  pool_rev = [i[0] for i in mycursor.fetchall()]
  all_songs = songs + pool_rev

#Create styles
  #Main numbers
  number_style = NamedStyle(name="number_style")
  number_style.alignment = Alignment(horizontal="right", vertical="center")
  number_style.font = Font(name="Calibri", size="11")
  number_style.number_format = '#,##0.00'

  #Title cells
  title_style = NamedStyle(name="title_style")
  title_style.alignment = Alignment(horizontal="center", vertical="center")
  title_style.font = Font(name="Calibri", size="11", bold=True)
  title_style.fill = PatternFill("solid", fgColor="A6ACAF")

  #Header cells
  header_style = NamedStyle(name="header_style")
  header_style.alignment = Alignment(horizontal="center", vertical="center")
  header_style.font = Font(name="Calibri", size="11", bold=True)
  header_style.fill = PatternFill("solid", fgColor="C9CDCF")

  #Sub header cells
  sub_header_style = NamedStyle(name="sub_header_style")
  sub_header_style.alignment = Alignment(horizontal="left", vertical="center")
  sub_header_style.font = Font(name="Calibri", size="11", bold=True, italic=True)

  #Name cells
  name_style = NamedStyle(name="name_style")
  name_style.alignment = Alignment(horizontal="left", vertical="center")
  name_style.font = Font(name="Calibri", size="11")

  #Total cells
  total_style = NamedStyle(name="total_style")
  total_style.alignment = Alignment(horizontal="right", vertical="center")
  total_style.font = Font(name="Calibri", size="11", bold=True)
  total_style.number_format = '#,##0.00'

  #Lined total row
  thin = Side(border_style="thin", color="000000")
  lined_total_style = NamedStyle(name="lined_total_style")
  lined_total_style.alignment = Alignment(horizontal="right", vertical="center")
  lined_total_style.font = Font(name="Calibri", size="11", bold=True)
  lined_total_style.number_format = '#,##0.00'
  lined_total_style.border = Border(top=thin, left=None, right=None, bottom=thin)

  #Total label style
  total_label_style = NamedStyle(name="total_label_style")
  total_label_style.alignment = Alignment(horizontal="center", vertical="center")
  total_label_style.font = Font(name="Calibri", size="11", bold=True)
  total_label_style.fill = PatternFill("solid", fgColor="C9CDCF")
  total_label_style.border = Border(top=thin, left=None, right=None, bottom=thin)

  #Publisher/PRO label style
  thick = Side(border_style="medium", color="000000")
  publisher_label_style = NamedStyle(name="publisher_label_style")
  publisher_label_style.alignment = Alignment(horizontal="center", vertical="center")
  publisher_label_style.font = Font(name="Calibri", size="11", bold=True, italic=True)
  publisher_label_style.border = Border(bottom=thick)

  #Not available style
  not_available_style = NamedStyle(name="not_available_style")
  not_available_style.alignment = Alignment(horizontal="center", vertical="center")
  not_available_style.font = Font(name="Calibri", size="11")

  #Add styles to workbook
  wb.add_named_style(number_style)
  wb.add_named_style(header_style)
  wb.add_named_style(title_style)
  wb.add_named_style(name_style)
  wb.add_named_style(total_style)
  wb.add_named_style(sub_header_style)
  wb.add_named_style(lined_total_style)
  wb.add_named_style(total_label_style)
  wb.add_named_style(publisher_label_style)
  wb.add_named_style(not_available_style)

#Column letters
  column_letters = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T',
                    'U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO',
                    'AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY','AZ']

#Build sheet
  ws.title = 'Song x Income x Plays x Rev'

  #Create list of column names
  column_names_1 = ['Song Title', 'Income Type', 'Plays', 'Revenue', 'Revenue Per Play']
  column_names_final_1 = [(column_names_1)]

  #Build tables
  total_row_no = 0
  ws.insert_cols(1, len(column_names_1))
  for s in all_songs:
    print(s)
    s_string = escape_string(s)
    select_table_1 = '''SELECT Song_Name_9LC, Normalized_Income_Type_9LC, sum( CASE WHEN Statement_Period_Half_9LC >= "{}" THEN Plays_SB ELSE "" END) 
    AS `Plays`, sum( CASE WHEN Statement_Period_Half_9LC >= "{}" THEN Adjusted_Royalty_SB ELSE "" END) AS `Revenue` FROM Master WHERE Song_Name_9LC = "{}"
    AND Statement_Period_Half_9LC <> "" GROUP BY Song_Name_9LC, Normalized_Income_Type_9LC ORDER BY `Revenue` DESC'''.format(cut_off, cut_off, s_string)
    mycursor.execute(select_table_1)
    table_1 = mycursor.fetchall()
    song_table_1 = column_names_final_1 + table_1

    #Add tables to sheet
    total_row_no += len(song_table_1) + 1
    space_row_no = total_row_no + 1
    header_row_no = total_row_no - len(song_table_1)
    total_column_no = len(column_names_1) - 1
    ws.append(range(len(song_table_1) + 2))
    song_table_rows = ws[header_row_no:space_row_no]
    for (row, l) in zip(song_table_rows, song_table_1):
      for (cell, m) in zip(row, range(len(l))):
        cell.value = l[m]
    for a in range(header_row_no + 1, total_row_no + 1):
      ws.cell(row=a, column=3).number_format = '#,##0'
      ws.cell(row=a, column=4).number_format = '#,##0.00'

    #Add total row
    ws.cell(row=total_row_no, column=2).value = 'Total'
    for n in range(3, total_column_no+1):
      ws.cell(row=total_row_no, column=n).value = '=SUM({}{}:{}{})'.format(column_letters[n - 1], header_row_no + 1,
                                                                           column_letters[n - 1], total_row_no - 1)
      ws.cell(row=total_row_no, column=n).font = Font(bold=True)

    #Add revenue per play column
    total_column_letter = column_letters[len(column_names_1) - 3]
    percentage_column_letter = column_letters[len(column_names_1) - 2]
    cumulative_column_letter = column_letters[len(column_names_1) - 1]
    for o in range(header_row_no + 1, total_row_no):
      ws.cell(row=o, column=total_column_no+1).value = '=IFERROR(D{}/C{},0)'.format(o,o)
      ws.cell(row=o, column=total_column_no+1).number_format = '0.00000'
      ws.cell(row=o, column=total_column_no+1).font = Font(name="Calibri", size="11")

    #Format header row
    for b in range(1, len(column_names_1) + 1):
      ws.cell(row=header_row_no, column=b).fill = PatternFill("solid", fgColor="A6ACAF")
      ws.cell(row=header_row_no, column=b).alignment = Alignment(horizontal="center", vertical="center")
      ws.cell(row=header_row_no, column=b).font = Font(bold=True)

    #Format total row header
    ws.cell(row=total_row_no, column=2).fill = PatternFill("solid", fgColor="A6ACAF")
    ws.cell(row=total_row_no, column=2).font = Font(bold=True)
    ws.cell(row=total_row_no, column=2).alignment = Alignment(horizontal="center", vertical="center")

    #Format total row and column
    for f in range(3, total_column_no):
      ws.cell(row=total_row_no, column=f).font = Font(bold=True)

    #Insert row between tables
    ws.insert_rows(total_row_no + 1)
    total_row_no += 1

  return wb.save(filename)

#songxincomexrevxhalf('Havoc_61c0ae68c94b68a289900e85')