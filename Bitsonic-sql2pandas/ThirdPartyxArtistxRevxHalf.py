from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, NamedStyle
from datetime import date
from pandas_utils.pandas_cursor import pandas_cursor


def thirdpartyxartistxrevxhalf(database, df, filename):
  mycursor = pandas_cursor(df=df)

#Create sheets
  wb = Workbook()
  ws = wb.active
  ws.title = "Payor x Artist x Rev x Half"

# Current year
  todays_date = date.today()
  current_year = todays_date.year

  # Find most recent year in data
  find_recent_year = '''SELECT Year_Statement_9LC FROM Master GROUP BY Year_Statement_9LC'''
  mycursor.execute(find_recent_year)
  recent_years = [i[0] for i in mycursor.fetchall()]
  recent_year = recent_years[-1]
  print(recent_year)
  print(current_year)
  if recent_year == current_year:
    find_cut_off_year = current_year - 1
  else:
    find_cut_off_year = int(recent_year) - 1

  # Find cut off
  find_period = 'SELECT DISTINCT Statement_Period_Half_9LC FROM Master WHERE Year_Statement_9LC <= "{}" ORDER BY Statement_Period_Half_9LC'.format(
    find_cut_off_year)
  mycursor.execute(find_period)
  statement_period_half_blank = [i[0] for i in mycursor.fetchall()]

  def check_blank(period):
    if period == '':
      return False
    else:
      return True

  remove_blank = filter(check_blank, statement_period_half_blank)
  statement_period_minus_blank = list(remove_blank)
  if len(statement_period_minus_blank) > 10:
    statement_period_half = statement_period_minus_blank[-10:]
    cut_off = statement_period_minus_blank[-10]
  else:
    statement_period_half = statement_period_minus_blank
    cut_off = statement_period_minus_blank[0]
  print(statement_period_half)
  print(cut_off)

#Get list of statement half periods
  find_half_period = '''SELECT Year_Statement_9LC, Half_Statement_9LC FROM Master 
                                       WHERE Half_Statement_9LC <> "" AND Statement_Period_Half_9LC >= "{}" GROUP BY Year_Statement_9LC, Half_Statement_9LC'''.format(
    cut_off)
  mycursor.execute(find_half_period)
  complete_list = [i for i in mycursor.fetchall()]
  statement_period_half_list = [i[0] + ' ' + i[1] for i in complete_list]
  year_statement_list_half = [i[0] for i in complete_list]
  half_statement_list = [i[1] for i in complete_list]
  year_list = list(dict.fromkeys(year_statement_list_half))

#Get list of statement quarter periods
  find_quarter_period = '''SELECT Year_Statement_9LC, Quarter_Statement_9LC FROM Master 
                                     WHERE Quarter_Statement_9LC <> "" AND Statement_Period_Half_9LC >= "{}" GROUP BY Year_Statement_9LC, Quarter_Statement_9LC'''.format(
    cut_off)
  mycursor.execute(find_quarter_period)
  complete_list = [i for i in mycursor.fetchall()]
  statement_period_quarter_list = [i[0] + ' ' + i[1] for i in complete_list]
  year_statement_list_quarter = [i[0] for i in complete_list]
  quarter_statement_list = [i[1] for i in complete_list]

#Find types of third party
  find_type = '''SELECT Rights_Type_9LC, Third_Party_9LC FROM Master GROUP BY Rights_Type_9LC, Third_Party_9LC'''
  mycursor.execute(find_type)
  type_list = list(dict.fromkeys([i[0] for i in mycursor.fetchall()]))
  test = [i[1] for i in mycursor.fetchall()]
  type_string = ', '.join('"{}"'.format(str(x)) for x in type_list)
  print(type_list)
  print(test)

#Find third parties for each type
  third_party_type_list = []
  for a in type_list:
    third_party_sub_list = [a]
    find_third_party_type = '''SELECT Third_Party_9LC FROM Master WHERE Rights_Type_9LC = "{}" GROUP BY Third_Party_9LC'''.format(
      a)
    mycursor.execute(find_third_party_type)
    test_list = [i[0] for i in mycursor.fetchall()]
    for b in test_list:
      third_party_sub_list.append(b)
    third_party_type_list.append(third_party_sub_list)
    third_party_sub_list = []

#Publisher list
  find_publishers = '''SELECT Third_Party_9LC FROM Master WHERE Third_Party_9LC NOT IN ('ASCAP', 'BMI', 'PRS', 'MCPS', 
                           'SoundExchange', 'SOCAN', 'APRA', 'AMCOS', 'SESAC', 'AMRA', 'GEMA', 'SUISA', 'ZAiKS') GROUP BY Third_Party_9LC'''
  mycursor.execute(find_publishers)
  publisher_list = [i[0] for i in mycursor.fetchall()]
  publisher_string = ', '.join('"{}"'.format(str(x)) for x in publisher_list)
  print(publisher_list)

  #PRO list
  find_PRO = '''SELECT Third_Party_9LC FROM Master WHERE Third_Party_9LC IN ('ASCAP', 'BMI', 'PRS', 'MCPS', 
                             'SoundExchange', 'SOCAN', 'APRA', 'AMCOS', 'SESAC', 'AMRA', 'GEMA', 'SUISA', 'ZAiKS') GROUP BY Third_Party_9LC'''
  mycursor.execute(find_PRO)
  PRO_list = [i[0] for i in mycursor.fetchall()]
  PRO_string = ', '.join('"{}"'.format(str(x)) for x in PRO_list)
  print(PRO_list)

#Third party list
  third_party_list = publisher_list + PRO_list

#Create list of column names
  half_column_names = ['Payor', 'Release Artist']
  for k in statement_period_half:
    half_column_names.append(k)
  half_column_names.append('Total')
  half_column_names.append('% Of Revenue')
  half_column_names.append('Cumulative %')
  half_column_names_final = [(half_column_names)]

  quarter_column_names = ['Payor', 'Release Artist']
  for k in statement_period_quarter_list:
    quarter_column_names.append(k)
  quarter_column_names.append('Total')
  quarter_column_names.append('% Of Revenue')
  quarter_column_names.append('Cumulative %')
  quarter_column_names_final = [(quarter_column_names)]

#Column letters
  column_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S',
                    'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK',
                    'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ']

#Find number of artists
  find_songs = '''SELECT Release_Artist_9LC
                  FROM Master GROUP BY Release_Artist_9LC'''
  mycursor.execute(find_songs)
  artists = [i[0] for i in mycursor.fetchall()]
  artist_count = len(artists)
  print(artist_count)
  print(artists)

#Create styles
#Main numbers
  number_style = NamedStyle(name="number_style")
  number_style.alignment = Alignment(horizontal="right", vertical="center")
  number_style.font = Font(name="Calibri", size="11")
  number_style.number_format = '#,##0.00'

#Title cells
  title_style = NamedStyle(name="title_style")
  title_style.alignment = Alignment(horizontal="center", vertical="center")
  title_style.font = Font(name="Calibri", size="11", bold=True)
  title_style.fill = PatternFill("solid", fgColor="A6ACAF")

#Header cells
  header_style = NamedStyle(name="header_style")
  header_style.alignment = Alignment(horizontal="center", vertical="center")
  header_style.font = Font(name="Calibri", size="11", bold=True)
  header_style.fill = PatternFill("solid", fgColor="C9CDCF")

#Sub header cells
  sub_header_style = NamedStyle(name="sub_header_style")
  sub_header_style.alignment = Alignment(horizontal="left", vertical="center")
  sub_header_style.font = Font(name="Calibri", size="11", bold=True, italic=True)

#Name cells
  name_style = NamedStyle(name="name_style")
  name_style.alignment = Alignment(horizontal="left", vertical="center")
  name_style.font = Font(name="Calibri", size="11")

#Total cells
  total_style = NamedStyle(name="total_style")
  total_style.alignment = Alignment(horizontal="right", vertical="center")
  total_style.font = Font(name="Calibri", size="11", bold=True)
  total_style.number_format = '#,##0.00'

#Lined total row
  thin = Side(border_style="thin", color="000000")
  lined_total_style = NamedStyle(name="lined_total_style")
  lined_total_style.alignment = Alignment(horizontal="right", vertical="center")
  lined_total_style.font = Font(name="Calibri", size="11", bold=True)
  lined_total_style.number_format = '#,##0.00'
  lined_total_style.border = Border(top=thin, left=None, right=None, bottom=thin)

#Total label style
  total_label_style = NamedStyle(name="total_label_style")
  total_label_style.alignment = Alignment(horizontal="center", vertical="center")
  total_label_style.font = Font(name="Calibri", size="11", bold=True)
  total_label_style.fill = PatternFill("solid", fgColor="C9CDCF")
  total_label_style.border = Border(top=thin, left=None, right=None, bottom=thin)

#Publisher/PRO label style
  thick = Side(border_style="medium", color="000000")
  publisher_label_style = NamedStyle(name="publisher_label_style")
  publisher_label_style.alignment = Alignment(horizontal="center", vertical="center")
  publisher_label_style.font = Font(name="Calibri", size="11", bold=True, italic=True)
  publisher_label_style.border = Border(bottom=thick)

#Not available style
  not_available_style = NamedStyle(name="not_available_style")
  not_available_style.alignment = Alignment(horizontal="center", vertical="center")
  not_available_style.font = Font(name="Calibri", size="11")

#Add styles to workbook
  wb.add_named_style(number_style)
  wb.add_named_style(header_style)
  wb.add_named_style(title_style)
  wb.add_named_style(name_style)
  wb.add_named_style(total_style)
  wb.add_named_style(sub_header_style)
  wb.add_named_style(lined_total_style)
  wb.add_named_style(total_label_style)
  wb.add_named_style(publisher_label_style)
  wb.add_named_style(not_available_style)

  print(third_party_type_list)

#Build tables
  total_row_no = 0
  for y in range(1, len(quarter_column_names)):
    ws.cell(row=1, column=y)
  for r in third_party_type_list:
    for s in r[1:]:
      find_quarterly_period = '''SELECT Third_Party_9LC FROM Master WHERE Quarter_Statement_9LC <> "" AND Third_Party_9LC = "{}" GROUP BY Third_Party_9LC'''.format(
        s)
      mycursor.execute(find_quarterly_period)
      quarterly = [i[0] for i in mycursor.fetchall()]
      if len(quarterly) == 1:
        smallest_period = 'Quarter'
        mySQL_column = 'Quarter_Statement_9LC'
        statement_list = quarter_statement_list
        year_statement_list = year_statement_list_quarter
        column_names_final_1 = quarter_column_names_final
        column_names_1 = quarter_column_names
      else:
        smallest_period = 'Half'
        mySQL_column = 'Half_Statement_9LC'
        statement_list = half_statement_list
        year_statement_list = year_statement_list_half
        column_names_final_1 = half_column_names_final
        column_names_1 = half_column_names

      select_table_1_1 = '''SELECT Third_Party_9LC, Release_Artist_9LC,'''
      select_table_1_2 = ""
      for j, k in zip(year_statement_list, statement_list):
        select_table_1_2 += (''' sum( CASE WHEN Year_Statement_9LC = "{}" AND {} = "{}" AND Release_Artist_9LC <> "" 
                               THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(j,mySQL_column,k,j,k))
      select_table_1_3 = '''sum( CASE WHEN Statement_Period_Half_9LC >= "{}" AND Release_Artist_9LC <> ""
                       THEN Adjusted_Royalty_SB ELSE "" END) AS `Total`
                       FROM Master WHERE Statement_Period_Half_9LC <> "" AND Release_Artist_9LC <> "" AND Third_Party_9LC = "{}"
                       GROUP BY Release_Artist_9LC ORDER BY `Total` DESC'''.format(cut_off, s)
      select_table_1 = select_table_1_1 + select_table_1_2 + select_table_1_3
      mycursor.execute(select_table_1)
      table_1 = mycursor.fetchall()

      #Add line for unknown artists
      unknown_artists_1_1 = '''SELECT Third_Party_9LC, Release_Artist_9LC,'''
      unknown_artists_1_2 = ""
      for l,m in zip(year_statement_list, statement_list):
        unknown_artists_1_2 += ('''sum( CASE WHEN Year_Statement_9LC = "{}" AND {} = "{}" AND Release_Artist_9LC = ""
                                 THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(l,mySQL_column,m,l,m))
      unknown_artists_1_3 = '''sum( CASE WHEN Release_Artist_9LC = "" AND Statement_Period_Half_9LC >= "{}"
                             THEN Adjusted_Royalty_SB ELSE "" END) AS `Total`
                             FROM Master WHERE Statement_Period_Half_9LC <> "" AND Release_Artist_9LC = "" AND Third_Party_9LC = "{}"
                             GROUP BY Release_Artist_9LC ORDER BY `Total` DESC'''.format(cut_off, s)
      unknown_artists_1 = unknown_artists_1_1 + unknown_artists_1_2 + unknown_artists_1_3
      mycursor.execute(unknown_artists_1)
      unknown_artist_line_1 = mycursor.fetchall()
      final_table_1 = column_names_final_1 + table_1 + unknown_artist_line_1

      #Add tables to sheet
      total_row_no += len(final_table_1) + 1
      space_row_no = total_row_no + 1
      header_row_no = total_row_no - len(final_table_1)
      total_column_no = len(column_names_1) - 1
      third_party_table_rows = ws[header_row_no:space_row_no]
      for (row, l) in zip(third_party_table_rows, final_table_1):
        for (cell, m) in zip(row, range(len(l))):
          cell.value = l[m]
        for cell, m in zip(row, range(len(l)-1)):
          cell.style = 'number_style'
      for row in range(header_row_no, space_row_no+1):
        ws.cell(row=row, column=1).style = 'name_style'
        ws.cell(row=row, column=2).style = 'name_style'
        ws.cell(row=row, column=len(column_names_1)-2).style = 'total_style'
      ws.cell(row=total_row_no-1, column=2).value = 'Unknown Artists'
      for a in range(1, len(column_names_1)+1):
        ws.cell(row=header_row_no, column=a).style = 'title_style'

      #Add total row
      ws.cell(row=total_row_no, column=2).value = 'Total'
      ws.cell(row=total_row_no, column=2).style = 'title_style'
      for n in range(3, total_column_no):
        ws.cell(row=total_row_no, column=n).value = '=SUM({}{}:{}{})'.format(column_letters[n - 1], header_row_no + 1,
                                                                             column_letters[n - 1], total_row_no - 1)
        ws.cell(row=total_row_no, column=n).style = 'total_style'

      #Add % of revenue column
      total_column_letter = column_letters[len(column_names_1) - 3]
      for n in range(header_row_no+1, total_row_no):
        ws.cell(row=n, column=len(column_names_1) - 1).value = "=({}{}/{}{})".format(total_column_letter, n,
                                                                                   total_column_letter, total_row_no)
        ws.cell(row=n, column=len(column_names_1) - 1).style = 'Percent'

      #Add cumulative % column
      percent_rev_column_letter = column_letters[len(column_names_1) - 2]
      cumulative_rev_column_letter = column_letters[len(column_names_1) - 1]
      ws.cell(row=header_row_no+1, column=len(column_names_1)).value = "=({}{})".format(percent_rev_column_letter, header_row_no+1)
      ws.cell(row=header_row_no+1, column=len(column_names_1)).style = 'Percent'
      for o in range(header_row_no+2, total_row_no):
        ws.cell(row=o, column=len(column_names_1)).value = "=({}{}+{}{})".format(cumulative_rev_column_letter, o - 1,
                                                                               percent_rev_column_letter, o)
        ws.cell(row=o, column=len(column_names_1)).style = 'Percent'
      ws.cell(row=header_row_no, column=len(column_names_1)).value = 'Cumulative %'

      #Insert row between tables
      ws.insert_rows(total_row_no + 1)
      total_row_no += 1

    #Save workbook
    return wb.save(filename)
