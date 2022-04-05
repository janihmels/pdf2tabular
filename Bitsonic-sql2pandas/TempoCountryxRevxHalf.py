from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, NamedStyle
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, NamedStyle
from datetime import date
from pandas_utils.pandas_cursor import pandas_cursor


def tempocountryxrevxhalf(database, df, filename):
    mycursor = pandas_cursor(df=df)

    # Third party list
    find_third_parties = '''SELECT Third_Party_9LC FROM Master GROUP BY Third_Party_9LC'''
    mycursor.execute(find_third_parties)
    third_party_list = [i[0] for i in mycursor.fetchall()]
    third_party_string = ', '.join('"{}"'.format(str(x)) for x in third_party_list)

    # Create sheets
    wb = Workbook()
    ws = wb.active

    # Current year
    todays_date = date.today()
    current_year = todays_date.year

    # Find most recent year in data
    find_recent_year = '''SELECT Year_Statement_9LC FROM Master GROUP BY Year_Statement_9LC'''
    mycursor.execute(find_recent_year)
    recent_years = [i[0] for i in mycursor.fetchall()]
    recent_year = recent_years[-1]
    print(recent_year)
    print(current_year)
    if recent_year == current_year:
        find_cut_off_year = current_year - 1
    else:
        find_cut_off_year = int(recent_year) - 1

    # Find cut off
    find_period = 'SELECT DISTINCT Statement_Period_Half_9LC FROM Master WHERE Year_Statement_9LC <= "{}" ORDER BY Statement_Period_Half_9LC'.format(
        find_cut_off_year)
    mycursor.execute(find_period)
    statement_period_half_blank = [i[0] for i in mycursor.fetchall()]

    def check_blank(period):
        if period == '':
            return False
        else:
            return True

    remove_blank = filter(check_blank, statement_period_half_blank)
    statement_period_minus_blank = list(remove_blank)
    print(statement_period_minus_blank)
    if len(statement_period_minus_blank) > 10:
        statement_period_half = statement_period_minus_blank[-10:]
        standard_cut_off = statement_period_minus_blank[-10]
    else:
        statement_period_half = statement_period_minus_blank
        standard_cut_off = statement_period_minus_blank[0]
    print(statement_period_half)
    print(standard_cut_off)
    year_cut_off = standard_cut_off[0:4]
    print(year_cut_off)

    # Get list of statement year periods
    find_year_period = '''SELECT Year_Statement_9LC FROM Master 
                                        WHERE Year_Statement_9LC >= "{}" GROUP BY Year_Statement_9LC'''.format(
        year_cut_off)
    mycursor.execute(find_year_period)
    complete_list = [i for i in mycursor.fetchall()]
    statement_period_year_list = [i[0] for i in complete_list]
    year_statement_list_year = [i[0] for i in complete_list]
    yearly_statement_list = [i[0] for i in complete_list]
    year_list = list(dict.fromkeys(year_statement_list_year))

    # Get list of statement half periods
    find_half_period = '''SELECT Year_Statement_9LC, Half_Statement_9LC FROM Master 
                                     WHERE Half_Statement_9LC <> "" AND Statement_Period_Half_9LC >= "{}" GROUP BY Year_Statement_9LC, Half_Statement_9LC'''.format(
        standard_cut_off)
    mycursor.execute(find_half_period)
    complete_list = [i for i in mycursor.fetchall()]
    statement_period_half_list = [i[0] + ' ' + i[1] for i in complete_list]
    year_statement_list_half = [i[0] for i in complete_list]
    half_statement_list = [i[1] for i in complete_list]

    # Get list of statement quarter periods
    find_quarter_period = '''SELECT Year_Statement_9LC, Quarter_Statement_9LC FROM Master 
                                   WHERE Quarter_Statement_9LC <> "" AND Statement_Period_Half_9LC >= "{}" GROUP BY Year_Statement_9LC, Quarter_Statement_9LC'''.format(
        standard_cut_off)
    mycursor.execute(find_quarter_period)
    complete_list = [i for i in mycursor.fetchall()]
    statement_period_quarter_list = [i[0] + ' ' + i[1] for i in complete_list]
    year_statement_list_quarter = [i[0] for i in complete_list]
    quarter_statement_list = [i[1] for i in complete_list]

    # Get list of statement month periods
    find_month_period = '''SELECT Year_Statement_9LC, Month_Statement_9LC FROM Master 
                                         WHERE Month_Statement_9LC <> "" AND Statement_Period_Half_9LC >= "{}" GROUP BY Year_Statement_9LC, Month_Statement_9LC'''.format(
        standard_cut_off)
    mycursor.execute(find_month_period)
    complete_list = [i for i in mycursor.fetchall()]
    statement_period_month_list = [i[0] + ' ' + i[1] for i in complete_list]
    year_statement_list_month = [i[0] for i in complete_list]
    month_statement_list = [i[1] for i in complete_list]

    # Most recent year
    test_year = int(year_statement_list_half[-1]) - 1
    test_third_parties = []
    for b in third_party_list:
        find_complete_year = '''SELECT Statement_Period_Half_9LC FROM Master WHERE Year_Statement_9LC = "{}" GROUP BY Statement_Period_Half_9LC'''.format(
            test_year)
        mycursor.execute(find_complete_year)
        complete_year = [i[0] for i in mycursor.fetchall()]
        if len(complete_year) / 2 == 1:
            test_third_parties.append(b)
    if len(test_third_parties) != len(third_party_list):
        base_year_value = int(test_year) - 2
    else:
        base_year_value = test_year

    # Find number of territories
    find_territories = '''SELECT Region_DD, Normalized_Country_9LC
                  FROM Master GROUP BY Region_DD, Normalized_Country_9LC '''
    mycursor.execute(find_territories)
    territories = [i[0] for i in mycursor.fetchall()]
    territory_count = len(territories)

    # Column letters
    column_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S',
                      'T',
                      'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK',
                      'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ']

    # Create styles
    # Main numbers
    number_style = NamedStyle(name="number_style")
    number_style.alignment = Alignment(horizontal="right", vertical="center")
    number_style.font = Font(name="Calibri", size="11")
    number_style.number_format = '#,##0.00'

    # Title cells
    title_style = NamedStyle(name="title_style")
    title_style.alignment = Alignment(horizontal="center", vertical="center")
    title_style.font = Font(name="Calibri", size="11", bold=True)
    title_style.fill = PatternFill("solid", fgColor="A6ACAF")

    # Header cells
    header_style = NamedStyle(name="header_style")
    header_style.alignment = Alignment(horizontal="center", vertical="center")
    header_style.font = Font(name="Calibri", size="11", bold=True)
    header_style.fill = PatternFill("solid", fgColor="C9CDCF")

    # Sub header cells
    sub_header_style = NamedStyle(name="sub_header_style")
    sub_header_style.alignment = Alignment(horizontal="left", vertical="center")
    sub_header_style.font = Font(name="Calibri", size="11", bold=True, italic=True)

    # Name cells
    name_style = NamedStyle(name="name_style")
    name_style.alignment = Alignment(horizontal="left", vertical="center")
    name_style.font = Font(name="Calibri", size="11")

    # Total cells
    total_style = NamedStyle(name="total_style")
    total_style.alignment = Alignment(horizontal="right", vertical="center")
    total_style.font = Font(name="Calibri", size="11", bold=True)
    total_style.number_format = '#,##0.00'

    # Lined total row
    thin = Side(border_style="thin", color="000000")
    lined_total_style = NamedStyle(name="lined_total_style")
    lined_total_style.alignment = Alignment(horizontal="right", vertical="center")
    lined_total_style.font = Font(name="Calibri", size="11", bold=True)
    lined_total_style.number_format = '#,##0.00'
    lined_total_style.border = Border(top=thin, left=None, right=None, bottom=thin)

    # Total label style
    total_label_style = NamedStyle(name="total_label_style")
    total_label_style.alignment = Alignment(horizontal="center", vertical="center")
    total_label_style.font = Font(name="Calibri", size="11", bold=True)
    total_label_style.fill = PatternFill("solid", fgColor="C9CDCF")
    total_label_style.border = Border(top=thin, left=None, right=None, bottom=thin)

    # Publisher/PRO label style
    thick = Side(border_style="medium", color="000000")
    publisher_label_style = NamedStyle(name="publisher_label_style")
    publisher_label_style.alignment = Alignment(horizontal="center", vertical="center")
    publisher_label_style.font = Font(name="Calibri", size="11", bold=True, italic=True)
    publisher_label_style.border = Border(bottom=thick)

    # Not available style
    not_available_style = NamedStyle(name="not_available_style")
    not_available_style.alignment = Alignment(horizontal="center", vertical="center")
    not_available_style.font = Font(name="Calibri", size="11")

    # Add styles to workbook
    wb.add_named_style(number_style)
    wb.add_named_style(header_style)
    wb.add_named_style(title_style)
    wb.add_named_style(name_style)
    wb.add_named_style(total_style)
    wb.add_named_style(sub_header_style)
    wb.add_named_style(lined_total_style)
    wb.add_named_style(total_label_style)
    wb.add_named_style(publisher_label_style)
    wb.add_named_style(not_available_style)

    # Build sheet
    monthly_third_parties = []
    quarterly_third_parties = []
    half_third_parties = []
    for a in third_party_list:
        find_statement_number_months = '''SELECT Statement_Period_9LC, Month_Statement_9LC FROM Master WHERE Third_Party_9LC = "{}" GROUP BY Statement_Period_9LC, Month_Statement_9LC'''.format(
            a)
        mycursor.execute(find_statement_number_months)
        full_list_months = mycursor.fetchall()
        statement_number_blanks = [i[0] for i in full_list_months]
        statement_number_months_blanks = [i[1] for i in full_list_months]
        statement_number = len(list(filter(check_blank, statement_number_blanks)))
        statement_number_months = len(list(filter(check_blank, statement_number_months_blanks)))
        find_months_test = '''SELECT Month_Statement_9LC FROM Master WHERE Month_Statement_9LC <> "" AND Third_Party_9LC = "{}" GROUP BY Month_Statement_9LC'''.format(
            base_year_value, a)
        mycursor.execute(find_months_test)
        months_test = [i[0] for i in mycursor.fetchall()]
        if statement_number == statement_number_months:
            monthly = True
        else:
            monthly = False

        find_statement_number_quarters = '''SELECT Statement_Period_9LC, Quarter_Statement_9LC FROM Master WHERE Third_Party_9LC = "{}" GROUP BY Statement_Period_9LC, Quarter_Statement_9LC'''.format(
            a)
        mycursor.execute(find_statement_number_quarters)
        full_list_quarters = mycursor.fetchall()
        statement_number_blanks = [i[0] for i in full_list_quarters]
        statement_number_quarters_blanks = [i[1] for i in full_list_quarters]
        statement_number = len(list(filter(check_blank, statement_number_blanks)))
        statement_number_quarters = len(list(filter(check_blank, statement_number_quarters_blanks)))
        print(statement_number)
        print(statement_number_quarters)
        find_quarters_test = '''SELECT Quarter_Statement_9LC FROM Master WHERE Quarter_Statement_9LC <> "" AND Year_Statement_9LC = "{}" AND Third_Party_9LC = "{}" GROUP BY Quarter_Statement_9LC'''.format(
            base_year_value, a)
        mycursor.execute(find_quarters_test)
        quarters_test = [i[0] for i in mycursor.fetchall()]
        print(quarters_test)
        if statement_number == statement_number_quarters:
            quarterly = True
        else:
            quarterly = False

        find_statement_number_halves = '''SELECT Statement_Period_9LC, Half_Statement_9LC FROM Master WHERE Third_Party_9LC = "{}" GROUP BY Statement_Period_9LC, Half_Statement_9LC'''.format(
            a)
        mycursor.execute(find_statement_number_halves)
        full_list_halves = mycursor.fetchall()
        statement_number_blanks = [i[0] for i in full_list_halves]
        statement_number_halves_blanks = [i[1] for i in full_list_halves]
        statement_number = len(list(filter(check_blank, statement_number_blanks)))
        statement_number_halves = len(list(filter(check_blank, statement_number_halves_blanks)))
        find_halves_test = '''SELECT Half_Statement_9LC FROM Master WHERE Half_Statement_9LC <> "" AND Year_Statement_9LC = "{}" AND Third_Party_9LC = "{}" GROUP BY Half_Statement_9LC'''.format(
            base_year_value, a)
        mycursor.execute(find_halves_test)
        halves_test = [i[0] for i in mycursor.fetchall()]
        if statement_number == statement_number_halves:
            half = True
        else:
            half = False

        if monthly == True:
            monthly_third_parties.append(a)
        if quarterly == True:
            quarterly_third_parties.append(a)
        if half == True:
            half_third_parties.append(a)

    monthly_data = False
    quarterly_data = False
    half_data = False

    if len(monthly_third_parties) == len(third_party_list):
        monthly_data = True
    if len(quarterly_third_parties) == len(third_party_list):
        quarterly_data = True
    if len(half_third_parties) == len(third_party_list):
        half_data = True

    print(monthly_third_parties)
    print(quarterly_third_parties)
    print(half_third_parties)

    # Smallest period criteria
    if monthly_data:
        smallest_period = 'monthly'
        period_name = 'Month'
        mySQL_column = 'Month_Statement_9LC'
        statement_list = month_statement_list
        year_statement_list = year_statement_list_month
        statement_period_list = statement_period_month_list
        cut_off_field = 'Statement_Period_Half_9LC'
        cut_off = standard_cut_off
    elif quarterly_data:
        smallest_period = 'quarterly'
        period_name = 'Quarter'
        mySQL_column = 'Quarter_Statement_9LC'
        statement_list = quarter_statement_list
        year_statement_list = year_statement_list_quarter
        statement_period_list = statement_period_quarter_list
        cut_off_field = 'Statement_Period_Half_9LC'
        cut_off = standard_cut_off
    elif half_data:
        smallest_period = 'half'
        period_name = 'Half'
        mySQL_column = 'Half_Statement_9LC'
        statement_list = half_statement_list
        year_statement_list = year_statement_list_half
        statement_period_list = statement_period_half_list
        cut_off_field = 'Statement_Period_Half_9LC'
        cut_off = year_cut_off
    else:
        smallest_period = 'Year'
        mySQL_column = 'Year_Statement_9LC'
        statement_list = yearly_statement_list
        year_statement_list = year_statement_list_year
        statement_period_list = statement_period_year_list
        cut_off_field = 'Year_Statement_9LC'
        cut_off = year_cut_off

    # Worksheet title
    ws.title = "Region, Country"

    # Build table
    select_table_1_1 = '''SELECT Region_DD,Normalized_Country_9LC,'''
    select_table_1_2 = ""
    for j, k in zip(year_statement_list, statement_list):
        select_table_1_2 += (''' sum( CASE WHEN Year_Statement_9LC = "{}" AND {} = "{}" 
                           THEN Adjusted_Royalty_SB ELSE "" END) AS `{} {}`,'''.format(j, mySQL_column, k, j, k))
    select_table_1_3 = '''sum( CASE WHEN {} >= "{}"
                   THEN Adjusted_Royalty_SB ELSE "" END) AS `Total`
                   FROM Master
                   GROUP BY Region_DD, Normalized_Country_9LC ORDER BY `Total` DESC'''.format(cut_off_field, cut_off)
    select_table_1 = select_table_1_1 + select_table_1_2 + select_table_1_3
    mycursor.execute(select_table_1)
    final_table_1 = mycursor.fetchall()

    # Create list of column names
    column_names_1 = ['Region', 'Country']
    for k in statement_period_list:
        column_names_1.append(k)
    column_names_1.append('Total')
    column_names_1.append('% Of Revenue')
    column_names_1.append('Cumulative %')

    # Size of worksheet
    for column_no in range(1, len(column_names_1) + 1):
        for row_no in range(1, territory_count + 1):
            ws.cell(row=row_no, column=column_no)

    # Add column names to worksheet
    header_row = ws[1]
    for (header_cell, i) in zip(header_row, range(len(column_names_1))):
        header_cell.value = column_names_1[i]
    for cell in header_row:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="A6ACAF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add table to worksheet
    table_rows = ws[2:territory_count + 1]
    for (row, j) in zip(table_rows, final_table_1):
        for (cell, k) in zip(row, range(len(j))):
            cell.value = j[k]
            cell.style = 'Comma'
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.font = Font(name="Calibri", size="11")
    for song_no in range(2, territory_count + 2):
        ws.cell(row=song_no, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=song_no, column=2).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=song_no, column=(len(column_names_1) - 2)).font = Font(bold=True)

    # Add total row
    for (l, m) in zip(range(3, len(column_names_1) - 1), column_letters[3:]):
        ws.cell(row=territory_count + 2, column=l).value = "=SUM({}{}:{}{})".format(m, 2, m, territory_count + 1)
        ws.cell(row=territory_count + 2, column=l).style = 'Comma'
        ws.cell(row=territory_count + 2, column=l).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=territory_count + 2, column=l).font = Font(bold=True)
    ws.cell(row=territory_count + 2, column=1).value = 'Total'
    ws.cell(row=territory_count + 2, column=1).font = Font(bold='True')
    ws.cell(row=territory_count + 2, column=1).fill = PatternFill("solid", fgColor="A6ACAF")
    ws.cell(row=territory_count + 2, column=1).alignment = Alignment(horizontal="center", vertical="center")

    # Add % of revenue column
    total_column_letter = column_letters[len(column_names_1) - 3]
    for n in range(2, territory_count + 2):
        ws.cell(row=n, column=len(column_names_1) - 1).value = "=({}{}/{}{})".format(total_column_letter, n,
                                                                                     total_column_letter,
                                                                                     territory_count + 2)
        ws.cell(row=n, column=len(column_names_1) - 1).number_format = '0.00%'

    # Add cumulative % column
    percent_rev_column_letter = column_letters[len(column_names_1) - 2]
    cumulative_rev_column_letter = column_letters[len(column_names_1) - 1]
    ws.cell(row=2, column=len(column_names_1)).value = "=({}2)".format(percent_rev_column_letter)
    ws.cell(row=2, column=len(column_names_1)).number_format = '0.00%'
    for o in range(3, territory_count + 2):
        ws.cell(row=o, column=len(column_names_1)).value = "=({}{}+{}{})".format(cumulative_rev_column_letter, o - 1,
                                                                                 percent_rev_column_letter, o)
        ws.cell(row=o, column=len(column_names_1)).number_format = '0.00%'

    # Save workbook
    return wb.save(filename)















